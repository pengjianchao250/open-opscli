#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""通用 Amazon 广告诊断 Excel 生成脚本。

输入为 ops-dataset-query 导出的 JSON，输出为多工作表 .xlsx。
依赖：openpyxl
"""

from __future__ import annotations

import argparse
import json
import math
from copy import deepcopy
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]

DEFAULT_THRESHOLDS = {
    "highAcos": 0.30,
    "warnAcos": 0.22,
    "goodAcos": 0.12,
    "lowCtr": 0.003,
    "lowCvr": 0.01,
    "goodCvr": 0.02,
    "highAdShare": 0.12,
    "weakMargin": 0.05,
    "rowSpend": {
        "channel": 100000,
        "adType": 50000,
        "category": 50000,
        "asin": 10000,
        "campaign": 10000,
    },
    "minClicks": {
        "channel": 5000,
        "adType": 3000,
        "category": 2000,
        "asin": 1000,
        "campaign": 1000,
    },
    "minImpressions": {
        "category": 300000,
        "asin": 300000,
        "campaign": 300000,
    },
}

THRESHOLDS = deepcopy(DEFAULT_THRESHOLDS)

DIMENSIONS = {
    "largeTeam": ["platform", "large_team"],
    "team": ["platform", "large_team", "team"],
    "channel": ["platform", "channel"],
    "adType": ["platform", "ad_type"],
    "category": ["platform", "amazon_cat", "category"],
    "asin": ["platform", "channel", "asin", "product", "amazon_cat", "category"],
    "campaign": ["platform", "channel", "ad_type", "campaign"],
}

AD_METRICS = [
    "ad_spend",
    "ad_sales",
    "ad_orders",
    "clicks",
    "impressions",
    "add_to_cart",
    "detail_page_views",
    "new_to_brand_sales",
]
PROFIT_METRICS = [
    "total_sales",
    "total_orders",
    "gross_profit",
    "profit_ad_spend",
    "sessions",
    "page_views",
]
COMPARE_METRICS = [
    "ad_spend",
    "ad_sales",
    "ad_orders",
    "impressions",
    "clicks",
    "ctr",
    "cpc",
    "cvr",
    "acos",
    "roas",
    "total_sales",
    "total_orders",
    "gross_profit",
    "gross_margin",
    "ad_share_total",
]
RATE_METRICS = {"ctr", "cvr", "acos", "gross_margin", "ad_share_total"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成 Amazon 广告多维诊断 Excel。")
    parser.add_argument("--ad-source", default=ROOT / "data" / "ad_raw_detail.json")
    parser.add_argument("--profit-source", default=ROOT / "data" / "profit_raw_detail.json")
    parser.add_argument("--campaign-source", default=ROOT / "data" / "campaign_raw_detail.json")
    parser.add_argument("--compare-ad-source")
    parser.add_argument("--compare-profit-source")
    parser.add_argument("--compare-campaign-source")
    parser.add_argument("--threshold-config", default=ROOT / "references" / "thresholds.default.json")
    parser.add_argument("--output", default=ROOT / "outputs" / "Amazon广告多维诊断.xlsx")
    parser.add_argument("--title", default="Amazon 广告多维诊断")
    parser.add_argument("--period", default="当前分析周期")
    parser.add_argument("--compare-period", default="对比周期")
    return parser.parse_args()


def to_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        n = float(value)
    except (TypeError, ValueError):
        return 0.0
    return n if math.isfinite(n) else 0.0


def safe_div(num: Any, den: Any) -> float:
    den_n = to_number(den)
    return to_number(num) / den_n if den_n else 0.0


def pct(value: Any) -> str:
    return f"{to_number(value) * 100:.2f}%"


def money(value: Any) -> str:
    return f"{round(to_number(value)):,.0f}"


def pct_change(current: Any, last: Any) -> float | None:
    current_n = to_number(current)
    last_n = to_number(last)
    if not last_n:
        return None if current_n else 0.0
    return (current_n - last_n) / abs(last_n)


def text(value: Any) -> str:
    if value in (None, ""):
        return "未归类"
    value = str(value).strip()
    return value or "未归类"


def merge_deep(base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
    result = deepcopy(base)
    for key, value in (override or {}).items():
        if isinstance(value, dict) and isinstance(result.get(key), dict):
            result[key] = merge_deep(result[key], value)
        elif value is not None:
            result[key] = value
    return result


def load_thresholds(config_path: str | Path | None) -> None:
    global THRESHOLDS
    THRESHOLDS = deepcopy(DEFAULT_THRESHOLDS)
    if not config_path:
        return
    path = Path(config_path)
    if not path.exists():
        return
    config = json.loads(path.read_text(encoding="utf-8-sig"))
    THRESHOLDS = merge_deep(THRESHOLDS, config.get("thresholds", config))


def load_rows(file_path: str | Path | None) -> List[Dict[str, Any]]:
    if not file_path:
        return []
    raw = json.loads(Path(file_path).read_text(encoding="utf-8-sig"))
    if isinstance(raw, list):
        return raw
    result_rows = raw.get("data", {}).get("result", {}).get("data")
    if isinstance(result_rows, list):
        return result_rows
    merged_rows = raw.get("merged", {}).get("rows")
    if isinstance(merged_rows, list):
        return merged_rows
    raise ValueError(f"无法从 {file_path} 读取数据行")


def key_of(row: Dict[str, Any], dims: Sequence[str]) -> str:
    return "\x1f".join(text(row.get(dim)) for dim in dims)


def aggregate(rows: Iterable[Dict[str, Any]], dims: Sequence[str], metrics: Sequence[str]) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        key = key_of(row, dims)
        if key not in grouped:
            grouped[key] = {dim: text(row.get(dim)) for dim in dims}
            for metric in metrics:
                grouped[key][metric] = 0.0
        target = grouped[key]
        for metric in metrics:
            target[metric] += to_number(row.get(metric))
    return list(grouped.values())


def derive_ad(row: Dict[str, Any]) -> Dict[str, Any]:
    row["ctr"] = safe_div(row.get("clicks"), row.get("impressions"))
    row["cpc"] = safe_div(row.get("ad_spend"), row.get("clicks"))
    row["cvr"] = safe_div(row.get("ad_orders"), row.get("clicks"))
    row["cpa"] = safe_div(row.get("ad_spend"), row.get("ad_orders"))
    row["acos"] = safe_div(row.get("ad_spend"), row.get("ad_sales"))
    row["roas"] = safe_div(row.get("ad_sales"), row.get("ad_spend"))
    return row


def derive_profit(row: Dict[str, Any]) -> Dict[str, Any]:
    row["gross_margin"] = safe_div(row.get("gross_profit"), row.get("total_sales"))
    row["ad_share_total"] = safe_div(row.get("profit_ad_spend") or row.get("ad_spend"), row.get("total_sales"))
    row["site_cvr"] = safe_div(row.get("total_orders"), row.get("sessions"))
    return row


def merge_profit(ad_agg: List[Dict[str, Any]], profit_agg: List[Dict[str, Any]], dims: Sequence[str]) -> List[Dict[str, Any]]:
    profit_map = {key_of(row, dims): row for row in profit_agg}
    merged = []
    for row in ad_agg:
        profit = profit_map.get(key_of(row, dims), {})
        next_row = {
            **row,
            "total_sales": to_number(profit.get("total_sales")),
            "total_orders": to_number(profit.get("total_orders")),
            "gross_profit": to_number(profit.get("gross_profit")),
            "profit_ad_spend": to_number(profit.get("profit_ad_spend")),
            "sessions": to_number(profit.get("sessions")),
            "page_views": to_number(profit.get("page_views")),
        }
        merged.append(derive_profit(derive_ad(next_row)))
    return merged


def attach_comparison(rows: List[Dict[str, Any]], compare_rows: List[Dict[str, Any]], dims: Sequence[str]) -> List[Dict[str, Any]]:
    compare_map = {key_of(row, dims): row for row in compare_rows}
    output = []
    for row in rows:
        last = compare_map.get(key_of(row, dims), {})
        next_row = dict(row)
        for metric in COMPARE_METRICS:
            if metric not in row and metric not in last:
                continue
            current_value = to_number(row.get(metric))
            last_value = to_number(last.get(metric))
            next_row[f"last_{metric}"] = last_value
            next_row[f"{metric}_mom"] = current_value - last_value if metric in RATE_METRICS else pct_change(current_value, last_value)
        output.append(next_row)
    return output


def analyze_row(row: Dict[str, Any], level: str) -> Dict[str, str]:
    spend = to_number(row.get("ad_spend"))
    clicks = to_number(row.get("clicks"))
    impressions = to_number(row.get("impressions"))
    has_profit = "gross_profit" in row and to_number(row.get("total_sales")) > 0
    problems: List[str] = []
    basis: List[str] = []
    priority = "短期观察"
    action = "保留观察，继续累计样本后再调整。"
    decision = "观察"
    problem_type = "样本或结构观察"
    is_low_sample = spend < THRESHOLDS["rowSpend"].get(level, 0)

    if is_low_sample:
        problems.append("样本量较小")
        basis.append(f"广告费 {money(spend)}，低于本层级判断阈值")
        priority = "暂不建议调整"
        problem_type = "样本不足"
        action = "保留观察，不做激进调价或停投。"

    if has_profit and to_number(row.get("gross_margin")) < 0:
        problems.append("经营亏损")
        basis.append(f"毛利率 {pct(row.get('gross_margin'))}，毛利 {money(row.get('gross_profit'))}")
        priority = "立即执行"
        decision = "收缩"
        problem_type = "Listing/产品竞争力问题"
        action = "优先收缩预算或暂停低效活动；同步检查售价、优惠、评论、配送和成本结构。"
    elif has_profit and to_number(row.get("gross_margin")) < THRESHOLDS["weakMargin"] and to_number(row.get("ad_share_total")) > 0.08:
        problems.append("毛利承接弱")
        basis.append(f"毛利率 {pct(row.get('gross_margin'))}，广告费占总销售 {pct(row.get('ad_share_total'))}")
        priority = "立即执行"
        decision = "控量"
        problem_type = "预算分配问题"
        action = "下调高消耗广告，预算转向高 ROAS 活动或高毛利 ASIN。"

    if to_number(row.get("acos")) > THRESHOLDS["highAcos"]:
        problems.append("ACOS 高")
        basis.append(f"ACOS {pct(row.get('acos'))}，ROAS {to_number(row.get('roas')):.2f}")
        priority = "立即执行" if priority != "立即执行" else priority
        decision = "收缩" if decision == "收缩" else "控量"
        if problem_type == "样本或结构观察":
            problem_type = "预算分配问题"
        if action == "保留观察，继续累计样本后再调整。":
            action = "降低出价/预算，排查高花费低转化流量。"
    elif to_number(row.get("acos")) > THRESHOLDS["warnAcos"]:
        problems.append("投产偏弱")
        basis.append(f"ACOS {pct(row.get('acos'))}")
        decision = "控量" if decision == "观察" else decision

    if clicks >= THRESHOLDS["minClicks"].get(level, 0) and to_number(row.get("cvr")) < THRESHOLDS["lowCvr"]:
        problems.append("点击后转化弱")
        basis.append(f"点击 {money(clicks)}，CVR {pct(row.get('cvr'))}")
        priority = "立即执行"
        decision = "收缩" if decision == "收缩" else "控量"
        problem_type = "流量质量或Listing承接问题"
        action = "降低泛流量出价，暂停低转化投放；同步检查主图、价格、评价、优惠和库存配送。"

    min_impressions = THRESHOLDS["minImpressions"].get(level, 500000)
    if impressions >= min_impressions and to_number(row.get("ctr")) < THRESHOLDS["lowCtr"]:
        problems.append("高曝光低点击")
        basis.append(f"曝光 {money(impressions)}，CTR {pct(row.get('ctr'))}")
        problem_type = "点击吸引力问题"
        if priority != "立即执行":
            priority = "短期观察"
        if decision == "观察":
            decision = "优化承接"
        action = "优化主图、价格和标题卖点；若相关性弱则降低对应投放出价。"

    if to_number(row.get("last_ad_spend")) > 0 and to_number(row.get("ad_spend_mom")) > 0.2 and to_number(row.get("acos_mom")) > 0.03:
        problems.append("广告费扩张但投产走弱")
        basis.append(f"广告费环比 {pct(row.get('ad_spend_mom'))}，ACOS 变化 {pct(row.get('acos_mom'))}")
        priority = "立即执行"
        decision = "收缩" if decision == "收缩" else "控量"
        problem_type = "预算分配问题"
        action = "暂停继续加预算，优先拆分检查新增花费来源；低转化活动降预算，高转化活动单独保留。"
    elif to_number(row.get("last_ad_sales")) > 0 and to_number(row.get("ad_sales_mom")) < -0.2 and spend >= THRESHOLDS["rowSpend"].get(level, 0):
        problems.append("销售额环比下滑")
        basis.append(f"广告销售额环比 {pct(row.get('ad_sales_mom'))}")
        if priority != "立即执行":
            priority = "短期观察"
        if problem_type == "样本或结构观察":
            problem_type = "有效流量不足或转化承接问题"
        if action == "保留观察，继续累计样本后再调整。":
            action = "核查核心词排名、预算断档和 Listing 转化变化；优先恢复历史高转化流量。"

    only_low_sample = all(problem == "样本量较小" for problem in problems)
    if (
        to_number(row.get("acos")) > 0
        and to_number(row.get("acos")) <= THRESHOLDS["goodAcos"]
        and to_number(row.get("cvr")) >= THRESHOLDS["goodCvr"]
        and (not has_profit or to_number(row.get("gross_margin")) > THRESHOLDS["weakMargin"])
        and only_low_sample
    ):
        problems = ["低样本高效，具备观察放量价值" if is_low_sample else "高效可放量"]
        basis = [f"ACOS {pct(row.get('acos'))}，CVR {pct(row.get('cvr'))}，ROAS {to_number(row.get('roas')):.2f}"]
        priority = "短期观察"
        decision = "放量"
        problem_type = "有效流量不足"
        action = "先转入独立观察或精准收割池，小幅提高预算/出价，确认 7 天转化稳定后再放大。" if is_low_sample else "逐步增加预算或复制到独立收割活动，放量幅度建议每次 10%-20%。"

    if len(problems) == 1 and problems[0] == "样本量较小":
        return {
            "priority": "暂不建议调整",
            "decision": "观察",
            "problem_type": "样本不足",
            "problem": "；".join(problems),
            "basis": "；".join(basis),
            "action": "保留观察，不做激进调价或停投。",
        }

    if not problems:
        problems.append("暂无明显异常")
        basis.append(f"ACOS {pct(row.get('acos'))}，CVR {pct(row.get('cvr'))}，CTR {pct(row.get('ctr'))}")
        action = "保持预算，持续监控 7 天趋势。"

    return {
        "priority": priority,
        "decision": decision,
        "problem_type": problem_type,
        "problem": "；".join(problems),
        "basis": "；".join(basis),
        "action": action,
    }


def enrich_rows(rows: List[Dict[str, Any]], level: str) -> List[Dict[str, Any]]:
    return [{**row, **analyze_row(row, level)} for row in rows]


def sort_by_spend(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return sorted(rows, key=lambda row: to_number(row.get("ad_spend")), reverse=True)


def summarize(rows: Iterable[Dict[str, Any]]) -> Dict[str, Any]:
    total: Dict[str, Any] = {}
    for row in rows:
        for key in ["ad_spend", "ad_sales", "ad_orders", "clicks", "impressions", "total_sales", "total_orders", "gross_profit", "sessions"]:
            total[key] = total.get(key, 0.0) + to_number(row.get(key))
    derive_profit(derive_ad(total))
    return total


def normalize_ad_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {
            "platform": text(r.get("platform")),
            "channel": text(r.get("channel")),
            "large_team": text(r.get("large_team")),
            "team": text(r.get("team")),
            "ad_type": text(r.get("ad_type")),
            "asin": text(r.get("asin")),
            "product": text(r.get("product")),
            "amazon_cat": text(r.get("amazon_cat")),
            "category": text(r.get("category")),
            "ad_spend": to_number(r.get("ad_spend")),
            "ad_sales": to_number(r.get("ad_sales")),
            "ad_orders": to_number(r.get("ad_orders")),
            "clicks": to_number(r.get("clicks")),
            "impressions": to_number(r.get("impressions")),
            "add_to_cart": to_number(r.get("add_to_cart")),
            "detail_page_views": to_number(r.get("detail_page_views")),
            "new_to_brand_sales": to_number(r.get("new_to_brand_sales")),
        }
        for r in rows
    ]


def normalize_profit_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {
            "platform": text(r.get("platform")),
            "channel": text(r.get("channel")),
            "large_team": text(r.get("large_team")),
            "team": text(r.get("team")),
            "asin": text(r.get("asin")),
            "product": text(r.get("product")),
            "amazon_cat": text(r.get("amazon_cat")),
            "category": text(r.get("category")),
            "total_sales": to_number(r.get("total_sales")),
            "total_orders": to_number(r.get("total_orders")),
            "gross_profit": to_number(r.get("gross_profit")),
            "profit_ad_spend": to_number(r.get("ad_spend")),
            "sessions": to_number(r.get("sessions")),
            "page_views": to_number(r.get("page_views")),
        }
        for r in rows
    ]


def normalize_campaign_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {
            "platform": text(r.get("platform")),
            "channel": text(r.get("channel")),
            "ad_type": text(r.get("ad_type")),
            "campaign": text(r.get("campaign")),
            "ad_spend": to_number(r.get("ad_spend")),
            "ad_sales": to_number(r.get("ad_sales")),
            "ad_orders": to_number(r.get("ad_orders")),
            "clicks": to_number(r.get("clicks")),
            "impressions": to_number(r.get("impressions")),
            "add_to_cart": to_number(r.get("add_to_cart")),
            "detail_page_views": to_number(r.get("detail_page_views")),
            "new_to_brand_sales": to_number(r.get("new_to_brand_sales")),
        }
        for r in rows
    ]


def build_profit_merged_level(ad_rows, profit_rows, compare_ad_rows, compare_profit_rows, dims, level):
    current = merge_profit(aggregate(ad_rows, dims, AD_METRICS), aggregate(profit_rows, dims, PROFIT_METRICS), dims)
    compare = merge_profit(aggregate(compare_ad_rows, dims, AD_METRICS), aggregate(compare_profit_rows, dims, PROFIT_METRICS), dims)
    return sort_by_spend(enrich_rows(attach_comparison(current, compare, dims), level))


def build_ad_only_level(ad_rows, compare_ad_rows, dims, level):
    current = [derive_ad(row) for row in aggregate(ad_rows, dims, AD_METRICS)]
    compare = [derive_ad(row) for row in aggregate(compare_ad_rows, dims, AD_METRICS)]
    return sort_by_spend(enrich_rows(attach_comparison(current, compare, dims), level))


def build_analysis(ad_rows, profit_rows, campaign_rows, compare_ad_rows, compare_profit_rows, compare_campaign_rows):
    return {
        "largeTeam": build_profit_merged_level(ad_rows, profit_rows, compare_ad_rows, compare_profit_rows, DIMENSIONS["largeTeam"], "channel"),
        "team": build_profit_merged_level(ad_rows, profit_rows, compare_ad_rows, compare_profit_rows, DIMENSIONS["team"], "channel"),
        "channel": build_profit_merged_level(ad_rows, profit_rows, compare_ad_rows, compare_profit_rows, DIMENSIONS["channel"], "channel"),
        "adType": build_ad_only_level(ad_rows, compare_ad_rows, DIMENSIONS["adType"], "adType"),
        "category": build_profit_merged_level(ad_rows, profit_rows, compare_ad_rows, compare_profit_rows, DIMENSIONS["category"], "category"),
        "asin": build_profit_merged_level(ad_rows, profit_rows, compare_ad_rows, compare_profit_rows, DIMENSIONS["asin"], "asin"),
        "campaign": build_ad_only_level(campaign_rows, compare_campaign_rows, DIMENSIONS["campaign"], "campaign"),
    }


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUMMARY_FILL = PatternFill("solid", fgColor="D9EAF7")
RED_FILL = PatternFill("solid", fgColor="FECACA")
YELLOW_FILL = PatternFill("solid", fgColor="FEF3C7")
GRAY_FILL = PatternFill("solid", fgColor="E5E7EB")


def cell_value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def write_sheet(wb: Workbook, name: str, headers: List[Dict[str, str]], rows: List[Dict[str, Any]], color_scale_keys: Sequence[str] = ()) -> None:
    ws = wb.create_sheet(name)
    ws.append([header["label"] for header in headers])
    for row in rows:
        ws.append([cell_value(row.get(header["key"])) for header in headers])

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row != 1:
                cell.font = Font(color="111827")

    for idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(idx)
        typ = header.get("type")
        if typ in ("money", "int"):
            number_format = "#,##0"
        elif typ == "pct":
            number_format = "0.00%"
        elif typ == "decimal":
            number_format = "0.00"
        else:
            number_format = None
        if number_format:
            for cell in ws[col_letter][1:]:
                cell.number_format = number_format

        width = max(len(str(header["label"])) + 2, 10)
        sample_values = [str(ws.cell(row=r, column=idx).value or "") for r in range(2, min(ws.max_row, 80) + 1)]
        if sample_values:
            width = max(width, min(max(len(v) for v in sample_values) + 2, 60))
        ws.column_dimensions[col_letter].width = width

    if rows:
        priority_col = next((i + 1 for i, h in enumerate(headers) if h["key"] == "priority"), None)
        if priority_col:
            col = get_column_letter(priority_col)
            ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", FormulaRule(formula=[f'ISNUMBER(SEARCH("立即执行",{col}2))'], fill=RED_FILL))
            ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", FormulaRule(formula=[f'ISNUMBER(SEARCH("短期观察",{col}2))'], fill=YELLOW_FILL))
            ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", FormulaRule(formula=[f'ISNUMBER(SEARCH("暂不建议调整",{col}2))'], fill=GRAY_FILL))

        problem_col = next((i + 1 for i, h in enumerate(headers) if h["key"] == "problem"), None)
        if problem_col:
            col = get_column_letter(problem_col)
            ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", FormulaRule(formula=[f'ISNUMBER(SEARCH("亏损",{col}2))'], fill=PatternFill("solid", fgColor="FEE2E2")))

        for key in color_scale_keys:
            idx = next((i + 1 for i, h in enumerate(headers) if h["key"] == key), None)
            if idx:
                col = get_column_letter(idx)
                ws.conditional_formatting.add(
                    f"{col}2:{col}{ws.max_row}",
                    ColorScaleRule(start_type="min", start_color="DCFCE7", mid_type="percentile", mid_value=50, mid_color="FEF3C7", end_type="max", end_color="FECACA"),
                )


COMMON_HEADERS = [
    {"key": "platform", "label": "平台"},
    {"key": "channel", "label": "渠道"},
    {"key": "ad_spend", "label": "广告费", "type": "money"},
    {"key": "last_ad_spend", "label": "对比期广告费", "type": "money"},
    {"key": "ad_spend_mom", "label": "广告费环比", "type": "pct"},
    {"key": "ad_sales", "label": "广告销售额", "type": "money"},
    {"key": "last_ad_sales", "label": "对比期广告销售额", "type": "money"},
    {"key": "ad_sales_mom", "label": "广告销售额环比", "type": "pct"},
    {"key": "ad_orders", "label": "广告订单", "type": "int"},
    {"key": "impressions", "label": "曝光", "type": "int"},
    {"key": "clicks", "label": "点击", "type": "int"},
    {"key": "ctr", "label": "CTR", "type": "pct"},
    {"key": "ctr_mom", "label": "CTR变化", "type": "pct"},
    {"key": "cpc", "label": "CPC", "type": "decimal"},
    {"key": "cvr", "label": "CVR", "type": "pct"},
    {"key": "cvr_mom", "label": "CVR变化", "type": "pct"},
    {"key": "acos", "label": "ACOS", "type": "pct"},
    {"key": "acos_mom", "label": "ACOS变化", "type": "pct"},
    {"key": "roas", "label": "ROAS", "type": "decimal"},
    {"key": "total_sales", "label": "总销售额", "type": "money"},
    {"key": "last_total_sales", "label": "对比期总销售额", "type": "money"},
    {"key": "total_sales_mom", "label": "总销售额环比", "type": "pct"},
    {"key": "gross_profit", "label": "毛利", "type": "money"},
    {"key": "last_gross_profit", "label": "对比期毛利", "type": "money"},
    {"key": "gross_profit_mom", "label": "毛利环比", "type": "pct"},
    {"key": "gross_margin", "label": "毛利率", "type": "pct"},
    {"key": "gross_margin_mom", "label": "毛利率变化", "type": "pct"},
    {"key": "ad_share_total", "label": "广告费占总销售", "type": "pct"},
    {"key": "ad_share_total_mom", "label": "广告费占比变化", "type": "pct"},
    {"key": "priority", "label": "优先级"},
    {"key": "decision", "label": "判断"},
    {"key": "problem_type", "label": "问题归因"},
    {"key": "problem", "label": "提取问题"},
    {"key": "basis", "label": "判断依据"},
    {"key": "action", "label": "动作建议"},
]

AD_ONLY_TAIL = [
    {"key": "last_ad_spend", "label": "对比期广告费", "type": "money"},
    {"key": "ad_spend_mom", "label": "广告费环比", "type": "pct"},
    {"key": "last_ad_sales", "label": "对比期广告销售额", "type": "money"},
    {"key": "ad_sales_mom", "label": "广告销售额环比", "type": "pct"},
    {"key": "ctr_mom", "label": "CTR变化", "type": "pct"},
    {"key": "cvr_mom", "label": "CVR变化", "type": "pct"},
    {"key": "acos_mom", "label": "ACOS变化", "type": "pct"},
    {"key": "priority", "label": "优先级"},
    {"key": "decision", "label": "判断"},
    {"key": "problem_type", "label": "问题归因"},
    {"key": "problem", "label": "提取问题"},
    {"key": "basis", "label": "判断依据"},
    {"key": "action", "label": "动作建议"},
]


def dashboard_metric(name: str, current: Any, last: Any, typ: str, note: str) -> List[Any]:
    change = to_number(current) - to_number(last) if typ == "rate" else pct_change(current, last)
    return [name, current, last, change, note]


def infer_core_conclusion(total_ad: Dict[str, Any], total_profit: Dict[str, Any], compare_ad: Dict[str, Any], compare_profit: Dict[str, Any]) -> str:
    ad_sales_mom = pct_change(total_ad.get("ad_sales"), compare_ad.get("ad_sales"))
    gross_profit_mom = pct_change(total_profit.get("gross_profit"), compare_profit.get("gross_profit"))
    if to_number(total_profit.get("gross_margin")) < 0:
        return "当前优化目标应优先提升利润，先收缩亏损 ASIN/活动，再保留可验证的有效流量。"
    if to_number(total_ad.get("acos")) > THRESHOLDS["highAcos"] or to_number(total_profit.get("ad_share_total")) > THRESHOLDS["highAdShare"]:
        return "当前优化目标更偏向降低广告费占比与提升投产比，预算应从高 ACOS 流量转向高毛利转化对象。"
    if ad_sales_mom is not None and ad_sales_mom < -0.15 and to_number(total_ad.get("acos")) <= THRESHOLDS["warnAcos"]:
        return "当前优化目标更偏向扩大有效流量，重点恢复对比期有效曝光和高转化活动。"
    if gross_profit_mom is not None and gross_profit_mom < -0.15:
        return "当前优化目标更偏向利润修复，需要同步处理广告效率与 Listing/价格承接。"
    return "当前广告整体投产可控，核心目标可放在稳定利润基础上的有效流量放大。"


def create_dashboard(wb: Workbook, analysis: Dict[str, Any], args: argparse.Namespace, ad_rows, profit_rows, compare_ad_rows, compare_profit_rows) -> None:
    ws = wb.create_sheet("00_总览")
    total_ad = summarize(ad_rows)
    total_profit = summarize({**r, "ad_spend": r.get("profit_ad_spend")} for r in profit_rows)
    compare_ad = summarize(compare_ad_rows)
    compare_profit = summarize({**r, "ad_spend": r.get("profit_ad_spend")} for r in compare_profit_rows)
    all_rows = analysis["largeTeam"] + analysis["team"] + analysis["channel"] + analysis["category"] + analysis["asin"] + analysis["campaign"]
    immediate = sum(1 for r in all_rows if r.get("priority") == "立即执行")
    scale = sum(1 for r in all_rows if r.get("decision") == "放量")
    rows = [
        [args.title, "", "", "", ""],
        ["分析周期", args.period, "对比周期", args.compare_period, ""],
        ["整体判断", infer_core_conclusion(total_ad, total_profit, compare_ad, compare_profit), "", "", ""],
        ["", "", "", "", ""],
        ["指标", "当前期", "对比期", "环比/变化", "说明"],
        dashboard_metric("广告费", total_ad.get("ad_spend"), compare_ad.get("ad_spend"), "amount", "广告底稿汇总"),
        dashboard_metric("广告销售额", total_ad.get("ad_sales"), compare_ad.get("ad_sales"), "amount", "广告归因销售"),
        dashboard_metric("ACOS", total_ad.get("acos"), compare_ad.get("acos"), "rate", "广告费/广告销售额"),
        dashboard_metric("ROAS", total_ad.get("roas"), compare_ad.get("roas"), "rate", "广告销售额/广告费"),
        dashboard_metric("CTR", total_ad.get("ctr"), compare_ad.get("ctr"), "rate", "点击/曝光"),
        dashboard_metric("CVR", total_ad.get("cvr"), compare_ad.get("cvr"), "rate", "广告订单/点击"),
        dashboard_metric("总销售额", total_profit.get("total_sales"), compare_profit.get("total_sales"), "amount", "经营底稿汇总"),
        dashboard_metric("毛利", total_profit.get("gross_profit"), compare_profit.get("gross_profit"), "amount", "经营底稿汇总"),
        dashboard_metric("毛利率", total_profit.get("gross_margin"), compare_profit.get("gross_margin"), "rate", "毛利/总销售额"),
        dashboard_metric("广告费占总销售", total_profit.get("ad_share_total"), compare_profit.get("ad_share_total"), "rate", "经营广告费/总销售额"),
        ["立即执行项", immediate, "", "", "所有维度规则触发数"],
        ["可放量项", scale, "", "", "所有维度规则触发数"],
    ]
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True, size=16)
    for cell in ws[5]:
        cell.fill = SUMMARY_FILL
        cell.font = Font(bold=True)

    for row in range(6, 18):
        ws.cell(row=row, column=2).number_format = "#,##0.00"
        ws.cell(row=row, column=3).number_format = "#,##0.00"
        ws.cell(row=row, column=4).number_format = "#,##0.00"
    for row in [8, 10, 11, 14, 15]:
        for col in [2, 3, 4]:
            ws.cell(row=row, column=col).number_format = "0.00%"

    platform_data: Dict[str, Dict[str, float]] = {}
    for row in analysis["channel"]:
        key = text(row.get("platform"))
        platform_data.setdefault(key, {"spend": 0, "lastSpend": 0, "sales": 0, "lastSales": 0})
        platform_data[key]["spend"] += to_number(row.get("ad_spend"))
        platform_data[key]["lastSpend"] += to_number(row.get("last_ad_spend"))
        platform_data[key]["sales"] += to_number(row.get("ad_sales"))
        platform_data[key]["lastSales"] += to_number(row.get("last_ad_sales"))

    chart_rows = [["平台", "当前广告费", "对比期广告费", "当前广告销售额", "对比期广告销售额"]]
    chart_rows.extend([key, v["spend"], v["lastSpend"], v["sales"], v["lastSales"]] for key, v in platform_data.items())
    for r_idx, row in enumerate(chart_rows, start=2):
        for c_idx, value in enumerate(row, start=8):
            ws.cell(row=r_idx, column=c_idx, value=value)
    for cell in ws[2][7:12]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    if len(chart_rows) > 1:
        chart = BarChart()
        chart.title = "平台广告费与广告销售额对比"
        chart.y_axis.title = "金额"
        chart.x_axis.title = "平台"
        data = Reference(ws, min_col=9, max_col=12, min_row=2, max_row=1 + len(chart_rows))
        cats = Reference(ws, min_col=8, min_row=3, max_row=1 + len(chart_rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws.add_chart(chart, "K2")

    action_rows = [["优先级", "对象", "问题", "建议"]]
    action_rows.extend(
        [r.get("priority"), f"{r.get('platform')}/{r.get('channel')}/{r.get('asin')}", r.get("problem"), r.get("action")]
        for r in analysis["asin"]
        if r.get("priority") == "立即执行"
    )
    for r_idx, row in enumerate(action_rows[:9], start=21):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    for cell in ws[21]:
        cell.fill = PatternFill("solid", fgColor="991B1B")
        cell.font = HEADER_FONT

    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 18 if col < 5 else 16
    ws.column_dimensions["B"].width = 60
    ws.freeze_panes = "A5"


def build_workbook(args, analysis, ad_rows, profit_rows, campaign_rows, compare_ad_rows, compare_profit_rows):
    wb = Workbook()
    wb.remove(wb.active)
    create_dashboard(wb, analysis, args, ad_rows, profit_rows, compare_ad_rows, compare_profit_rows)

    total_ad = summarize(ad_rows)
    total_profit_for_ad = summarize({**r, "ad_spend": r.get("profit_ad_spend")} for r in profit_rows)
    compare_ad = summarize(compare_ad_rows)
    compare_profit_for_ad = summarize({**r, "ad_spend": r.get("profit_ad_spend")} for r in compare_profit_rows)

    write_sheet(
        wb,
        "01_口径校验",
        [
            {"key": "item", "label": "校验项"},
            {"key": "ad_value", "label": "当前广告底稿", "type": "money"},
            {"key": "profit_value", "label": "当前经营底稿", "type": "money"},
            {"key": "diff", "label": "当前差异", "type": "money"},
            {"key": "last_ad_value", "label": "对比期广告底稿", "type": "money"},
            {"key": "last_profit_value", "label": "对比期经营底稿", "type": "money"},
            {"key": "last_diff", "label": "对比期差异", "type": "money"},
            {"key": "note", "label": "说明"},
        ],
        [
            {
                "item": "广告费",
                "ad_value": total_ad.get("ad_spend"),
                "profit_value": total_profit_for_ad.get("ad_spend"),
                "diff": to_number(total_ad.get("ad_spend")) - to_number(total_profit_for_ad.get("ad_spend")),
                "last_ad_value": compare_ad.get("ad_spend"),
                "last_profit_value": compare_profit_for_ad.get("ad_spend"),
                "last_diff": to_number(compare_ad.get("ad_spend")) - to_number(compare_profit_for_ad.get("ad_spend")),
                "note": "用于识别广告类型底稿与经营广告费口径差异",
            },
            {"item": "广告销售额", "ad_value": total_ad.get("ad_sales"), "last_ad_value": compare_ad.get("ad_sales"), "note": "仅广告底稿具备广告归因销售"},
            {"item": "广告底稿行数", "ad_value": len(ad_rows), "last_ad_value": len(compare_ad_rows), "note": "按广告类型+ASIN+类目+渠道明细"},
            {"item": "经营底稿行数", "profit_value": len(profit_rows), "last_profit_value": len(compare_profit_rows), "note": "按 ASIN+类目+渠道明细"},
        ],
    )

    write_sheet(wb, "02_大组分析", [{"key": "platform", "label": "平台"}, {"key": "large_team", "label": "大组"}, *COMMON_HEADERS[2:]], analysis["largeTeam"], ["acos", "ad_share_total"])
    write_sheet(wb, "03_小组分析", [{"key": "platform", "label": "平台"}, {"key": "large_team", "label": "大组"}, {"key": "team", "label": "销售小组"}, *COMMON_HEADERS[2:]], analysis["team"], ["acos", "ad_share_total"])
    write_sheet(wb, "04_渠道分析", COMMON_HEADERS, analysis["channel"], ["acos", "ad_share_total"])
    write_sheet(wb, "05_广告类型分析", [
        {"key": "platform", "label": "平台"},
        {"key": "ad_type", "label": "广告类型"},
        {"key": "ad_spend", "label": "广告费", "type": "money"},
        {"key": "ad_sales", "label": "广告销售额", "type": "money"},
        {"key": "ad_orders", "label": "广告订单", "type": "int"},
        {"key": "impressions", "label": "曝光", "type": "int"},
        {"key": "clicks", "label": "点击", "type": "int"},
        {"key": "ctr", "label": "CTR", "type": "pct"},
        {"key": "cpc", "label": "CPC", "type": "decimal"},
        {"key": "cvr", "label": "CVR", "type": "pct"},
        {"key": "acos", "label": "ACOS", "type": "pct"},
        {"key": "roas", "label": "ROAS", "type": "decimal"},
        *AD_ONLY_TAIL,
    ], analysis["adType"], ["acos"])
    write_sheet(wb, "06_类目分析", [{"key": "platform", "label": "平台"}, {"key": "amazon_cat", "label": "平台类目"}, {"key": "category", "label": "品类"}, *COMMON_HEADERS[2:]], analysis["category"], ["acos", "ad_share_total"])
    write_sheet(wb, "07_ASIN分析", [{"key": "platform", "label": "平台"}, {"key": "channel", "label": "渠道"}, {"key": "asin", "label": "ASIN"}, {"key": "product", "label": "产品名称"}, {"key": "amazon_cat", "label": "平台类目"}, {"key": "category", "label": "品类"}, *COMMON_HEADERS[2:]], analysis["asin"], ["acos", "ad_share_total"])
    write_sheet(wb, "08_活动分析", [
        {"key": "platform", "label": "平台"},
        {"key": "channel", "label": "渠道"},
        {"key": "ad_type", "label": "广告类型"},
        {"key": "campaign", "label": "广告活动"},
        {"key": "ad_spend", "label": "广告费", "type": "money"},
        {"key": "ad_sales", "label": "广告销售额", "type": "money"},
        {"key": "ad_orders", "label": "广告订单", "type": "int"},
        {"key": "impressions", "label": "曝光", "type": "int"},
        {"key": "clicks", "label": "点击", "type": "int"},
        {"key": "ctr", "label": "CTR", "type": "pct"},
        {"key": "cpc", "label": "CPC", "type": "decimal"},
        {"key": "cvr", "label": "CVR", "type": "pct"},
        {"key": "acos", "label": "ACOS", "type": "pct"},
        {"key": "roas", "label": "ROAS", "type": "decimal"},
        *AD_ONLY_TAIL,
    ], analysis["campaign"], ["acos"])

    write_sheet(wb, "09_规则说明", [{"key": "rule", "label": "规则"}, {"key": "threshold", "label": "阈值"}, {"key": "meaning", "label": "含义"}], [
        {"rule": "高 ACOS", "threshold": pct(THRESHOLDS["highAcos"]), "meaning": "优先控量或降价"},
        {"rule": "预警 ACOS", "threshold": pct(THRESHOLDS["warnAcos"]), "meaning": "投产偏弱，短期观察或控量"},
        {"rule": "低 CTR", "threshold": pct(THRESHOLDS["lowCtr"]), "meaning": "高曝光低点击，检查主图/价格/相关性"},
        {"rule": "低 CVR", "threshold": pct(THRESHOLDS["lowCvr"]), "meaning": "点击后不转化，检查流量质量和 Listing"},
        {"rule": "高广告费占比", "threshold": pct(THRESHOLDS["highAdShare"]), "meaning": "广告费可能侵蚀毛利"},
        {"rule": "弱毛利率", "threshold": pct(THRESHOLDS["weakMargin"]), "meaning": "产品承接利润不足"},
    ])

    ad_raw_headers = [
        {"key": "platform", "label": "平台"}, {"key": "channel", "label": "渠道"}, {"key": "large_team", "label": "大组"}, {"key": "team", "label": "销售小组"},
        {"key": "ad_type", "label": "广告类型"}, {"key": "asin", "label": "ASIN"}, {"key": "product", "label": "产品名称"}, {"key": "amazon_cat", "label": "平台类目"}, {"key": "category", "label": "品类"},
        {"key": "ad_spend", "label": "广告费", "type": "money"}, {"key": "ad_sales", "label": "广告销售额", "type": "money"}, {"key": "ad_orders", "label": "广告订单", "type": "int"}, {"key": "clicks", "label": "点击", "type": "int"}, {"key": "impressions", "label": "曝光", "type": "int"},
    ]
    profit_raw_headers = [
        {"key": "platform", "label": "平台"}, {"key": "channel", "label": "渠道"}, {"key": "large_team", "label": "大组"}, {"key": "team", "label": "销售小组"},
        {"key": "asin", "label": "ASIN"}, {"key": "product", "label": "产品名称"}, {"key": "amazon_cat", "label": "平台类目"}, {"key": "category", "label": "品类"},
        {"key": "total_sales", "label": "总销售额", "type": "money"}, {"key": "total_orders", "label": "总订单", "type": "int"}, {"key": "gross_profit", "label": "毛利", "type": "money"}, {"key": "profit_ad_spend", "label": "广告费", "type": "money"}, {"key": "sessions", "label": "Sessions", "type": "int"}, {"key": "page_views", "label": "Page Views", "type": "int"},
    ]
    write_sheet(wb, "10_广告底稿_当前期", ad_raw_headers, ad_rows)
    write_sheet(wb, "11_经营底稿_当前期", profit_raw_headers, profit_rows)
    if compare_ad_rows:
        write_sheet(wb, "12_广告底稿_对比期", ad_raw_headers, compare_ad_rows)
    if compare_profit_rows:
        write_sheet(wb, "13_经营底稿_对比期", profit_raw_headers, compare_profit_rows)
    return wb


def main() -> None:
    args = parse_args()
    load_thresholds(args.threshold_config)
    ad_rows = normalize_ad_rows(load_rows(args.ad_source))
    profit_rows = normalize_profit_rows(load_rows(args.profit_source))
    campaign_rows = normalize_campaign_rows(load_rows(args.campaign_source))
    compare_ad_rows = normalize_ad_rows(load_rows(args.compare_ad_source))
    compare_profit_rows = normalize_profit_rows(load_rows(args.compare_profit_source))
    compare_campaign_rows = normalize_campaign_rows(load_rows(args.compare_campaign_source))
    analysis = build_analysis(ad_rows, profit_rows, campaign_rows, compare_ad_rows, compare_profit_rows, compare_campaign_rows)
    workbook = build_workbook(args, analysis, ad_rows, profit_rows, campaign_rows, compare_ad_rows, compare_profit_rows)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(json.dumps({
        "success": True,
        "output": str(output_path.resolve()),
        "thresholdConfig": str(Path(args.threshold_config).resolve()) if args.threshold_config else None,
        "sheets": workbook.sheetnames,
        "rows": {
            "ad": len(ad_rows),
            "profit": len(profit_rows),
            "campaign": len(campaign_rows),
            "compareAd": len(compare_ad_rows),
            "compareProfit": len(compare_profit_rows),
            "compareCampaign": len(compare_campaign_rows),
            "largeTeam": len(analysis["largeTeam"]),
            "team": len(analysis["team"]),
            "channel": len(analysis["channel"]),
            "adType": len(analysis["adType"]),
            "category": len(analysis["category"]),
            "asin": len(analysis["asin"]),
            "campaignAgg": len(analysis["campaign"]),
        },
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
