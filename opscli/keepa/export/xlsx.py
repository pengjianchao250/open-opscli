"""Keepa 格式化工作表模型与 XLSX 导出。"""

from __future__ import annotations

import json
from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from opscli.keepa.domain.exceptions import KeepaConfigError
from opscli.keepa.domain.models import KeepaExportResult


@dataclass(frozen=True)
class ExportColumn:
    """描述格式化工作表的一列及其来源字段。"""

    title: str
    source: str


@dataclass(frozen=True)
class FormattedWorksheet:
    """保存 XLSX 与 JSON 共用的工作表名称、列定义和原始行引用。"""

    name: str
    columns: list[ExportColumn]
    rows: list[Any]
    scenario: str

    def iter_values(self) -> Iterator[list[Any]]:
        """生成格式化行值；无需额外参数，返回按列顺序排列的行迭代器。"""
        for row in self.rows:
            normalized_row = _normalize_row(row, scenario=self.scenario)
            yield [
                _cell_value(normalized_row.get(column.source))
                for column in self.columns
            ]


EMPTY_COLUMNS = [ExportColumn("原始数据", "value")]

EXCLUDED_FIELDS = {
    "timestamp",
    "tokensLeft",
    "tokensConsumed",
    "refillIn",
    "refillRate",
    "rowSource",
}

FIELD_TITLES = {
    "asin": "ASIN",
    "asins": "ASIN列表",
    "parentAsin": "父ASIN",
    "title": "标题",
    "brand": "品牌",
    "manufacturer": "制造商",
    "productGroup": "产品组",
    "categoryTree": "类目树",
    "rootCategory": "根类目",
    "categories": "类目",
    "categoryId": "类目ID",
    "catId": "类目ID",
    "name": "名称",
    "parent": "父类目ID",
    "children": "子类目",
    "productCount": "产品数量",
    "highestRank": "最高排名",
    "sellerId": "Seller ID",
    "sellerName": "店铺名称",
    "businessName": "公司名称",
    "sellerBrand": "卖家品牌",
    "rating": "评分",
    "ratingCount": "评分数",
    "asinList": "ASIN列表",
    "sellerIdList": "Seller ID列表",
    "monthlySold": "月销量",
    "buyBoxSellerId": "Buy Box卖家ID",
    "imagesCSV": "图片CSV",
    "description": "描述",
    "features": "五点描述",
    "lastUpdate": "最近更新(Keepa分钟)",
    "lastUpdateUtc": "最近更新(UTC)",
    "lastUpdateUnixSeconds": "最近更新Unix秒",
    "lastUpdateUnixMilliseconds": "最近更新Unix毫秒",
    "listedSince": "上架时间(Keepa分钟)",
    "listedSinceUtc": "上架时间(UTC)",
    "lastPriceChange": "最近价格变化(Keepa分钟)",
    "lastPriceChangeUtc": "最近价格变化(UTC)",
    "stats": "统计数据",
    "csv": "CSV原始数据",
    "offers": "Offer数据",
    "fbaFees": "FBA费用",
    "variations": "变体",
    "eans": "EAN",
    "eanList": "EAN列表",
    "upcList": "UPC列表",
    "model": "型号",
    "color": "颜色",
    "size": "尺寸",
    "packageHeight": "包装高度",
    "packageLength": "包装长度",
    "packageWidth": "包装宽度",
    "packageWeight": "包装重量",
    "itemHeight": "商品高度",
    "itemLength": "商品长度",
    "itemWidth": "商品宽度",
    "itemWeight": "商品重量",
    "isAdultProduct": "成人产品",
    "isEligibleForTradeIn": "可Trade-In",
    "isEligibleForSuperSaverShipping": "可Super Saver配送",
    "trackingSince": "跟踪开始时间(Keepa分钟)",
    "totalResults": "总结果数",
    "timestamp": "响应时间戳",
    "tokensLeft": "剩余Token",
    "tokensConsumed": "消耗Token",
    "refillIn": "Token恢复倒计时",
    "refillRate": "Token恢复速率",
    "rowSource": "行来源",
    "productsRaw": "products原始数据",
    "sellersRaw": "sellers原始数据",
    "categoriesRaw": "categories原始数据",
    "dealsRaw": "deals原始数据",
    "bestSellersListRaw": "bestSellersList原始数据",
    "asinListRaw": "asinList原始数据",
    "sellerIdListRaw": "sellerIdList原始数据",
    "lightningDealsRaw": "lightningDeals原始数据",
    "trackingsRaw": "trackings原始数据",
    "notificationsRaw": "notifications原始数据",
    "value": "原始数据",
}

EXCEL_CELL_LIMIT = 32767


def export_rows_to_xlsx(
    *,
    rows: list[Any],
    output_path: Path,
    scenario: str,
    site: str = "US",
    params: dict[str, Any] | None = None,
    extra_sheets: dict[str, list[Any]] | None = None,
) -> KeepaExportResult:
    """把主表、场景信息和附加表导出为 XLSX，返回文件路径、格式和 MIME 元数据。"""
    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise KeepaConfigError("缺少 openpyxl 依赖，无法导出 XLSX") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(write_only=True)
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True)

    for formatted_sheet in build_formatted_worksheets(
        rows=rows,
        scenario=scenario,
        site=site,
        params=params,
        extra_sheets=extra_sheets,
    ):
        sheet = workbook.create_sheet(title=formatted_sheet.name)
        _write_rows_sheet(
            sheet=sheet,
            formatted_sheet=formatted_sheet,
            header_fill=header_fill,
            header_font=header_font,
            write_only_cell=WriteOnlyCell,
            get_column_letter=get_column_letter,
        )

    workbook.save(output_path)
    resolved = output_path.resolve()
    return KeepaExportResult(path=str(resolved), filename=resolved.name, url=resolved.as_uri())


def _write_rows_sheet(
    *,
    sheet: Any,
    formatted_sheet: FormattedWorksheet,
    header_fill: Any,
    header_font: Any,
    write_only_cell: Any,
    get_column_letter: Any,
) -> None:
    columns = formatted_sheet.columns
    sheet.freeze_panes = "A2"

    header_cells = []
    for column in columns:
        cell = write_only_cell(sheet, value=column.title)
        cell.font = header_font
        cell.fill = header_fill
        header_cells.append(cell)
    sheet.append(header_cells)

    for values in formatted_sheet.iter_values():
        sheet.append(values)

    for column_index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = _column_width(column.title)


def _columns_from_row_values(rows: list[Any], *, scenario: str) -> list[ExportColumn]:
    """扫描原始行确定导出列，避免为大结果额外复制整表数据。"""
    fields: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in _normalize_row(row, scenario=scenario):
            if key in EXCLUDED_FIELDS:
                continue
            if key not in seen:
                seen.add(key)
                fields.append(key)
    return [ExportColumn(_title_for_field(field), field) for field in fields] or EMPTY_COLUMNS


def build_formatted_worksheets(
    *,
    rows: list[Any],
    scenario: str,
    site: str = "US",
    params: dict[str, Any] | None = None,
    extra_sheets: dict[str, list[Any]] | None = None,
) -> list[FormattedWorksheet]:
    """根据结果行、场景、站点、参数和附加表构造唯一命名的有序工作表列表。"""
    inputs: list[tuple[str, list[Any]]] = [
        (
            _sheet_title(
                scenario=scenario,
                site=site,
                params=params or {},
                rows=rows,
            ),
            rows,
        )
    ]
    inputs.extend(
        (name, sheet_rows)
        for name, sheet_rows in (extra_sheets or {}).items()
        if sheet_rows
    )
    used_names: set[str] = set()
    worksheets: list[FormattedWorksheet] = []
    for requested_name, sheet_rows in inputs:
        name = _unique_sheet_title(requested_name, used_names)
        used_names.add(name.casefold())
        worksheets.append(
            FormattedWorksheet(
                name=name,
                columns=_columns_from_row_values(sheet_rows, scenario=scenario),
                rows=sheet_rows,
                scenario=scenario,
            )
        )
    return worksheets


def _normalize_row(row: Any, *, scenario: str) -> dict[str, Any]:
    if isinstance(row, dict):
        return row
    if isinstance(row, str):
        if scenario in {"seller", "seller-finder", "top-seller"}:
            return {"sellerId": row}
        return {"asin": row}
    return {"value": row}


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False)
    if isinstance(value, str) and len(value) > EXCEL_CELL_LIMIT:
        return value[:EXCEL_CELL_LIMIT]
    return value


def _sheet_title(*, scenario: str, site: str, params: dict[str, Any], rows: list[Any]) -> str:
    target = params.get("keyword") or params.get("term") or params.get("asin") or params.get("seller") or ""
    suffix = f"-{target}" if target else ""
    return f"Keepa-{site.upper()}-{scenario}{suffix}({len(rows)})"


def _safe_sheet_title(value: str) -> str:
    title = "".join(char for char in value if char not in r"[]:*?/\\")
    return (title or "Keepa")[:31]


def _unique_sheet_title(value: str, used_names: set[str]) -> str:
    """按 Excel 的大小写不敏感规则生成不超过 31 字符的唯一表名。"""
    base = _safe_sheet_title(value)
    candidate = base
    suffix = 1
    while candidate.casefold() in used_names:
        suffix_text = str(suffix)
        candidate = f"{base[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    return candidate


def _column_width(title: str) -> int:
    if any(key in title for key in ["标题", "原始数据", "类目树", "统计数据", "CSV原始数据", "Offer数据"]):
        return 48
    if any(key in title for key in ["ASIN", "品牌", "Seller"]):
        return 18
    return max(12, min(24, len(str(title)) * 2 + 4))


def _title_for_field(field: str) -> str:
    return FIELD_TITLES.get(field, field)
