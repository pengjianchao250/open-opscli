"""卖家精灵采集结果解析 Adapter。"""

from __future__ import annotations

import hashlib
import json
from collections.abc import Iterable, Iterator, Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Any

from opscli.collector_mcp.storage.models import (
    CollectionArtifact,
    CollectionDataset,
    CollectionRecord,
    CollectionSubmission,
    ParsedCollection,
    PermanentCollectionError,
)

# Parser 版本写入 collection_runs，便于未来格式升级后追踪解析口径。
PARSER_VERSION = "seller-sprite-v1"


class CollectionParseError(PermanentCollectionError):
    """采集结果文件缺失、越界或不符合持久化合同。"""


class SellerSpriteCollectionParser:
    """将卖家精灵成功任务转换为跨场景的逻辑 Dataset。"""

    source_system = "seller_sprite"
    parser_version = PARSER_VERSION

    def parse(self, submission: CollectionSubmission) -> ParsedCollection:
        """解析 SellerSprite JSON 或 XLSX 格式化导出。"""
        if submission.source_system != self.source_system:
            raise CollectionParseError(
                f"SellerSprite Parser 不支持来源：{submission.source_system}"
            )
        result_path = submission.result_path.resolve()
        root_dir = result_path.parent
        result = _read_json(result_path)
        if str(result.get("job_id") or "") != submission.source_job_id:
            raise CollectionParseError("result.json 的 job_id 与 Outbox 提交不一致")

        params_path = _artifact_path(result.get("params_path"), root_dir, "params")
        raw_path = _artifact_path(result.get("raw_path"), root_dir, "raw")
        export_payload = result.get("export")
        if not isinstance(export_payload, dict):
            raise CollectionParseError("卖家精灵成功结果缺少 export")
        export_path = _artifact_path(export_payload.get("path"), root_dir, "export")
        export_format = str(
            export_payload.get("format") or export_path.suffix.lstrip(".")
        ).lower()
        params = _read_json(params_path)
        if export_format == "json" or export_path.suffix.lower() == ".json":
            datasets = _json_datasets(_read_json(export_path))
        elif export_format in {"xls", "xlsx"} or export_path.suffix.lower() == ".xlsx":
            datasets = _xlsx_datasets(export_path)
        else:
            raise CollectionParseError(f"不支持的卖家精灵导出格式：{export_format}")
        artifacts = (
            _build_artifact("params", params_path, "application/json"),
            _build_artifact("raw", raw_path, "application/json"),
            _build_artifact("result", result_path, "application/json"),
            _build_artifact(
                "export",
                export_path,
                str(export_payload.get("mime_type") or "application/json"),
            ),
        )
        return ParsedCollection(
            submission=submission,
            parser_version=self.parser_version,
            request_params=params,
            artifacts=artifacts,
            datasets=datasets,
        )


def _json_datasets(payload: dict[str, Any]) -> tuple[CollectionDataset, ...]:
    sheets: list[tuple[str, dict[str, Any]]] = [
        (
            "main",
            {
                "name": payload.get("sheet_name") or "main",
                "columns": payload.get("columns"),
                "rows": payload.get("rows"),
            },
        )
    ]
    additional = payload.get("additional_sheets") or []
    if not isinstance(additional, list):
        raise CollectionParseError("additional_sheets 必须是数组")
    for index, sheet in enumerate(additional, start=1):
        if not isinstance(sheet, dict):
            raise CollectionParseError("additional_sheets 条目必须是对象")
        sheets.append((f"additional_{index}", sheet))
    return tuple(_dataset_from_sheet(code, sheet) for code, sheet in sheets)


def _xlsx_datasets(path: Path) -> tuple[CollectionDataset, ...]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise CollectionParseError("无法读取卖家精灵 XLSX 导出") from exc
    try:
        datasets: list[CollectionDataset] = []
        for index, worksheet in enumerate(workbook.worksheets):
            rows = iter(worksheet.iter_rows(values_only=True))
            try:
                header = next(rows)
            except StopIteration:
                continue
            columns = tuple(_normalized_columns(list(header)))
            sheet_name = worksheet.title or f"sheet_{index + 1}"
            code = "main" if not datasets else f"additional_{len(datasets)}"
            datasets.append(
                CollectionDataset(
                    dataset_code=code,
                    dataset_name=sheet_name,
                    source_sheet=sheet_name,
                    columns=columns,
                    records=_iter_xlsx_records(path, sheet_name, columns),
                )
            )
    finally:
        workbook.close()
    if not datasets:
        raise CollectionParseError("卖家精灵 XLSX 导出不包含可入库工作表")
    return tuple(datasets)


def _dataset_from_sheet(code: str, sheet: dict[str, Any]) -> CollectionDataset:
    raw_columns = sheet.get("columns")
    raw_rows = sheet.get("rows")
    if not isinstance(raw_columns, list) or not isinstance(raw_rows, list):
        raise CollectionParseError("格式化工作表必须包含 columns 和 rows 数组")
    columns = tuple(_normalized_columns(raw_columns))
    name = str(sheet.get("name") or code)
    return CollectionDataset(
        dataset_code=code,
        dataset_name=name,
        source_sheet=name,
        columns=columns,
        records=_iter_json_records(raw_rows, columns),
    )


def _iter_json_records(
    raw_rows: Iterable[Any],
    columns: Sequence[tuple[str, str]],
) -> Iterator[CollectionRecord]:
    """逐行转换格式化 JSON，避免再复制一份完整记录列表。"""
    for row_number, raw_row in enumerate(raw_rows, start=1):
        if not isinstance(raw_row, list) or len(raw_row) != len(columns):
            raise CollectionParseError("格式化工作表的行列数量不一致")
        yield _build_record(row_number, raw_row, columns)


def _iter_xlsx_records(
    path: Path,
    sheet_name: str,
    columns: Sequence[tuple[str, str]],
) -> Iterator[CollectionRecord]:
    """按 Sheet 重新打开只读工作簿并流式产出非空数据行。"""
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise CollectionParseError("无法读取卖家精灵 XLSX 导出") from exc
    try:
        worksheet = workbook[sheet_name]
        rows = iter(worksheet.iter_rows(values_only=True))
        next(rows, None)
        row_number = 0
        for raw_row in rows:
            if not any(value is not None for value in raw_row):
                continue
            row_number += 1
            values = [_json_safe_value(value) for value in raw_row]
            if len(values) != len(columns):
                raise CollectionParseError("格式化工作表的行列数量不一致")
            yield _build_record(row_number, values, columns)
    finally:
        workbook.close()


def _build_record(
    row_number: int,
    values: Sequence[Any],
    columns: Sequence[tuple[str, str]],
) -> CollectionRecord:
    """根据稳定列键构造一条可写入 MySQL JSON 的格式化记录。"""
    payload = {
        normalized_name: value
        for (_, normalized_name), value in zip(columns, values, strict=True)
    }
    return CollectionRecord(
        row_number=row_number,
        payload=payload,
        record_hash=_record_hash(payload),
    )


def _normalized_columns(raw_columns: list[Any]) -> list[tuple[str, str]]:
    counts: dict[str, int] = {}
    columns: list[tuple[str, str]] = []
    for index, value in enumerate(raw_columns, start=1):
        original = str(value or f"column_{index}")
        counts[original] = counts.get(original, 0) + 1
        occurrence = counts[original]
        normalized = original if occurrence == 1 else f"{original}__{occurrence}"
        columns.append((original, normalized))
    return columns


def _artifact_path(value: Any, root_dir: Path, artifact_type: str) -> Path:
    if not value:
        raise CollectionParseError(f"卖家精灵结果缺少 {artifact_type} 文件路径")
    path = Path(str(value)).expanduser().resolve()
    try:
        path.relative_to(root_dir)
    except ValueError as exc:
        raise CollectionParseError(f"{artifact_type} 文件不在成功任务目录内") from exc
    if not path.is_file():
        raise CollectionParseError(f"{artifact_type} 文件不存在")
    return path


def _build_artifact(
    artifact_type: str, path: Path, mime_type: str
) -> CollectionArtifact:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return CollectionArtifact(
        artifact_type=artifact_type,
        path=path,
        filename=path.name,
        mime_type=mime_type,
        size_bytes=path.stat().st_size,
        sha256=digest.hexdigest(),
    )


def _read_json(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise CollectionParseError(f"无法读取采集 JSON 文件：{path.name}") from exc
    if not isinstance(payload, dict):
        raise CollectionParseError(f"采集 JSON 文件必须是对象：{path.name}")
    return payload


def _record_hash(payload: dict[str, Any]) -> str:
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _json_safe_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)
