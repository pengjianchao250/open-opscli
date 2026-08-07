"""Collector 来源 Parser 共用的文件和 Dataset 解析工具。"""

from __future__ import annotations

import hashlib
import json
from collections.abc import Iterable, Iterator, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from opscli.collector_mcp.storage.models import (
    CollectionArtifact,
    CollectionDataset,
    CollectionRecord,
    CollectionSubmission,
    PermanentCollectionError,
)


class CollectionParseError(PermanentCollectionError):
    """采集结果文件缺失、越界或不符合持久化合同。"""


@dataclass(frozen=True)
class CollectionResultFiles:
    """一个来源成功合同解析出的标准文件集合。"""

    result: dict[str, Any]
    result_path: Path
    params: dict[str, Any]
    params_path: Path
    raw_path: Path
    export: dict[str, Any]
    export_path: Path


def load_result_files(
    submission: CollectionSubmission,
    *,
    source_name: str,
) -> CollectionResultFiles:
    """校验通用成功合同并解析 params/raw/result/export 文件。"""
    result_path = submission.result_path.resolve()
    root_dir = result_path.parent
    result = read_json_object(result_path, source_name=source_name)
    if str(result.get("job_id") or "") != submission.source_job_id:
        raise CollectionParseError("result.json 的 job_id 与 Outbox 提交不一致")
    params_path = resolve_artifact_path(
        result.get("params_path"),
        root_dir=root_dir,
        artifact_type="params",
        source_name=source_name,
    )
    raw_path = resolve_artifact_path(
        result.get("raw_path"),
        root_dir=root_dir,
        artifact_type="raw",
        source_name=source_name,
    )
    export = result.get("export")
    if not isinstance(export, dict):
        raise CollectionParseError(f"{source_name}成功结果缺少 export")
    export_path = resolve_artifact_path(
        export.get("path"),
        root_dir=root_dir,
        artifact_type="export",
        source_name=source_name,
    )
    return CollectionResultFiles(
        result=result,
        result_path=result_path,
        params=read_json_object(params_path, source_name=source_name),
        params_path=params_path,
        raw_path=raw_path,
        export=export,
        export_path=export_path,
    )


def standard_artifacts(
    files: CollectionResultFiles,
    *,
    default_export_mime_type: str,
) -> tuple[CollectionArtifact, ...]:
    """为标准成功合同构造 params/raw/result/export 四类 Artifact。"""
    return (
        build_artifact("params", files.params_path, "application/json"),
        build_artifact("raw", files.raw_path, "application/json"),
        build_artifact("result", files.result_path, "application/json"),
        build_artifact(
            "export",
            files.export_path,
            str(files.export.get("mime_type") or default_export_mime_type),
        ),
    )


def read_json_object(path: Path, *, source_name: str) -> dict[str, Any]:
    """读取来源 JSON 对象并把不可重试的格式错误归类。"""
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise CollectionParseError(f"无法读取{source_name}采集 JSON 文件：{path.name}") from exc
    if not isinstance(payload, dict):
        raise CollectionParseError(f"{source_name}采集 JSON 文件必须是对象：{path.name}")
    return payload


def resolve_artifact_path(
    value: Any,
    *,
    root_dir: Path,
    artifact_type: str,
    source_name: str,
) -> Path:
    """校验来源结果引用的制品存在且没有越出任务目录。"""
    if not value:
        raise CollectionParseError(f"{source_name}结果缺少 {artifact_type} 文件路径")
    path = Path(str(value)).expanduser().resolve()
    try:
        path.relative_to(root_dir)
    except ValueError as exc:
        raise CollectionParseError(f"{artifact_type} 文件不在{source_name}成功任务目录内") from exc
    if not path.is_file():
        raise CollectionParseError(f"{artifact_type} 文件不存在")
    return path


def build_artifact(
    artifact_type: str,
    path: Path,
    mime_type: str,
) -> CollectionArtifact:
    """流式计算文件摘要并构造通用 Artifact。"""
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return CollectionArtifact(
        artifact_type=artifact_type,
        path=path,
        filename=path.name,
        mime_type=mime_type,
        size_bytes=path.stat().st_size,
        sha256=digest.hexdigest(),
    )


def json_datasets(
    payload: dict[str, Any],
    *,
    source_name: str,
    business_key_fields: Sequence[str] = (),
) -> tuple[CollectionDataset, ...]:
    """把通用格式化 JSON 主表和附加表转换为 Dataset。"""
    sheets: list[tuple[str, dict[str, Any]]] = [
        (
            "main",
            {
                "name": payload.get("sheet_name") or "main",
                "columns": payload.get("columns"),
                "rows": payload.get("rows"),
            },
        )
    ]
    additional = payload.get("additional_sheets") or []
    if not isinstance(additional, list):
        raise CollectionParseError(f"{source_name} additional_sheets 必须是数组")
    for index, sheet in enumerate(additional, start=1):
        if not isinstance(sheet, dict):
            raise CollectionParseError(f"{source_name} additional_sheets 条目必须是对象")
        sheets.append((f"additional_{index}", sheet))
    return tuple(
        _dataset_from_sheet(
            code,
            sheet,
            source_name=source_name,
            business_key_fields=business_key_fields,
        )
        for code, sheet in sheets
    )


def xlsx_datasets(
    path: Path,
    *,
    source_name: str,
    business_key_fields: Sequence[str] = (),
) -> tuple[CollectionDataset, ...]:
    """把 XLSX 每个非空工作表转换为惰性记录 Dataset。"""
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise CollectionParseError(f"无法读取{source_name} XLSX 导出") from exc
    try:
        datasets: list[CollectionDataset] = []
        for worksheet in workbook.worksheets:
            rows = iter(worksheet.iter_rows(values_only=True))
            try:
                header = next(rows)
            except StopIteration:
                continue
            columns = tuple(normalized_columns(list(header)))
            sheet_name = worksheet.title or f"sheet_{len(datasets) + 1}"
            code = "main" if not datasets else f"additional_{len(datasets)}"
            datasets.append(
                CollectionDataset(
                    dataset_code=code,
                    dataset_name=sheet_name,
                    source_sheet=sheet_name,
                    columns=columns,
                    records=_iter_xlsx_records(
                        path,
                        sheet_name,
                        columns,
                        source_name=source_name,
                        business_key_fields=business_key_fields,
                    ),
                )
            )
    finally:
        workbook.close()
    if not datasets:
        raise CollectionParseError(f"{source_name} XLSX 导出不包含可入库工作表")
    return tuple(datasets)


def normalized_columns(raw_columns: list[Any]) -> list[tuple[str, str]]:
    """保留原始表头顺序并为重复列生成稳定 JSON key。"""
    counts: dict[str, int] = {}
    columns: list[tuple[str, str]] = []
    for index, value in enumerate(raw_columns, start=1):
        original = str(value or f"column_{index}")
        counts[original] = counts.get(original, 0) + 1
        occurrence = counts[original]
        normalized = original if occurrence == 1 else f"{original}__{occurrence}"
        columns.append((original, normalized))
    return columns


def _dataset_from_sheet(
    code: str,
    sheet: dict[str, Any],
    *,
    source_name: str,
    business_key_fields: Sequence[str],
) -> CollectionDataset:
    raw_columns = sheet.get("columns")
    raw_rows = sheet.get("rows")
    if not isinstance(raw_columns, list) or not isinstance(raw_rows, list):
        raise CollectionParseError(f"{source_name}格式化工作表必须包含 columns 和 rows 数组")
    columns = tuple(normalized_columns(raw_columns))
    name = str(sheet.get("name") or code)
    return CollectionDataset(
        dataset_code=code,
        dataset_name=name,
        source_sheet=name,
        columns=columns,
        records=_iter_json_records(
            raw_rows,
            columns,
            source_name=source_name,
            business_key_fields=business_key_fields,
        ),
    )


def _iter_json_records(
    raw_rows: Iterable[Any],
    columns: Sequence[tuple[str, str]],
    *,
    source_name: str,
    business_key_fields: Sequence[str],
) -> Iterator[CollectionRecord]:
    for row_number, raw_row in enumerate(raw_rows, start=1):
        if not isinstance(raw_row, list) or len(raw_row) != len(columns):
            raise CollectionParseError(f"{source_name}格式化工作表的行列数量不一致")
        yield _build_record(row_number, raw_row, columns, business_key_fields)


def _iter_xlsx_records(
    path: Path,
    sheet_name: str,
    columns: Sequence[tuple[str, str]],
    *,
    source_name: str,
    business_key_fields: Sequence[str],
) -> Iterator[CollectionRecord]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise CollectionParseError(f"无法读取{source_name} XLSX 导出") from exc
    try:
        worksheet = workbook[sheet_name]
        rows = iter(worksheet.iter_rows(values_only=True))
        next(rows, None)
        row_number = 0
        for raw_row in rows:
            if not any(value is not None for value in raw_row):
                continue
            row_number += 1
            values = [_json_safe_value(value) for value in raw_row]
            if len(values) != len(columns):
                raise CollectionParseError(f"{source_name}格式化工作表的行列数量不一致")
            yield _build_record(row_number, values, columns, business_key_fields)
    finally:
        workbook.close()


def _build_record(
    row_number: int,
    values: Sequence[Any],
    columns: Sequence[tuple[str, str]],
    business_key_fields: Sequence[str],
) -> CollectionRecord:
    payload = {
        normalized_name: value
        for (_, normalized_name), value in zip(columns, values, strict=True)
    }
    return CollectionRecord(
        row_number=row_number,
        payload=payload,
        record_hash=_record_hash(payload),
        business_key=_business_key(payload, business_key_fields),
    )


def _business_key(
    payload: dict[str, Any],
    preferred_fields: Sequence[str],
) -> str | None:
    by_normalized_name = {str(key).casefold(): value for key, value in payload.items()}
    for field in preferred_fields:
        value = by_normalized_name.get(field.casefold())
        if value is not None and str(value).strip():
            return str(value).strip()[:255]
    return None


def _record_hash(payload: dict[str, Any]) -> str:
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _json_safe_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)
