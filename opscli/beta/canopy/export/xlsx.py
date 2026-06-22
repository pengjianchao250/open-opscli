"""beta Canopy XLSX 导出。"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from opscli.beta.canopy.domain.exceptions import CanopyConfigError
from opscli.beta.canopy.domain.models import CanopyExportResult


@dataclass(frozen=True)
class ExportColumn:
    title: str
    source: str


EMPTY_COLUMNS = [ExportColumn("原始数据", "value")]
EXCEL_CELL_LIMIT = 32767

FIELD_TITLES = {
    "asin": "ASIN",
    "parentAsin": "父ASIN",
    "title": "标题",
    "brand": "品牌",
    "manufacturer": "制造商",
    "domain": "站点",
    "url": "URL",
    "image": "图片",
    "images": "图片",
    "price": "价格",
    "listPrice": "标价",
    "currency": "币种",
    "rating": "评分",
    "ratingsTotal": "评分数",
    "reviewsTotal": "评论数",
    "reviewCount": "评论数",
    "productTitle": "商品标题",
    "productRating": "商品评分",
    "reviewId": "评论ID",
    "reviewTitle": "评论标题",
    "reviewText": "评论内容",
    "reviewerName": "评论人",
    "reviewerId": "评论人ID",
    "reviewerUrl": "评论人URL",
    "reviewDate": "评论日期",
    "reviewUrl": "评论URL",
    "helpfulVotes": "有用票数",
    "verifiedPurchase": "已验证购买",
    "imageUrls": "评论图片",
    "videos": "评论视频",
    "currentPage": "当前页",
    "totalPages": "总页数",
    "totalResults": "总结果数",
    "sellerId": "Seller ID",
    "sellerName": "卖家名称",
    "buyBoxWinner": "Buy Box",
    "availability": "可售状态",
    "stock": "库存",
    "sales": "销量",
    "categoryId": "类目ID",
    "categoryName": "类目名称",
    "name": "名称",
    "rank": "排名",
    "gtin": "GTIN",
    "ean": "EAN",
    "upc": "UPC",
    "success": "成功",
    "value": "原始数据",
    "rowSource": "行来源",
}

PRIMARY_KEYS = (
    "reviews",
    "products",
    "searchResults",
    "results",
    "items",
    "offers",
    "variants",
    "variantList",
    "deals",
    "categories",
    "bestsellers",
    "bestSellers",
    "bestSellerCategories",
    "bestsellerCategories",
    "sellerProducts",
    "sellers",
    "authors",
    "asinList",
    "product",
    "amazonProduct",
    "seller",
    "author",
    "category",
    "stock",
    "sales",
)

COLLECTION_KEYS = {
    "reviews",
    "products",
    "searchResults",
    "results",
    "items",
    "offers",
    "variants",
    "variantList",
    "deals",
    "categories",
    "bestsellers",
    "bestSellers",
    "bestSellerCategories",
    "bestsellerCategories",
    "sellerProducts",
    "sellers",
    "authors",
    "asinList",
}

SINGLETON_KEYS = {"product", "amazonProduct", "seller", "author", "category", "stock", "sales"}


_FIELD_TITLE_SUFFIXES = {
    "asin": "ASIN",
    "title": "标题",
    "name": "名称",
    "price": "价格",
    "rating": "评分",
    "rank": "排名",
    "url": "URL",
    "id": "ID",
}


def export_rows_to_xlsx(
    *,
    rows: list[Any],
    output_path: Path,
    scenario: str,
    domain: str = "US",
    params: dict[str, Any] | None = None,
) -> CanopyExportResult:
    """将 Canopy rows 导出为 Excel 兼容 XLSX 文件。"""
    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise CanopyConfigError("缺少 openpyxl 依赖，无法导出 XLSX") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title=_safe_sheet_title(_sheet_title(scenario=scenario, domain=domain, params=params or {}, rows=rows)))
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True)

    columns = _columns_from_row_values(rows)
    sheet.freeze_panes = "A2"

    header_cells = []
    for column in columns:
        cell = WriteOnlyCell(sheet, value=column.title)
        cell.font = header_font
        cell.fill = header_fill
        header_cells.append(cell)
    sheet.append(header_cells)

    for row in rows:
        normalized = _normalize_row(row)
        sheet.append([_cell_value(normalized.get(column.source)) for column in columns])

    for column_index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = _column_width(column.title)

    workbook.save(output_path)
    resolved = output_path.resolve()
    return CanopyExportResult(path=str(resolved), filename=resolved.name, url=resolved.as_uri())


def response_to_export_rows(payload: Any) -> list[dict[str, Any]]:
    """将 Canopy 原始响应转换为可导出的行。"""
    if isinstance(payload, list):
        return [_normalize_row(item) for item in payload]
    if not isinstance(payload, dict):
        return [{"value": payload}]

    canopy_review_rows = _canopy_amazon_product_review_rows(payload)
    if canopy_review_rows is not None:
        return canopy_review_rows

    row_source_key, row_items = _primary_row_items(payload)
    if row_items:
        base_row = {key: value for key, value in payload.items() if key != row_source_key and key != "data"}
        rows: list[dict[str, Any]] = []
        for item in row_items:
            row = dict(base_row)
            if row_source_key:
                row["rowSource"] = row_source_key
            row.update(_normalize_row(item, row_source_key=row_source_key))
            rows.append(row)
        return rows

    data = payload.get("data")
    if isinstance(data, (dict, list)):
        data_rows = response_to_export_rows(data)
        if data_rows:
            base_row = {key: value for key, value in payload.items() if key != "data"}
            if base_row:
                return [{**base_row, **row} for row in data_rows]
            return data_rows

    return [_normalize_row(payload)]


def _canopy_amazon_product_review_rows(payload: dict[str, Any]) -> list[dict[str, Any]] | None:
    amazon_product = _amazon_product_from_payload(payload)
    if not isinstance(amazon_product, dict):
        return None

    reviews_paginated = amazon_product.get("reviewsPaginated")
    has_reviews_paginated = isinstance(reviews_paginated, dict) and "reviews" in reviews_paginated
    has_top_reviews = "topReviews" in amazon_product
    page_info = reviews_paginated.get("pageInfo") if isinstance(reviews_paginated, dict) else None
    review_items: list[tuple[str, Any]] = []

    if has_reviews_paginated:
        paginated_reviews = reviews_paginated.get("reviews")
        if not isinstance(paginated_reviews, list):
            return []
        review_items = [("reviewsPaginated.reviews", item) for item in paginated_reviews]
    elif has_top_reviews:
        top_reviews = amazon_product.get("topReviews")
        if not isinstance(top_reviews, list):
            return []
        review_items = [("topReviews", item) for item in top_reviews]
    else:
        return None

    if not review_items:
        return []

    base_row = {key: value for key, value in payload.items() if key not in {"data", "amazonProduct"}}
    product_context = _amazon_product_context(amazon_product)
    page_context = _page_info_context(page_info)
    rows: list[dict[str, Any]] = []
    for row_source, review in review_items:
        row = dict(base_row)
        row["rowSource"] = row_source
        row.update(product_context)
        row.update(_normalize_review_row(review))
        row.update(page_context)
        rows.append(row)
    return rows


def _amazon_product_from_payload(payload: dict[str, Any]) -> dict[str, Any] | None:
    amazon_product = payload.get("amazonProduct")
    if isinstance(amazon_product, dict):
        return amazon_product
    data = payload.get("data")
    if isinstance(data, dict) and isinstance(data.get("amazonProduct"), dict):
        return data["amazonProduct"]
    return None


def _amazon_product_context(amazon_product: dict[str, Any]) -> dict[str, Any]:
    context: dict[str, Any] = {}
    for source, target in (
        ("asin", "asin"),
        ("parentAsin", "parentAsin"),
        ("title", "productTitle"),
        ("brand", "brand"),
        ("manufacturer", "manufacturer"),
        ("rating", "productRating"),
        ("ratingsTotal", "ratingsTotal"),
        ("reviewsTotal", "reviewsTotal"),
        ("reviewCount", "reviewCount"),
        ("url", "url"),
    ):
        if source in amazon_product and amazon_product[source] is not None:
            context[target] = amazon_product[source]
    return context


def _page_info_context(page_info: Any) -> dict[str, Any]:
    if not isinstance(page_info, dict):
        return {}
    context: dict[str, Any] = {}
    for key in ("currentPage", "totalPages", "totalResults"):
        if key in page_info and page_info[key] is not None:
            context[key] = page_info[key]
    return context


def _normalize_review_row(review: Any) -> dict[str, Any]:
    if not isinstance(review, dict):
        return _normalize_row(review, row_source_key="reviews")

    mapped: dict[str, Any] = {}
    consumed: set[str] = set()
    for target, source_keys in {
        "reviewId": ("reviewId", "id"),
        "reviewTitle": ("reviewTitle", "title"),
        "reviewText": ("reviewText", "body", "text", "content"),
        "rating": ("rating",),
        "helpfulVotes": ("helpfulVotes", "helpfulVoteCount", "helpful_vote_count"),
        "verifiedPurchase": ("verifiedPurchase", "isVerifiedPurchase", "verified"),
        "reviewerName": ("reviewerName", "profileName", "author", "name"),
        "reviewerId": ("reviewerId", "profileId"),
        "reviewerUrl": ("reviewerUrl", "profileUrl"),
        "reviewDate": ("reviewDate", "date", "createdAt"),
        "reviewUrl": ("reviewUrl", "url", "permalink"),
        "imageUrls": ("imageUrls", "images"),
        "videos": ("videos",),
    }.items():
        found, key, value = _first_present(review, source_keys)
        if found:
            mapped[target] = value
            consumed.add(key)

    for key, value in review.items():
        if key not in consumed and key not in mapped:
            mapped[key] = value
    return mapped


def _first_present(payload: dict[str, Any], keys: tuple[str, ...]) -> tuple[bool, str, Any]:
    for key in keys:
        if key in payload and payload[key] is not None:
            return True, key, payload[key]
    return False, "", None


def _primary_row_items(payload: dict[str, Any]) -> tuple[str | None, list[Any]]:
    for key in PRIMARY_KEYS:
        if key not in payload:
            continue
        rows = _rows_from_response_value(key, payload.get(key))
        if rows:
            return key, rows
    return None, []


def _rows_from_response_value(key: str, value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if key in SINGLETON_KEYS and isinstance(value, dict):
        return [value]
    if isinstance(value, dict):
        for child_key in ("items", "results", "products", "reviews", "offers", "variants", "deals", "dr", "list"):
            child_value = value.get(child_key)
            if isinstance(child_value, list):
                return child_value
        if key in COLLECTION_KEYS:
            values = list(value.values())
            if values and all(isinstance(item, dict) for item in values):
                return values
            return [value]
        return [value]
    if value is not None and key in {"asinList"}:
        return [value]
    return []


def _normalize_row(row: Any, *, row_source_key: str | None = None) -> dict[str, Any]:
    if isinstance(row, dict):
        return row
    if isinstance(row, str):
        if row_source_key in {"sellers", "sellerProducts"}:
            return {"sellerId": row}
        if row_source_key in {"asinList", "bestsellers", "bestSellers"}:
            return {"asin": row}
    return {"value": row}


def _columns_from_row_values(rows: list[Any]) -> list[ExportColumn]:
    fields: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in _normalize_row(row):
            if key not in seen:
                seen.add(key)
                fields.append(key)
    return [ExportColumn(_title_for_field(field), field) for field in fields] or EMPTY_COLUMNS


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False)
    if isinstance(value, str) and len(value) > EXCEL_CELL_LIMIT:
        return value[:EXCEL_CELL_LIMIT]
    return value


def _sheet_title(*, scenario: str, domain: str, params: dict[str, Any], rows: list[Any]) -> str:
    target = params.get("asin") or params.get("searchTerm") or params.get("categoryId") or params.get("sellerId") or ""
    suffix = f"-{target}" if target else ""
    return f"Canopy-{domain.upper()}-{scenario}{suffix}({len(rows)})"


def _safe_sheet_title(value: str) -> str:
    title = "".join(char for char in value if char not in r"[]:*?/\\")
    return (title or "Canopy")[:31]


def _column_width(title: str) -> int:
    if any(key in title for key in ["标题", "内容", "原始数据", "图片", "URL"]):
        return 48
    if any(key in title for key in ["ASIN", "Seller", "类目", "日期"]):
        return 18
    return max(12, min(28, len(str(title)) * 2 + 4))


def _title_for_field(field: str) -> str:
    if field in FIELD_TITLES:
        return FIELD_TITLES[field]
    lower = field.lower()
    for suffix, title in _FIELD_TITLE_SUFFIXES.items():
        if lower.endswith(suffix.lower()):
            return title if field == suffix else field
    return field
