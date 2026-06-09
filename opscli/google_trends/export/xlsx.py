"""Google Trends XLSX export."""

from __future__ import annotations

import json
from dataclasses import dataclass
from numbers import Real
from pathlib import Path
from typing import Any

from opscli.google_trends.domain.exceptions import GoogleTrendsConfigError
from opscli.google_trends.domain.models import GoogleTrendsExportResult


@dataclass(frozen=True)
class ExportColumn:
    title: str
    source: str | None = None
    transform: str | None = None
    fallback_source: str | None = None


INTEREST_TIME_COLUMNS = [
    ExportColumn("日期", "date"),
    ExportColumn("是否部分数据", "isPartial", "bool"),
]

REGION_COLUMNS = [
    ExportColumn("地区", "geo_name"),
    ExportColumn("地区代码", "geoCode"),
]

RELATED_COLUMNS = [
    ExportColumn("关键词", "keyword"),
    ExportColumn("类型", "type"),
    ExportColumn("相关查询/主题", "query", fallback_source="topic_title"),
    ExportColumn("值", "value"),
    ExportColumn("topic_mid", "topic_mid"),
    ExportColumn("topic_type", "topic_type"),
]

SUGGESTION_COLUMNS = [
    ExportColumn("标题", "title"),
    ExportColumn("类型", "type"),
    ExportColumn("topic_mid", "mid"),
]

TRENDING_COLUMNS = [
    ExportColumn("排名", "rank"),
    ExportColumn("搜索词", "search_term"),
    ExportColumn("traffic", "traffic"),
]

GENERIC_COLUMNS = [
    ExportColumn("关键词", "keyword"),
    ExportColumn("日期", "date"),
    ExportColumn("地区", "geo_name"),
    ExportColumn("搜索词", "search_term"),
    ExportColumn("搜索指数", "value"),
    ExportColumn("原始数据", None, "rowJson"),
]


def export_rows_to_xlsx(
    *,
    rows: list[dict[str, Any]],
    output_path: Path,
    scenario: str,
    geo: str = "US",
    params: dict[str, Any] | None = None,
) -> GoogleTrendsExportResult:
    """Export Google Trends rows to a user-friendly XLSX workbook."""
    try:
        from openpyxl import Workbook  # type: ignore[import-not-found]
        from openpyxl.styles import Font, PatternFill  # type: ignore[import-not-found]
        from openpyxl.utils import get_column_letter  # type: ignore[import-not-found]
    except ModuleNotFoundError as exc:
        raise GoogleTrendsConfigError("缺少 openpyxl 依赖，无法导出 XLSX") from exc

    normalized_rows = [_normalize_row(row) for row in rows]
    columns = _columns_for_scenario(scenario, normalized_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_title(_sheet_title(scenario=scenario, geo=geo, params=params or {}, rows=rows))

    header_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True)
    for column_index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=column.title)
        cell.font = header_font
        cell.fill = header_fill

    for row_index, row in enumerate(normalized_rows, start=2):
        for column_index, column in enumerate(columns, start=1):
            value = _cell_value(_column_value(row, column))
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            _apply_number_format(cell)

    sheet.freeze_panes = "A2"
    for column_index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = _column_width(column.title)

    workbook.save(output_path)
    resolved = output_path.resolve()
    return GoogleTrendsExportResult(path=str(resolved), filename=resolved.name, url=resolved.as_uri())


def _columns_for_scenario(scenario: str, rows: list[dict[str, Any]]) -> list[ExportColumn]:
    if scenario == "interest-over-time":
        return _with_dynamic_keyword_columns(INTEREST_TIME_COLUMNS, rows, skip={"date", "isPartial"})
    if scenario == "interest-by-region":
        return _with_dynamic_keyword_columns(REGION_COLUMNS, rows, skip={"geo_name", "geoCode"})
    if scenario in {"related-queries", "related-topics"}:
        return RELATED_COLUMNS
    if scenario == "suggestions":
        return SUGGESTION_COLUMNS
    if scenario in {"trending-searches", "realtime-trending"}:
        return _merge_columns(TRENDING_COLUMNS, rows)
    return _generic_columns(rows)


def _with_dynamic_keyword_columns(base: list[ExportColumn], rows: list[dict[str, Any]], *, skip: set[str]) -> list[ExportColumn]:
    fields: list[str] = []
    for row in rows:
        for key, value in row.items():
            if key in skip or isinstance(value, (dict, list)):
                continue
            if key not in fields:
                fields.append(key)
    return [*base, *[ExportColumn(_title_for_field(field), field) for field in fields]]


def _merge_columns(base: list[ExportColumn], rows: list[dict[str, Any]]) -> list[ExportColumn]:
    seen = {column.source for column in base if column.source}
    extra = [ExportColumn(_title_for_field(field), field) for field in _fields(rows) if field not in seen]
    return [*base, *extra]


def _generic_columns(rows: list[dict[str, Any]]) -> list[ExportColumn]:
    fields = _fields(rows)
    return [ExportColumn(_title_for_field(field), field) for field in fields] or GENERIC_COLUMNS


def _fields(rows: list[dict[str, Any]]) -> list[str]:
    fields: list[str] = []
    for row in rows:
        for key, value in row.items():
            if isinstance(value, (dict, list)):
                continue
            if key not in fields:
                fields.append(key)
            if len(fields) >= 24:
                break
    return fields


def _normalize_row(row: Any) -> dict[str, Any]:
    if isinstance(row, dict):
        return row
    return {"value": row}


def _column_value(row: dict[str, Any], column: ExportColumn) -> Any:
    value = row if column.source is None else _get_value(row, column.source)
    fallback = column.fallback_source
    if _is_blank(value) and fallback:
        value = _get_value(row, fallback)
    if column.transform == "rowJson":
        return json.dumps(row, ensure_ascii=False)
    if column.transform == "bool":
        return "是" if bool(value) else "否"
    return value


def _get_value(row: dict[str, Any], field: str) -> Any:
    value: Any = row
    for part in field.split("."):
        if isinstance(value, dict):
            value = value.get(part)
        elif isinstance(value, list) and part.isdigit():
            index = int(part)
            value = value[index] if index < len(value) else None
        else:
            return None
    return value


def _cell_value(value: Any) -> Any:
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _apply_number_format(cell) -> None:
    value = cell.value
    if isinstance(value, bool) or not isinstance(value, Real):
        return
    cell.number_format = "#,##0" if float(value).is_integer() else "#,##0.00"


def _sheet_title(*, scenario: str, geo: str, params: dict[str, Any], rows: list[Any]) -> str:
    target = params.get("keyword") or _first(params.get("keywords")) or _first(params.get("kw_list")) or ""
    suffix = f"-{target}" if target else ""
    return f"Trends-{geo or 'GLOBAL'}-{scenario}{suffix}({len(rows)})"


def _safe_sheet_title(value: str) -> str:
    title = "".join(char for char in value if char not in r"[]:*?/\\")
    return (title or "GoogleTrends")[:31]


def _column_width(title: str) -> int:
    if any(key in title for key in ["相关", "原始", "搜索词", "标题"]):
        return 38
    if any(key in title for key in ["日期", "地区", "关键词", "topic"]):
        return 20
    return max(12, min(24, len(str(title)) * 2 + 4))


def _title_for_field(field: str) -> str:
    titles = {
        "date": "日期",
        "geo_name": "地区",
        "geoCode": "地区代码",
        "keyword": "关键词",
        "query": "相关查询",
        "topic_title": "相关主题",
        "value": "值",
        "rank": "排名",
        "search_term": "搜索词",
        "traffic": "traffic",
        "isPartial": "是否部分数据",
    }
    return titles.get(field, field)


def _first(value: Any) -> Any:
    if isinstance(value, list) and value:
        return value[0]
    return value


def _is_blank(value: Any) -> bool:
    return value is None or value == ""
