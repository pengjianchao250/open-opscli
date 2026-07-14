"""ASIN batch data collection CLI."""

from __future__ import annotations

import asyncio
import json
import shutil
import time
from io import BytesIO
from enum import Enum
from pathlib import Path
from typing import Any

import httpx
import typer

from opscli.asin_data.services.bi_report_data import (
    BASIC_REPORT_SOURCE_KEYS,
    BI_ONLY_REPORT_SOURCE_KEYS,
    LISTING_REPORT_SOURCE_KEYS,
)
from opscli.asin_data.services.category_top import AsinCategoryTopService
from opscli.asin_data.services.collector import (
    DEFAULT_LISTING_ANALYSIS_POLL_ATTEMPTS,
    DEFAULT_LISTING_ANALYSIS_POLL_INTERVAL_SECONDS,
    AsinDataCollector,
)
from opscli.asin_data.services.daily_pipeline import DailyAsinDataPipeline
from opscli.asin_data.services.live_data import AsinLiveDataService, fetch_split_file
from opscli.asin_data.services.report_file_submitter import (
    DEFAULT_REPORT_TYPE,
    DEFAULT_SOURCE,
    AsinReportFileSubmitter,
)
from opscli.asin_data.services.report_files import AsinReportFileClient, AsinReportFileNotFoundError
from opscli.asin_data.services.split_package_builder import (
    FILE_FIELD_MAP,
    SPLIT_FILE_KEYS,
)
from opscli.asin_data.services.yicopy_keyword_engine import (
    YicopyKeywordEngine,
    YicopyRunOptions,
    build_yicopy_ai_ready_response,
    load_source_tokens_from_file,
    normalize_yicopy_result_format,
    render_yicopy_result,
)
from opscli.shared.file_uploads import FileUploadClient



class KeywordSource(str, Enum):
    input_only = "input_only"
    reverse_top = "reverse_top"
    skip = "skip"


class FieldMode(str, Enum):
    full = "full"
    compatible = "compatible"


class DailyStage(str, Enum):
    query = "query"
    bi = "bi"
    basic = "basic"
    seller_keyword_reverse = "seller-keyword-reverse"
    seller_keyword_miner = "seller-keyword-miner"
    seller_listing_analysis = "seller-listing-analysis"
    rufus = "rufus"


class RufusBatchMode(str, Enum):
    fast = "fast"
    balanced = "balanced"
    safe = "safe"


class FileKey(str, Enum):
    """Split package file keys exposed to AI per-file delivery."""

    basic = "basic"
    bi = "bi"
    keyword_reverse = "keyword_reverse"
    keyword_miner = "keyword_miner"
    competitor = "competitor"
    rufus = "rufus"


class LiveDataScope(str, Enum):
    """实时数据范围。"""

    all = "all"
    basic = "basic"
    bi = "bi"
    listing = "listing"
    listing_basic = "listing_basic"


class LiveDataReturnMode(str, Enum):
    """实时取数返回模式。"""

    content = "content"
    url_only = "url_only"
    both = "both"
    ai_ready = "ai_ready"


app = typer.Typer(help="ASIN 批量取数服务")


@app.callback()
def main() -> None:
    """ASIN 批量取数命令组入口。"""


def _emit(payload: dict, pretty: bool) -> None:
    if pretty:
        typer.echo(json.dumps(payload, ensure_ascii=False, indent=2))
    else:
        typer.echo(json.dumps(payload, ensure_ascii=False))


def _error_payload(command: str, exc: Exception) -> dict:
    if hasattr(exc, "to_dict"):
        error = exc.to_dict()  # type: ignore[call-arg]
    else:
        error = {"code": "ASIN_DATA_ERROR", "message": str(exc)}
    return {"success": False, "command": command, "data": None, "error": error}


def _load_frontend_data(output_dir: str) -> dict[str, Any]:
    """读取本次实时采集生成的前端 JSON 数据。"""
    frontend_path = Path(output_dir) / "frontend-data.json"
    return json.loads(frontend_path.read_text(encoding="utf-8"))


def _load_live_split_files(result: dict[str, Any], file_keys: tuple[str, ...] = ("basic", "bi")) -> dict[str, dict[str, Any]]:
    """读取实时采集生成的 ASIN 拆包文件内容。"""
    manifest = result.get("manifest") if isinstance(result.get("manifest"), dict) else {}
    package = manifest.get("asin_data_package") if isinstance(manifest.get("asin_data_package"), dict) else {}
    items = package.get("items") if isinstance(package.get("items"), list) else []
    split_files: dict[str, dict[str, Any]] = {}
    for item in items:
        if not isinstance(item, dict):
            continue
        asin = str(item.get("asin") or "").strip().upper()
        files = item.get("files") if isinstance(item.get("files"), dict) else {}
        if not asin:
            continue
        asin_files: dict[str, Any] = {}
        for file_key in file_keys:
            path_text = files.get(file_key)
            if not isinstance(path_text, str) or not path_text.strip():
                continue
            path = Path(path_text)
            asin_files[file_key] = {
                "asin": asin,
                "file_key": file_key,
                "file_path": path.as_posix(),
                "content": _read_local_split_file_content(path, file_key),
            }
        if asin_files:
            split_files[asin] = asin_files
    return split_files


def _read_local_split_file_content(path: Path, file_key: str) -> Any:
    """读取本地拆包文件，并返回与 fetch-file 一致的结构化内容。"""
    if file_key == "rufus":
        return path.read_text(encoding="utf-8", errors="replace")
    return _read_xlsx_content(path)


def _upload_live_split_files(split_files: dict[str, dict[str, Any]], *, run_id: str) -> dict[str, Any]:
    """上传实时 basic/bi xlsx，并把 OSS 地址写回 split_files。"""
    client = FileUploadClient()
    items: list[dict[str, Any]] = []
    files_uploaded = 0
    for asin, files in split_files.items():
        uploaded: dict[str, Any] = {}
        for file_key in ("basic", "bi"):
            file_item = files.get(file_key)
            if not isinstance(file_item, dict):
                continue
            path_text = file_item.get("file_path")
            if not isinstance(path_text, str) or not path_text.strip():
                continue
            path = Path(path_text)
            upload_path = path.with_name(f"{asin}-{file_key}-live-data.xlsx")
            if upload_path != path:
                shutil.copyfile(path, upload_path)
            upload = client.upload(
                upload_path,
                purpose="asin_data_live_xlsx",
                folder="asin-data",
                public="1",
                metadata={
                    "run_id": run_id,
                    "asin": asin,
                    "file_key": file_key,
                    "report_filename": upload_path.name,
                    "source_filename": path.name,
                    "source": "asin-data live-data",
                },
            )
            file_item["file_url"] = upload.url
            file_item["upload"] = {"url": upload.url, "raw": upload.raw}
            uploaded[file_key] = {
                "url": upload.url,
                "file_name": upload_path.name,
                "file_path": path.as_posix(),
                "upload_path": upload_path.as_posix(),
            }
            files_uploaded += 1
        items.append({"asin": asin, "files": uploaded})
    return {"files_uploaded": files_uploaded, "items": items}


def _split_file_urls(split_files: dict[str, dict[str, Any]]) -> dict[str, dict[str, str]]:
    """从 split_files 中提取轻量 OSS URL 映射。"""
    urls: dict[str, dict[str, str]] = {}
    for asin, files in split_files.items():
        asin_urls = {
            file_key: str(file_item.get("file_url"))
            for file_key, file_item in files.items()
            if isinstance(file_item, dict) and file_item.get("file_url")
        }
        if asin_urls:
            urls[asin] = asin_urls
    return urls


def _live_data_source_keys(scope: LiveDataScope) -> tuple[str, ...] | None:
    """把 live-data 范围参数映射到实时 BI 接口 source key。"""
    if scope == LiveDataScope.basic:
        return BASIC_REPORT_SOURCE_KEYS
    if scope in {LiveDataScope.listing, LiveDataScope.listing_basic}:
        return LISTING_REPORT_SOURCE_KEYS
    if scope == LiveDataScope.bi:
        return BI_ONLY_REPORT_SOURCE_KEYS
    return None


def _live_data_file_keys(scope: LiveDataScope) -> tuple[str, ...]:
    """把 live-data 范围参数映射到需要返回/上传的拆包文件。"""
    if scope in {LiveDataScope.basic, LiveDataScope.listing, LiveDataScope.listing_basic}:
        return ("basic",)
    if scope == LiveDataScope.bi:
        return ("bi",)
    return ("basic", "bi")


@app.command("report-url")
def report_url(
    asin: str = typer.Option(..., "--asin", help="ASIN"),
    site: str = typer.Option("US", "--site", help="站点"),
    url_only: bool = typer.Option(False, "--url-only", help="只输出报告文件地址"),
    pretty: bool = typer.Option(False, "--pretty", help="格式化输出 JSON"),
) -> None:
    """只查询 ASIN 报告文件接口并返回报告地址。"""
    normalized_asin = asin.strip().upper()
    normalized_site = site.strip().upper()
    try:
        report_file = AsinReportFileClient().fetch(asin=normalized_asin, site=normalized_site)
        if not report_file.url:
            raise AsinReportFileNotFoundError(asin=normalized_asin, site=normalized_site)
    except Exception as exc:
        _emit(_error_payload("asin-data report-url", exc), pretty)
        raise typer.Exit(1)

    if url_only:
        typer.echo(report_file.url)
        return

    _emit(
        {
            "success": True,
            "command": "asin-data report-url",
            "data": {
                "asin": report_file.asin,
                "site": report_file.site,
                "report_file_url": report_file.url,
                "record": report_file.record,
                "raw": report_file.raw,
            },
            "error": None,
        },
        pretty,
    )


@app.command("abtest-url")
def abtest_url(
    asin: str = typer.Option(..., "--asin", help="ASIN"),
    site: str = typer.Option("US", "--site", help="站点"),
    data_type: str = typer.Option("file", "--data-type", help="返回数据类型，默认 file 取报告文件地址"),
    url_only: bool = typer.Option(False, "--url-only", help="只输出 ABTest 报告文件地址"),
    pretty: bool = typer.Option(False, "--pretty", help="格式化输出 JSON"),
) -> None:
    """只查询 ABTest 报告文件接口并返回报告地址。"""
    normalized_asin = asin.strip().upper()
    normalized_site = site.strip().upper()
    try:
        report_file = AsinReportFileClient().fetch_abtest(
            asin=normalized_asin, site=normalized_site, data_type=data_type
        )
        if not report_file.url:
            raise AsinReportFileNotFoundError(asin=normalized_asin, site=normalized_site)
    except Exception as exc:
        _emit(_error_payload("asin-data abtest-url", exc), pretty)
        raise typer.Exit(1)

    if url_only:
        typer.echo(report_file.url)
        return

    _emit(
        {
            "success": True,
            "command": "asin-data abtest-url",
            "data": {
                "asin": report_file.asin,
                "site": report_file.site,
                "data_type": data_type,
                "abtest_report_url": report_file.url,
                "record": report_file.record,
                "raw": report_file.raw,
            },
            "error": None,
        },
        pretty,
    )


@app.command("file-url")
def file_url(
    asin: str = typer.Option(..., "--asin", help="ASIN"),
    site: str = typer.Option("US", "--site", help="站点"),
    file: FileKey | None = typer.Option(
        None, "--file", help="文件类型；不传时配合 --list 列出全部"
    ),
    list_all: bool = typer.Option(False, "--list", help="列出该 ASIN 所有可用的拆分文件地址"),
    url_only: bool = typer.Option(False, "--url-only", help="只输出文件地址"),
    pretty: bool = typer.Option(False, "--pretty", help="格式化输出 JSON"),
) -> None:
    """查询 ASIN 拆分数据包中某个文件的 OSS 地址（按文件粒度）。"""
    normalized_asin = asin.strip().upper()
    normalized_site = site.strip().upper()
    client = AsinReportFileClient()
    try:
        result = client.fetch_split_files(asin=normalized_asin, site=normalized_site)
        files = result.get("files") or {}
        if list_all or file is None:
            data = [
                {"file_key": key, "url": value}
                for key, value in files.items()
            ]
            if url_only:
                lines: list[str] = []
                for item in data:
                    urls = item["url"]
                    if isinstance(urls, list):
                        lines.extend(urls)
                    elif urls:
                        lines.append(str(urls))
                typer.echo("\n".join(lines))
                return
            _emit(
                {
                    "success": True,
                    "command": "asin-data file-url",
                    "data": {
                        "asin": normalized_asin,
                        "site": normalized_site,
                        "files": data,
                        "record": result.get("record"),
                    },
                    "error": None,
                },
                pretty,
            )
            return

        urls = files.get(file.value)
        if not urls:
            raise AsinReportFileNotFoundError(asin=normalized_asin, site=normalized_site)
    except Exception as exc:
        _emit(_error_payload("asin-data file-url", exc), pretty)
        raise typer.Exit(1)

    if url_only:
        if isinstance(urls, list):
            typer.echo("\n".join(urls))
        else:
            typer.echo(urls)
        return

    _emit(
        {
            "success": True,
            "command": "asin-data file-url",
            "data": {
                "asin": normalized_asin,
                "site": normalized_site,
                "file_key": file.value,
                "file_url": urls,
                "record": result.get("record"),
            },
            "error": None,
        },
        pretty,
    )


@app.command("fetch-file")
def fetch_file(
    asin: str = typer.Option(..., "--asin", help="ASIN"),
    file: FileKey = typer.Option(..., "--file", help="文件类型"),
    site: str = typer.Option("US", "--site", help="站点"),
    pretty: bool = typer.Option(False, "--pretty", help="格式化输出 JSON"),
) -> None:
    """下载并返回 ASIN 拆分数据包中某个文件的内容（xlsx 转 JSON，md 输出文本）。"""
    normalized_asin = asin.strip().upper()
    normalized_site = site.strip().upper()
    try:
        data = fetch_split_file(
            asin=normalized_asin,
            site=normalized_site,
            file_key=file.value,
        )
    except Exception as exc:
        _emit(_error_payload("asin-data fetch-file", exc), pretty)
        raise typer.Exit(1)

    _emit(
        {
            "success": True,
            "command": "asin-data fetch-file",
            "data": data,
            "error": None,
        },
        pretty,
    )


@app.command("live-data")
def live_data(
    input_path: str | None = typer.Option(None, "--input", "-i", help="CSV/XLSX/JSON/JSONL 输入文件"),
    asin: str | None = typer.Option(None, "--asin", help="单个 ASIN；与 --input 二选一"),
    keywords: list[str] | None = typer.Option(None, "--keyword", help="单个 ASIN 的关键词，可重复传入"),
    asin_column: str = typer.Option("asin", "--asin-column", help="ASIN 列名"),
    keyword_column: str = typer.Option("keyword", "--keyword-column", help="关键词列名"),
    site_column: str = typer.Option("site", "--site-column", help="站点列名"),
    site: str = typer.Option("US", "--site", help="默认站点，同时作为 crawler-details 的 country"),
    output_dir: str = typer.Option("output/asin-data", "--output-dir", help="输出目录"),
    run_id: str | None = typer.Option(None, "--run-id", help="本次运行 ID"),
    dry_run: bool = typer.Option(False, "--dry-run", help="只生成计划与前端文件，不执行远程取数"),
    skip_query: bool = typer.Option(True, "--skip-query/--no-skip-query", help="跳过旧 BI query 取数；live-data 默认走实时接口"),
    skip_bi_report_data: bool = typer.Option(False, "--skip-bi-report-data", help="跳过 ASIN 报告 BI 接口取数"),
    skip_sales_query: bool = typer.Option(False, "--skip-sales-query", help="跳过销售数据 query"),
    skip_crawler_query: bool = typer.Option(False, "--skip-crawler-query", help="跳过旧爬虫 Listing query"),
    legacy_crawler_query: bool = typer.Option(False, "--legacy-crawler-query", help="启用旧爬虫 Listing query"),
    sales_table_id: int | None = typer.Option(None, "--sales-table-id", help="BI 销售数据 table_id"),
    sales_dataset_alias: str = typer.Option("ds_d35ac6f3910c", "--sales-dataset-alias", help="BI 销售数据 dataset alias"),
    sales_field_mode: FieldMode = typer.Option(FieldMode.full, "--sales-field-mode", help="销售字段模式"),
    sales_start: str | None = typer.Option(None, "--sales-start", help="销售开始日期"),
    sales_end: str | None = typer.Option(None, "--sales-end", help="销售结束日期"),
    query_chunk_size: int = typer.Option(100, "--query-chunk-size", min=1, help="query 每批 ASIN 数量"),
    crawler_table_id: int | None = typer.Option(None, "--crawler-table-id", help="爬虫 Listing table_id"),
    crawler_dataset_alias: str = typer.Option("ds_icw50TLOFu4F", "--crawler-dataset-alias", help="爬虫 Listing dataset alias"),
    crawler_field_mode: FieldMode = typer.Option(FieldMode.full, "--crawler-field-mode", help="爬虫 Listing 字段模式"),
    data_scope: LiveDataScope = typer.Option(LiveDataScope.all, "--data-scope", help="实时数据范围：all=基础+BI，basic=完整基础(刊登+爬虫)，listing/listing_basic=仅刊登，bi=仅BI"),
    upload_xlsx: bool = typer.Option(False, "--upload-xlsx/--no-upload-xlsx", help="上传实时生成的基础/BI xlsx 到 OSS 并返回 file_url"),
    return_mode: LiveDataReturnMode = typer.Option(
        LiveDataReturnMode.content,
        "--return-mode",
        help="返回模式：content=返回内联 JSON，url_only=仅返回 xlsx URL，both=内容和 URL 都返回，ai_ready=返回文件索引、数据集预览和诊断",
    ),
    pretty: bool = typer.Option(False, "--pretty", help="格式化输出 JSON"),
) -> None:
    """实时获取 ASIN 基础刊登数据与 BI 数据，并直接返回前端 JSON。"""
    try:
        result = AsinLiveDataService(
            collector=AsinDataCollector(),
            file_upload_client_factory=FileUploadClient,
        ).run(
            input_path=input_path,
            asin=asin,
            keywords=keywords,
            asin_column=asin_column,
            keyword_column=keyword_column,
            site_column=site_column,
            site=site,
            output_dir=output_dir,
            run_id=run_id,
            dry_run=dry_run,
            skip_query=skip_query,
            skip_bi_report_data=skip_bi_report_data,
            skip_sales_query=skip_sales_query,
            skip_crawler_query=skip_crawler_query,
            legacy_crawler_query=legacy_crawler_query,
            sales_table_id=sales_table_id,
            sales_dataset_alias=sales_dataset_alias,
            sales_field_mode=sales_field_mode.value,
            sales_start=sales_start,
            sales_end=sales_end,
            query_chunk_size=query_chunk_size,
            crawler_table_id=crawler_table_id,
            crawler_dataset_alias=crawler_dataset_alias,
            crawler_field_mode=crawler_field_mode.value,
            data_scope=data_scope.value,
            upload_xlsx=upload_xlsx,
            return_mode=return_mode.value,
        )
    except Exception as exc:
        _emit(_error_payload("asin-data live-data", exc), pretty)
        raise typer.Exit(1)

    _emit({"success": True, "command": "asin-data live-data", "data": result, "error": None}, pretty)


@app.command("abtest-url")
def abtest_url(
    asin: str = typer.Option(..., "--asin", help="ASIN"),
    site: str = typer.Option("US", "--site", help="站点"),
    data_type: str = typer.Option("file", "--data-type", help="返回数据类型，默认 file 取报告文件地址"),
    url_only: bool = typer.Option(False, "--url-only", help="只输出 ABTest 报告文件地址"),
    pretty: bool = typer.Option(False, "--pretty", help="格式化输出 JSON"),
) -> None:
    """只查询 ABTest 报告文件接口并返回报告地址。"""
    normalized_asin = asin.strip().upper()
    normalized_site = site.strip().upper()
    try:
        report_file = AsinReportFileClient().fetch_abtest(
            asin=normalized_asin, site=normalized_site, data_type=data_type
        )
        if not report_file.url:
            raise AsinReportFileNotFoundError(asin=normalized_asin, site=normalized_site)
    except Exception as exc:
        _emit(_error_payload("asin-data abtest-url", exc), pretty)
        raise typer.Exit(1)

    if url_only:
        typer.echo(report_file.url)
        return

    _emit(
        {
            "success": True,
            "command": "asin-data abtest-url",
            "data": {
                "asin": report_file.asin,
                "site": report_file.site,
                "data_type": data_type,
                "abtest_report_url": report_file.url,
                "record": report_file.record,
                "raw": report_file.raw,
            },
            "error": None,
        },
        pretty,
    )


@app.command("file-url")
def file_url(
    asin: str = typer.Option(..., "--asin", help="ASIN"),
    site: str = typer.Option("US", "--site", help="站点"),
    file: FileKey | None = typer.Option(
        None, "--file", help="文件类型；不传时配合 --list 列出全部"
    ),
    list_all: bool = typer.Option(False, "--list", help="列出该 ASIN 所有可用的拆分文件地址"),
    url_only: bool = typer.Option(False, "--url-only", help="只输出文件地址"),
    pretty: bool = typer.Option(False, "--pretty", help="格式化输出 JSON"),
) -> None:
    """查询 ASIN 拆分数据包中某个文件的 OSS 地址（按文件粒度）。"""
    normalized_asin = asin.strip().upper()
    normalized_site = site.strip().upper()
    client = AsinReportFileClient()
    try:
        result = client.fetch_split_files(asin=normalized_asin, site=normalized_site)
        files = result.get("files") or {}
        if list_all or file is None:
            data = [
                {"file_key": key, "url": value}
                for key, value in files.items()
            ]
            if url_only:
                lines: list[str] = []
                for item in data:
                    urls = item["url"]
                    if isinstance(urls, list):
                        lines.extend(urls)
                    elif urls:
                        lines.append(str(urls))
                typer.echo("\n".join(lines))
                return
            _emit(
                {
                    "success": True,
                    "command": "asin-data file-url",
                    "data": {
                        "asin": normalized_asin,
                        "site": normalized_site,
                        "files": data,
                        "record": result.get("record"),
                    },
                    "error": None,
                },
                pretty,
            )
            return

        urls = files.get(file.value)
        if not urls:
            raise AsinReportFileNotFoundError(asin=normalized_asin, site=normalized_site)
    except Exception as exc:
        _emit(_error_payload("asin-data file-url", exc), pretty)
        raise typer.Exit(1)

    if url_only:
        if isinstance(urls, list):
            typer.echo("\n".join(urls))
        else:
            typer.echo(urls)
        return

    _emit(
        {
            "success": True,
            "command": "asin-data file-url",
            "data": {
                "asin": normalized_asin,
                "site": normalized_site,
                "file_key": file.value,
                "file_url": urls,
                "record": result.get("record"),
            },
            "error": None,
        },
        pretty,
    )


@app.command("fetch-file")
def fetch_file(
    asin: str = typer.Option(..., "--asin", help="ASIN"),
    file: FileKey = typer.Option(..., "--file", help="文件类型"),
    site: str = typer.Option("US", "--site", help="站点"),
    pretty: bool = typer.Option(False, "--pretty", help="格式化输出 JSON"),
) -> None:
    """下载并返回 ASIN 拆分数据包中某个文件的内容（xlsx 转 JSON，md 输出文本）。"""
    normalized_asin = asin.strip().upper()
    normalized_site = site.strip().upper()
    try:
        result = AsinReportFileClient().fetch_split_files(
            asin=normalized_asin, site=normalized_site
        )
        files = result.get("files") or {}
        urls = files.get(file.value)
        if not urls:
            raise AsinReportFileNotFoundError(asin=normalized_asin, site=normalized_site)
        first_url = urls[0] if isinstance(urls, list) else urls
        content = _download_file_content(first_url, file.value)
    except Exception as exc:
        _emit(_error_payload("asin-data fetch-file", exc), pretty)
        raise typer.Exit(1)

    _emit(
        {
            "success": True,
            "command": "asin-data fetch-file",
            "data": {
                "asin": normalized_asin,
                "site": normalized_site,
                "file_key": file.value,
                "file_url": urls,
                "content": content,
            },
            "error": None,
        },
        pretty,
    )


def _download_file_content(url: str, file_key: str) -> Any:
    """下载拆包文件，并返回结构化内容。"""
    response = httpx.get(url, timeout=60, follow_redirects=True)
    response.raise_for_status()
    raw = response.content
    if file_key == "rufus":
        return raw.decode("utf-8", errors="replace")
    return _read_xlsx_content(BytesIO(raw))


def _read_xlsx_content(source: Any) -> dict[str, list[list[Any]]]:
    """将 xlsx 转为 fetch-file 使用的 {sheet_name: [rows]} JSON 结构。"""
    from openpyxl import load_workbook

    wb = load_workbook(source, read_only=True, data_only=True)
    sheets: dict[str, list[list[Any]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        sheets[sheet_name] = rows
    wb.close()
    return sheets


@app.command("live-data")
def live_data(
    input_path: str | None = typer.Option(None, "--input", "-i", help="CSV/XLSX/JSON/JSONL 输入文件"),
    asin: str | None = typer.Option(None, "--asin", help="单个 ASIN；与 --input 二选一"),
    keywords: list[str] | None = typer.Option(None, "--keyword", help="单个 ASIN 的关键词，可重复传入"),
    asin_column: str = typer.Option("asin", "--asin-column", help="ASIN 列名"),
    keyword_column: str = typer.Option("keyword", "--keyword-column", help="关键词列名"),
    site_column: str = typer.Option("site", "--site-column", help="站点列名"),
    site: str = typer.Option("US", "--site", help="默认站点，同时作为 crawler-details 的 country"),
    output_dir: str = typer.Option("output/asin-data", "--output-dir", help="输出目录"),
    run_id: str | None = typer.Option(None, "--run-id", help="本次运行 ID"),
    dry_run: bool = typer.Option(False, "--dry-run", help="只生成计划与前端文件，不执行远程取数"),
    skip_query: bool = typer.Option(True, "--skip-query/--no-skip-query", help="跳过旧 BI query 取数；live-data 默认走实时接口"),
    skip_bi_report_data: bool = typer.Option(False, "--skip-bi-report-data", help="跳过 ASIN 报告 BI 接口取数"),
    skip_sales_query: bool = typer.Option(False, "--skip-sales-query", help="跳过销售数据 query"),
    skip_crawler_query: bool = typer.Option(False, "--skip-crawler-query", help="跳过旧爬虫 Listing query"),
    legacy_crawler_query: bool = typer.Option(False, "--legacy-crawler-query", help="启用旧爬虫 Listing query"),
    sales_table_id: int | None = typer.Option(None, "--sales-table-id", help="BI 销售数据 table_id"),
    sales_dataset_alias: str = typer.Option("ds_d35ac6f3910c", "--sales-dataset-alias", help="BI 销售数据 dataset alias"),
    sales_field_mode: FieldMode = typer.Option(FieldMode.full, "--sales-field-mode", help="销售字段模式"),
    sales_start: str | None = typer.Option(None, "--sales-start", help="销售开始日期"),
    sales_end: str | None = typer.Option(None, "--sales-end", help="销售结束日期"),
    query_chunk_size: int = typer.Option(100, "--query-chunk-size", min=1, help="query 每批 ASIN 数量"),
    crawler_table_id: int | None = typer.Option(None, "--crawler-table-id", help="爬虫 Listing table_id"),
    crawler_dataset_alias: str = typer.Option("ds_icw50TLOFu4F", "--crawler-dataset-alias", help="爬虫 Listing dataset alias"),
    crawler_field_mode: FieldMode = typer.Option(FieldMode.full, "--crawler-field-mode", help="爬虫 Listing 字段模式"),
    data_scope: LiveDataScope = typer.Option(LiveDataScope.all, "--data-scope", help="实时数据范围：all=基础+BI，basic=完整基础(刊登+爬虫)，listing/listing_basic=仅刊登，bi=仅BI"),
    upload_xlsx: bool = typer.Option(False, "--upload-xlsx/--no-upload-xlsx", help="上传实时生成的基础/BI xlsx 到 OSS 并返回 file_url"),
    return_mode: LiveDataReturnMode = typer.Option(
        LiveDataReturnMode.content,
        "--return-mode",
        help="返回模式：content=返回内联 JSON，url_only=仅返回 xlsx URL，both=内容和 URL 都返回，ai_ready=返回文件索引、数据集预览和诊断",
    ),
    pretty: bool = typer.Option(False, "--pretty", help="格式化输出 JSON"),
) -> None:
    """实时获取 ASIN 基础刊登数据与 BI 数据，并直接返回前端 JSON。"""
    try:
        result = AsinLiveDataService(
            collector=AsinDataCollector(),
            file_upload_client_factory=FileUploadClient,
        ).run(
            input_path=input_path,
            asin=asin,
            keywords=keywords,
            asin_column=asin_column,
            keyword_column=keyword_column,
            site_column=site_column,
            site=site,
            output_dir=output_dir,
            run_id=run_id,
            dry_run=dry_run,
            skip_query=skip_query,
            skip_bi_report_data=skip_bi_report_data,
            skip_sales_query=skip_sales_query,
            skip_crawler_query=skip_crawler_query,
            legacy_crawler_query=legacy_crawler_query,
            sales_table_id=sales_table_id,
            sales_dataset_alias=sales_dataset_alias,
            sales_field_mode=sales_field_mode.value,
            sales_start=sales_start,
            sales_end=sales_end,
            query_chunk_size=query_chunk_size,
            crawler_table_id=crawler_table_id,
            crawler_dataset_alias=crawler_dataset_alias,
            crawler_field_mode=crawler_field_mode.value,
            data_scope=data_scope.value,
            upload_xlsx=upload_xlsx,
            return_mode=return_mode.value,
        )
    except Exception as exc:
        _emit(_error_payload("asin-data live-data", exc), pretty)
        raise typer.Exit(1)

    _emit({"success": True, "command": "asin-data live-data", "data": result, "error": None}, pretty)


@app.command("category-top")
def category_top(
    category: str = typer.Option(..., "--category", help="平台类目名称，精确匹配 amazon_cat"),
    date_from: str | None = typer.Option(None, "--date-from", help="起始日期 YYYY-MM-DD，默认由后端使用当月 1 日"),
    date_to: str | None = typer.Option(None, "--date-to", help="截止日期 YYYY-MM-DD，默认由后端使用当天"),
    limit: int = typer.Option(10, "--limit", min=1, max=100, help="返回 Top ASIN 数量，范围 1-100"),
    site: str = typer.Option(
        "US",
        "--site",
        help="无法从渠道推断站点时使用的默认站点，同时作为 crawler-details 的 country",
    ),
    output_dir: str = typer.Option("output/asin-data", "--output-dir", help="输出目录"),
    run_id: str | None = typer.Option(None, "--run-id", help="本次运行 ID"),
    upload: bool = typer.Option(True, "--upload/--no-upload", help="是否上传合并后的 JSON 文件到 OSS"),
    enrich: bool = typer.Option(True, "--enrich/--no-enrich", help="是否补充查询刊登基础数据和爬虫详情数据"),
    return_content: bool = typer.Option(False, "--return-content", help="是否在命令结果中返回完整 JSON 内容"),
    pretty: bool = typer.Option(False, "--pretty", help="格式化输出 JSON"),
) -> None:
    """查询内部类目 Top ASIN，并合并刊登基础数据和爬虫详情为单个 OSS JSON 文件。"""
    try:
        result = AsinCategoryTopService().run(
            category=category,
            date_from=date_from,
            date_to=date_to,
            limit=limit,
            site=site,
            output_dir=output_dir,
            run_id=run_id,
            upload=upload,
            enrich=enrich,
            return_content=return_content,
        )
    except Exception as exc:
        _emit(_error_payload("asin-data category-top", exc), pretty)
        raise typer.Exit(1)

    _emit({"success": True, "command": "asin-data category-top", "data": result, "error": None}, pretty)


@app.command("yicopy-keyword-engine")
def yicopy_keyword_engine(
    asin: list[str] | None = typer.Option(None, "--asin", "-a", help="ASIN 或包含 ASIN 的文本，可重复传入。"),
    url: list[str] | None = typer.Option(None, "--url", "-u", help="Amazon 商品详情页 URL，可重复传入。"),
    input_file: Path | None = typer.Option(None, "--input-file", "-i", help="包含 ASIN/URL 的 JSON、JSON 数组或文本文件。"),
    site: str = typer.Option("US", "--site", help="Amazon 站点代码，默认 US。"),
    locale: str = typer.Option("en_US", "--locale", help="Amazon completion API locale，默认 en_US。"),
    result_format: str = typer.Option(
        "keyword-reverse",
        "--result-format",
        help="输出格式：keyword-reverse 输出示例一致纯数组；full 输出全链路调试数据。",
    ),
    max_asins: int | None = typer.Option(None, "--max-asins", help="最多处理多少个 ASIN。"),
    max_prefixes_per_asin: int | None = typer.Option(None, "--max-prefixes-per-asin", help="每个 ASIN 最多查询多少个标题前缀。"),
    completion_limit: int = typer.Option(11, "--completion-limit", help="Amazon 自动补全每次返回上限。"),
    timeout_seconds: float = typer.Option(30.0, "--timeout-seconds", help="HTTP 请求超时秒数。"),
    request_delay_seconds: float = typer.Option(0.0, "--request-delay-seconds", help="每次补全请求后的等待秒数。"),
    output_file: Path | None = typer.Option(None, "--output-file", "--output", "-o", help="把 JSON 结果写入 UTF-8 文件。"),
    pretty: bool = typer.Option(False, "--pretty", help="格式化输出 JSON"),
) -> None:
    """执行 yicopy 销词引擎流程，返回关键词反查和词频结果。"""

    started = time.perf_counter()
    try:
        sources: list[str] = []
        sources.extend(asin or [])
        sources.extend(url or [])
        if input_file is not None:
            sources.extend(load_source_tokens_from_file(input_file))
        if not sources:
            raise ValueError("请通过 --asin、--url 或 --input-file 传入至少一个 ASIN 或 URL。")

        normalized_format = normalize_yicopy_result_format(result_format)
        result = asyncio.run(
            YicopyKeywordEngine().run(
                sources,
                YicopyRunOptions(
                    site=site,
                    locale=locale,
                    timeout_seconds=timeout_seconds,
                    request_delay_seconds=request_delay_seconds,
                    max_asins=max_asins,
                    max_prefixes_per_asin=max_prefixes_per_asin,
                    completion_limit=completion_limit,
                ),
            )
        )
        rendered = render_yicopy_result(result, normalized_format)
        output_path: str | None = None
        if output_file is not None:
            output_file.parent.mkdir(parents=True, exist_ok=True)
            output_file.write_text(json.dumps(rendered, indent=2, ensure_ascii=False), encoding="utf-8")
            output_path = str(output_file)

        request = {
            "asin": asin or [],
            "url": url or [],
            "input_file": str(input_file) if input_file is not None else None,
            "site": site,
            "locale": locale,
            "result_format": normalized_format,
            "max_asins": max_asins,
            "max_prefixes_per_asin": max_prefixes_per_asin,
            "completion_limit": completion_limit,
            "timeout_seconds": timeout_seconds,
            "request_delay_seconds": request_delay_seconds,
            "output_file": output_path,
        }
        data = build_yicopy_ai_ready_response(
            tool_name="asin-data yicopy-keyword-engine",
            request=request,
            result=result,
            rendered_result=rendered,
            result_format=normalized_format,
            site=site,
            output_file=output_path,
            elapsed_seconds=time.perf_counter() - started,
        )
        if output_file is None:
            data["result"] = rendered
        else:
            data.pop("result", None)
    except Exception as exc:
        _emit(_error_payload("asin-data yicopy-keyword-engine", exc), pretty)
        raise typer.Exit(1)

    _emit({"success": True, "command": "asin-data yicopy-keyword-engine", "data": data, "error": None}, pretty)


@app.command("collect")
def collect(
    input_path: str | None = typer.Option(None, "--input", "-i", help="CSV/XLSX/JSON/JSONL 输入文件"),
    asin: str | None = typer.Option(None, "--asin", help="单个 ASIN；与 --input 二选一"),
    keywords: list[str] | None = typer.Option(None, "--keyword", help="单个 ASIN 的关键词，可重复传入"),
    asin_column: str = typer.Option("asin", "--asin-column", help="ASIN 列名"),
    keyword_column: str = typer.Option("keyword", "--keyword-column", help="关键词列名"),
    site_column: str = typer.Option("site", "--site-column", help="站点列名"),
    site: str = typer.Option("US", "--site", help="默认站点"),
    output_dir: str = typer.Option("output/asin-data", "--output-dir", help="输出目录"),
    run_id: str | None = typer.Option(None, "--run-id", help="本次运行 ID"),
    dry_run: bool = typer.Option(False, "--dry-run", help="只生成计划与前端文件，不执行远程取数"),
    skip_seller_sprite: bool = typer.Option(False, "--skip-seller-sprite", help="跳过卖家精灵数据"),
    skip_keyword_miner: bool = typer.Option(False, "--skip-keyword-miner", help="跳过关键词挖掘"),
    skip_listing_analysis: bool = typer.Option(False, "--skip-listing-analysis", help="跳过卖家精灵 AI 全景分析"),
    skip_amazon: bool = typer.Option(False, "--skip-amazon", help="跳过 Amazon 页面抓取"),
    skip_query: bool = typer.Option(False, "--skip-query", help="跳过 BI query 取数"),
    skip_bi_report_data: bool = typer.Option(False, "--skip-bi-report-data", help="跳过 ASIN 报告 BI 接口取数"),
    skip_sales_query: bool = typer.Option(False, "--skip-sales-query", help="跳过销售数据 query"),
    skip_crawler_query: bool = typer.Option(False, "--skip-crawler-query", help="跳过旧爬虫 Listing query；默认已跳过，爬虫详情改走 crawler-details 接口"),
    legacy_crawler_query: bool = typer.Option(False, "--legacy-crawler-query", help="兼容旧逻辑：启用旧爬虫 Listing query"),
    skip_rufus: bool = typer.Option(False, "--skip-rufus", help="跳过 Rufus 优化建议"),
    seller_sprite_period: str = typer.Option("30d", "--seller-sprite-period", help="卖家精灵周期"),
    seller_sprite_page_size: int = typer.Option(100, "--seller-sprite-page-size", min=1, help="卖家精灵分页大小"),
    keyword_source: KeywordSource = typer.Option(KeywordSource.reverse_top, "--keyword-source", help="关键词来源策略"),
    max_miner_keywords: int = typer.Option(1, "--max-miner-keywords", min=1, help="最多挖掘关键词数量"),
    listing_analysis_station: str = typer.Option("GLOBAL", "--listing-analysis-station", help="AI 全景分析站点参数"),
    listing_analysis_poll_attempts: int | None = typer.Option(
        DEFAULT_LISTING_ANALYSIS_POLL_ATTEMPTS,
        "--listing-analysis-poll-attempts",
        min=1,
        help="AI 全景分析轮询次数",
    ),
    listing_analysis_poll_interval_seconds: float | None = typer.Option(
        DEFAULT_LISTING_ANALYSIS_POLL_INTERVAL_SECONDS,
        "--listing-analysis-poll-interval-seconds",
        min=0,
        help="AI 全景分析轮询间隔秒数",
    ),
    rufus_country: str | None = typer.Option(None, "--rufus-country", help="覆盖 Rufus 国家站点；默认跟随输入站点"),
    rufus_questions: list[str] | None = typer.Option(
        None,
        "--rufus-question",
        help="Rufus 问题，可重复传入；支持 {{asin}} 占位符",
    ),
    rufus_skills_dir: str = typer.Option(".agents/skills", "--rufus-skills-dir", help="Rufus Skill 根目录"),
    rufus_timeout_seconds: int = typer.Option(180, "--rufus-timeout-seconds", min=1, help="Rufus 单题超时秒数"),
    rufus_login_timeout_seconds: int = typer.Option(180, "--rufus-login-timeout-seconds", min=1, help="Rufus 登录恢复超时秒数"),
    rufus_parallel: bool = typer.Option(False, "--rufus-parallel", help="Rufus 多题并发获取"),
    rufus_concurrency: int = typer.Option(3, "--rufus-concurrency", min=1, help="Rufus 并发数，仅 --rufus-parallel 生效"),
    rufus_retry: int = typer.Option(0, "--rufus-retry", min=0, help="Rufus 单题无效回答重试次数"),
    rufus_strict_answer: bool = typer.Option(False, "--rufus-strict-answer", help="Rufus 无效回答重试后仍失败时中止"),
    skip_rufus_login_recovery: bool = typer.Option(False, "--skip-rufus-login-recovery", help="登录态不可用时不自动恢复"),
    sales_table_id: int | None = typer.Option(None, "--sales-table-id", help="BI 销售数据 table_id"),
    sales_dataset_alias: str = typer.Option("ds_d35ac6f3910c", "--sales-dataset-alias", help="BI 销售数据 dataset alias"),
    sales_field_mode: FieldMode = typer.Option(FieldMode.full, "--sales-field-mode", help="销售字段模式"),
    sales_start: str | None = typer.Option(None, "--sales-start", help="销售开始日期"),
    sales_end: str | None = typer.Option(None, "--sales-end", help="销售结束日期"),
    query_chunk_size: int = typer.Option(100, "--query-chunk-size", min=1, help="query 每批 ASIN 数量"),
    crawler_table_id: int | None = typer.Option(None, "--crawler-table-id", help="爬虫 Listing table_id"),
    crawler_dataset_alias: str = typer.Option("ds_icw50TLOFu4F", "--crawler-dataset-alias", help="爬虫 Listing dataset alias"),
    crawler_field_mode: FieldMode = typer.Option(FieldMode.full, "--crawler-field-mode", help="爬虫 Listing 字段模式"),
    fetch_report_files: bool = typer.Option(
        True,
        "--fetch-report-files/--no-fetch-report-files",
        help="从 ASIN 报告文件接口获取报告地址",
    ),
    upload: bool = typer.Option(True, "--upload/--no-upload", help="上传 ASIN 拆分数据包 zip 并返回阿里云文件地址"),
    submit_report_files: bool = typer.Option(False, "--submit-report-files/--no-submit-report-files", help="采集完成后提交报告文件记录到 /dataMetrics/v1/asin-report-files"),
    submit_file_records: bool = typer.Option(False, "--submit-file-records/--no-submit-file-records", help="额外按文件粒度提交拆分文件记录（每个 xlsx/md 一条，report_type 区分）"),
    report_date: str | None = typer.Option(None, "--report-date", help="报告日期，默认当天"),
    report_type: str = typer.Option(DEFAULT_REPORT_TYPE, "--report-type", help="报告类型"),
    report_source: str = typer.Option(DEFAULT_SOURCE, "--report-source", help="报告记录来源"),
    register_endpoint: str | None = typer.Option(None, "--register-endpoint", help="报告文件保存接口地址"),
    include_report_content: bool = typer.Option(False, "--include-report-content/--no-include-report-content", help="提交时是否包含报告 txt 内容和明细 JSON"),
    url_only: bool = typer.Option(False, "--url-only", help="只输出报告文件地址"),
    pretty: bool = typer.Option(False, "--pretty", help="格式化输出 JSON"),
) -> None:
    """执行 ASIN 批量取数，并输出标准前端数据文件。"""
    try:
        effective_fetch_report_files = False if submit_report_files else fetch_report_files
        result = AsinDataCollector().collect(
            input=input_path,
            asin=asin,
            keywords=keywords,
            asin_column=asin_column,
            keyword_column=keyword_column,
            site_column=site_column,
            site=site,
            output_dir=output_dir,
            run_id=run_id,
            dry_run=dry_run,
            skip_seller_sprite=skip_seller_sprite,
            skip_keyword_miner=skip_keyword_miner,
            skip_listing_analysis=skip_listing_analysis,
            skip_amazon=skip_amazon,
            skip_query=skip_query,
            skip_bi_report_data=skip_bi_report_data,
            skip_sales_query=skip_sales_query,
            skip_crawler_query=(skip_crawler_query or not legacy_crawler_query),
            skip_rufus=skip_rufus,
            seller_sprite_period=seller_sprite_period,
            seller_sprite_page_size=seller_sprite_page_size,
            keyword_source=keyword_source.value,
            max_miner_keywords=max_miner_keywords,
            listing_analysis_station=listing_analysis_station,
            listing_analysis_poll_attempts=listing_analysis_poll_attempts,
            listing_analysis_poll_interval_seconds=listing_analysis_poll_interval_seconds,
            rufus_country=rufus_country,
            rufus_questions=rufus_questions,
            rufus_skills_dir=rufus_skills_dir,
            rufus_timeout_seconds=rufus_timeout_seconds,
            rufus_login_timeout_seconds=rufus_login_timeout_seconds,
            rufus_parallel=rufus_parallel,
            rufus_concurrency=rufus_concurrency,
            rufus_retry=rufus_retry,
            rufus_strict_answer=rufus_strict_answer,
            skip_rufus_login_recovery=skip_rufus_login_recovery,
            sales_table_id=sales_table_id,
            sales_dataset_alias=sales_dataset_alias,
            sales_field_mode=sales_field_mode.value,
            sales_start=sales_start,
            sales_end=sales_end,
            query_chunk_size=query_chunk_size,
            crawler_table_id=crawler_table_id,
            crawler_dataset_alias=crawler_dataset_alias,
            crawler_field_mode=crawler_field_mode.value,
            fetch_report_files=effective_fetch_report_files,
            upload=upload,
        )
        if submit_report_files:
            if dry_run:
                result["report_file_submit"] = {
                    "submitted": False,
                    "reason": "dry_run",
                }
            else:
                client = AsinReportFileClient(endpoint=register_endpoint) if register_endpoint else AsinReportFileClient()
                result["report_file_submit"] = AsinReportFileSubmitter(client=client).submit(
                    result,
                    report_date=report_date,
                    report_type=report_type,
                    source=report_source,
                    include_content=include_report_content,
                )
        if submit_file_records:
            if dry_run:
                result["file_record_submit"] = {
                    "submitted": False,
                    "reason": "dry_run",
                }
            elif not result.get("asin_data_files"):
                result["file_record_submit"] = {
                    "submitted": False,
                    "reason": "no per-file uploads (require --upload)",
                }
            else:
                client = AsinReportFileClient(endpoint=register_endpoint) if register_endpoint else AsinReportFileClient()
                result["file_record_submit"] = AsinReportFileSubmitter(client=client).submit(
                    result,
                    report_date=report_date,
                    source=report_source,
                    file_mode=True,
                )
    except Exception as exc:
        _emit(_error_payload("asin-data collect", exc), pretty)
        raise typer.Exit(1)

    if url_only:
        url = result.get("aliyun_url")
        if not url:
            _emit(_error_payload("asin-data collect", ValueError("No Aliyun file URL returned. Use --upload or check upload configuration.")), pretty)
            raise typer.Exit(1)
        typer.echo(url)
        return

    _emit({"success": True, "command": "asin-data collect", "data": result, "error": None}, pretty)


def _daily_pipeline_kwargs(
    *,
    input_path: str | None,
    asin: str | None,
    keywords: list[str] | None,
    asin_column: str,
    keyword_column: str,
    site_column: str,
    site: str,
    output_dir: str,
    run_id: str | None,
    dry_run: bool,
    sales_start: str | None,
    sales_end: str | None,
    seller_sprite_period: str,
    seller_sprite_page_size: int,
    keyword_source: KeywordSource,
    max_miner_keywords: int,
    listing_analysis_station: str,
    rufus_country: str | None,
    rufus_questions: list[str] | None,
    rufus_skills_dir: str,
    rufus_batch_mode: RufusBatchMode,
    rufus_asin_concurrency: int | None,
    rufus_timeout_seconds: int,
    rufus_concurrency: int,
    rufus_retry: int,
    rufus_resume: bool,
    legacy_crawler_query: bool,
    upload: bool,
    fetch_report_files: bool,
) -> dict:
    return {
        "input": input_path,
        "asin": asin,
        "keywords": keywords,
        "asin_column": asin_column,
        "keyword_column": keyword_column,
        "site_column": site_column,
        "site": site,
        "output_dir": output_dir,
        "run_id": run_id,
        "dry_run": dry_run,
        "seller_sprite_period": seller_sprite_period,
        "seller_sprite_page_size": seller_sprite_page_size,
        "keyword_source": keyword_source.value,
        "max_miner_keywords": max_miner_keywords,
        "listing_analysis_station": listing_analysis_station,
        "rufus_country": rufus_country,
        "rufus_questions": rufus_questions,
        "rufus_skills_dir": rufus_skills_dir,
        "rufus_batch_mode": rufus_batch_mode.value,
        "rufus_asin_concurrency": rufus_asin_concurrency,
        "rufus_timeout_seconds": rufus_timeout_seconds,
        "rufus_parallel": False,
        "rufus_concurrency": rufus_concurrency,
        "rufus_retry": rufus_retry,
        "rufus_strict_answer": True,
        "rufus_resume": rufus_resume,
        "sales_start": sales_start,
        "sales_end": sales_end,
        "skip_crawler_query": not legacy_crawler_query,
        "fetch_report_files": fetch_report_files,
        "upload": upload,
    }


@app.command("stage-collect")
def stage_collect(
    stage: DailyStage = typer.Option(..., "--stage", help="Daily stage to run"),
    input_path: str | None = typer.Option(None, "--input", "-i", help="CSV/XLSX/JSON/JSONL input file"),
    asin: str | None = typer.Option(None, "--asin", help="Single ASIN; mutually exclusive with --input"),
    keywords: list[str] | None = typer.Option(None, "--keyword", help="Keyword for single ASIN; repeatable"),
    asin_column: str = typer.Option("asin", "--asin-column"),
    keyword_column: str = typer.Option("keyword", "--keyword-column"),
    site_column: str = typer.Option("site", "--site-column"),
    site: str = typer.Option("US", "--site"),
    output_dir: str = typer.Option("output/asin-data", "--output-dir"),
    run_id: str | None = typer.Option(None, "--run-id"),
    dry_run: bool = typer.Option(False, "--dry-run"),
    sales_start: str | None = typer.Option(None, "--sales-start"),
    sales_end: str | None = typer.Option(None, "--sales-end"),
    seller_sprite_period: str = typer.Option("30d", "--seller-sprite-period"),
    seller_sprite_page_size: int = typer.Option(100, "--seller-sprite-page-size", min=1),
    keyword_source: KeywordSource = typer.Option(KeywordSource.reverse_top, "--keyword-source"),
    max_miner_keywords: int = typer.Option(1, "--max-miner-keywords", min=1),
    listing_analysis_station: str = typer.Option("GLOBAL", "--listing-analysis-station"),
    rufus_country: str | None = typer.Option(None, "--rufus-country"),
    rufus_questions: list[str] | None = typer.Option(None, "--rufus-question"),
    rufus_skills_dir: str = typer.Option(".agents/skills", "--rufus-skills-dir"),
    rufus_batch_mode: RufusBatchMode = typer.Option(RufusBatchMode.balanced, "--rufus-batch-mode"),
    rufus_asin_concurrency: int | None = typer.Option(2, "--rufus-asin-concurrency", min=1),
    rufus_timeout_seconds: int = typer.Option(240, "--rufus-timeout-seconds", min=1),
    rufus_concurrency: int = typer.Option(2, "--rufus-concurrency", min=1),
    rufus_retry: int = typer.Option(1, "--rufus-retry", min=0),
    rufus_resume: bool = typer.Option(True, "--rufus-resume/--no-rufus-resume"),
    legacy_crawler_query: bool = typer.Option(False, "--legacy-crawler-query"),
    pretty: bool = typer.Option(False, "--pretty"),
) -> None:
    """Run one stage and write it under output/asin-data/<run-id>/stages."""
    try:
        result = DailyAsinDataPipeline().run_stage(
            stage.value,
            **_daily_pipeline_kwargs(
                input_path=input_path,
                asin=asin,
                keywords=keywords,
                asin_column=asin_column,
                keyword_column=keyword_column,
                site_column=site_column,
                site=site,
                output_dir=output_dir,
                run_id=run_id,
                dry_run=dry_run,
                sales_start=sales_start,
                sales_end=sales_end,
                seller_sprite_period=seller_sprite_period,
                seller_sprite_page_size=seller_sprite_page_size,
                keyword_source=keyword_source,
                max_miner_keywords=max_miner_keywords,
                listing_analysis_station=listing_analysis_station,
                rufus_country=rufus_country,
                rufus_questions=rufus_questions,
                rufus_skills_dir=rufus_skills_dir,
                rufus_batch_mode=rufus_batch_mode,
                rufus_asin_concurrency=rufus_asin_concurrency,
                rufus_timeout_seconds=rufus_timeout_seconds,
                rufus_concurrency=rufus_concurrency,
                rufus_retry=rufus_retry,
                rufus_resume=rufus_resume,
                legacy_crawler_query=legacy_crawler_query,
                upload=False,
                fetch_report_files=False,
            ),
        )
    except Exception as exc:
        _emit(_error_payload("asin-data stage-collect", exc), pretty)
        raise typer.Exit(1)
    _emit(result, pretty)


@app.command("merge-stages")
def merge_stages(
    input_path: str | None = typer.Option(None, "--input", "-i", help="CSV/XLSX/JSON/JSONL input file"),
    asin: str | None = typer.Option(None, "--asin", help="Single ASIN; mutually exclusive with --input"),
    keywords: list[str] | None = typer.Option(None, "--keyword", help="Keyword for single ASIN; repeatable"),
    asin_column: str = typer.Option("asin", "--asin-column"),
    keyword_column: str = typer.Option("keyword", "--keyword-column"),
    site_column: str = typer.Option("site", "--site-column"),
    site: str = typer.Option("US", "--site"),
    output_dir: str = typer.Option("output/asin-data", "--output-dir"),
    run_id: str = typer.Option(..., "--run-id"),
    upload: bool = typer.Option(True, "--upload/--no-upload"),
    fetch_report_files: bool = typer.Option(False, "--fetch-report-files/--no-fetch-report-files"),
    url_only: bool = typer.Option(False, "--url-only"),
    pretty: bool = typer.Option(False, "--pretty"),
) -> None:
    """Merge cached stages, build frontend files, split package, and optional upload."""
    try:
        result = DailyAsinDataPipeline().merge(
            input=input_path,
            asin=asin,
            keywords=keywords,
            asin_column=asin_column,
            keyword_column=keyword_column,
            site_column=site_column,
            site=site,
            output_dir=output_dir,
            run_id=run_id,
            upload=upload,
            fetch_report_files=fetch_report_files,
        )
    except Exception as exc:
        _emit(_error_payload("asin-data merge-stages", exc), pretty)
        raise typer.Exit(1)

    if url_only:
        url = result.get("aliyun_url")
        if not url:
            _emit(_error_payload("asin-data merge-stages", ValueError("No Aliyun file URL returned. Use --upload or check upload configuration.")), pretty)
            raise typer.Exit(1)
        typer.echo(url)
        return
    _emit({"success": True, "command": "asin-data merge-stages", "data": result, "error": None}, pretty)


@app.command("daily-collect")
def daily_collect(
    input_path: str | None = typer.Option(None, "--input", "-i", help="CSV/XLSX/JSON/JSONL input file"),
    asin: str | None = typer.Option(None, "--asin", help="Single ASIN; mutually exclusive with --input"),
    keywords: list[str] | None = typer.Option(None, "--keyword", help="Keyword for single ASIN; repeatable"),
    asin_column: str = typer.Option("asin", "--asin-column"),
    keyword_column: str = typer.Option("keyword", "--keyword-column"),
    site_column: str = typer.Option("site", "--site-column"),
    site: str = typer.Option("US", "--site"),
    output_dir: str = typer.Option("output/asin-data", "--output-dir"),
    run_id: str | None = typer.Option(None, "--run-id"),
    dry_run: bool = typer.Option(False, "--dry-run"),
    sales_start: str | None = typer.Option(None, "--sales-start"),
    sales_end: str | None = typer.Option(None, "--sales-end"),
    upload: bool = typer.Option(True, "--upload/--no-upload"),
    pretty: bool = typer.Option(False, "--pretty"),
) -> None:
    """Run all daily stages in a conservative sequence, then merge."""
    try:
        result = DailyAsinDataPipeline().run_all(
            input=input_path,
            asin=asin,
            keywords=keywords,
            asin_column=asin_column,
            keyword_column=keyword_column,
            site_column=site_column,
            site=site,
            output_dir=output_dir,
            run_id=run_id,
            dry_run=dry_run,
            sales_start=sales_start,
            sales_end=sales_end,
            rufus_batch_mode="balanced",
            rufus_asin_concurrency=2,
            rufus_timeout_seconds=240,
            rufus_concurrency=2,
            rufus_retry=1,
            rufus_strict_answer=True,
            rufus_resume=True,
            upload=upload,
            fetch_report_files=False,
        )
    except Exception as exc:
        _emit(_error_payload("asin-data daily-collect", exc), pretty)
        raise typer.Exit(1)
    _emit({"success": True, "command": "asin-data daily-collect", "data": result, "error": None}, pretty)


__all__ = ["app"]
