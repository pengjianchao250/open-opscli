"""ASIN 实时取数与拆包文件读取服务。

本模块承载 `asin-data live-data` 的可复用业务编排，使 CLI 与 MCP
入口共享同一套实时取数、xlsx 解析和 OSS 上传逻辑。
"""

from __future__ import annotations

import json
import time
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from pathlib import Path
from typing import Any, Callable

import httpx

from opscli.asin_data.services.ai_response import build_ai_ready_response
from opscli.asin_data.services.bi_report_data import (
    BASIC_REPORT_SOURCE_KEYS,
    BI_ONLY_REPORT_SOURCE_KEYS,
    LISTING_REPORT_SOURCE_KEYS,
)
from opscli.asin_data.services.collector import AsinDataCollector
from opscli.asin_data.services.report_files import AsinReportFileClient, AsinReportFileNotFoundError
from opscli.shared.file_uploads import FileUploadClient

VALID_FILE_KEYS = {
    "basic",
    "bi",
    "keyword_reverse",
    "keyword_miner",
    "competitor",
    "rufus",
}
VALID_LIVE_DATA_SCOPES = {"all", "basic", "bi", "listing", "listing_basic"}
VALID_LIVE_DATA_RETURN_MODES = {"content", "url_only", "both", "ai_ready"}


def normalize_live_data_scope(scope: str | None) -> str:
    """规范化实时取数范围参数。"""
    value = (scope or "all").strip().lower().replace("-", "_")
    if value not in VALID_LIVE_DATA_SCOPES:
        allowed = ", ".join(sorted(VALID_LIVE_DATA_SCOPES))
        raise ValueError(f"不支持的 data_scope：{scope}，可选值：{allowed}")
    return value


def normalize_live_data_return_mode(return_mode: str | None) -> str:
    """规范化实时取数返回模式。"""
    value = (return_mode or "content").strip().lower().replace("-", "_")
    if value not in VALID_LIVE_DATA_RETURN_MODES:
        allowed = ", ".join(sorted(VALID_LIVE_DATA_RETURN_MODES))
        raise ValueError(f"不支持的 return_mode：{return_mode}，可选值：{allowed}")
    return value


def live_data_source_keys(scope: str | None) -> tuple[str, ...] | None:
    """把 live-data 范围参数映射到实时 BI 接口 source key。"""
    normalized = normalize_live_data_scope(scope)
    if normalized == "basic":
        return BASIC_REPORT_SOURCE_KEYS
    if normalized in {"listing", "listing_basic"}:
        return LISTING_REPORT_SOURCE_KEYS
    if normalized == "bi":
        return BI_ONLY_REPORT_SOURCE_KEYS
    return None


def live_data_file_keys(scope: str | None) -> tuple[str, ...]:
    """把 live-data 范围参数映射到需要返回或上传的拆包文件。"""
    normalized = normalize_live_data_scope(scope)
    if normalized in {"basic", "listing", "listing_basic"}:
        return ("basic",)
    if normalized == "bi":
        return ("bi",)
    return ("basic", "bi")


def read_xlsx_content(source: Any) -> dict[str, list[list[Any]]]:
    """将 xlsx 转为 `{sheet_name: [rows]}` JSON 结构。"""
    from openpyxl import load_workbook

    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        sheets: dict[str, list[list[Any]]] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            sheets[sheet_name] = rows
        return sheets
    finally:
        wb.close()


def read_local_split_file_content(path: Path, file_key: str) -> Any:
    """读取本地拆包文件，并返回与 fetch-file 一致的结构化内容。"""
    if file_key == "rufus":
        return path.read_text(encoding="utf-8", errors="replace")
    return read_xlsx_content(path)


def load_frontend_data(output_dir: str) -> dict[str, Any]:
    """读取本次实时采集生成的前端 JSON 数据。"""
    frontend_path = Path(output_dir) / "frontend-data.json"
    return json.loads(frontend_path.read_text(encoding="utf-8"))


def load_live_split_files(
    result: dict[str, Any],
    *,
    file_keys: tuple[str, ...] = ("basic", "bi"),
) -> dict[str, dict[str, Any]]:
    """读取实时采集生成的 ASIN 拆包文件内容。"""
    split_files = collect_live_split_file_paths(result, file_keys=file_keys)
    for asin_files in split_files.values():
        for file_key, file_item in asin_files.items():
            path_text = file_item.get("file_path")
            if isinstance(path_text, str) and path_text.strip():
                file_item["content"] = read_local_split_file_content(Path(path_text), file_key)
    return split_files


def collect_live_split_file_paths(
    result: dict[str, Any],
    *,
    file_keys: tuple[str, ...] = ("basic", "bi"),
) -> dict[str, dict[str, Any]]:
    """只收集实时拆包文件路径，供 URL-only 快速返回和上传复用。"""
    manifest = result.get("manifest") if isinstance(result.get("manifest"), dict) else {}
    package = manifest.get("asin_data_package") if isinstance(manifest.get("asin_data_package"), dict) else {}
    items = package.get("items") if isinstance(package.get("items"), list) else []
    split_files: dict[str, dict[str, Any]] = {}
    for item in items:
        if not isinstance(item, dict):
            continue
        asin = str(item.get("asin") or "").strip().upper()
        files = item.get("files") if isinstance(item.get("files"), dict) else {}
        if not asin:
            continue
        asin_files: dict[str, Any] = {}
        for file_key in file_keys:
            path_text = files.get(file_key)
            if not isinstance(path_text, str) or not path_text.strip():
                continue
            path = Path(path_text)
            asin_files[file_key] = {
                "asin": asin,
                "file_key": file_key,
                "file_path": path.as_posix(),
            }
        if asin_files:
            split_files[asin] = asin_files
    return split_files


def upload_live_split_files(
    split_files: dict[str, dict[str, Any]],
    *,
    run_id: str,
    file_upload_client: Any | None = None,
) -> dict[str, Any]:
    """上传实时 basic/bi xlsx，并把 OSS 地址写回 split_files。"""
    client = file_upload_client or FileUploadClient()
    upload_tasks: list[tuple[str, str, dict[str, Any]]] = []
    for asin, files in split_files.items():
        for file_key in ("basic", "bi"):
            file_item = files.get(file_key)
            if not isinstance(file_item, dict):
                continue
            path_text = file_item.get("file_path")
            if not isinstance(path_text, str) or not path_text.strip():
                continue
            upload_tasks.append((asin, file_key, file_item))

    def upload_one(asin: str, file_key: str, file_item: dict[str, Any]) -> tuple[str, str, dict[str, Any]]:
        path = Path(str(file_item["file_path"]))
        upload_filename = f"{asin}-{file_key}-live-data.xlsx"
        upload = client.upload(
            path,
            purpose="asin_data_live_xlsx",
            folder="asin-data",
            public="1",
            filename=upload_filename,
            metadata={
                "run_id": run_id,
                "asin": asin,
                "file_key": file_key,
                "report_filename": upload_filename,
                "source_filename": path.name,
                "source": "asin-data live-data",
            },
        )
        file_item["file_url"] = upload.url
        file_item["upload"] = {"url": upload.url, "raw": upload.raw}
        return asin, file_key, {
            "url": upload.url,
            "file_name": upload_filename,
            "file_path": path.as_posix(),
            "upload_path": path.as_posix(),
        }

    uploaded_by_asin: dict[str, dict[str, Any]] = {asin: {} for asin in split_files}
    if upload_tasks:
        with ThreadPoolExecutor(max_workers=min(8, len(upload_tasks))) as executor:
            futures = [executor.submit(upload_one, asin, file_key, file_item) for asin, file_key, file_item in upload_tasks]
            for future in futures:
                asin, file_key, uploaded = future.result()
                uploaded_by_asin.setdefault(asin, {})[file_key] = uploaded

    items: list[dict[str, Any]] = []
    for asin in split_files:
        items.append({"asin": asin, "files": uploaded_by_asin.get(asin, {})})
    return {"files_uploaded": len(upload_tasks), "items": items}


def split_file_urls(split_files: dict[str, dict[str, Any]]) -> dict[str, dict[str, str]]:
    """从 split_files 中提取轻量 OSS URL 映射。"""
    urls: dict[str, dict[str, str]] = {}
    for asin, files in split_files.items():
        asin_urls = {
            file_key: str(file_item.get("file_url"))
            for file_key, file_item in files.items()
            if isinstance(file_item, dict) and file_item.get("file_url")
        }
        if asin_urls:
            urls[asin] = asin_urls
    return urls


def download_file_content(url: str, file_key: str) -> Any:
    """下载拆包文件，并返回结构化内容。"""
    response = httpx.get(url, timeout=60, follow_redirects=True)
    response.raise_for_status()
    raw = response.content
    if file_key == "rufus":
        return raw.decode("utf-8", errors="replace")
    return read_xlsx_content(BytesIO(raw))


def fetch_split_file(
    *,
    asin: str,
    site: str = "US",
    file_key: str,
    report_file_client: Any | None = None,
) -> dict[str, Any]:
    """读取历史 ASIN 拆包文件地址并下载内容。"""
    normalized_asin = asin.strip().upper()
    normalized_site = site.strip().upper()
    normalized_file_key = file_key.strip().lower()
    if normalized_file_key not in VALID_FILE_KEYS:
        allowed = ", ".join(sorted(VALID_FILE_KEYS))
        raise ValueError(f"不支持的 file_key：{file_key}，可选值：{allowed}")

    client = report_file_client or AsinReportFileClient()
    result = client.fetch_split_files(asin=normalized_asin, site=normalized_site)
    files = result.get("files") or {}
    urls = files.get(normalized_file_key)
    if not urls:
        raise AsinReportFileNotFoundError(asin=normalized_asin, site=normalized_site)
    first_url = urls[0] if isinstance(urls, list) else urls
    content = download_file_content(first_url, normalized_file_key)
    return {
        "asin": normalized_asin,
        "site": normalized_site,
        "file_key": normalized_file_key,
        "file_url": urls,
        "content": content,
    }


class AsinLiveDataService:
    """ASIN 实时取数服务，供 CLI 和 MCP 共用。"""

    def __init__(
        self,
        *,
        collector: Any | None = None,
        file_upload_client_factory: Callable[[], Any] | None = None,
    ) -> None:
        """创建实时取数服务。"""
        self._collector = collector or AsinDataCollector()
        self._file_upload_client_factory = file_upload_client_factory or FileUploadClient

    def run(
        self,
        *,
        input_path: str | None = None,
        asin: str | None = None,
        keywords: list[str] | None = None,
        asin_column: str = "asin",
        keyword_column: str = "keyword",
        site_column: str = "site",
        site: str = "US",
        output_dir: str = "output/asin-data",
        run_id: str | None = None,
        dry_run: bool = False,
        skip_query: bool = True,
        skip_bi_report_data: bool = False,
        skip_sales_query: bool = False,
        skip_crawler_query: bool = False,
        legacy_crawler_query: bool = False,
        sales_table_id: int | None = None,
        sales_dataset_alias: str = "ds_d35ac6f3910c",
        sales_field_mode: str = "full",
        sales_start: str | None = None,
        sales_end: str | None = None,
        query_chunk_size: int = 100,
        crawler_table_id: int | None = None,
        crawler_dataset_alias: str = "ds_icw50TLOFu4F",
        crawler_field_mode: str = "full",
        data_scope: str = "all",
        upload_xlsx: bool = False,
        return_mode: str = "content",
    ) -> dict[str, Any]:
        """执行实时 ASIN 取数并返回 CLI/MCP 共用结果结构。"""
        started_at = time.perf_counter()
        normalized_scope = normalize_live_data_scope(data_scope)
        normalized_return_mode = normalize_live_data_return_mode(return_mode)
        source_keys = live_data_source_keys(normalized_scope)
        file_keys = live_data_file_keys(normalized_scope)
        lean_output = normalized_return_mode in {"url_only", "ai_ready"}
        request = {
            "asin": asin,
            "input_path": input_path,
            "sales_start": sales_start,
            "sales_end": sales_end,
            "upload_xlsx": upload_xlsx,
            "return_mode": normalized_return_mode,
            "data_scope": normalized_scope,
            "site": site,
            "keywords": keywords,
            "asin_column": asin_column,
            "keyword_column": keyword_column,
            "site_column": site_column,
            "output_dir": output_dir,
            "run_id": run_id,
            "query_chunk_size": query_chunk_size,
        }

        result = self._collector.collect(
            input=input_path,
            asin=asin,
            keywords=keywords,
            asin_column=asin_column,
            keyword_column=keyword_column,
            site_column=site_column,
            site=site,
            output_dir=output_dir,
            run_id=run_id,
            dry_run=dry_run,
            skip_seller_sprite=True,
            skip_keyword_miner=True,
            skip_listing_analysis=True,
            skip_amazon=True,
            skip_query=skip_query,
            skip_bi_report_data=skip_bi_report_data,
            skip_sales_query=skip_sales_query,
            skip_crawler_query=(skip_crawler_query or not legacy_crawler_query),
            skip_rufus=True,
            sales_table_id=sales_table_id,
            sales_dataset_alias=sales_dataset_alias,
            sales_field_mode=sales_field_mode,
            sales_start=sales_start,
            sales_end=sales_end,
            bi_report_source_keys=source_keys,
            query_chunk_size=query_chunk_size,
            crawler_table_id=crawler_table_id,
            crawler_dataset_alias=crawler_dataset_alias,
            crawler_field_mode=crawler_field_mode,
            fetch_report_files=False,
            upload=False,
            split_file_keys=file_keys if lean_output else None,
            build_split_package_zip=not lean_output,
            write_frontend_outputs=not lean_output,
        )
        result["data_scope"] = normalized_scope
        result["return_mode"] = normalized_return_mode
        if normalized_return_mode in {"content", "both"}:
            result["frontend_data"] = load_frontend_data(str(result["output_dir"]))
            split_files = load_live_split_files(result, file_keys=file_keys)
            result["split_files"] = split_files
        else:
            split_files = collect_live_split_file_paths(result, file_keys=file_keys)
            if not upload_xlsx or normalized_return_mode == "ai_ready":
                result["split_file_paths"] = split_files
        if upload_xlsx:
            manifest = result.get("manifest") if isinstance(result.get("manifest"), dict) else {}
            run_id_for_upload = str(manifest.get("run_id") or Path(str(result["output_dir"])).name)
            result["split_file_uploads"] = upload_live_split_files(
                split_files,
                run_id=run_id_for_upload,
                file_upload_client=self._file_upload_client_factory(),
            )
            result["split_file_urls"] = split_file_urls(split_files)
        if normalized_return_mode == "ai_ready":
            return build_ai_ready_response(
                tool_name="asin_data_live_data",
                request=request,
                result=result,
                data_scope=normalized_scope,
                site=site,
                split_files=split_files,
                elapsed_seconds=time.perf_counter() - started_at,
            )
        return result
