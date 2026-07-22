"""构建 ASIN 实时取数的 AI Ready 返回协议。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROTOCOL = "asin_data_ai_response"
PROTOCOL_VERSION = "1.0"

DEPRECATED_FIELDS = ["split_file_urls", "split_file_paths", "manifest"]
PREFERRED_FIELDS = ["items[].artifacts", "items[].datasets", "items[].diagnostics"]

PREVIEW_LIMITS = {
    "sales_traffic": 5,
    "sp_search_term": 10,
    "sqp": 10,
    "deals": 5,
    "turnover_inventory": 5,
    "listing_basic": 1,
    "crawler_details": 5,
    "product_detail": 5,
    "bullets": 10,
    "image_links": 10,
    "qa": 5,
    "reviews": 5,
}

BI_SOURCE_KEYS_BY_POSITION = (
    "sales_traffic",
    "sp_search_term",
    "deals",
    "turnover_inventory",
)

BASIC_SOURCE_KEYS_BY_POSITION = (
    "listing_basic",
    "crawler_details",
    "product_detail",
    "bullets",
    "image_links",
    "qa",
    "reviews",
)

BASIC_SHEET_SOURCE_KEYS = {
    "基础汇总": "listing_basic",
    "刊登数据": "listing_basic",
    "爬虫数据": "crawler_details",
    "商品详情": "product_detail",
    "五点描述": "bullets",
    "图片链接": "image_links",
    "QA": "qa",
    "评论": "reviews",
}

MOJIBAKE_MARKERS = ("鎼", "鍒", "骞", "憡", "閿", "鏁", "嵁", "鍩", "姹", "€", "�")


def build_ai_ready_response(
    *,
    tool_name: str,
    request: dict[str, Any],
    result: dict[str, Any],
    data_scope: str,
    site: str,
    split_files: dict[str, dict[str, Any]],
    elapsed_seconds: float | None = None,
) -> dict[str, Any]:
    """把实时取数结果转换为轻量、稳定的 AI Ready 协议。"""
    manifest = result.get("manifest") if isinstance(result.get("manifest"), dict) else {}
    package = manifest.get("asin_data_package") if isinstance(manifest.get("asin_data_package"), dict) else {}
    manifest_items = package.get("items") if isinstance(package.get("items"), list) else []
    requested_asins = _requested_asins(manifest_items, split_files, request)
    items: list[dict[str, Any]] = []

    for asin in requested_asins:
        asin_files = split_files.get(asin, {})
        artifacts, artifact_diagnostics = build_artifacts_for_asin(
            asin=asin,
            split_files=asin_files,
            request=request,
        )
        datasets: list[dict[str, Any]] = []
        item_diagnostics = list(artifact_diagnostics)
        if not artifacts:
            item_diagnostics.append(
                diagnostic(
                    "error",
                    "SOURCE_ERROR",
                    f"No live-data artifacts were generated for {asin}.",
                    action="rerun_live_data",
                )
            )
        for artifact in artifacts:
            artifact_datasets, workbook_diagnostics = build_datasets_from_artifact(
                artifact=artifact,
                asin=asin,
                request=request,
            )
            datasets.extend(artifact_datasets)
            item_diagnostics.extend(workbook_diagnostics)
        item_diagnostics.extend(_collect_dataset_diagnostics(datasets))
        items.append(
            {
                "asin": asin,
                "site": _site_for_asin(manifest_items, asin, site),
                "status": "success" if artifacts and not _has_error(item_diagnostics) else "failed",
                "artifacts": artifacts,
                "datasets": datasets,
                "diagnostics": item_diagnostics,
            }
        )

    diagnostics = _global_diagnostics(data_scope=data_scope, request=request, items=items)
    run_id = str(manifest.get("run_id") or Path(str(result.get("output_dir") or "")).name)
    response = {
        "success": bool(result.get("success", True)),
        "metadata": {
            "protocol": PROTOCOL,
            "protocol_version": PROTOCOL_VERSION,
            "tool": tool_name,
            "data_scope": data_scope,
            "site": site,
            "request": _sanitize_request(request),
        },
        "run": {
            "run_id": run_id,
            "output_dir": str(result.get("output_dir") or ""),
            "elapsed_seconds": round(elapsed_seconds, 3) if elapsed_seconds is not None else None,
            "cache_hit": False,
        },
        "summary": _summary(result, items),
        "items": items,
        "diagnostics": diagnostics,
        "deprecated_fields": DEPRECATED_FIELDS,
        "preferred_fields": PREFERRED_FIELDS,
    }
    _attach_legacy_fields(response, result)
    return response


def build_artifacts_for_asin(
    *,
    asin: str,
    split_files: dict[str, Any],
    request: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """为单个 ASIN 构建完整文件索引。"""
    artifacts: list[dict[str, Any]] = []
    diagnostics: list[dict[str, Any]] = []
    for file_key, file_item in split_files.items():
        if not isinstance(file_item, dict):
            continue
        local_path = str(file_item.get("file_path") or "")
        path = Path(local_path) if local_path else None
        uri = str(file_item.get("file_url") or "") or None
        file_type = "md" if file_key == "rufus" else "xlsx"
        source_filename = path.name if path else ""
        artifact = {
            "artifact_id": f"{asin}_{file_key}_{file_type}",
            "file_key": file_key,
            "type": file_type,
            "uri": uri,
            "local_path": path.as_posix() if path else "",
            "complete": bool(path and path.exists()),
            "source_filename": source_filename,
            "report_filename": _report_filename(asin, file_key, file_type, source_filename),
        }
        artifacts.append(artifact)
        if not artifact["complete"]:
            diagnostics.append(
                diagnostic(
                    "error",
                    "SOURCE_ERROR",
                    f"{file_key} artifact file is missing.",
                    source_key=file_key,
                    action="rerun_live_data",
                )
            )
        elif not uri:
            diagnostics.append(
                diagnostic(
                    "warning",
                    "LOCAL_ONLY" if not request.get("upload_xlsx") else "UPLOAD_FAILED",
                    f"{file_key} artifact has no remote URI.",
                    source_key=file_key,
                    action="upload_artifact_if_needed",
                )
            )
    return artifacts, diagnostics


def build_datasets_from_artifact(
    *,
    artifact: dict[str, Any],
    asin: str,
    request: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """从 xlsx artifact 中读取 sheet 索引、字段和预览行。"""
    if artifact.get("type") != "xlsx" or not artifact.get("complete"):
        return [], []

    path = Path(str(artifact.get("local_path") or ""))
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return [], [
            diagnostic(
                "error",
                "SOURCE_ERROR",
                f"Failed to read workbook: {exc}",
                source_key=str(artifact.get("file_key") or ""),
                action="rerun_live_data",
            )
        ]

    datasets: list[dict[str, Any]] = []
    workbook_diagnostics: list[dict[str, Any]] = []
    try:
        for index, sheet_name in enumerate(workbook.sheetnames):
            source_key = _source_key(str(artifact.get("file_key") or ""), sheet_name, index)
            dataset = _dataset_from_sheet(
                workbook[sheet_name],
                asin=asin,
                artifact_id=str(artifact.get("artifact_id") or ""),
                source_key=source_key,
                request=request,
            )
            datasets.append(dataset)
    except Exception as exc:
        workbook_diagnostics.append(
            diagnostic(
                "error",
                "SOURCE_ERROR",
                f"Failed to build workbook manifest: {exc}",
                source_key=str(artifact.get("file_key") or ""),
                action="rerun_live_data",
            )
        )
    finally:
        workbook.close()
    return datasets, workbook_diagnostics


def diagnostic(
    level: str,
    code: str,
    message: str,
    *,
    source_key: str | None = None,
    action: str | None = None,
) -> dict[str, Any]:
    """创建统一诊断对象。"""
    item = {
        "level": level,
        "code": code,
        "message": message,
    }
    if source_key:
        item["source_key"] = source_key
    if action:
        item["action"] = action
    return item


def _dataset_from_sheet(
    sheet: Any,
    *,
    asin: str,
    artifact_id: str,
    source_key: str,
    request: dict[str, Any],
) -> dict[str, Any]:
    headers: list[str] = []
    preview_rows: list[dict[str, Any]] = []
    row_count = 0
    preview_limit = PREVIEW_LIMITS.get(source_key, 5)

    for raw_row in sheet.iter_rows(values_only=True):
        values = [_json_value(value) for value in raw_row]
        if not _has_row_value(values):
            continue
        if not headers:
            headers = _headers(values)
            continue
        row = _row_dict(headers, values)
        row_count += 1
        if len(preview_rows) < preview_limit:
            preview_rows.append(row)

    empty = row_count == 0 or (row_count == 1 and preview_rows and _is_no_data_row(preview_rows[0]))
    if empty:
        row_count = 0
        preview_rows = []

    dataset_diagnostics: list[dict[str, Any]] = []
    if empty:
        dataset_diagnostics.append(
            diagnostic(
                "warning",
                "EMPTY_DATASET",
                f"{source_key} has no data rows.",
                source_key=source_key,
                action="check_source_or_date_range",
            )
        )
    if row_count > 500:
        dataset_diagnostics.append(
            diagnostic(
                "warning",
                "LARGE_DATASET",
                f"{source_key} has {row_count} rows. Full data is available in artifact.",
                source_key=source_key,
                action="read_artifact_if_needed",
            )
        )

    encoding_suspected = _encoding_suspected([sheet.title, *headers])
    if encoding_suspected:
        dataset_diagnostics.append(
            diagnostic(
                "warning",
                "ENCODING_SUSPECTED",
                f"{source_key} sheet or columns may contain mojibake characters.",
                source_key=source_key,
                action="use_source_key_and_artifact",
            )
        )

    dataset = {
        "dataset_id": f"{asin}_{source_key}",
        "source_key": source_key,
        "semantic_type": source_key,
        "sheet_name": sheet.title,
        "artifact_id": artifact_id,
        "row_count": row_count,
        "column_count": len(headers),
        "columns": headers,
        "preview_rows": preview_rows,
        "quality": {
            "empty": empty,
            "large_sheet": row_count > 500,
            "encoding_ok": not encoding_suspected,
            "encoding_suspected": encoding_suspected,
            "has_warnings": bool(dataset_diagnostics),
        },
        "diagnostics": dataset_diagnostics,
    }
    if source_key == "sp_search_term":
        dataset["filter"] = _sp_search_term_filter(asin, request)
        dataset["diagnostics"].append(
            diagnostic(
                "warning",
                "ASIN_FILTER_UNVERIFIED",
                "sp_search_term may contain ASIN group level data. Verify before using it for single-ASIN conclusions.",
                source_key=source_key,
                action="verify_filter_before_conclusion",
            )
        )
        dataset["quality"]["has_warnings"] = True
    return dataset


def _requested_asins(
    manifest_items: list[Any],
    split_files: dict[str, dict[str, Any]],
    request: dict[str, Any],
) -> list[str]:
    asins: list[str] = []
    for item in manifest_items:
        if isinstance(item, dict):
            asin = str(item.get("asin") or "").strip().upper()
            if asin and asin not in asins:
                asins.append(asin)
    for asin in split_files:
        normalized = str(asin or "").strip().upper()
        if normalized and normalized not in asins:
            asins.append(normalized)
    requested = str(request.get("asin") or "").strip().upper()
    if requested and requested not in asins:
        asins.append(requested)
    return asins


def _site_for_asin(manifest_items: list[Any], asin: str, default_site: str) -> str:
    for item in manifest_items:
        if isinstance(item, dict) and str(item.get("asin") or "").strip().upper() == asin:
            site = str(item.get("site") or "").strip().upper()
            if site:
                return site
    return default_site


def _source_key(file_key: str, sheet_name: str, index: int) -> str:
    if file_key == "bi":
        return BI_SOURCE_KEYS_BY_POSITION[index] if index < len(BI_SOURCE_KEYS_BY_POSITION) else f"bi_sheet_{index + 1}"
    if file_key == "basic":
        if sheet_name in BASIC_SHEET_SOURCE_KEYS:
            return BASIC_SHEET_SOURCE_KEYS[sheet_name]
        return BASIC_SOURCE_KEYS_BY_POSITION[index] if index < len(BASIC_SOURCE_KEYS_BY_POSITION) else f"basic_sheet_{index + 1}"
    return f"{file_key}_sheet_{index + 1}"


def _headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        header = str(value).strip() if value not in (None, "") else f"column_{index}"
        count = used.get(header, 0)
        used[header] = count + 1
        headers.append(header if count == 0 else f"{header}_{count + 1}")
    return headers


def _row_dict(headers: list[str], values: list[Any]) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for index, header in enumerate(headers):
        row[header] = values[index] if index < len(values) else None
    return row


def _has_row_value(values: list[Any]) -> bool:
    return any(value not in (None, "") for value in values)


def _is_no_data_row(row: dict[str, Any]) -> bool:
    if len(row) != 1:
        return False
    key, value = next(iter(row.items()))
    return str(key).strip() == "说明" and str(value or "").strip().startswith("无")


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _encoding_suspected(values: list[Any]) -> bool:
    text = " ".join(str(value) for value in values if value not in (None, ""))
    return any(marker in text for marker in MOJIBAKE_MARKERS)


def _sp_search_term_filter(asin: str, request: dict[str, Any]) -> dict[str, Any]:
    return {
        "requested_asin": asin,
        "requested_date_range": {
            "start": request.get("sales_start"),
            "end": request.get("sales_end"),
        },
        "backend_request_body": {
            "start_date": request.get("sales_start"),
            "end_date": request.get("sales_end"),
            "asin": asin,
        },
        "effective_filter_verified": False,
        "verification_notes": ["sp_search_term backend filter semantics not confirmed"],
    }


def _summary(result: dict[str, Any], items: list[dict[str, Any]]) -> dict[str, Any]:
    original = result.get("summary") if isinstance(result.get("summary"), dict) else {}
    failed_asins = [str(item.get("asin")) for item in items if item.get("status") != "success"]
    return {
        **original,
        "input_count": original.get("input_count", original.get("asin_count", len(items))),
        "asin_count": original.get("asin_count", len(items)),
        "source_error_count": original.get("source_error_count", 0),
        "failed_asin_count": len(failed_asins),
        "failed_asins": failed_asins,
    }


def _global_diagnostics(
    *,
    data_scope: str,
    request: dict[str, Any],
    items: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    diagnostics: list[dict[str, Any]] = []
    if data_scope in {"all", "bi"} and (not request.get("sales_start") or not request.get("sales_end")):
        diagnostics.append(
            diagnostic(
                "warning",
                "DATE_RANGE_MISSING",
                "BI data was requested without a complete sales_start/sales_end date range.",
                action="pass_sales_start_and_sales_end",
            )
        )
    for item in items:
        diagnostics.extend(item.get("diagnostics") if isinstance(item.get("diagnostics"), list) else [])
    return diagnostics


def _collect_dataset_diagnostics(datasets: list[dict[str, Any]]) -> list[dict[str, Any]]:
    diagnostics: list[dict[str, Any]] = []
    for dataset in datasets:
        values = dataset.get("diagnostics") if isinstance(dataset.get("diagnostics"), list) else []
        diagnostics.extend(values)
    return diagnostics


def _has_error(diagnostics: list[dict[str, Any]]) -> bool:
    return any(item.get("level") == "error" for item in diagnostics)


def _report_filename(asin: str, file_key: str, file_type: str, source_filename: str) -> str:
    if file_type == "xlsx" and file_key in {"basic", "bi"}:
        return f"{asin}-{file_key}-live-data.xlsx"
    return source_filename


def _sanitize_request(request: dict[str, Any]) -> dict[str, Any]:
    blocked_keys = {"jwt", "session_id", "authorization", "cookie", "password", "token"}
    return {key: value for key, value in request.items() if key.lower() not in blocked_keys}


def _attach_legacy_fields(response: dict[str, Any], result: dict[str, Any]) -> None:
    for key in ("manifest", "split_file_paths", "split_file_urls", "split_file_uploads"):
        if key in result:
            response[key] = result[key]
