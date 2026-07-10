"""Build split ASIN data packages for operator-facing downloads."""

from __future__ import annotations

import json
import re
import shutil
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from opscli.amazon_rufus.services.answer_report_formatter import AnswerReportFormatter
from opscli.asin_data.services.merged_report_renderer import (
    BI_REPORT_SOURCE_ORDER,
    append_rufus_display_blocks,
    rufus_summary,
)


EXCEL_CELL_LIMIT = 32000
BI_SOURCE_KEYS = tuple(key for key, _label in BI_REPORT_SOURCE_ORDER)

FILE_BASIC = "01-\u57fa\u7840\u6570\u636e.xlsx"
FILE_BI = "02-BI\u6570\u636e.xlsx"
FILE_KEYWORD_REVERSE = "03-\u5356\u5bb6\u7cbe\u7075\u5173\u952e\u8bcd\u6570\u636e.xlsx"
FILE_KEYWORD_MINER = "04-\u5356\u5bb6\u7cbe\u7075\u5173\u952e\u8bcd\u6316\u6398.xlsx"
FILE_COMPETITOR = "05-\u5356\u5bb6\u7cbe\u7075\u7ade\u54c1\u5206\u6790.xlsx"
FILE_RUFUS = "06-Rufus\u6570\u636e\u5206\u6790.md"

SHEET_NO_DATA = "\u65e0\u6570\u636e"
SHEET_BASIC = "\u57fa\u7840\u6c47\u603b"
SHEET_LISTING = "\u520a\u767b\u6570\u636e"
SHEET_CRAWLER = "\u722c\u866b\u6570\u636e"
SHEET_PRODUCT = "\u5546\u54c1\u8be6\u60c5"
SHEET_BULLETS = "\u4e94\u70b9\u63cf\u8ff0"
SHEET_IMAGES = "\u56fe\u7247\u94fe\u63a5"
SHEET_QA = "QA"
SHEET_REVIEWS = "\u8bc4\u8bba"

LISTING_SHEET_ALWAYS_OMITTED_FIELDS = {
    "asin",
    "\u5173\u952e\u8bcd\u641c\u7d22",
    "generic_keyword.value",
    "\u8f93\u5165\u5173\u952e\u8bcd",
    "\u8f93\u5165\u5173\u952e\u8bcd\u5217\u8868",
    "\u5173\u952e\u8bcd\u6570\u91cf",
    "\u5173\u952e\u8bcd\u6765\u6e90",
    "\u8f93\u5165\u884c\u53f7",
    "\u6765\u6e90\u6587\u4ef6",
    "\u4e94\u70b9\u63cf\u8ff0",
    "\u5e97\u94fa/\u90e8\u95e8",
    "\u8d1f\u8d23\u4eba",
    "listid",
}

LISTING_SHEET_DUPLICATE_FIELDS = {
    "\u5546\u54c1\u6807\u9898": "\u4ea7\u54c1\u6807\u9898",
    "\u54c1\u724c": "\u54c1\u724c\u540d",
    "\u4e3b\u56fe\u94fe\u63a5": "\u4e3b\u56fe",
}

# file_key -> (db_column, is_multi) 映射，用于逐文件交付。
# 每个 ASIN 的拆包文件会单独上传，OSS URL 写入 ops_asin_data_report_files
# 中对应的字段。
# - 单文件类型使用 varchar 字段（一个 URL）
# - 多文件类型（keyword_miner / competitor）使用 json 字段（URL 数组）
FILE_FIELD_MAP = {
    "basic": ("basic_data_url", False),
    "bi": ("bi_data_url", False),
    "keyword_reverse": ("keyword_reverse_url", False),
    "keyword_miner": ("keyword_miner_urls", True),
    "competitor": ("competitor_urls", True),
    "rufus": ("rufus_report_url", False),
}

# 与 01~06 拆包文件顺序一致的 file_key 列表。
SPLIT_FILE_KEYS = ("basic", "bi", "keyword_reverse", "keyword_miner", "competitor", "rufus")


def build_split_package(
    *,
    output_root: Path,
    asin_results: list[dict[str, Any]],
    summary: dict[str, Any],
    file_keys: tuple[str, ...] = SPLIT_FILE_KEYS,
    include_zip: bool = True,
) -> dict[str, Any]:
    """按需生成当前拆包 Excel/Markdown 文件，并可选压缩为 zip。"""
    selected_keys = tuple(key for key in SPLIT_FILE_KEYS if key in set(file_keys))
    package_dir = output_root / "asin-data-packages"
    if package_dir.exists():
        shutil.rmtree(package_dir)
    package_dir.mkdir(parents=True, exist_ok=True)

    items: list[dict[str, Any]] = []
    for asin_result in asin_results:
        asin = normalize_asin(asin_result.get("asin"))
        if not asin:
            continue
        asin_dir = package_dir / asin
        asin_dir.mkdir(parents=True, exist_ok=True)
        all_files = {
            "basic": asin_dir / FILE_BASIC,
            "bi": asin_dir / FILE_BI,
            "keyword_reverse": asin_dir / FILE_KEYWORD_REVERSE,
            "keyword_miner": asin_dir / FILE_KEYWORD_MINER,
            "competitor": asin_dir / FILE_COMPETITOR,
            "rufus": asin_dir / FILE_RUFUS,
        }
        files = {key: all_files[key] for key in selected_keys}
        if "basic" in files:
            write_basic_workbook(files["basic"], asin_result)
        if "bi" in files:
            write_bi_workbook(files["bi"], asin_result)
        seller_sprite = asin_result.get("seller_sprite") if isinstance(asin_result.get("seller_sprite"), dict) else {}
        if "keyword_reverse" in files:
            write_seller_sprite_workbook(files["keyword_reverse"], seller_sprite.get("keyword_reverse"), title="\u5173\u952e\u8bcd\u53cd\u67e5")
        if "keyword_miner" in files:
            write_keyword_miner_workbook(files["keyword_miner"], seller_sprite.get("keyword_miner"))
        if "competitor" in files:
            write_competitor_workbook(files["competitor"], seller_sprite.get("competitor"))
        if "rufus" in files:
            write_rufus_markdown(files["rufus"], asin_result)
        items.append(
            {
                "asin": asin,
                "dir": asin_dir.as_posix(),
                "files": {key: path.as_posix() for key, path in files.items()},
            }
        )

    zip_path: Path | None = None
    if include_zip:
        readme_path = package_dir / "README.md"
        readme_path.write_text(build_readme(summary, items), encoding="utf-8")
        zip_path = output_root / split_package_zip_name(asin_results)
        if zip_path.exists():
            zip_path.unlink()
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for path in sorted(package_dir.rglob("*")):
                if path.is_file():
                    archive.write(path, path.relative_to(package_dir).as_posix())
    return {
        "package_dir": package_dir.as_posix(),
        "zip_path": zip_path.as_posix() if zip_path else None,
        "items": items,
    }


def split_package_zip_name(asin_results: list[dict[str, Any]]) -> str:
    asins = [normalize_asin(item.get("asin")) for item in asin_results if normalize_asin(item.get("asin"))]
    if len(asins) == 1:
        return f"{asins[0]}-asin-data-package.zip"
    return "asin-data-packages.zip"


def write_basic_workbook(path: Path, asin_result: dict[str, Any]) -> None:
    wb = new_workbook()
    query = asin_result.get("query") if isinstance(asin_result.get("query"), dict) else {}
    frontend = asin_result.get("frontend_data") if isinstance(asin_result.get("frontend_data"), dict) else {}
    base = frontend.get("\u57fa\u7840\u6570\u636e") if isinstance(frontend.get("\u57fa\u7840\u6570\u636e"), dict) else {}
    sales_rows = rows_from_source(query.get("sales"))
    legacy_crawler_rows = rows_from_source(query.get("crawler_listing"))
    listing_basic_row = listing_basic_row_from_bi(asin_result)
    crawler_row = crawler_row_from_bi(asin_result) or (legacy_crawler_rows[0] if legacy_crawler_rows else {})
    product_details = parse_jsonish(first_value(crawler_row, "product_details", "f_product_details", "\u5546\u54c1\u8be6\u60c5")) or {}
    if not isinstance(product_details, dict):
        product_details = {}
    listing_main_image = first_value(
        listing_basic_row,
        "\u4e3b\u56fe\u94fe\u63a5",
        "main_product_image_locator.media_location",
        "main_image_url",
    )
    listing_other_images = first_value(listing_basic_row, "\u5176\u4ed6\u9644\u56fe\u94fe\u63a5")
    listing_bullets = first_value(listing_basic_row, "\u4e94\u70b9\u63cf\u8ff0")

    listing_rows = []
    listing_omitted_fields = listing_sheet_omitted_fields(listing_basic_row)
    if listing_basic_row:
        listing_rows.extend(
            {"\u6570\u636e\u7c7b\u578b": "\u520a\u767b\u57fa\u7840\u6570\u636e", "\u5b57\u6bb5": key, "\u503c": value}
            for key, value in listing_basic_row.items()
            if key not in listing_omitted_fields
        )
    if isinstance(base, dict):
        excluded_base_fields = {
            "BI\u9500\u552e\u6570\u636e",
            "\u722c\u866bListing\u6570\u636e",
            "BI\u63a5\u53e3\u6570\u636e",
            "\u9519\u8bef\u5217\u8868",
            *listing_omitted_fields,
        }
        listing_rows.extend({"\u5b57\u6bb5": key, "\u503c": value} for key, value in base.items() if key not in excluded_base_fields)
    if sales_rows:
        listing_rows.extend(add_row_type(sales_rows, "\u9500\u552e/\u520a\u767b\u5173\u8054\u6570\u636e"))
    write_rows(wb.create_sheet(SHEET_LISTING), listing_rows)

    write_rows(wb.create_sheet(SHEET_CRAWLER), legacy_crawler_rows or ([crawler_row] if crawler_row else []))
    write_key_values(wb.create_sheet(SHEET_PRODUCT), product_details)
    bullet_values = as_list(first_value(crawler_row, "five_point_description", "f_five_point_description", "\u4e94\u70b9\u63cf\u8ff0"))
    if not bullet_values:
        bullet_values = as_list(listing_bullets)
    write_rows(wb.create_sheet(SHEET_BULLETS), numbered_rows(bullet_values), ["\u5e8f\u53f7", "\u5185\u5bb9"])
    image_rows: list[dict[str, Any]] = []
    append_image_rows(image_rows, "\u4e3b\u56fe", first_value(crawler_row, "image", "f_image", "\u4e3b\u56fe") or listing_main_image)
    append_image_rows(image_rows, "\u9644\u56fe", first_value(crawler_row, "subplot", "f_subplot", "\u9644\u56fe") or listing_other_images)
    append_image_rows(image_rows, "A+\u56fe\u7247", first_value(crawler_row, "a_image", "f_a_image", "A+\u56fe\u7247"))
    write_rows(wb.create_sheet(SHEET_IMAGES), image_rows, ["\u7c7b\u578b", "\u5e8f\u53f7", "URL"])
    write_rows(wb.create_sheet(SHEET_QA), normalize_rows(first_value(crawler_row, "qa", "f_qa", "QA")))
    write_rows(wb.create_sheet(SHEET_REVIEWS), normalize_rows(first_value(crawler_row, "review_list", "f_review_list", "\u8bc4\u8bba")))
    save_workbook(wb, path)


def write_bi_workbook(path: Path, asin_result: dict[str, Any]) -> None:
    wb = new_workbook()
    sources = bi_sources(asin_result)
    wrote = False
    for key, label in BI_REPORT_SOURCE_ORDER:
        source = sources.get(key) if isinstance(sources.get(key), dict) else {"label": label, "rows": []}
        rows = strip_hidden_rows(source.get("rows") if isinstance(source.get("rows"), list) else [])
        if not rows:
            rows = [{"\u8bf4\u660e": "\u65e0\u6570\u636e"}]
        write_rows(wb.create_sheet(safe_sheet_name(str(source.get("label") or label))), rows)
        wrote = True
    if not wrote:
        write_rows(wb.create_sheet(SHEET_NO_DATA), [{"\u8bf4\u660e": "\u65e0BI\u6570\u636e"}])
    save_workbook(wb, path)


def write_seller_sprite_workbook(path: Path, payload: Any, *, title: str) -> None:
    export_paths = seller_sprite_export_paths(payload)
    if export_paths:
        write_export_workbook(path, export_paths)
        return
    rows = seller_sprite_rows(payload)
    wb = new_workbook()
    write_rows(wb.create_sheet(safe_sheet_name(title)), rows or [{"\u8bf4\u660e": f"\u65e0{title}\u6570\u636e"}])
    save_workbook(wb, path)


def write_keyword_miner_workbook(path: Path, payload: Any) -> None:
    export_paths = seller_sprite_export_paths(payload)
    if export_paths:
        write_export_workbook(path, export_paths)
        return
    rows = seller_sprite_rows(payload)
    wb = new_workbook()
    write_rows(wb.create_sheet("\u5173\u952e\u8bcd\u6316\u6398"), rows or [{"\u8bf4\u660e": "\u65e0\u5173\u952e\u8bcd\u6316\u6398\u6570\u636e"}])
    save_workbook(wb, path)


def write_competitor_workbook(path: Path, payload: Any) -> None:
    export_paths = seller_sprite_export_paths(payload)
    if export_paths:
        write_export_workbook(path, export_paths)
        return
    rows = seller_sprite_rows(payload)
    wb = new_workbook()
    write_rows(wb.create_sheet("\u7ade\u54c1\u5206\u6790"), rows or [{"\u8bf4\u660e": "\u65e0\u7ade\u54c1\u5206\u6790\u6570\u636e"}])
    save_workbook(wb, path)


def write_rufus_markdown(path: Path, asin_result: dict[str, Any]) -> None:
    rufus = asin_result.get("rufus") if isinstance(asin_result.get("rufus"), dict) else {}
    formatted = format_rufus_diagnosis_report(asin_result, rufus)
    if formatted:
        path.write_text(formatted, encoding="utf-8")
        return

    report_path = rufus.get("report_path")
    if isinstance(report_path, str) and report_path.strip():
        source = Path(report_path)
        if source.exists() and source.resolve() != path.resolve():
            shutil.copyfile(source, path)
            return
    summary = rufus_summary(rufus)
    asin = normalize_asin(asin_result.get("asin"))
    site = normalize_asin(asin_result.get("site"))
    lines = [f"# Rufus 数据 - {asin}", ""]
    lines.append(f"- ASIN: {asin}")
    lines.append(f"- 站点: {site}")
    lines.append(f"- 状态: {summary.get('status')}")
    lines.append(f"- 问题数量: {summary.get('question_count')}")
    lines.append(f"- 商品URL: {summary.get('page_url') or ''}")
    lines.append(f"- 原始报告: {summary.get('report_path') or ''}")
    for answer in summary.get("answers") or []:
        if not isinstance(answer, dict):
            continue
        question = str(answer.get("question") or "")
        lines.extend(["", f"## 第 {answer.get('index')} 题", "", "问题:"])
        lines.extend(question.splitlines() or [question])
        display_lines: list[str] = []
        append_rufus_display_blocks(display_lines, answer.get("display_blocks"), answer.get("answer"))
        if display_lines[:2] == ["#### Rufus 展示内容", ""]:
            display_lines = display_lines[2:]
        lines.extend(["", "#### Rufus 展示内容", "", *display_lines])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def format_rufus_diagnosis_report(asin_result: dict[str, Any], rufus: dict[str, Any]) -> str:
    """Render Rufus JSON into the operator-facing diagnosis report format."""
    if not isinstance(rufus, dict):
        return ""
    has_structured_report = any(
        isinstance(rufus.get(key), dict)
        for key in (
            "diagnosis_report",
            "diagnosisReport",
            "listing_optimization_report",
            "listingOptimizationReport",
        )
    )
    answers = rufus.get("answers")
    if not has_structured_report and not (isinstance(answers, list) and answers):
        return ""

    payload = dict(rufus)
    payload.setdefault("asin", asin_result.get("asin"))
    payload.setdefault("country", asin_result.get("site") or rufus.get("country"))
    rendered = AnswerReportFormatter().format_data(payload).strip()
    if not rendered.startswith("# ASIN "):
        return ""
    return rendered + "\n"


def write_export_workbook(path: Path, export_paths: list[Path]) -> None:
    wb = new_workbook()
    wrote = False
    for export_index, export_path in enumerate(export_paths, start=1):
        try:
            source_wb = load_workbook(export_path, read_only=True, data_only=True)
        except Exception:
            continue
        try:
            for sheet_name in source_wb.sheetnames:
                if is_hidden_sheet_name(sheet_name):
                    continue
                source_ws = source_wb[sheet_name]
                target_name = sheet_name if len(export_paths) == 1 else f"{export_index}-{sheet_name}"
                ws = wb.create_sheet(safe_sheet_name(target_name))
                for row in source_ws.iter_rows(values_only=True):
                    ws.append([cell_value(value) for value in row])
                wrote = True
        finally:
            source_wb.close()
    if not wrote:
        write_rows(wb.create_sheet(SHEET_NO_DATA), [{"\u8bf4\u660e": "\u65e0\u53ef\u8bfb\u53d6\u7684Excel\u6570\u636e"}])
    save_workbook(wb, path)


def seller_sprite_export_paths(payload: Any) -> list[Path]:
    paths: list[Path] = []

    def collect(item: Any) -> None:
        if not isinstance(item, dict):
            return
        for key in ("export_path", "path"):
            value = item.get(key)
            if isinstance(value, str) and value.strip():
                add_export_path(paths, value)
        export = item.get("export")
        if isinstance(export, dict):
            value = export.get("path")
            if isinstance(value, str) and value.strip():
                add_export_path(paths, value)

    collect(payload)
    if isinstance(payload, dict):
        for job in payload.get("jobs") or []:
            collect(job)
    return paths


def add_export_path(paths: list[Path], value: str) -> None:
    path = Path(value)
    if path.exists() and path.suffix.lower() in {".xlsx", ".xlsm"} and path not in paths:
        paths.append(path)


def seller_sprite_rows(payload: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if isinstance(payload, dict):
        raw_rows = payload.get("rows")
        if isinstance(raw_rows, list):
            rows.extend(item for item in raw_rows if isinstance(item, dict))
        full_result = payload.get("full_result")
        if isinstance(full_result, dict) and isinstance(full_result.get("data"), list):
            rows.extend(item for item in full_result["data"] if isinstance(item, dict))
        for job in payload.get("jobs") or []:
            if isinstance(job, dict):
                rows.extend(seller_sprite_rows(job))
    return rows


def crawler_row_from_bi(asin_result: dict[str, Any]) -> dict[str, Any]:
    source = bi_sources(asin_result).get("crawler_details")
    if not isinstance(source, dict):
        return {}
    rows = source.get("rows")
    return rows[0] if isinstance(rows, list) and rows and isinstance(rows[0], dict) else {}


def listing_basic_row_from_bi(asin_result: dict[str, Any]) -> dict[str, Any]:
    source = bi_sources(asin_result).get("listing_basic")
    if not isinstance(source, dict):
        return {}
    rows = source.get("rows")
    return rows[0] if isinstance(rows, list) and rows and isinstance(rows[0], dict) else {}


def bi_sources(asin_result: dict[str, Any]) -> dict[str, Any]:
    payload = asin_result.get("bi_report_data") if isinstance(asin_result.get("bi_report_data"), dict) else {}
    sources = payload.get("sources") if isinstance(payload.get("sources"), dict) else {}
    return sources


def rows_from_source(payload: Any) -> list[dict[str, Any]]:
    if not isinstance(payload, dict):
        return []
    rows = payload.get("rows")
    return [row for row in rows if isinstance(row, dict)] if isinstance(rows, list) else []


def strip_hidden_rows(rows: list[Any]) -> list[dict[str, Any]]:
    return [strip_hidden_keys(row) for row in rows if isinstance(row, dict)]


def strip_hidden_keys(value: Any) -> Any:
    hidden = {"endpoint", "raw", "request_endpoint", "request_url", "headers", "cookies"}
    if isinstance(value, dict):
        return {key: strip_hidden_keys(item) for key, item in value.items() if str(key) not in hidden}
    if isinstance(value, list):
        return [strip_hidden_keys(item) for item in value]
    return value


def new_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def save_workbook(wb: Workbook, path: Path) -> None:
    if not wb.worksheets:
        write_rows(wb.create_sheet(SHEET_NO_DATA), [{"\u8bf4\u660e": "\u65e0\u6570\u636e"}])
    for ws in wb.worksheets:
        format_sheet(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_rows(ws: Any, rows: Any, headers: list[str] | None = None) -> None:
    normalized = normalize_rows(rows)
    if headers is None:
        headers = collect_headers(normalized)
    if not headers:
        headers = ["\u8bf4\u660e"]
        normalized = [{"\u8bf4\u660e": "\u65e0\u6570\u636e"}]
    ws.append(headers)
    for row_data in normalized:
        ws.append([cell_value(row_data.get(header)) for header in headers])


def write_key_values(ws: Any, payload: dict[str, Any]) -> None:
    write_rows(ws, [{"\u5b57\u6bb5": key, "\u503c": value} for key, value in payload.items()], ["\u5b57\u6bb5", "\u503c"])


def normalize_rows(value: Any) -> list[dict[str, Any]]:
    value = parse_jsonish(value)
    if isinstance(value, list):
        return [item if isinstance(item, dict) else {"\u503c": item} for item in value]
    if isinstance(value, dict):
        return [value]
    if value in (None, ""):
        return []
    return [{"\u503c": value}]


def collect_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for row in rows:
        for key in row.keys():
            if str(key) not in headers:
                headers.append(str(key))
    return headers


def cell_value(value: Any) -> Any:
    value = parse_jsonish(value)
    if isinstance(value, list):
        if all(not isinstance(item, (dict, list)) for item in value):
            return trim_cell("\n".join(str(item) for item in value))
        return trim_cell(json.dumps(value, ensure_ascii=False, default=str))
    if isinstance(value, dict):
        return trim_cell(json.dumps(value, ensure_ascii=False, default=str))
    if value is None:
        return ""
    return trim_cell(value) if isinstance(value, str) else value


def trim_cell(text: str) -> str:
    if len(text) <= EXCEL_CELL_LIMIT:
        return text
    return text[: EXCEL_CELL_LIMIT - 20] + "\n...[truncated]"


def parse_jsonish(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text or text[0] not in "[{":
        return value
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return value


def as_list(value: Any) -> list[Any]:
    value = parse_jsonish(value)
    if isinstance(value, list):
        return value
    if value in (None, ""):
        return []
    return [value]


def numbered_rows(values: list[Any]) -> list[dict[str, Any]]:
    return [{"\u5e8f\u53f7": index, "\u5185\u5bb9": value} for index, value in enumerate(values, start=1)]


def append_image_rows(rows: list[dict[str, Any]], image_type: str, value: Any) -> None:
    for index, url in enumerate(as_list(value), start=1):
        rows.append({"\u7c7b\u578b": image_type, "\u5e8f\u53f7": index, "URL": url})


def add_row_type(rows: list[dict[str, Any]], row_type: str) -> list[dict[str, Any]]:
    return [{"\u7c7b\u578b": row_type, **row} for row in rows]


def listing_sheet_omitted_fields(row: dict[str, Any]) -> set[str]:
    """计算刊登数据 sheet 中需要隐藏的输入、内部和重复字段。"""
    omitted = set(LISTING_SHEET_ALWAYS_OMITTED_FIELDS)
    for source_field, template_field in LISTING_SHEET_DUPLICATE_FIELDS.items():
        if has_cell_value(row.get(template_field)):
            omitted.add(source_field)
    if any(has_cell_value(row.get(f"\u526f\u56fe{index}")) for index in range(1, 9)):
        omitted.add("\u5176\u4ed6\u9644\u56fe\u94fe\u63a5")
    return omitted


def has_cell_value(value: Any) -> bool:
    if isinstance(value, str):
        return bool(value.strip())
    return value not in (None, "", [], {})


def flatten_payload(payload: Any, prefix: str = "") -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if isinstance(payload, dict):
        for key, value in payload.items():
            path = f"{prefix}.{key}" if prefix else str(key)
            if isinstance(value, (dict, list)):
                rows.extend(flatten_payload(value, path))
            else:
                rows.append({"\u8def\u5f84": path, "\u503c": value})
    elif isinstance(payload, list):
        for index, value in enumerate(payload, start=1):
            path = f"{prefix}[{index}]"
            if isinstance(value, (dict, list)):
                rows.extend(flatten_payload(value, path))
            else:
                rows.append({"\u8def\u5f84": path, "\u503c": value})
    elif payload not in (None, ""):
        rows.append({"\u8def\u5f84": prefix or "root", "\u503c": payload})
    return rows


def first_value(payload: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in payload and payload.get(key) not in (None, ""):
            return payload.get(key)
    return None


def kv_row(key: str, value: Any) -> dict[str, Any]:
    return {"\u6570\u636e\u9879": key, "\u503c": value}


def safe_sheet_name(name: str) -> str:
    safe = re.sub(r"[\[\]\:\*\?\/\\]", "_", name).strip() or "Sheet"
    return safe[:31]


def is_hidden_sheet_name(name: str) -> bool:
    upper = name.upper()
    return name in {"\u6458\u8981", "\u5b8c\u6574JSON"} or "JSON" in upper


def format_sheet(ws: Any) -> None:
    if ws.max_row >= 1:
        fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_len = 10
        for cell in column_cells[:200]:
            value = "" if cell.value is None else str(cell.value)
            first_line = value.splitlines()[0] if value else ""
            max_len = max(max_len, min(60, len(first_line)))
        ws.column_dimensions[column_letter].width = max_len + 2


def normalize_asin(value: Any) -> str:
    return str(value or "").strip().upper()


def build_readme(summary: dict[str, Any], items: list[dict[str, Any]]) -> str:
    lines = [
        "# ASIN \u6570\u636e\u5305",
        "",
        f"\u8fd0\u884cID: {summary.get('run_id') or ''}",
        "",
        "\u6bcf\u4e2a ASIN \u6587\u4ef6\u5939\u5305\u542b:",
        "",
        f"1. `{FILE_BASIC}`",
        f"2. `{FILE_BI}`",
        f"3. `{FILE_KEYWORD_REVERSE}`",
        f"4. `{FILE_KEYWORD_MINER}`",
        f"5. `{FILE_COMPETITOR}`",
        f"6. `{FILE_RUFUS}`",
        "",
        "## ASIN",
        "",
    ]
    for item in items:
        lines.append(f"- {item.get('asin')}")
    lines.append("")
    return "\n".join(lines)


__all__ = ["build_split_package", "split_package_zip_name"]
