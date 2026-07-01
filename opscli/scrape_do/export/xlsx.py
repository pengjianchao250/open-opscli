"""Scrape.do XLSX 导出。"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from opscli.scrape_do.domain.exceptions import ScrapeDoConfigError
from opscli.scrape_do.domain.models import ScrapeDoExportResult

EXCEL_CELL_LIMIT = 32767


@dataclass(frozen=True)
class ExportColumn:
    title: str
    source: str


FIELD_TITLES = {
    "asin": "ASIN",
    "site": "站点",
    "keyword": "关键词",
    "page": "页码",
    "position": "页内排名",
    "brand": "品牌",
    "title": "标题",
    "url": "URL",
    "thumbnail": "主图",
    "image_url": "图片",
    "price": "价格",
    "list_price": "标价",
    "shipping_price": "运费",
    "total_price": "总价",
    "currency": "币种",
    "rating": "评分",
    "rating_count": "评分数",
    "total_ratings": "总评分数",
    "review_count_text": "评论数文本",
    "is_sponsored": "广告位",
    "is_prime": "Prime",
    "badge": "Badge",
    "seller_id": "Seller ID",
    "merchant_name": "卖家名称",
    "condition": "商品状态",
    "is_buybox_winner": "Buy Box",
    "is_fba": "FBA",
    "ships_from": "发货地",
    "delivery_date": "配送日期",
    "quantity": "库存数量",
    "description": "描述",
    "shipping_info": "配送信息",
    "best_seller_rankings": "BSR",
    "technical_details": "技术规格",
}


def extract_rows(scenario: str, payload: dict[str, Any]) -> list[dict[str, Any]]:
    if scenario == "amazon-pdp":
        return [_pdp_row(payload)] if payload else []
    if scenario == "amazon-offer-listing":
        asin = str(payload.get("asin") or "")
        offers = payload.get("offers")
        if not isinstance(offers, list):
            return []
        return [_offer_row(asin, offer) for offer in offers if isinstance(offer, dict)]
    if scenario == "amazon-search":
        keyword = str(payload.get("keyword") or "")
        page = payload.get("page")
        products = payload.get("products")
        if not isinstance(products, list):
            return []
        return [_search_row(keyword, page, product) for product in products if isinstance(product, dict)]
    return [_normalize_row(payload)]


def export_rows_to_xlsx(
    *,
    rows: list[dict[str, Any]],
    output_path: Path,
    scenario: str,
    site: str = "US",
    params: dict[str, Any] | None = None,
    raw_payload: dict[str, Any] | None = None,
) -> ScrapeDoExportResult:
    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise ScrapeDoConfigError("缺少 openpyxl 依赖，无法导出 XLSX") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(write_only=True)
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True)

    _write_rows_sheet(
        workbook=workbook,
        title=_safe_sheet_title(f"{scenario}-{site}"),
        rows=rows,
        header_fill=header_fill,
        header_font=header_font,
        write_only_cell=WriteOnlyCell,
        get_column_letter=get_column_letter,
    )
    if raw_payload:
        _write_raw_payload_sheets(
            workbook=workbook,
            scenario=scenario,
            site=site,
            raw_payload=raw_payload,
            header_fill=header_fill,
            header_font=header_font,
            write_only_cell=WriteOnlyCell,
            get_column_letter=get_column_letter,
        )

    workbook.save(output_path)
    resolved = output_path.resolve()
    return ScrapeDoExportResult(path=str(resolved), filename=resolved.name, url=resolved.as_uri())


def _write_rows_sheet(
    *,
    workbook,
    title: str,
    rows: list[dict[str, Any]],
    header_fill,
    header_font,
    write_only_cell,
    get_column_letter,
) -> None:
    sheet = workbook.create_sheet(title=title)
    columns = _columns_from_rows(rows)
    _append_header(sheet, [column.title for column in columns], header_fill, header_font, write_only_cell)
    for row in rows:
        normalized = _normalize_row(row)
        sheet.append([_cell_value(normalized.get(column.source)) for column in columns])
    _set_column_widths(sheet, [column.title for column in columns], get_column_letter)


def _write_raw_payload_sheets(
    *,
    workbook,
    scenario: str,
    site: str,
    raw_payload: dict[str, Any],
    header_fill,
    header_font,
    write_only_cell,
    get_column_letter,
) -> None:
    asin = str(raw_payload.get("asin") or "")
    fields = []
    array_fields: list[tuple[str, list[Any]]] = []
    for key, value in raw_payload.items():
        fields.append({"field": key, "type": _value_type(value), "value": _json_cell(value)})
        if isinstance(value, list) and value:
            array_fields.append((key, value))
    _write_named_rows_sheet(
        workbook=workbook,
        title="Raw Fields",
        rows=fields,
        columns=[ExportColumn("字段", "field"), ExportColumn("类型", "type"), ExportColumn("值", "value")],
        header_fill=header_fill,
        header_font=header_font,
        write_only_cell=write_only_cell,
        get_column_letter=get_column_letter,
    )
    for key, values in array_fields:
        rows = _array_rows(asin=asin, values=values)
        _write_named_rows_sheet(
            workbook=workbook,
            title=_raw_array_sheet_title(key),
            rows=rows,
            columns=_raw_columns_from_rows(rows),
            header_fill=header_fill,
            header_font=header_font,
            write_only_cell=write_only_cell,
            get_column_letter=get_column_letter,
        )


def _write_named_rows_sheet(
    *,
    workbook,
    title: str,
    rows: list[dict[str, Any]],
    columns: list[ExportColumn],
    header_fill,
    header_font,
    write_only_cell,
    get_column_letter,
) -> None:
    sheet = workbook.create_sheet(title=_safe_sheet_title(title))
    _append_header(sheet, [column.title for column in columns], header_fill, header_font, write_only_cell)
    for row in rows:
        normalized = _normalize_row(row)
        sheet.append([_cell_value(normalized.get(column.source)) for column in columns])
    _set_column_widths(sheet, [column.title for column in columns], get_column_letter)


def _append_header(sheet, titles: list[str], header_fill, header_font, write_only_cell) -> None:
    header_cells = []
    for title in titles:
        cell = write_only_cell(sheet, value=title)
        cell.font = header_font
        cell.fill = header_fill
        header_cells.append(cell)
    sheet.append(header_cells)


def _set_column_widths(sheet, titles: list[str], get_column_letter) -> None:
    for column_index, title in enumerate(titles, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(len(str(title)) + 2, 12), 48)


def _array_rows(*, asin: str, values: list[Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in values:
        if isinstance(item, dict):
            row = dict(item)
            if asin and "asin" not in {str(key).lower() for key in row.keys()}:
                row = {"asin": asin, **row}
        else:
            row = {"value": item}
            if asin:
                row = {"asin": asin, **row}
        rows.append(row)
    return rows


def _raw_array_sheet_title(key: str) -> str:
    words = [part.capitalize() for part in key.replace("-", "_").split("_") if part]
    return _safe_sheet_title("Raw " + (" ".join(words) or "Items"))


def _value_type(value: Any) -> str:
    if isinstance(value, dict):
        return "dict"
    if isinstance(value, list):
        return "list"
    if value is None:
        return "null"
    return type(value).__name__


def _pdp_row(payload: dict[str, Any]) -> dict[str, Any]:
    row = {
        "asin": payload.get("asin"),
        "brand": payload.get("brand"),
        "title": payload.get("name") or payload.get("title"),
        "url": payload.get("url"),
        "thumbnail": payload.get("thumbnail"),
        "price": payload.get("price"),
        "list_price": payload.get("list_price"),
        "currency": payload.get("currency"),
        "rating": payload.get("rating"),
        "total_ratings": payload.get("total_ratings"),
        "is_prime": payload.get("is_prime"),
        "description": payload.get("description"),
        "shipping_info": _json_cell(payload.get("shipping_info")),
        "best_seller_rankings": _json_cell(payload.get("best_seller_rankings")),
        "technical_details": _json_cell(payload.get("technical_details")),
    }
    return {key: value for key, value in row.items() if value is not None and value != ""}


def _offer_row(asin: str, offer: dict[str, Any]) -> dict[str, Any]:
    listing_price = _money_amount(offer.get("listingPrice"))
    shipping_price = _money_amount(offer.get("shipping"))
    currency = _money_currency(offer.get("listingPrice")) or _money_currency(offer.get("shipping"))
    prime = offer.get("primeInformation") if isinstance(offer.get("primeInformation"), dict) else {}
    shipping_time = offer.get("shippingTime") if isinstance(offer.get("shippingTime"), dict) else {}
    row = {
        "asin": asin,
        "seller_id": offer.get("sellerId"),
        "merchant_name": offer.get("merchantName"),
        "condition": offer.get("condition"),
        "listing_price": listing_price,
        "shipping_price": shipping_price,
        "total_price": _sum_money(listing_price, shipping_price),
        "currency": currency,
        "is_buybox_winner": offer.get("isBuyBoxWinner"),
        "is_fba": offer.get("isFulfilledByAmazon"),
        "is_prime": prime.get("isPrime"),
        "ships_from": offer.get("shipsFrom"),
        "delivery_date": shipping_time.get("deliveryDate"),
        "quantity": offer.get("quantity"),
    }
    return {key: value for key, value in row.items() if value is not None and value != ""}


def _search_row(keyword: str, page: Any, product: dict[str, Any]) -> dict[str, Any]:
    rating = product.get("rating") if isinstance(product.get("rating"), dict) else {}
    price = product.get("price")
    row = {
        "keyword": keyword,
        "page": page,
        "position": product.get("position"),
        "asin": product.get("asin"),
        "title": product.get("title"),
        "url": product.get("url"),
        "image_url": product.get("imageUrl"),
        "price": _money_amount(price),
        "currency": _money_currency(price),
        "rating": rating.get("value"),
        "rating_count": rating.get("count"),
        "review_count_text": product.get("reviewCount"),
        "is_sponsored": product.get("isSponsored"),
        "is_prime": product.get("isPrime"),
        "badge": product.get("badge"),
    }
    return {key: value for key, value in row.items() if value is not None and value != ""}


def _money_amount(value: Any) -> Any:
    return value.get("amount") if isinstance(value, dict) else None


def _money_currency(value: Any) -> Any:
    return value.get("currencyCode") if isinstance(value, dict) else None


def _sum_money(left: Any, right: Any) -> Any:
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return left + right
    if isinstance(left, (int, float)):
        return left
    return None


def _columns_from_rows(rows: list[dict[str, Any]]) -> list[ExportColumn]:
    keys = _keys_from_rows(rows)
    return [ExportColumn(FIELD_TITLES.get(key, key), key) for key in keys]


def _raw_columns_from_rows(rows: list[dict[str, Any]]) -> list[ExportColumn]:
    keys = _keys_from_rows(rows)
    return [ExportColumn(key, key) for key in keys]


def _keys_from_rows(rows: list[dict[str, Any]]) -> list[str]:
    keys: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in keys:
                keys.append(key)
    if not keys:
        keys = ["value"]
    return keys


def _normalize_row(row: Any) -> dict[str, Any]:
    if isinstance(row, dict):
        return {str(key): _json_cell(value) for key, value in row.items()}
    return {"value": _json_cell(row)}


def _json_cell(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return value


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        value = _json_cell(value)
    if isinstance(value, str) and len(value) > EXCEL_CELL_LIMIT:
        return value[: EXCEL_CELL_LIMIT - 3] + "..."
    return value


def _safe_sheet_title(value: str) -> str:
    text = "".join("-" if char in r"[]:*?/\\" else char for char in value)
    return (text or "scrape-do")[:31]
