"""图表数据 Excel 导出脚本（MCP 无状态模式）。

从 MCP query_chart 获取并保存的 chart run 结果中提取明细、小计、总计数据，
按透视表格式写入 Excel 文件（.xlsx）。
自动将 global_alias 列名映射为可读的 verbose_name。

================================================================================
MCP 使用指南
================================================================================

【前置要求】
1. 先检查 session：auth_is_authenticated(session_id="xxx")

【获取图表数据】
通过 MCP query_chart Tool 获取并执行查询：

    query_chart(
        chart_uuid="4NQ5f66sU9",
        run=True,
        session_id="860b0636485b5188a2b9b4ed5210e736"
    )

将返回的 JSON 保存到文件（如 /tmp/chart_result.json），然后通过 --input 传入本脚本。

【本地数据缺失时】
如字段映射失败，调用 MCP skills_upgrade Tool 更新本地索引：

    skills_upgrade(name="ops-dataset-query", skills_dir="/Users/mask/.config/opencode/skills")

================================================================================

用法：
    python excel_export_mcp.py --input /tmp/chart_result.json --output /tmp/output.xlsx
    python excel_export_mcp.py --input /tmp/chart_result.json --output /tmp/output.xlsx --sheet-name 销售数据

前置依赖：
    pip install openpyxl

格式规范：
    - 表头：蓝色背景（4472C4）白色粗体字，冻结首行
    - 明细行：数值列千分位格式（#,##0.00），百分比列 0.00% 格式
    - 小计行：灰色背景（D9E2F3），粗体字
    - 总计行：深蓝背景（4472C4）白色粗体字
    - 负值：红色字体（FF0000）标注亏损
    - 列宽：根据内容自动调整（最大 50 字符）
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from chart_map import map_chart_queries, map_query_results
from core import discover_data_dir, load_local_index, to_float

# ---------------------------------------------------------------------------
# 百分比列检测关键词
# ---------------------------------------------------------------------------

_PCT_ZH = ("率", "占比", "比率", "%")
_PCT_EN = ("rate", "ratio", "pct", "percent")

# ---------------------------------------------------------------------------
# Excel 样式常量
# ---------------------------------------------------------------------------

_HEADER_BG = "4472C4"
_SUBTOTAL_BG = "D9E2F3"
_TOTAL_BG = "4472C4"
_NEG_COLOR = "FF0000"
_FMT_NUMBER = "#,##0.00"
_FMT_PCT = "0.00%"


# ---------------------------------------------------------------------------
# 列定义构建
# ---------------------------------------------------------------------------


def _is_pct_col(verbose_name: str, field_name: str) -> bool:
    """判断字段是否为百分比类型列。"""
    vn = verbose_name.lower()
    fn = field_name.lower()
    return any(k in vn for k in _PCT_ZH) or any(k in fn for k in _PCT_EN)


def _build_col_layout(mapped_q0: dict) -> list[dict]:
    """从 queries[0] 的映射信息中提取有序列定义。"""
    cols = []
    for fm in mapped_q0.get("_mapping", {}).get("field_mappings", []):
        fi = fm.get("field_info", {})
        alias = fm.get("alias", "")
        name = fm.get("mapped_name", alias)
        field_type = fi.get("field_type", "")
        field_name = fi.get("field_name", "")
        cols.append({
            "alias": alias,
            "name": name,
            "field_type": field_type,
            "is_pct": _is_pct_col(name, field_name),
        })
    return cols


# ---------------------------------------------------------------------------
# 行类型判断
# ---------------------------------------------------------------------------


def _get_row_type(qi: int, total_queries: int) -> str:
    """根据 query 索引判断行类型。"""
    if total_queries == 1:
        return "detail"
    if qi == 0:
        return "detail"
    if qi == total_queries - 1:
        return "total"
    return "subtotal"


def _find_marker_col(cols: list[dict], q_groupby: set[str], row_type: str) -> str | None:
    """找到放置小计/总计标记文本的列别名。"""
    if row_type == "detail":
        return None
    for col in cols:
        if col["field_type"] == "dimension" and col["alias"] not in q_groupby:
            return col["alias"]
    for col in cols:
        if col["field_type"] == "dimension":
            return col["alias"]
    return None


# ---------------------------------------------------------------------------
# Excel 写入核心
# ---------------------------------------------------------------------------


def export_to_excel(
    queries: list[dict],
    mapped_queries: list[dict],
    output_path: str,
    sheet_name: str = "数据透视表",
) -> dict:
    """将图表查询结果写入 Excel 文件。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("缺少依赖：请先执行 pip install openpyxl", file=sys.stderr)
        raise SystemExit(1)

    if not mapped_queries:
        return {"success": False, "error": "映射结果为空，无数据可导出"}

    cols = _build_col_layout(mapped_queries[0])
    if not cols:
        return {"success": False, "error": "无法从 queries[0] 提取列定义，请检查字段映射"}

    wb = Workbook()
    ws = wb.active
    assert ws is not None, "Workbook 初始化失败，未获得活动 Sheet"
    ws.title = sheet_name

    # 1. 写入表头
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_font = Font(bold=True, color="FFFFFF")
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col["name"])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    # 2. 遍历所有 query，按行类型写入数据
    n_queries = len(queries)
    subtotal_fill = PatternFill("solid", fgColor=_SUBTOTAL_BG)
    total_fill = PatternFill("solid", fgColor=_TOTAL_BG)

    marker_text_map = {"subtotal": "小计", "total": "总计"}

    for qi, q in enumerate(queries):
        row_type = _get_row_type(qi, n_queries)
        result_data = (q.get("result") or {}).get("data", [])
        if not result_data:
            continue

        q_groupby: set[str] = set(
            (q.get("payload") or {}).get("query", {}).get("groupBy", [])
        )
        marker_alias = _find_marker_col(cols, q_groupby, row_type)
        marker_text = marker_text_map.get(row_type, "")

        for row_data in result_data:
            ri = ws.max_row + 1

            for ci, col in enumerate(cols, 1):
                alias = col["alias"]
                raw = row_data.get(alias)

                if row_type != "detail":
                    if alias == marker_alias:
                        val: object = marker_text
                    elif col["field_type"] == "dimension" and alias not in q_groupby:
                        val = ""
                    else:
                        val = to_float(raw) if col["field_type"] == "metric" and raw not in (None, "") else raw
                else:
                    val = to_float(raw) if col["field_type"] == "metric" and raw not in (None, "") else raw

                cell = ws.cell(row=ri, column=ci, value=val)

                is_negative = isinstance(val, (int, float)) and val < 0

                if row_type == "subtotal":
                    cell.fill = subtotal_fill
                    font_color = _NEG_COLOR if is_negative else "000000"
                    cell.font = Font(bold=True, color=font_color)
                elif row_type == "total":
                    cell.fill = total_fill
                    font_color = _NEG_COLOR if is_negative else "FFFFFF"
                    cell.font = Font(bold=True, color=font_color)
                elif is_negative:
                    cell.font = Font(color=_NEG_COLOR)

                if col["field_type"] == "metric" and isinstance(val, (int, float)):
                    cell.number_format = _FMT_PCT if col["is_pct"] else _FMT_NUMBER

    # 3. 列宽自适应
    for ci in range(1, len(cols) + 1):
        max_len = 0
        for ri in range(1, ws.max_row + 1):
            v = ws.cell(row=ri, column=ci).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 50)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    total_rows = ws.max_row - 1
    col_names = [col["name"] for col in cols]
    return {
        "success": True,
        "output": str(out.resolve()),
        "rows": total_rows,
        "columns": col_names,
    }


# ---------------------------------------------------------------------------
# 本地工具函数（不依赖 chart_analyze.py）
# ---------------------------------------------------------------------------


def _load_chart_data_from_file(input_path: str) -> tuple[list[dict], str | None, dict | None]:
    """加载图表数据，返回 (queries, chart_uuid, merged_data)。仅从文件读取。"""
    with open(input_path, "r", encoding="utf-8") as f:
        content = json.load(f)

    if isinstance(content, dict) and "data" in content:
        data = content["data"]
        if "queries" in data:
            queries = data["queries"]
            chart_uuid = data.get("chart_uuid")
            merged = data.get("merged")
            return queries, chart_uuid, merged
        elif "result" in data:
            rows = data["result"].get("data", [])
            return [{"index": 0, "result": data["result"]}], None, {"rows": rows}
    elif isinstance(content, list):
        return content, None, None

    return [], None, None


def _check_mapping_hit(mapped_queries: list[dict]) -> bool:
    """检查映射结果中是否有任意字段命中了本地索引。"""
    for q in mapped_queries:
        for fm in q.get("_mapping", {}).get("field_mappings", []):
            fi = fm.get("field_info", {})
            if fi and fi.get("field_name"):
                return True
    return False


# ---------------------------------------------------------------------------
# 入口
# ---------------------------------------------------------------------------


def main() -> None:
    """脚本入口：解析参数，加载数据，执行 Excel 导出。"""
    parser = argparse.ArgumentParser(description="图表数据 Excel 导出工具（MCP 无状态模式）")
    parser.add_argument("--input", required=True, help="已保存的 chart run JSON 文件路径（由 MCP query_chart 获取）")
    parser.add_argument("--output", default="/tmp/output.xlsx", help="输出 Excel 文件路径（默认 /tmp/output.xlsx）")
    parser.add_argument("--sheet-name", default="数据透视表", help="Sheet 名称（默认：数据透视表）")
    parser.add_argument("--skills-dir", help="指定 Skill 安装根目录")
    parser.add_argument("--data-dir", help="直接指定数据目录路径（最高优先级）")
    parser.add_argument("--pretty", action="store_true", help="格式化 JSON 输出")
    args = parser.parse_args()

    # 1. 确定本地数据目录
    if args.data_dir:
        data_dir = Path(args.data_dir)
    else:
        discovered = discover_data_dir(skills_dir=args.skills_dir)
        if discovered is None:
            err = {
                "success": False,
                "error": "未找到 ops-dataset-query 数据目录。请先通过 MCP skills_upgrade 更新数据。",
                "mcp_hint": "调用 skills_upgrade(name='ops-dataset-query') 后重试",
            }
            print(json.dumps(err, ensure_ascii=False), file=sys.stderr)
            raise SystemExit(1)
        data_dir = discovered

    # 2. 加载本地字段索引
    dataset_index, field_index = load_local_index(data_dir)

    # 3. 加载图表查询数据（仅从文件）
    queries, chart_uuid, _merged = _load_chart_data_from_file(args.input)

    if not queries:
        err = {"success": False, "error": "未获取到任何 query 数据，请检查 --input 参数"}
        print(json.dumps(err, ensure_ascii=False, indent=2 if args.pretty else None))
        raise SystemExit(1)

    # 4. 字段别名映射
    mapped_queries = map_chart_queries(queries, dataset_index, field_index)

    # 字段映射全部失败时提示
    if not _check_mapping_hit(mapped_queries):
        print(
            json.dumps(
                {
                    "success": False,
                    "error": "本地字段映射未命中任何字段，可能本地数据已过期。",
                    "mcp_hint": "调用 skills_upgrade(name='ops-dataset-query', force=True) 后重试",
                },
                ensure_ascii=False,
            ),
            file=sys.stderr,
        )
        raise SystemExit(1)

    # 5. 导出 Excel
    result = export_to_excel(
        queries=queries,
        mapped_queries=mapped_queries,
        output_path=args.output,
        sheet_name=args.sheet_name,
    )

    if chart_uuid:
        result["chart_uuid"] = chart_uuid

    print(json.dumps(result, ensure_ascii=False, indent=2 if args.pretty else None))


if __name__ == "__main__":
    main()
