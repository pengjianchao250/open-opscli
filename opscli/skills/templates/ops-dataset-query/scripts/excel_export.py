"""图表数据 Excel 导出脚本。

从 opscli query chart --run 的结果中提取明细、小计、总计数据，
按透视表格式写入 Excel 文件（.xlsx）。
自动将 global_alias 列名映射为可读的 verbose_name。

用法：
    # 从已保存的 chart run 结果导出
    python excel_export.py --input /tmp/chart_result.json --output /tmp/output.xlsx

    # 通过 UUID 直接获取并导出（需要 opscli 已认证）
    python excel_export.py --uuid <chart_uuid> --output /tmp/output.xlsx

    # 自定义 Sheet 名称
    python excel_export.py --input /tmp/chart_result.json --output /tmp/output.xlsx --sheet-name 销售数据

    # 显式指定 Skill 数据目录（安装在非标准位置时）
    python excel_export.py --input /tmp/chart_result.json --output /tmp/output.xlsx --skills-dir ~/.openclaw/skills

前置依赖：
    pip install openpyxl

格式规范：
    - 表头：蓝色背景（4472C4）白色粗体字，冻结首行
    - 明细行：数值列千分位格式（#,##0.00），百分比列 0.00% 格式
    - 小计行：灰色背景（D9E2F3），粗体字
    - 总计行：深蓝背景（4472C4）白色粗体字
    - 负值：红色字体（FF0000）标注亏损
    - 列宽：根据内容自动调整（最大 50 字符）
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from chart_map import discover_data_dir, load_local_index, map_chart_queries
from chart_analyze import _check_mapping_hit, _try_upgrade, load_chart_data
from core import to_float

# ---------------------------------------------------------------------------
# 百分比列检测关键词
# ---------------------------------------------------------------------------

# verbose_name 包含以下中文关键词时，视为百分比列
_PCT_ZH = ("率", "占比", "比率", "%")
# field_name 包含以下英文关键词时，视为百分比列
_PCT_EN = ("rate", "ratio", "pct", "percent")

# ---------------------------------------------------------------------------
# Excel 样式常量
# ---------------------------------------------------------------------------

# 表头背景色（蓝色）
_HEADER_BG = "4472C4"
# 小计行背景色（浅灰蓝）
_SUBTOTAL_BG = "D9E2F3"
# 总计行背景色（深蓝，与表头相同）
_TOTAL_BG = "4472C4"
# 负值字体颜色（红色）
_NEG_COLOR = "FF0000"
# 数值列格式（千分位两位小数）
_FMT_NUMBER = "#,##0.00"
# 百分比列格式（两位小数百分比）
_FMT_PCT = "0.00%"


# ---------------------------------------------------------------------------
# 列定义构建
# ---------------------------------------------------------------------------


def _is_pct_col(verbose_name: str, field_name: str) -> bool:
    """判断字段是否为百分比类型列。

    通过 verbose_name（中文名）和 field_name（英文名）关键词匹配。

    Args:
        verbose_name: 字段中文业务名
        field_name: 字段英文名

    Returns:
        True 表示该字段为百分比类型
    """
    vn = verbose_name.lower()
    fn = field_name.lower()
    return any(k in vn for k in _PCT_ZH) or any(k in fn for k in _PCT_EN)


def _build_col_layout(mapped_q0: dict) -> list[dict]:
    """从 queries[0] 的映射信息中提取有序列定义。

    保持 queries[0] select 字段的原始顺序，每列包含：
    - alias: global_alias，用于从 result.data 行中取值
    - name: verbose_name，用于 Excel 列头显示；映射失败时回退为 alias
    - field_type: "dimension"（维度）或 "metric"（指标）
    - is_pct: 是否为百分比列（影响 Excel 数字格式）

    Args:
        mapped_q0: map_chart_queries 返回的 queries[0] 条目（含 _mapping）

    Returns:
        列定义列表，顺序与 queries[0] select 保持一致
    """
    cols = []
    for fm in mapped_q0.get("_mapping", {}).get("field_mappings", []):
        fi = fm.get("field_info", {})
        alias = fm.get("alias", "")
        # 有映射时使用 verbose_name，否则回退为 global_alias
        name = fm.get("mapped_name", alias)
        field_type = fi.get("field_type", "")
        field_name = fi.get("field_name", "")
        cols.append({
            "alias": alias,
            "name": name,
            "field_type": field_type,
            "is_pct": _is_pct_col(name, field_name),
        })
    return cols


# ---------------------------------------------------------------------------
# 行类型判断
# ---------------------------------------------------------------------------


def _get_row_type(qi: int, total_queries: int) -> str:
    """根据 query 索引判断行类型。

    约定（与 SKILL.md 规范保持一致）：
    - queries[0]：明细行（detail）
    - queries[1] 到 queries[-2]：小计行（subtotal）
    - queries[-1]：总计行（total）；仅 1 个 query 时全为明细

    Args:
        qi: 当前 query 的索引（0-based）
        total_queries: 所有 queries 的数量

    Returns:
        "detail" / "subtotal" / "total"
    """
    if total_queries == 1:
        return "detail"
    if qi == 0:
        return "detail"
    if qi == total_queries - 1:
        return "total"
    return "subtotal"


def _find_marker_col(cols: list[dict], q_groupby: set[str], row_type: str) -> str | None:
    """找到放置小计/总计标记文本的列别名。

    规则：取第一个维度列且其 alias 不在当前 query 的 groupBy 中。
    如果所有维度都在 groupBy（不应发生于小计/总计行），返回 None。

    Args:
        cols: 列定义列表（来自 _build_col_layout）
        q_groupby: 当前 query 的 groupBy alias 集合
        row_type: 行类型（"subtotal" / "total"）

    Returns:
        应放置标记文本的列 alias，或 None（明细行不需要标记）
    """
    if row_type == "detail":
        return None
    for col in cols:
        if col["field_type"] == "dimension" and col["alias"] not in q_groupby:
            return col["alias"]
    # 兜底：取第一个维度列（总计行所有维度均不在 groupBy）
    for col in cols:
        if col["field_type"] == "dimension":
            return col["alias"]
    return None


# ---------------------------------------------------------------------------
# Excel 写入核心
# ---------------------------------------------------------------------------


def export_to_excel(
    queries: list[dict],
    mapped_queries: list[dict],
    output_path: str,
    sheet_name: str = "数据透视表",
) -> dict:
    """将图表查询结果写入 Excel 文件。

    明细行（queries[0]）、小计行（queries[1:-1]）、总计行（queries[-1]）
    合并写入同一 Sheet，各行类型通过背景色区分。
    百分比列自动识别并应用 Excel 百分比格式，负值标红。

    Args:
        queries: 原始 queries 列表，每项含 payload 和 result
        mapped_queries: map_chart_queries 的返回值，含 _mapping 字段别名映射
        output_path: 输出 Excel 文件路径（.xlsx）
        sheet_name: Sheet 名称，默认"数据透视表"

    Returns:
        结果信息字典，含 success / output / rows / columns
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("缺少依赖：请先执行 pip install openpyxl", file=sys.stderr)
        raise SystemExit(1)

    if not mapped_queries:
        return {"success": False, "error": "映射结果为空，无数据可导出"}

    # 从 queries[0] 提取列定义（维度 + 指标，保持 select 原始顺序）
    cols = _build_col_layout(mapped_queries[0])
    if not cols:
        return {"success": False, "error": "无法从 queries[0] 提取列定义，请检查字段映射"}

    wb = Workbook()
    ws = wb.active
    assert ws is not None, "Workbook 初始化失败，未获得活动 Sheet"
    ws.title = sheet_name

    # ------------------------------------------------------------------
    # 1. 写入表头（蓝底白字粗体，冻结首行）
    # ------------------------------------------------------------------
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_font = Font(bold=True, color="FFFFFF")
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col["name"])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # 2. 遍历所有 query，按行类型写入数据
    # ------------------------------------------------------------------
    n_queries = len(queries)
    subtotal_fill = PatternFill("solid", fgColor=_SUBTOTAL_BG)
    total_fill = PatternFill("solid", fgColor=_TOTAL_BG)

    # 标记文本映射
    marker_text_map = {"subtotal": "小计", "total": "总计"}

    for qi, q in enumerate(queries):
        row_type = _get_row_type(qi, n_queries)
        result_data = (q.get("result") or {}).get("data", [])
        if not result_data:
            continue

        # 当前 query 的 groupBy alias 集合（用于判断哪些维度列留空）
        q_groupby: set[str] = set(
            (q.get("payload") or {}).get("query", {}).get("groupBy", [])
        )
        marker_alias = _find_marker_col(cols, q_groupby, row_type)
        marker_text = marker_text_map.get(row_type, "")

        for row_data in result_data:
            ri = ws.max_row + 1

            for ci, col in enumerate(cols, 1):
                alias = col["alias"]
                raw = row_data.get(alias)

                # 小计/总计行：在 marker 列填入标记文本，非 groupBy 维度列留空
                if row_type != "detail":
                    if alias == marker_alias:
                        val: object = marker_text
                    elif col["field_type"] == "dimension" and alias not in q_groupby:
                        val = ""
                    else:
                        # 指标列或在 groupBy 中的维度列：使用原始值
                        val = to_float(raw) if col["field_type"] == "metric" and raw not in (None, "") else raw
                else:
                    # 明细行：指标列转 float，其余保留原值
                    val = to_float(raw) if col["field_type"] == "metric" and raw not in (None, "") else raw

                cell = ws.cell(row=ri, column=ci, value=val)

                # 判断是否为负值（影响字体颜色）
                is_negative = isinstance(val, (int, float)) and val < 0

                # 根据行类型和负值状态决定字体颜色
                if row_type == "subtotal":
                    cell.fill = subtotal_fill
                    font_color = _NEG_COLOR if is_negative else "000000"
                    cell.font = Font(bold=True, color=font_color)
                elif row_type == "total":
                    cell.fill = total_fill
                    # 总计行负值显示红色，正值显示白色（深蓝背景下可读）
                    font_color = _NEG_COLOR if is_negative else "FFFFFF"
                    cell.font = Font(bold=True, color=font_color)
                elif is_negative:
                    # 明细行负值标红
                    cell.font = Font(color=_NEG_COLOR)

                # 数字格式（仅指标列且值为数值时应用）
                if col["field_type"] == "metric" and isinstance(val, (int, float)):
                    cell.number_format = _FMT_PCT if col["is_pct"] else _FMT_NUMBER

    # ------------------------------------------------------------------
    # 3. 列宽自适应（最大 50 字符）
    # ------------------------------------------------------------------
    for ci in range(1, len(cols) + 1):
        max_len = 0
        for ri in range(1, ws.max_row + 1):
            v = ws.cell(row=ri, column=ci).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 50)

    # 写入文件
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    total_rows = ws.max_row - 1  # 减去表头行
    col_names = [col["name"] for col in cols]
    return {
        "success": True,
        "output": str(out.resolve()),
        "rows": total_rows,
        "columns": col_names,
    }


# ---------------------------------------------------------------------------
# 入口
# ---------------------------------------------------------------------------


def main() -> None:
    """脚本入口：解析参数，加载数据，执行 Excel 导出。"""
    parser = argparse.ArgumentParser(description="图表数据 Excel 导出工具")
    parser.add_argument("--uuid", help="图表 UUID，自动调用 opscli 获取并执行查询")
    parser.add_argument("--input", help="已保存的 chart run JSON 文件路径")
    parser.add_argument("--output", default="/tmp/output.xlsx", help="输出 Excel 文件路径（默认 /tmp/output.xlsx）")
    parser.add_argument("--sheet-name", default="数据透视表", help="Sheet 名称（默认：数据透视表）")
    parser.add_argument("--skills-dir", help="指定 Skill 安装根目录")
    parser.add_argument("--data-dir", help="直接指定数据目录路径（最高优先级）")
    parser.add_argument("--no-auto-upgrade", action="store_true",
                        help="禁用自动升级兜底（字段映射失败时不自动调用 opscli skills upgrade）")
    parser.add_argument("--pretty", action="store_true", help="格式化 JSON 输出")
    args = parser.parse_args()

    if not args.uuid and not args.input:
        parser.error("必须提供 --uuid 或 --input 之一")

    # ------------------------------------------------------------------
    # 1. 确定本地数据目录
    # ------------------------------------------------------------------
    if args.data_dir:
        data_dir = Path(args.data_dir)
    else:
        discovered = discover_data_dir(skills_dir=args.skills_dir)
        if discovered is None:
            # 未找到数据目录，尝试自动升级兜底
            if not args.no_auto_upgrade and _try_upgrade(
                Path.home() / ".claude" / "skills" / "ops-dataset-query" / "data"
            ):
                discovered = discover_data_dir(skills_dir=args.skills_dir)
            if discovered is None:
                err = {
                    "success": False,
                    "error": "未找到 ops-dataset-query 数据目录，请先执行 opscli skills install ops-dataset-query",
                }
                print(json.dumps(err, ensure_ascii=False), file=sys.stderr)
                raise SystemExit(1)
        data_dir = discovered

    # ------------------------------------------------------------------
    # 2. 加载本地字段索引
    # ------------------------------------------------------------------
    dataset_index, field_index = load_local_index(data_dir)

    # ------------------------------------------------------------------
    # 3. 加载图表查询数据（来自文件或直接调用 opscli）
    # ------------------------------------------------------------------
    queries, chart_uuid, _merged = load_chart_data(args.uuid, args.input)

    if not queries:
        err = {"success": False, "error": "未获取到任何 query 数据，请检查 --uuid 或 --input 参数"}
        print(json.dumps(err, ensure_ascii=False, indent=2 if args.pretty else None))
        raise SystemExit(1)

    # ------------------------------------------------------------------
    # 4. 字段别名映射
    # ------------------------------------------------------------------
    mapped_queries = map_chart_queries(queries, dataset_index, field_index)

    # 字段映射全部失败时，尝试自动升级本地索引并重试一次
    if not args.no_auto_upgrade and not _check_mapping_hit(mapped_queries):
        print("[excel_export] 字段映射未命中，尝试 opscli skills upgrade 更新数据...", file=sys.stderr)
        if _try_upgrade(data_dir):
            dataset_index, field_index = load_local_index(data_dir)
            mapped_queries = map_chart_queries(queries, dataset_index, field_index)

    # ------------------------------------------------------------------
    # 5. 导出 Excel
    # ------------------------------------------------------------------
    result = export_to_excel(
        queries=queries,
        mapped_queries=mapped_queries,
        output_path=args.output,
        sheet_name=args.sheet_name,
    )

    if chart_uuid:
        result["chart_uuid"] = chart_uuid

    print(json.dumps(result, ensure_ascii=False, indent=2 if args.pretty else None))


if __name__ == "__main__":
    main()
