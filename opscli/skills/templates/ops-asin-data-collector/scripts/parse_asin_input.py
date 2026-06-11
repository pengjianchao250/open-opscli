#!/usr/bin/env python3
"""Parse ASIN input files for ops-asin-data-collector."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from typing import Any


ASIN_PATTERN = re.compile(r"^[A-Z0-9]{10}$")
KEYWORD_SPLIT_PATTERN = re.compile(r"[\n\r,，;；|]+")
ASIN_COLUMN_ALIASES = ("ASIN", "asin", "亚马逊ASIN", "商品ASIN")
SITE_COLUMN_ALIASES = ("site", "站点", "country", "国家", "marketplace", "市场")
KEYWORD_COLUMN_ALIASES = ("keyword", "keywords", "关键词", "核心关键词", "搜索词", "种子关键词")


def load_asin_records(
    input_path: str | Path,
    *,
    asin_column: str = "asin",
    keyword_column: str | None = "keyword",
    site_column: str | None = "site",
    default_site: str = "US",
    include_source_row: bool = True,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Load and normalize ASIN records.

    Returns (records, errors). Records are de-duplicated by ASIN + site.
    """
    path = Path(input_path).expanduser()
    if not path.exists():
        raise FileNotFoundError(f"Input file does not exist: {path}")

    raw_rows = _load_rows(path)
    records: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    records_by_key: dict[tuple[str, str], dict[str, Any]] = {}

    for index, row in enumerate(raw_rows, start=1):
        if isinstance(row, str):
            row = {asin_column: row}
        if not isinstance(row, dict):
            errors.append({"row_index": index, "error": "row is not an object", "row": row})
            continue

        asin_raw = _pick(row, asin_column, aliases=ASIN_COLUMN_ALIASES)
        asin = str(asin_raw or "").strip().upper()
        if not asin:
            errors.append({"row_index": index, "error": "missing asin", "row": row})
            continue
        if not ASIN_PATTERN.match(asin):
            errors.append({"row_index": index, "asin": asin, "error": "invalid asin format", "row": row})
            continue

        site = str(_pick(row, site_column, aliases=SITE_COLUMN_ALIASES) or default_site or "US").strip().upper()
        keywords = normalize_keywords(_pick(row, keyword_column, aliases=KEYWORD_COLUMN_ALIASES)) if keyword_column else []
        keyword = keywords[0] if keywords else ""
        key = (asin, site)
        existing = records_by_key.get(key)
        if existing:
            existing["keywords"] = merge_keywords(existing.get("keywords") or [], keywords)
            existing["keyword"] = existing.get("keyword") or keyword
            continue

        record = {
            "asin": asin,
            "site": site,
            "keyword": keyword,
            "keywords": keywords,
            "row_index": index,
            "source_file": str(path),
        }
        if include_source_row:
            record["source_row"] = row
        records_by_key[key] = record
        records.append(record)

    return records, errors


def _pick(row: dict[str, Any], column: str | None, *, aliases: tuple[str, ...] = ()) -> Any:
    if not column:
        return None
    lower_map = {str(k).strip().lower(): k for k in row.keys()}
    empty_value: Any = None
    found_empty = False
    for candidate in _candidate_columns(column, aliases):
        key = candidate if candidate in row else lower_map.get(candidate.strip().lower())
        if key is None:
            continue
        value = row.get(key)
        if value is None or (isinstance(value, str) and not value.strip()):
            empty_value = value
            found_empty = True
            continue
        return value
    return empty_value if found_empty else None


def _candidate_columns(column: str, aliases: tuple[str, ...]) -> list[str]:
    candidates: list[str] = []
    for candidate in (column, *aliases):
        text = str(candidate or "").strip()
        if text and text.lower() not in {item.lower() for item in candidates}:
            candidates.append(text)
    return candidates


def normalize_keywords(value: Any) -> list[str]:
    """Normalize one or many keyword values while preserving phrase spaces."""
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        merged: list[str] = []
        for item in value:
            merged = merge_keywords(merged, normalize_keywords(item))
        return merged
    text = str(value).strip()
    if not text:
        return []
    if text.startswith("[") and text.endswith("]"):
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            parsed = None
        if isinstance(parsed, list):
            return normalize_keywords(parsed)
    return merge_keywords([], [part.strip() for part in KEYWORD_SPLIT_PATTERN.split(text) if part.strip()])


def merge_keywords(existing: list[str], incoming: list[str]) -> list[str]:
    seen = {keyword.casefold() for keyword in existing}
    merged = list(existing)
    for keyword in incoming:
        text = str(keyword or "").strip()
        if not text:
            continue
        key = text.casefold()
        if key in seen:
            continue
        seen.add(key)
        merged.append(text)
    return merged


def _load_rows(path: Path) -> list[Any]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _load_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx(path)
    if suffix == ".jsonl":
        return _load_jsonl(path)
    if suffix == ".json":
        return _load_json(path)
    raise ValueError(f"Unsupported input format: {suffix}. Use CSV, XLSX, JSON, or JSONL.")


def _load_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _load_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX input") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or "").strip() for cell in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        result.append(item)
    return result


def _load_json(path: Path) -> list[Any]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("items", "asins", "data", "records"):
            value = payload.get(key)
            if isinstance(value, list):
                return value
        return [payload]
    raise ValueError("JSON input must be an object or array")


def _load_jsonl(path: Path) -> list[Any]:
    rows: list[Any] = []
    with path.open("r", encoding="utf-8-sig") as f:
        for line_no, line in enumerate(f, start=1):
            text = line.strip()
            if not text:
                continue
            try:
                rows.append(json.loads(text))
            except json.JSONDecodeError as exc:
                raise ValueError(f"Invalid JSONL at line {line_no}: {exc}") from exc
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse ASIN input files.")
    parser.add_argument("--input", required=True, help="CSV/XLSX/JSON/JSONL input file")
    parser.add_argument("--asin-column", default="asin")
    parser.add_argument("--keyword-column", default="keyword")
    parser.add_argument("--site-column", default="site")
    parser.add_argument("--default-site", default="US")
    parser.add_argument("--output", help="Optional output JSON file")
    parser.add_argument("--pretty", action="store_true")
    args = parser.parse_args()

    try:
        records, errors = load_asin_records(
            args.input,
            asin_column=args.asin_column,
            keyword_column=args.keyword_column,
            site_column=args.site_column,
            default_site=args.default_site,
        )
        payload = {
            "success": True,
            "data": records,
            "summary": {
                "record_count": len(records),
                "error_count": len(errors),
            },
            "errors": errors,
        }
        text = json.dumps(payload, ensure_ascii=False, indent=2 if args.pretty else None)
        if args.output:
            output_path = Path(args.output).expanduser()
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(text, encoding="utf-8")
        print(text)
    except Exception as exc:
        error = {"success": False, "error": {"type": type(exc).__name__, "message": str(exc)}}
        print(json.dumps(error, ensure_ascii=False, indent=2), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
