"""西柚洞察接口直连任务编排。"""

from __future__ import annotations

import json
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote
from uuid import uuid4

from opscli.xiyou.api.client import XiyouApiClient
from opscli.xiyou.api.payloads import (
    SUPPORTED_SITES,
    normalize_reverse_keyword_view,
    normalize_view_mode,
)
from opscli.xiyou.api.scenarios import (
    get_resource_scenario,
    get_scenario,
    list_resource_scenarios,
    list_scenarios,
)
from opscli.xiyou.config import DEFAULT_PROVIDER, XiyouSettings, load_settings
from opscli.xiyou.credentials import XiyouCredentialProvider
from opscli.xiyou.domain.exceptions import (
    XiyouApiError,
    XiyouConfigError,
    XiyouUnsupportedExportError,
)
from opscli.xiyou.domain.models import XiyouExportResult, XiyouRankingRequest, XiyouRankingResult
from opscli.xiyou.export.xlsx import export_rows_to_xlsx
from opscli.shared.file_uploads import FileUploadClient, FileUploadError


TWO_STEP_RESOURCE_FUNCTIONS = {"flow-insight", "ad-insight", "flow-weekly"}
UNSUPPORTED_EXPORT_FUNCTIONS: set[str] = set()
FUNCTION_ALIASES = {
    "流量诊断仪": "flow-diagnosis",
    "流量诊断": "flow-diagnosis",
}


class XiyouApiManager:
    """执行西柚洞察接口场景并落盘任务结果。"""

    def __init__(
        self,
        *,
        settings: XiyouSettings | None = None,
        credential_provider: XiyouCredentialProvider | None = None,
        jwt: str | None = None,
        session_id: str | None = None,
    ) -> None:
        self.settings = settings or load_settings()
        self.credential_provider = credential_provider or XiyouCredentialProvider(
            self.settings,
            jwt=jwt,
            session_id=session_id,
        )
        self.jwt = jwt
        self.session_id = session_id

    def scenarios(self) -> list[dict[str, Any]]:
        """列出支持的接口场景。"""
        return [
            {
                "function": "ranking",
                "provider": DEFAULT_PROVIDER,
                "targets": list_scenarios(),
                "periods": ["week", "month"],
                "sites": sorted(SUPPORTED_SITES),
            },
            *list_resource_scenarios(),
        ]

    async def run(self, request: XiyouRankingRequest) -> XiyouRankingResult:
        """执行接口场景。"""
        self._normalize_request(request)
        self._validate_request(request)
        if (request.function or "").lower() != "ranking":
            return await self._run_resource(request)
        return await self._run_ranking(request)

    async def _run_ranking(self, request: XiyouRankingRequest) -> XiyouRankingResult:
        """执行排行榜接口场景。"""
        scenario = get_scenario(request.target)
        rank_pattern = scenario.normalize_rank_pattern(request.rank_pattern)
        job_id = request.job_id or _build_job_id(request.function)
        root_dir = self._build_root_dir(request, job_id)
        root_dir.mkdir(parents=True, exist_ok=True)

        site = (request.site or self.settings.default_site).upper()
        period = request.period or self.settings.default_period
        page_size = request.page_size or self.settings.page_size
        payload = scenario.build_payload(
            site=site,
            period=period,
            rank_pattern=rank_pattern,
            query=request.query,
            page=request.page,
            page_size=page_size,
        )

        params_path = root_dir / "params.json"
        raw_path = root_dir / "raw.json"
        result_path = root_dir / "result.json"
        _write_json(
            params_path,
            {
                "request": request.to_dict(),
                "endpoint": scenario.endpoint,
                "payload": payload,
            },
        )

        credential = self.credential_provider.get_default()
        async with XiyouApiClient(credential=credential, settings=self.settings) as client:
            response = await client.post_json(scenario.endpoint, payload)

        raw = {
            "job_id": job_id,
            "function": request.function,
            "provider": request.provider,
            "target": request.target,
            "endpoint": scenario.endpoint,
            "payload": payload,
            "response": response,
        }
        _write_json(raw_path, raw)

        rows = _extract_items(response)
        export_format = _normalize_export_format(request.export_format)
        warnings: list[dict[str, Any]] = []
        if export_format == "xlsx":
            export = export_rows_to_xlsx(
                rows=rows,
                output_path=root_dir / f"{job_id}.xlsx",
                target=request.target,
                site=site,
                period=period,
            )
        else:
            export = _export_rows_to_json(
                output_path=root_dir / f"{job_id}.json",
                job_id=job_id,
                request=request,
                site=site,
                period=period,
                rank_pattern=rank_pattern,
                rows=rows,
            )
        _upload_export_if_enabled(
            export=export,
            job_id=job_id,
            request=request,
            site=site,
            period=period,
            rank_pattern=rank_pattern,
            warnings=warnings,
            jwt=self.jwt,
            session_id=self.session_id,
        )

        result = XiyouRankingResult(
            job_id=job_id,
            function=request.function,
            provider=request.provider,
            target=request.target,
            site=site,
            period=period,
            rank_pattern=rank_pattern,
            row_count=len(rows),
            root_dir=str(root_dir),
            params_path=str(params_path),
            raw_path=str(raw_path),
            result_path=str(result_path),
            dataset=request.dataset,
            data_mode="rows",
            export=export,
            data=rows,
            warnings=warnings,
        )
        _write_json(result_path, result.to_dict())
        return result

    async def _run_resource(self, request: XiyouRankingRequest) -> XiyouRankingResult:
        """执行 resource 导出场景。"""
        normalized_function = (request.function or "").lower()
        if normalized_function in UNSUPPORTED_EXPORT_FUNCTIONS:
            raise XiyouUnsupportedExportError(function=normalized_function)

        scenario = get_resource_scenario(request.function)
        if (
            normalized_function == "keyword-analysis"
            and _normalize_export_format(request.export_format) == "json"
        ):
            return await self._run_keyword_analysis_rows(request, scenario)
        if normalized_function in TWO_STEP_RESOURCE_FUNCTIONS:
            return await self._run_two_step_resource(request, scenario)
        if scenario.mode == "rows":
            return await self._run_rows_resource(request, scenario)
        job_id = request.job_id or _build_job_id(request.function)
        root_dir = self._build_root_dir(request, job_id)
        root_dir.mkdir(parents=True, exist_ok=True)

        site = (request.site or self.settings.default_site).upper()
        page_size = request.page_size or self.settings.page_size
        dataset = request.dataset or scenario.default_dataset
        request_url = _build_resource_request_url(
            function=request.function,
            site=site,
            asin=request.asin,
            asins=request.asins,
            keyword=request.keyword or "",
            view_mode=request.view_mode,
            keyword_type=request.keyword_type,
            replay_type=request.replay_type,
        )

        params_path = root_dir / "params.json"
        raw_path = root_dir / "raw.json"
        result_path = root_dir / "result.json"
        prequery_endpoint = _build_resource_prequery_endpoint(request)
        prequery_payload: dict[str, Any] | None = None
        # warnings/row_count 在 async with 之前初始化，使得 xlsx 分支
        # 既能填充真实行数，也能在 page_size 与全量导出不符时追加提示
        warnings: list[dict[str, Any]] = []
        row_count = 0

        credential = self.credential_provider.get_default()
        async with XiyouApiClient(credential=credential, settings=self.settings) as client:
            await self._hydrate_variation_context(
                request,
                site=site,
                client=client,
                request_url=request_url,
            )
            payload = scenario.build_payload(
                site=site,
                asin=request.asin,
                asins=request.asins,
                keyword=request.keyword,
                query=request.query,
                parent_asin=request.parent_asin,
                cycle_period=request.cycle_period,
                start_month=request.start_month,
                end_month=request.end_month,
                start_date=request.start_date,
                end_date=request.end_date,
                report_date=request.report_date,
                search_terms=request.search_terms,
                view_mode=request.view_mode,
                keyword_type=request.keyword_type,
                page=request.page,
                page_size=page_size,
            )
            if prequery_endpoint:
                prequery_payload = _build_resource_prequery_payload(request=request, payload=payload)
            _write_json(
                params_path,
                {
                    "request": request.to_dict(),
                    "endpoint": scenario.endpoint,
                    "status_endpoint": scenario.status_endpoint,
                    "prequery_endpoint": prequery_endpoint,
                    "prequery_payload": prequery_payload,
                    "payload": payload,
                    "request_url": request_url,
                },
            )
            prequery_response: dict[str, Any] | None = None
            if prequery_endpoint and prequery_payload:
                prequery_response = await client.post_json(
                    prequery_endpoint,
                    prequery_payload,
                    request_url=request_url,
                )
            submit_response = await client.post_json(scenario.endpoint, payload, request_url=request_url)
            resource_id = _extract_resource_id(submit_response)
            status_payload = _build_resource_status_payload(
                function=request.function,
                payload=payload,
                resource_id=resource_id,
            )
            status_response = await _poll_resource_status(
                client=client,
                endpoint=scenario.status_endpoint,
                payload=status_payload,
                request_url=request_url,
            )
            resource_url = _extract_resource_url(status_response)
            export_format = _normalize_export_format(request.export_format)
            if export_format == "xlsx":
                export_path = root_dir / f"{job_id}.xlsx"
                try:
                    export_path.write_bytes(await client.get_bytes(resource_url))
                except XiyouApiError as exc:
                    # 透传 OSS 错误响应（含 <StringToSign>、<SignatureProvided>
                    # 等关键字段），便于反向定位西柚预签名链接的签名问题
                    raise XiyouConfigError(
                        f"西柚导出文件下载失败 status={exc.status_code} "
                        f"url={resource_url} detail={exc.response_excerpt}"
                    ) from exc
                export = _resource_xlsx_export(export_path)
                # resource 导出是西柚后端生成的成品 xlsx，opscli 拿不到 rows 数组，
                # 必须读文件才能反映真实行数到 row_count
                row_count = _count_xlsx_rows(export_path)
                # 西柚 resource 接口固定返回全量数据，pageSize 仅对分页查询生效。
                # 当调用方明确传了 page_size 且与实际行数不符时追加 warning，
                # 避免 Agent / 用户误判为「参数没生效」或「导出截断失败」
                if request.page_size and row_count > request.page_size:
                    warnings.append(
                        {
                            "stage": "resource_export",
                            "message": (
                                f"西柚 resource 导出接口固定返回全量数据，本次返回 "
                                f"{row_count} 行；请求的 page_size={request.page_size} "
                                f"仅对分页查询生效，已被西柚后端忽略。"
                            ),
                        }
                    )
            else:
                export = _export_resource_to_json(
                    output_path=root_dir / f"{job_id}.json",
                    job_id=job_id,
                    request=request,
                    site=site,
                    dataset=dataset,
                    resource_id=resource_id,
                    resource_url=resource_url,
                    status_response=status_response,
                )

        raw = {
            "job_id": job_id,
            "function": request.function,
            "provider": request.provider,
            "dataset": dataset,
            "endpoint": scenario.endpoint,
            "status_endpoint": scenario.status_endpoint,
            "payload": payload,
            "request_url": request_url,
            "submit_response": submit_response,
            "status_response": status_response,
            "resource_id": resource_id,
            "resource_url": resource_url,
        }
        _write_json(raw_path, raw)

        _upload_export_if_enabled(
            export=export,
            job_id=job_id,
            request=request,
            site=site,
            period=request.period,
            rank_pattern=request.rank_pattern or "",
            warnings=warnings,
            jwt=self.jwt,
            session_id=self.session_id,
        )

        result = XiyouRankingResult(
            job_id=job_id,
            function=request.function,
            provider=request.provider,
            target=request.target,
            site=site,
            period=request.period,
            rank_pattern=request.rank_pattern or "",
            row_count=row_count,
            root_dir=str(root_dir),
            params_path=str(params_path),
            raw_path=str(raw_path),
            result_path=str(result_path),
            dataset=dataset,
            data_mode="resource_export",
            resource_id=resource_id,
            resource_url=resource_url,
            export=export,
            data=[],
            warnings=warnings,
        )
        _write_json(result_path, result.to_dict())
        return result

    async def _run_keyword_analysis_rows(
        self,
        request: XiyouRankingRequest,
        scenario,
    ) -> XiyouRankingResult:
        """关键词分析在 JSON 模式下直接走列表接口，返回页面明细。"""
        job_id = request.job_id or _build_job_id(request.function)
        root_dir = self._build_root_dir(request, job_id)
        root_dir.mkdir(parents=True, exist_ok=True)

        site = (request.site or self.settings.default_site).upper()
        page_size = request.page_size or self.settings.page_size
        dataset = request.dataset or scenario.default_dataset
        payload = scenario.build_payload(
            site=site,
            asin=request.asin,
            asins=request.asins,
            keyword=request.keyword,
            query=request.query,
            parent_asin=request.parent_asin,
            cycle_period=request.cycle_period,
            start_month=request.start_month,
            end_month=request.end_month,
            start_date=request.start_date,
            end_date=request.end_date,
            report_date=request.report_date,
            search_terms=request.search_terms,
            view_mode=request.view_mode,
            keyword_type=request.keyword_type,
            page=request.page,
            page_size=page_size,
        )
        endpoint = "/v4/searchTerms/analysis/list"
        request_url = _build_resource_request_url(
            function=request.function,
            site=site,
            asin=request.asin,
            asins=request.asins,
            keyword=request.keyword or "",
            view_mode=request.view_mode,
            keyword_type=request.keyword_type,
            replay_type=request.replay_type,
            request_mode="rows",
        )

        params_path = root_dir / "params.json"
        raw_path = root_dir / "raw.json"
        result_path = root_dir / "result.json"
        _write_json(
            params_path,
            {
                "request": request.to_dict(),
                "endpoint": endpoint,
                "payload": payload,
                "request_url": request_url,
            },
        )

        credential = self.credential_provider.get_default()
        async with XiyouApiClient(credential=credential, settings=self.settings) as client:
            response = await client.post_json(endpoint, payload, request_url=request_url)

        raw = {
            "job_id": job_id,
            "function": request.function,
            "provider": request.provider,
            "dataset": dataset,
            "endpoint": endpoint,
            "payload": payload,
            "request_url": request_url,
            "response": response,
        }
        _write_json(raw_path, raw)

        rows = _extract_items(response)
        warnings: list[dict[str, Any]] = []
        export = _export_rows_to_json(
            output_path=root_dir / f"{job_id}.json",
            job_id=job_id,
            request=request,
            site=site,
            period=request.period,
            rank_pattern=request.rank_pattern or "",
            rows=rows,
        )
        _upload_export_if_enabled(
            export=export,
            job_id=job_id,
            request=request,
            site=site,
            period=request.period,
            rank_pattern=request.rank_pattern or "",
            warnings=warnings,
            jwt=self.jwt,
            session_id=self.session_id,
        )

        result = XiyouRankingResult(
            job_id=job_id,
            function=request.function,
            provider=request.provider,
            target=request.target,
            site=site,
            period=request.period,
            rank_pattern=request.rank_pattern or "",
            row_count=len(rows),
            root_dir=str(root_dir),
            params_path=str(params_path),
            raw_path=str(raw_path),
            result_path=str(result_path),
            dataset=dataset,
            data_mode="rows",
            export=export,
            data=rows,
            warnings=warnings,
        )
        _write_json(result_path, result.to_dict())
        return result

    async def _run_rows_resource(
        self,
        request: XiyouRankingRequest,
        scenario,
    ) -> XiyouRankingResult:
        """执行直接返回列表数据的资源场景。"""
        job_id = request.job_id or _build_job_id(request.function)
        root_dir = self._build_root_dir(request, job_id)
        root_dir.mkdir(parents=True, exist_ok=True)

        site = (request.site or self.settings.default_site).upper()
        page_size = request.page_size or self.settings.page_size
        dataset = request.dataset or scenario.default_dataset
        payload = scenario.build_payload(
            site=site,
            asin=request.asin,
            asins=request.asins,
            keyword=request.keyword,
            query=request.query,
            parent_asin=request.parent_asin,
            cycle_period=request.cycle_period,
            start_month=request.start_month,
            end_month=request.end_month,
            start_date=request.start_date,
            end_date=request.end_date,
            report_date=request.report_date,
            search_terms=request.search_terms,
            view_mode=request.view_mode,
            keyword_type=request.keyword_type,
            page=request.page,
            page_size=page_size,
        )
        request_url = _build_resource_request_url(
            function=request.function,
            site=site,
            asin=request.asin,
            asins=request.asins,
            keyword=request.keyword or "",
            view_mode=request.view_mode,
            keyword_type=request.keyword_type,
            replay_type=request.replay_type,
        )

        params_path = root_dir / "params.json"
        raw_path = root_dir / "raw.json"
        result_path = root_dir / "result.json"
        _write_json(
            params_path,
            {
                "request": request.to_dict(),
                "endpoint": scenario.endpoint,
                "payload": payload,
                "request_url": request_url,
            },
        )

        credential = self.credential_provider.get_default()
        async with XiyouApiClient(credential=credential, settings=self.settings) as client:
            response = await client.post_json(scenario.endpoint, payload, request_url=request_url)

        raw = {
            "job_id": job_id,
            "function": request.function,
            "provider": request.provider,
            "dataset": dataset,
            "endpoint": scenario.endpoint,
            "payload": payload,
            "request_url": request_url,
            "response": response,
        }
        _write_json(raw_path, raw)

        rows = _extract_items(response)
        warnings: list[dict[str, Any]] = []
        export_format = _normalize_export_format(request.export_format)
        if export_format == "xlsx":
            export = export_rows_to_xlsx(
                rows=rows,
                output_path=root_dir / f"{job_id}.xlsx",
                target=request.function,
                site=site,
                period=request.period or "",
            )
        else:
            export = _export_rows_to_json(
                output_path=root_dir / f"{job_id}.json",
                job_id=job_id,
                request=request,
                site=site,
                period=request.period,
                rank_pattern=request.rank_pattern or "",
                rows=rows,
            )
        _upload_export_if_enabled(
            export=export,
            job_id=job_id,
            request=request,
            site=site,
            period=request.period,
            rank_pattern=request.rank_pattern or "",
            warnings=warnings,
            jwt=self.jwt,
            session_id=self.session_id,
        )

        result = XiyouRankingResult(
            job_id=job_id,
            function=request.function,
            provider=request.provider,
            target=request.target,
            site=site,
            period=request.period,
            rank_pattern=request.rank_pattern or "",
            row_count=len(rows),
            root_dir=str(root_dir),
            params_path=str(params_path),
            raw_path=str(raw_path),
            result_path=str(result_path),
            dataset=dataset,
            data_mode="rows",
            export=export,
            data=rows,
            warnings=warnings,
        )
        _write_json(result_path, result.to_dict())
        return result

    async def _run_two_step_resource(
        self,
        request: XiyouRankingRequest,
        scenario,
    ) -> XiyouRankingResult:
        job_id = request.job_id or _build_job_id(request.function)
        root_dir = self._build_root_dir(request, job_id)
        root_dir.mkdir(parents=True, exist_ok=True)

        site = (request.site or self.settings.default_site).upper()
        dataset = request.dataset or scenario.default_dataset
        payload = scenario.build_payload(
            site=site,
            asin=request.asin,
            asins=request.asins,
            keyword=request.keyword,
            query=request.query,
            parent_asin=request.parent_asin,
            cycle_period=request.cycle_period,
            start_month=request.start_month,
            end_month=request.end_month,
            start_date=request.start_date,
            end_date=request.end_date,
            report_date=request.report_date,
            search_terms=request.search_terms,
            view_mode=request.view_mode,
            keyword_type=request.keyword_type,
            page=request.page,
            page_size=request.page_size or self.settings.page_size,
        )
        request_url = _build_resource_request_url(
            function=request.function,
            site=site,
            asin=request.asin,
            asins=request.asins,
            keyword=request.keyword or "",
            view_mode=request.view_mode,
            keyword_type=request.keyword_type,
            replay_type=request.replay_type,
        )

        params_path = root_dir / "params.json"
        raw_path = root_dir / "raw.json"
        result_path = root_dir / "result.json"
        _write_json(
            params_path,
            {
                "request": request.to_dict(),
                "endpoint": scenario.endpoint,
                "status_endpoint": scenario.status_endpoint,
                "payload": payload,
                "request_url": request_url,
            },
        )

        warnings: list[dict[str, Any]] = []
        credential = self.credential_provider.get_default()
        async with XiyouApiClient(credential=credential, settings=self.settings) as client:
            submit_response = await client.post_json(scenario.endpoint, payload, request_url=request_url)
            resource_id = _extract_resource_id(submit_response)
            status_payload = {
                "resource": _resource_payload(payload),
                "biz": {"resourceId": resource_id},
            }
            status_response = await _poll_resource_status(
                client=client,
                endpoint=scenario.status_endpoint,
                payload=status_payload,
                request_url=request_url,
            )
            resource_url = _extract_resource_url(status_response)
            export = _export_resource_to_json(
                output_path=root_dir / f"{job_id}.json",
                job_id=job_id,
                request=request,
                site=site,
                dataset=dataset,
                resource_id=resource_id,
                resource_url=resource_url,
                status_response=status_response,
            )

        raw = {
            "job_id": job_id,
            "function": request.function,
            "provider": request.provider,
            "dataset": dataset,
            "endpoint": scenario.endpoint,
            "status_endpoint": scenario.status_endpoint,
            "payload": payload,
            "request_url": request_url,
            "submit_response": submit_response,
            "status_response": status_response,
            "resource_id": resource_id,
            "resource_url": resource_url,
        }
        _write_json(raw_path, raw)

        _upload_export_if_enabled(
            export=export,
            job_id=job_id,
            request=request,
            site=site,
            period=request.period,
            rank_pattern=request.rank_pattern or "",
            warnings=warnings,
            jwt=self.jwt,
            session_id=self.session_id,
        )

        result = XiyouRankingResult(
            job_id=job_id,
            function=request.function,
            provider=request.provider,
            target=request.target,
            site=site,
            period=request.period,
            rank_pattern=request.rank_pattern or "",
            row_count=0,
            root_dir=str(root_dir),
            params_path=str(params_path),
            raw_path=str(raw_path),
            result_path=str(result_path),
            dataset=dataset,
            data_mode="resource_export",
            resource_id=resource_id,
            resource_url=resource_url,
            export=export,
            data=[],
            warnings=warnings,
        )
        _write_json(result_path, result.to_dict())
        return result

    def job_status(self, job_id: str) -> dict[str, Any]:
        """读取已落盘任务状态。"""
        root_dir = self.settings.output_dir / job_id
        result_path = root_dir / "result.json"
        if not result_path.exists():
            raise XiyouConfigError(f"任务不存在：{job_id}")
        return json.loads(result_path.read_text(encoding="utf-8"))

    def _validate_request(self, request: XiyouRankingRequest) -> None:
        if (request.provider or DEFAULT_PROVIDER).lower() != DEFAULT_PROVIDER:
            raise XiyouConfigError("opscli xiyou run 目前仅支持 provider：xiyou")
        function = (request.function or "").lower()
        if function == "ranking":
            scenario = get_scenario(request.target)
            scenario.normalize_period(request.period)
        else:
            get_resource_scenario(function)
        if request.page <= 0:
            raise XiyouConfigError("page 必须为正整数")
        if request.page_size <= 0:
            raise XiyouConfigError("page_size 必须为正整数")

    def _normalize_request(self, request: XiyouRankingRequest) -> None:
        request.function = _normalize_function_name(request.function)
        function = (request.function or "").lower()
        if function == "ranking":
            if not (request.query or "").strip():
                target = (request.target or "").lower()
                if target == "asin" and request.asin:
                    request.query = str(request.asin).strip().upper()
                elif target == "keyword" and request.keyword:
                    request.query = str(request.keyword).strip()
            return

        self._normalize_resource_period(request)

        if function != "keyword-historical-traffic":
            return

        if request.start_date or request.end_date:
            raise XiyouConfigError(
                "keyword-historical-traffic 不支持用户自定义时间范围；固定导出最近一个月（不包含今天和昨天）"
            )

        period = (request.period or "").strip().lower()
        if period not in {"", "week", "month", "last1month"}:
            raise XiyouConfigError(
                "keyword-historical-traffic 不支持用户自定义时间范围；固定导出最近一个月（不包含今天和昨天）"
            )

        cycle_period = (request.cycle_period or "").strip().lower()
        if cycle_period or request.start_month or request.end_month:
            raise XiyouConfigError(
                "keyword-historical-traffic 不支持用户自定义时间范围；固定导出最近一个月（不包含今天和昨天）"
            )

        # `period` 在通用请求里默认是 week，这里统一归一化为固定窗口标签，
        # 兼容旧调用默认值，同时避免结果元数据继续误导为周榜/月榜语义。
        request.period = "last1month"
        request.cycle_period = None

    def _build_root_dir(self, request: XiyouRankingRequest, job_id: str) -> Path:
        base_dir = Path(request.output_dir).expanduser() if request.output_dir else self.settings.output_dir
        if not base_dir.is_absolute():
            base_dir = Path.cwd() / base_dir
        return base_dir.resolve() / job_id

    def _normalize_resource_period(self, request: XiyouRankingRequest) -> None:
        function = (request.function or "").lower()
        if function in {"", "ranking"}:
            return
        scenario = get_resource_scenario(function)
        scenario.normalize_request_defaults(request, self.settings.page_size)

    async def _hydrate_variation_context(
        self,
        request: XiyouRankingRequest,
        *,
        site: str,
        client: XiyouApiClient,
        request_url: str | None,
    ) -> None:
        function = (request.function or "").lower()
        if function not in {"ad-analysis", "parent-analysis", "sales-analysis"}:
            return

        needs_parent_asin = not _normalize_optional_asin(request.parent_asin)
        needs_related_asins = function in {"ad-analysis", "parent-analysis"} and not _has_request_asins(
            request.asins
        )
        if not needs_parent_asin and not needs_related_asins:
            return

        asin = _normalize_optional_asin(request.asin)
        if not asin:
            return

        variation_response = await client.post_json(
            "/v4/asins/variations",
            {"asin": asin, "country": site},
            request_url=request_url,
        )
        variation_id = _extract_variation_id(variation_response)
        status_response = await client.get_json(
            f"/v4/asins/variations/status?variationId={quote(variation_id, safe='')}",
            request_url=request_url,
        )

        resolved_parent_asin = _extract_parent_asin_from_variation_status(status_response)
        if needs_parent_asin and resolved_parent_asin:
            request.parent_asin = resolved_parent_asin

        parent_asin = resolved_parent_asin or _normalize_optional_asin(request.parent_asin)
        resolved_asins = _extract_related_asins_from_variation_status(
            status_response,
            primary_asin=asin,
            parent_asin=parent_asin,
        )
        if needs_related_asins and resolved_asins:
            request.asins = resolved_asins

        if needs_parent_asin and not _normalize_optional_asin(request.parent_asin):
            raise XiyouConfigError(
                f"{function} 自动补齐 parent_asin 失败：/v4/asins/variations/status 未返回可识别的 parentAsin"
            )
        if needs_related_asins and not _has_request_asins(request.asins):
            raise XiyouConfigError(
                f"{function} 自动补齐 asins 失败：/v4/asins/variations/status 未返回可识别的变体 ASIN 列表"
            )


def _extract_items(response: dict[str, Any]) -> list[dict[str, Any]]:
    """从常见响应结构中提取 rows。"""
    candidates: list[Any] = []
    if isinstance(response, dict):
        data = response.get("data")
        biz = response.get("biz")
        candidates.extend([response.get(key) for key in ["list", "items", "records", "rows", "data"]])
        candidates.extend([data, biz])
        if isinstance(data, dict):
            candidates.extend(data.get(key) for key in ["items", "list", "records", "rows", "data"])
        if isinstance(biz, dict):
            candidates.extend(biz.get(key) for key in ["items", "list", "records", "rows", "data"])
    for candidate in candidates:
        if isinstance(candidate, list):
            return [item for item in candidate if isinstance(item, dict)]
    return []


def _build_job_id(function: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    suffix = uuid4().hex[:6]
    return f"xiyou-{function}-{timestamp}-{suffix}"


def _normalize_function_name(function: str | None) -> str:
    text = str(function or "").strip()
    if not text:
        return ""
    return FUNCTION_ALIASES.get(text, text)


def _normalize_export_format(value: str) -> str:
    text = (value or "json").lower()
    if text in {"xls", "xlsx"}:
        return "xlsx"
    if text == "json":
        return "json"
    raise XiyouConfigError(f"不支持的导出格式：{value}")


def _build_resource_prequery_endpoint(request: XiyouRankingRequest) -> str | None:
    function = (request.function or "").lower()
    if function == "reverse-keyword" and (request.query or "").strip():
        return "/v3/asins/research/list"
    return None


def _build_resource_prequery_payload(
    *,
    request: XiyouRankingRequest,
    payload: dict[str, Any],
) -> dict[str, Any] | None:
    function = (request.function or "").lower()
    if function != "reverse-keyword":
        return None

    prequery_payload = deepcopy(payload)
    biz = prequery_payload.get("biz")
    if isinstance(biz, dict):
        # 列表预查询用来让西柚先按筛选词命中服务端结果集，不需要资源导出专用 tableType。
        biz.pop("tableType", None)
    return prequery_payload


def _build_resource_request_url(
    *,
    function: str,
    site: str,
    asin: str | None,
    asins: list[str] | str | None,
    keyword: str,
    view_mode: str | None,
    keyword_type: str | None,
    replay_type: str | None,
    request_mode: str = "resource",
) -> str | None:
    normalized_function = (function or "").lower()
    if normalized_function == "reverse-keyword":
        normalized_view_mode = normalize_reverse_keyword_view(view_mode)
        normalized_asin = str(asin or "").strip().upper()
        suffix = ""
        if normalized_view_mode == "trends":
            suffix = "?listType=trendsViewList"
        elif normalized_view_mode == "top10":
            suffix = "?listType=topDataList"
        elif str(keyword_type or "").strip().lower() in {"organic", "advertising"}:
            suffix = "?listType=dataList"
        return f"/detail/asin/look_up/{site}/{normalized_asin}{suffix}"
    if normalized_function == "asin-compare":
        joined_asins = _join_asins_for_request_url(asins)
        return f"/detail/asin_compare/look_up/{site}/{joined_asins}"
    if normalized_function == "ad-analysis":
        normalized_asin = str(asin or "").strip().upper()
        return f"/detail/asin/ad_analysis/{site}/{normalized_asin}"
    if normalized_function == "parent-analysis":
        normalized_asin = str(asin or "").strip().upper()
        return f"/detail/asin/variations_insight/{site}/{normalized_asin}?listType=dataList"
    if normalized_function == "sales-analysis":
        normalized_asin = str(asin or "").strip().upper()
        return f"/detail/asin/sales_analysis/{site}/{normalized_asin}?listType=dataList"
    if normalized_function == "flow-insight":
        normalized_asin = str(asin or "").strip().upper()
        return f"/detail/asin/flow_insight/{site}/{normalized_asin}?listType=dataList"
    if normalized_function == "ad-insight":
        normalized_asin = str(asin or "").strip().upper()
        return f"/detail/asin/ad_insight/{site}/{normalized_asin}?listType=dataList"
    if normalized_function == "flow-weekly":
        normalized_asin = str(asin or "").strip().upper()
        return f"/detail/asin/flow_weekly/{site}/{normalized_asin}?listType=dataList"
    if normalized_function == "flow-diagnosis":
        normalized_asin = str(asin or "").strip().upper()
        return f"/detail/asin/flow_diagnosis/{site}/{normalized_asin}"

    normalized_view_mode = normalize_view_mode(view_mode)
    list_type = "dataList" if normalized_view_mode == "data" else "trendsViewList"
    encoded_keyword = quote(keyword.strip()) if keyword else ""
    if normalized_function == "keyword-explorer":
        return f"/detail/searchTerm_explorer/{site}/{encoded_keyword}"
    if normalized_function == "keyword-analysis":
        if request_mode == "rows":
            return f"/detail/search_term/look_up/{site}/{encoded_keyword}?listType={list_type}"
        return f"/detail/search_term/look_up/{site}/{encoded_keyword}"
    if normalized_function == "keyword-historical-traffic":
        return f"/detail/search_term/historical_traffic_analysis/{site}/{encoded_keyword}"
    if normalized_function == "keyword-ad-replay":
        return f"/detail/search_term/ad_once_more/{site}/{encoded_keyword}"
    if normalized_function == "keyword-organic-replay":
        return f"/detail/search_term/na_once_more/{site}/{encoded_keyword}?adOnceMoreType=na"
    if normalized_function == "keyword-ad-toppers":
        return f"/detail/search_term/ad_display_insight/{site}/{encoded_keyword}?adOnceMoreType=na"
    return None


def _join_asins_for_request_url(asins: list[str] | str | None) -> str:
    if isinstance(asins, str):
        raw_values = asins.split(",")
    elif isinstance(asins, list):
        raw_values = asins
    else:
        raw_values = []
    normalized: list[str] = []
    seen: set[str] = set()
    for raw_value in raw_values:
        asin = str(raw_value or "").strip().upper()
        if asin and asin not in seen:
            seen.add(asin)
            normalized.append(asin)
    return ",".join(normalized)


def _extract_resource_id(response: dict[str, Any]) -> str:
    resource_id = response.get("resourceId") if isinstance(response, dict) else None
    if not resource_id and isinstance(response.get("data"), dict):
        resource_id = response["data"].get("resourceId")
    if not resource_id:
        raise XiyouConfigError("西柚洞察 resource 接口未返回 resourceId")
    return str(resource_id)


def _extract_resource_url(response: dict[str, Any]) -> str:
    resource_url = response.get("resourceUrl") if isinstance(response, dict) else None
    if not resource_url and isinstance(response.get("data"), dict):
        resource_url = response["data"].get("resourceUrl")
    if not resource_url:
        raise XiyouConfigError("西柚洞察 resource 任务未返回 resourceUrl")
    return str(resource_url).replace("\\u0026", "&")


def _resource_payload(payload: dict[str, Any]) -> dict[str, Any]:
    resource = payload.get("resource") if isinstance(payload, dict) else None
    return resource if isinstance(resource, dict) else {}


def _build_resource_status_payload(
    *,
    function: str,
    payload: dict[str, Any],
    resource_id: str,
) -> dict[str, Any]:
    """Build status polling payloads according to the documented Xiyou scenario contract."""
    resource = _resource_payload(payload)
    normalized_function = (function or "").lower()
    if normalized_function in {"keyword-organic-replay", "keyword-ad-toppers"}:
        return {
            "resource": resource,
            "biz": {"resourceId": resource_id},
        }
    if normalized_function == "ad-analysis":
        return {
            "resource": resource,
            "biz": {"resourceId": resource_id},
        }
    return {
        "resource": resource,
        "resourceId": resource_id,
    }


def _has_request_asins(value: list[str] | str | None) -> bool:
    return bool(_normalize_asin_collection(value))


def _normalize_optional_asin(value: Any) -> str | None:
    text = str(value or "").strip().upper()
    if len(text) == 10 and text.isalnum():
        return text
    return None


def _normalize_asin_collection(value: Any) -> list[str]:
    if isinstance(value, str):
        raw_values = value.split(",")
    elif isinstance(value, list):
        raw_values = value
    else:
        raw_values = []

    normalized: list[str] = []
    seen: set[str] = set()
    for raw_value in raw_values:
        asin = _normalize_optional_asin(raw_value)
        if asin and asin not in seen:
            seen.add(asin)
            normalized.append(asin)
    return normalized


def _extract_variation_id(response: dict[str, Any]) -> str:
    value = _find_first_by_key(response, {"variationid"})
    variation_id = str(value or "").strip()
    if not variation_id:
        raise XiyouConfigError("西柚 variations 接口未返回 variationId")
    return variation_id


def _extract_parent_asin_from_variation_status(response: dict[str, Any]) -> str | None:
    value = _find_first_by_key(response, {"parentasin", "parent_asin"})
    return _normalize_optional_asin(value)


def _extract_related_asins_from_variation_status(
    response: dict[str, Any],
    *,
    primary_asin: str,
    parent_asin: str | None,
) -> list[str]:
    normalized_parent_asin = _normalize_optional_asin(parent_asin)
    related_asins: list[str] = []
    seen: set[str] = set()

    def add_asin(candidate: Any) -> None:
        asin = _normalize_optional_asin(candidate)
        if not asin or asin == normalized_parent_asin or asin in seen:
            return
        seen.add(asin)
        related_asins.append(asin)

    def walk(node: Any) -> None:
        if isinstance(node, dict):
            for key, value in node.items():
                lowered = key.lower()
                if lowered in {
                    "asins",
                    "asinlist",
                    "childasins",
                    "variationasins",
                    "variationasinlist",
                    "relatedasins",
                }:
                    for asin in _normalize_asin_collection(value):
                        add_asin(asin)
                    continue
                if lowered in {"asin", "childasin", "variationasin", "subasin"}:
                    add_asin(value)
                    continue
                walk(value)
            return
        if isinstance(node, list):
            for item in node:
                walk(item)

    walk(response)
    add_asin(primary_asin)
    return related_asins


def _find_first_by_key(node: Any, keys: set[str]) -> Any:
    if isinstance(node, dict):
        for key, value in node.items():
            if key.lower() in keys:
                return value
        for value in node.values():
            found = _find_first_by_key(value, keys)
            if found is not None:
                return found
    elif isinstance(node, list):
        for item in node:
            found = _find_first_by_key(item, keys)
            if found is not None:
                return found
    return None


async def _poll_resource_status(
    *,
    client: XiyouApiClient,
    endpoint: str,
    payload: dict[str, Any],
    request_url: str | None = None,
    max_attempts: int = 30,
) -> dict[str, Any]:
    import asyncio

    last_response: dict[str, Any] | None = None
    for attempt in range(max_attempts):
        last_response = await client.post_json(endpoint, payload, request_url=request_url)
        status = str(last_response.get("status") or "").lower()
        resource_url = last_response.get("resourceUrl")
        if status == "done" and resource_url:
            return last_response
        if attempt < max_attempts - 1:
            await asyncio.sleep(2)
    raise XiyouConfigError(f"资源导出任务超时：{last_response}")


def _count_xlsx_rows(path: Path) -> int:
    """读取 xlsx 真实数据行数（不含表头）。

    resource 导出场景下，xlsx 由西柚后端生成并通过 OSS 下载，opscli
    无法从响应里直接拿到 rows 数组，必须打开文件才能给 row_count
    填充真实值。文件损坏或非标准 xlsx 时返回 0，避免拖垮主任务。
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            return max((ws.max_row or 0) - 1, 0)
        finally:
            wb.close()
    except Exception:
        return 0


def _resource_xlsx_export(output_path: Path) -> XiyouExportResult:
    resolved_output = output_path.resolve()
    return XiyouExportResult(
        path=str(resolved_output),
        filename=resolved_output.name,
        url=resolved_output.as_uri(),
        format="xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _export_rows_to_json(
    *,
    output_path: Path,
    job_id: str,
    request: XiyouRankingRequest,
    site: str,
    period: str,
    rank_pattern: str,
    rows: list[dict[str, Any]],
) -> XiyouExportResult:
    payload = {
        "job_id": job_id,
        "function": request.function,
        "provider": request.provider,
        "target": request.target,
        "site": site,
        "period": period,
        "rank_pattern": rank_pattern,
        "row_count": len(rows),
        "rows": rows,
        "warnings": [],
    }
    _write_json(output_path, payload)
    resolved_output = output_path.resolve()
    return XiyouExportResult(
        path=str(resolved_output),
        filename=resolved_output.name,
        url=resolved_output.as_uri(),
        format="json",
        mime_type="application/json",
    )


def _export_resource_to_json(
    *,
    output_path: Path,
    job_id: str,
    request: XiyouRankingRequest,
    site: str,
    dataset: str,
    resource_id: str,
    resource_url: str,
    status_response: dict[str, Any],
) -> XiyouExportResult:
    payload = {
        "job_id": job_id,
        "function": request.function,
        "provider": request.provider,
        "site": site,
        "dataset": dataset,
        "resource_id": resource_id,
        "resource_url": resource_url,
        "status_response": status_response,
        "warnings": [],
    }
    _write_json(output_path, payload)
    resolved_output = output_path.resolve()
    return XiyouExportResult(
        path=str(resolved_output),
        filename=resolved_output.name,
        url=resolved_output.as_uri(),
        format="json",
        mime_type="application/json",
    )


def _upload_export_if_enabled(
    *,
    export: XiyouExportResult,
    job_id: str,
    request: XiyouRankingRequest,
    site: str,
    period: str,
    rank_pattern: str,
    warnings: list[dict[str, Any]],
    jwt: str | None = None,
    session_id: str | None = None,
) -> None:
    client = FileUploadClient(jwt=jwt, session_id=session_id)
    if not client.enabled:
        return
    try:
        upload = client.upload(
            export.path,
            purpose="xiyou_export",
            folder="xiyou/exports",
            public="1",
            metadata={
                "job_id": job_id,
                "function": request.function,
                "provider": request.provider,
                "target": request.target,
                "site": site,
                "period": period,
                "rank_pattern": rank_pattern,
                "filename": export.filename,
            },
        )
        export.url = upload.url
    except FileUploadError as exc:
        warnings.append(
            {
                "stage": "file_upload",
                "message": "西柚洞察导出文件上传失败，已保留本地文件",
                "error": exc.to_dict(),
            }
        )
    except Exception as exc:
        warnings.append(
            {
                "stage": "file_upload",
                "message": "西柚洞察导出文件上传失败，已保留本地文件",
                "error": {
                    "code": type(exc).__name__,
                    "message": str(exc),
                },
            }
        )


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
