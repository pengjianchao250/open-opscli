"""西柚洞察 XLSX 导出。"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from opscli.xiyou.domain.exceptions import XiyouConfigError
from opscli.xiyou.domain.models import XiyouExportResult
from opscli.xiyou.export.columns import ExportColumn, columns_for_target


def export_rows_to_xlsx(
    *,
    rows: list[dict[str, Any]],
    output_path: Path,
    target: str,
    site: str = "US",
    period: str = "week",
) -> XiyouExportResult:
    """将接口 rows 导出为 XLSX。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise XiyouConfigError("缺少 openpyxl 依赖，无法导出 XLSX") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    columns = _merge_columns(columns_for_target(target), rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_title(f"Xiyou-{site.upper()}-{target}-{period}")

    header_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True)
    for column_index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=column.title)
        cell.font = header_font
        cell.fill = header_fill

    for row_index, row in enumerate(rows, start=2):
        for column_index, column in enumerate(columns, start=1):
            sheet.cell(row=row_index, column=column_index, value=_cell_value(_column_value(row, column)))

    sheet.freeze_panes = "A2"
    for column_index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = _column_width(column.title)

    workbook.save(output_path)
    resolved_output = output_path.resolve()
    return XiyouExportResult(
        path=str(resolved_output),
        filename=resolved_output.name,
        url=resolved_output.as_uri(),
    )


def _merge_columns(preferred: list[ExportColumn], rows: list[dict[str, Any]]) -> list[ExportColumn]:
    fields = _collect_fields(rows)
    used = {column.source for column in preferred if column.source}
    used.update(column.fallback for column in preferred if column.fallback)
    used_roots = {field.split(".", 1)[0] for field in used}
    extra = [ExportColumn(field, field) for field in fields if field not in used and field not in used_roots]
    return [*preferred, *extra] if preferred else extra


def _collect_fields(rows: list[dict[str, Any]]) -> list[str]:
    fields: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                fields.append(key)
    return fields


def _column_value(row: dict[str, Any], column: ExportColumn) -> Any:
    if column.source is None:
        return ""
    value = _get_value(row, column.source)
    if _is_blank(value) and column.fallback:
        value = _get_value(row, column.fallback)
    return value


def _get_value(row: dict[str, Any], field: str) -> Any:
    value: Any = row
    for part in field.split("."):
        if isinstance(value, dict):
            value = value.get(part)
        elif isinstance(value, list) and part.isdigit():
            index = int(part)
            value = value[index] if index < len(value) else None
        else:
            return None
    return value


def _cell_value(value: Any) -> Any:
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _safe_sheet_title(value: str) -> str:
    title = "".join(char for char in value if char not in r"[]:*?/\\")
    return (title or "xiyou")[:31]


def _column_width(title: str) -> int:
    if any(key in title for key in ["标题", "主图"]):
        return 42
    if any(key in title for key in ["关键词", "ASIN"]):
        return 24
    return max(12, min(22, len(str(title)) * 2 + 4))


def _is_blank(value: Any) -> bool:
    return value is None or value == ""
