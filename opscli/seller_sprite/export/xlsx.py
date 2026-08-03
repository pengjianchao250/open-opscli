"""卖家精灵 XLSX 导出。"""

from __future__ import annotations

import json
import re
from datetime import datetime, timedelta, timezone
from numbers import Real
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from opscli.seller_sprite.api.payloads import split_association_traffic_asins
from opscli.seller_sprite.domain.exceptions import SellerSpriteConfigError
from opscli.seller_sprite.domain.models import SellerSpriteExportResult
from opscli.seller_sprite.export.columns import ExportColumn, columns_for_scenario, currency_label
from opscli.seller_sprite.export.worksheet import SellerSpriteWorksheet


def build_export_worksheets(
    *,
    rows: list[dict[str, Any]],
    scenario: str,
    site: str = "US",
    period: str = "30d",
    params: dict[str, Any] | None = None,
    high_frequency_rows: list[dict[str, Any]] | None = None,
) -> list[SellerSpriteWorksheet]:
    """构造 XLSX 与 JSON 共用的格式化工作表。

    参数：
        rows: 接口返回的原始业务行。
        scenario: SellerSprite 场景标识。
        site: Amazon 站点代码，用于币种和标题格式化。
        period: 查询周期，用于动态列名和工作表名称。
        params: 场景参数，用于工作表名称和辅助表。
        high_frequency_rows: 可选高频词辅助表数据。

    返回：
        主表在首位、辅助表按 XLSX 顺序排列的格式化工作表列表。
    """
    if scenario in {
        "traffic-extend",
        "keyword-conversion-rate",
        "real-time-bidding",
    }:
        rows = rows[:100]
    normalized_params = params or {}
    columns = columns_for_scenario(scenario, site, period)
    if not columns:
        columns = [ExportColumn(field, field) for field in _collect_fields(rows)]

    worksheets = [
        SellerSpriteWorksheet(
            name=_main_sheet_title(
                scenario=scenario,
                site=site,
                period=period,
                params=normalized_params,
                rows=rows,
            ),
            columns=[column.title for column in columns],
            number_formats=_number_formats_for_scenario(scenario, len(columns)),
            rows=[
                [_cell_value(_column_value(row, column, site=site)) for column in columns]
                for row in rows
            ],
        )
    ]
    if scenario == "traffic-extend":
        worksheets.append(
            _high_frequency_worksheet(
                _traffic_extend_word_frequency(rows),
                percentage_format="0.00%",
            )
        )
        worksheets.append(
            _asin_worksheet(normalized_params.get("asins") or normalized_params.get("asin"))
        )
    elif high_frequency_rows:
        worksheets.append(_high_frequency_worksheet(high_frequency_rows))
    return worksheets


def export_rows_to_xlsx(
    *,
    rows: list[dict[str, Any]],
    output_path: Path,
    scenario: str,
    site: str = "US",
    period: str = "30d",
    params: dict[str, Any] | None = None,
    high_frequency_rows: list[dict[str, Any]] | None = None,
) -> SellerSpriteExportResult:
    """将接口 rows 导出为 XLSX。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise SellerSpriteConfigError("缺少 openpyxl 依赖，无法导出 XLSX") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    worksheets = build_export_worksheets(
        rows=rows,
        scenario=scenario,
        site=site,
        period=period,
        params=params,
        high_frequency_rows=high_frequency_rows,
    )
    rows = rows[:100] if scenario in {"traffic-extend", "keyword-conversion-rate", "real-time-bidding"} else rows
    columns = columns_for_scenario(scenario, site, period)
    if not columns:
        columns = [ExportColumn(dictionary_title, dictionary_title) for dictionary_title in _collect_fields(rows)]

    workbook = Workbook()
    sheet = workbook.active
    main_worksheet = worksheets[0]
    sheet.title = main_worksheet.name

    keyword_research = scenario == "keyword-research"
    aba_research = scenario == "aba-research"
    association_traffic = scenario == "association-traffic"
    traffic_extend = scenario == "traffic-extend"
    keyword_conversion_rate = scenario == "keyword-conversion-rate"
    real_time_bidding = scenario == "real-time-bidding"
    official_orange_header = (
        keyword_research
        or aba_research
        or association_traffic
        or keyword_conversion_rate
        or real_time_bidding
    )
    header_fill = PatternFill(
        "solid",
        fgColor=(
            "FFC00000"
            if traffic_extend
            else ("FFE98A00" if official_orange_header else "EAF2F8")
        ),
    )
    header_font = (
        Font(
            name="宋体",
            size=10,
            bold=False,
            color="FFFFFFFF",
        )
        if real_time_bidding
        else
        Font(
            name="Calibri",
            size=10,
            bold=False,
            color="FFFFFFFF" if keyword_conversion_rate else None,
        )
        if official_orange_header
        else Font(
            bold=True,
            color="FFFFFFFF" if traffic_extend else None,
        )
    )
    for column_index, title in enumerate(main_worksheet.columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=title)
        cell.font = header_font
        cell.fill = header_fill
        if official_orange_header:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, values in enumerate(main_worksheet.rows, start=2):
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            if aba_research:
                _apply_aba_research_style(cell, column_index)
            number_format = main_worksheet.number_formats[column_index - 1]
            if number_format and isinstance(value, Real) and not isinstance(value, bool):
                cell.number_format = number_format
            elif not number_format:
                _apply_number_format(cell)
        if association_traffic:
            _apply_association_traffic_hyperlinks(sheet, row_index, rows[row_index - 2])

    sheet.freeze_panes = "A2"
    for column_index, column in enumerate(columns, start=1):
        if keyword_research:
            width = KEYWORD_RESEARCH_COLUMN_WIDTHS[column_index - 1]
        elif aba_research:
            width = ABA_RESEARCH_COLUMN_WIDTHS[column_index - 1]
        elif association_traffic:
            width = ASSOCIATION_TRAFFIC_COLUMN_WIDTHS[column_index - 1]
        elif traffic_extend:
            width = TRAFFIC_EXTEND_COLUMN_WIDTHS[column_index - 1]
        elif keyword_conversion_rate:
            width = KEYWORD_CONVERSION_RATE_COLUMN_WIDTHS[column_index - 1]
        elif real_time_bidding:
            width = REAL_TIME_BIDDING_COLUMN_WIDTHS[column_index - 1]
        else:
            width = _column_width(column.title)
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    if aba_research:
        _apply_aba_research_sheet_style(sheet)
    for worksheet in worksheets[1:]:
        if worksheet.name == "Unique Words":
            _add_high_frequency_sheet(
                workbook,
                worksheet,
                traffic_extend_style=traffic_extend,
            )
        elif worksheet.name == "Asin":
            _add_asin_sheet(workbook, worksheet)
    workbook.save(output_path)
    resolved_output = output_path.resolve()
    return SellerSpriteExportResult(
        path=str(resolved_output),
        filename=resolved_output.name,
        url=resolved_output.as_uri(),
    )


def _collect_fields(rows: list[dict[str, Any]]) -> list[str]:
    fields: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                fields.append(key)
    return fields


def _get_value(row: dict[str, Any], field: str) -> Any:
    value: Any = row
    for part in field.split("."):
        if isinstance(value, dict):
            value = value.get(part)
        elif isinstance(value, list) and part.isdigit():
            index = int(part)
            value = value[index] if index < len(value) else None
        else:
            return None
    return value


def _column_value(row: dict[str, Any], column: ExportColumn, *, site: str) -> Any:
    if column.source is None:
        return ""
    value = _get_value(row, column.source)
    if _is_blank(value) and column.fallback:
        value = _get_value(row, column.fallback)
    return _apply_transform(value, column.transform, row, site=site)


def _apply_transform(value: Any, transform: str | None, row: dict[str, Any], *, site: str) -> Any:
    if not transform:
        return value
    if transform == "emptyIfNegative":
        return "" if _is_number(value) and float(value) < 0 else value
    if transform == "jsonObjectLines":
        return _json_object_lines(value)
    if transform == "amazonProductUrl":
        return _amazon_product_url(value, row)
    if transform == "amazonSellerUrl":
        return _amazon_seller_url(value, row)
    if transform == "badgeFlag":
        return "" if _is_blank(value) else "Y"
    if transform == "amazonChoiceKeyword":
        return "" if _is_blank(value) else "Amazon's Choice"
    if transform == "booleanY":
        return "Y" if bool(value) else ""
    if transform == "departmentsJoin":
        return _departments_join(value)
    if transform == "percentage":
        return _percentage(value)
    if transform == "percentSuffix":
        return "" if _is_blank(value) else f"{value}%"
    if transform == "dateMillis":
        return _date_millis(value)
    if transform == "keywordReverseUpdatedTime":
        return _keyword_reverse_updated_time(value, site=site)
    if transform == "realTimeBiddingQueryTime":
        return _real_time_bidding_query_time(value)
    if transform == "rankPosition":
        return _rank_position(value)
    if transform == "rankPage":
        return _rank_page(value)
    if transform == "divide10":
        return "" if _is_blank(value) else float(value) / 10
    if transform == "divide10Text":
        return "" if _is_blank(value) else f"{float(value) / 10:.1f}"
    if transform == "divide100":
        return "" if _is_blank(value) else float(value) / 100
    if transform == "sellerNation":
        return _seller_nation(value)
    if transform == "sellerAddress":
        return _seller_address(value)
    if transform in {"currency", "yen"}:
        return "" if _is_blank(value) else f"{currency_label(site)}{float(value):.2f}"
    if transform == "bidRange":
        return _bid_range(row, site=site)
    if transform == "abaBidRange":
        return _aba_bid_range(row, site=site)
    if transform == "abaHistoricalRanks":
        return _aba_period_lines(row, "SearchRank", value_kind="rank")
    if transform == "abaRankGrowthValues":
        return _aba_period_lines(row, "RankGrowthValue", value_kind="signed")
    if transform == "abaRankGrowthRates":
        return _aba_period_lines(row, "RankGrowthRate", value_kind="percentage")
    if transform == "abaClickShares":
        return _aba_share_lines(value, field="clickRate")
    if transform == "abaConversionShares":
        return _aba_share_lines(value, field="conversionRate")
    if transform == "abaTopAsins":
        return _aba_top_asins(value)
    if transform == "abaBrands":
        return _aba_list_text(value, separator="、")
    if transform == "abaDepartments":
        return _aba_list_text(value, separator="; ")
    if transform == "asinList":
        return _asin_list(value)
    if transform == "listLength":
        return len(value) if isinstance(value, list) else 0
    if transform == "listJoin":
        return _list_join(value)
    if transform == "badgeLabels":
        return _enum_list_join(value, BADGE_LABELS)
    if transform == "trafficSourceLabels":
        return _enum_list_join(value, TRAFFIC_SOURCE_LABELS)
    if transform == "trafficKeywordTypeLabels":
        return _enum_list_join(value, TRAFFIC_KEYWORD_TYPE_LABELS)
    if transform == "conversionKeywordTypeLabels":
        return _enum_list_join(value, CONVERSION_KEYWORD_TYPE_LABELS)
    if transform == "relationLabels":
        return _enum_list_join(value, ASSOCIATION_RELATION_LABELS)
    return value


def _cell_value(value: Any) -> Any:
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _apply_number_format(cell) -> None:
    value = cell.value
    if isinstance(value, bool) or not isinstance(value, Real):
        return
    cell.number_format = "#,##0" if float(value).is_integer() else "#,##0.00"


def _keyword_research_number_format(column_index: int) -> str:
    """返回关键词选品指定列的官方数字格式。"""
    # 列序号与官方 28 列主表一一对应，保留百分比、小数位和整数的显示口径。
    formats = {
        5: "#,##0.00%",
        7: "#,##0.00%",
        8: "#,##0_ ",
        9: "#,##0_ ",
        11: "#,##0.0",
        14: "#,##0.00%",
        15: "#,##0.00%",
        16: "#,##0.0%",
        19: "#,##0.0",
        20: "#,##0.00_ ",
        21: "#,##0.00_ ",
        22: "#,##0.00_ ",
        24: "#,##0.00%",
        26: "#,##0.00%",
    }
    return formats.get(column_index, "#,##0")


def _apply_aba_research_style(cell, column_index: int) -> None:
    """按官方 ABA 数据选品主表设置数据样式。"""
    from openpyxl.styles import Alignment, Font

    cell.font = Font(name="等线", size=10, bold=False)
    cell.alignment = Alignment(
        horizontal="justify" if column_index in {5, 6, 7, 14, 15} else (
            "right" if column_index in {3, 4, 8, 10, 11, 12} else (
                "center" if column_index in {9, 13} else "left"
            )
        ),
        vertical="center",
    )


def _aba_research_number_format(column_index: int) -> str:
    """返回 ABA 数据选品指定列的官方数字格式。"""
    formats = {
        3: "#,##0_ ",
        4: "#,##0_ ",
        5: "#,##0",
        6: "#,##0",
        7: "#,##0.00%",
        8: "0.00_ ",
        9: "#,##0",
        10: "#,##0_ ",
        11: "#,##0_ ",
        12: "#,##0",
        13: "#,##0",
        14: "#,##0.0%",
        15: "#,##0.0%",
    }
    return formats.get(column_index, "General")


def _apply_aba_research_sheet_style(sheet) -> None:
    """补齐 ABA 数据选品主表行高及官方表头字体差异。"""
    from openpyxl.styles import Font

    sheet.row_dimensions[1].height = 20
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 20
    # 官方模板中部分表头使用白色字体，其余沿用 indexed 颜色；统一白色可保证橙底可读性。
    for cell in sheet[1]:
        cell.font = Font(name="Calibri", size=10, bold=False, color="FFFFFFFF")


def _association_traffic_number_format(column_index: int) -> str:
    """返回关联流量指定列的官方数字格式。"""
    formats = {
        2: "#,##0_ ",
        13: "#,##0",
        14: "#,##0",
        15: "#,##0%",
        18: "#,##0",
        19: "#,##0%",
        20: "#,##0",
        21: "#,##0_ ",
        22: "#,##0_ ",
        23: "#,##0.00",
        24: "#,##0",
        25: "#,##0%",
        26: "#,##0.00",
        27: "#,##0",
        28: "#,##0.00%",
        29: "#,##0.00",
        30: "#,##0",
        33: "#,##0.00",
        34: "#,##0",
        35: "#,##0",
        36: "#,##0",
    }
    return formats.get(column_index, "#,##0")


def _traffic_extend_number_format(column_index: int) -> str:
    """返回拓展流量词指定列的官方数字格式。"""
    percentage_columns = {4, 12, 20, 21, 25, 26, 28, 29, 31, 32}
    if column_index in percentage_columns:
        return "0.00%"
    if column_index == 18:
        return "0.00"
    return "#,##0"


def _keyword_conversion_rate_number_format(column_index: int) -> str:
    """返回关键词转化率指定列的官方数字格式。"""
    if column_index in {
        7,
        8,
        18,
        19,
        20,
        22,
        23,
        25,
        26,
        28,
        29,
        31,
        32,
    }:
        return "0.00%"
    if column_index in {9, 10, 11, 12, 13, 14, 15, 16, 17, 21}:
        return "#,##0.00_ "
    if column_index in {3, 4, 5, 6}:
        return "#,##0_ "
    return "#,##0"


def _number_formats_for_scenario(scenario: str, column_count: int) -> list[str | None]:
    """返回与业务列对齐、供 XLSX 和 JSON 共用的数字格式。"""
    resolvers = {
        "keyword-research": _keyword_research_number_format,
        "aba-research": _aba_research_number_format,
        "association-traffic": _association_traffic_number_format,
        "traffic-extend": _traffic_extend_number_format,
        "keyword-conversion-rate": _keyword_conversion_rate_number_format,
    }
    resolver = resolvers.get(scenario)
    if resolver:
        return [resolver(index) for index in range(1, column_count + 1)]
    if scenario == "real-time-bidding":
        return ["0.00%" if index in {44, 45} else None for index in range(1, column_count + 1)]
    return [None] * column_count


def _apply_association_traffic_hyperlinks(sheet, row_index: int, row: dict[str, Any]) -> None:
    """补齐官方工作簿中 ASIN、主图、父体和大类目的可点击链接。"""
    asin = row.get("asin")
    if asin:
        cell = sheet.cell(row=row_index, column=1)
        cell.hyperlink = _amazon_product_url(asin, row)
        cell.style = "Hyperlink"
    image_url = row.get("bigImageUrl") or row.get("imageUrl")
    if image_url:
        cell = sheet.cell(row=row_index, column=9)
        cell.hyperlink = str(image_url)
        cell.style = "Hyperlink"
    parent = row.get("parent")
    if parent:
        cell = sheet.cell(row=row_index, column=10)
        cell.hyperlink = _amazon_product_url(parent, row)
        cell.style = "Hyperlink"
    category_url = _amazon_bestseller_url(row)
    if row.get("bsrLabel") and category_url:
        cell = sheet.cell(row=row_index, column=12)
        cell.hyperlink = category_url
        cell.style = "Hyperlink"


def _add_high_frequency_sheet(
    workbook,
    worksheet: SellerSpriteWorksheet,
    *,
    traffic_extend_style: bool = False,
) -> None:
    from openpyxl.styles import Font

    sheet = workbook.create_sheet(worksheet.name)
    for column_index, title in enumerate(worksheet.columns, start=1):
        sheet.cell(row=1, column=column_index, value=title)
        sheet.cell(row=1, column=column_index).font = Font(bold=True)
    for row_index, row in enumerate(worksheet.rows, start=2):
        sheet.cell(row=row_index, column=1, value=row[0])
        frequency_cell = sheet.cell(row=row_index, column=2, value=row[1])
        percentage_cell = sheet.cell(row=row_index, column=3, value=row[2])
        for column_index, cell in ((2, frequency_cell), (3, percentage_cell)):
            number_format = worksheet.number_formats[column_index - 1]
            if number_format and isinstance(cell.value, Real) and not isinstance(cell.value, bool):
                cell.number_format = number_format
            else:
                _apply_number_format(cell)
    sheet.column_dimensions["A"].width = 19 if traffic_extend_style else 24
    sheet.column_dimensions["B"].width = 18.6725663716814 if traffic_extend_style else 14
    sheet.column_dimensions["C"].width = 20 if traffic_extend_style else 14


def _traffic_extend_word_frequency(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """根据当前导出的 100 条关键词生成词频表。"""
    counts: dict[str, int] = {}
    total = 0
    for row in rows:
        keywords = str(row.get("keywords") or "").lower()
        for word in re.findall(r"[a-z0-9]+(?:['-][a-z0-9]+)*", keywords):
            counts[word] = counts.get(word, 0) + 1
            total += 1
    return [
        {
            "keyword": word,
            "frequency": frequency,
            "percentage": frequency / total if total else 0,
        }
        for word, frequency in sorted(
            counts.items(),
            key=lambda item: (-item[1], item[0]),
        )
    ]


def _add_asin_sheet(workbook, worksheet: SellerSpriteWorksheet) -> None:
    """写入拓展流量词的原始 ASIN 列表。"""
    from openpyxl.styles import Font

    sheet = workbook.create_sheet(worksheet.name)
    sheet.cell(row=1, column=1, value=worksheet.columns[0]).font = Font(bold=True)
    for row_index, row in enumerate(worksheet.rows, start=2):
        sheet.cell(row=row_index, column=1, value=row[0])
    sheet.column_dimensions["A"].width = 19


def _high_frequency_worksheet(
    rows: list[dict[str, Any]],
    *,
    percentage_format: str | None = None,
) -> SellerSpriteWorksheet:
    """构造高频词辅助工作表，字段取值与 XLSX 保持一致。"""
    return SellerSpriteWorksheet(
        name="Unique Words",
        columns=["词语", "出现频次", "百分比"],
        number_formats=[None, "#,##0", percentage_format],
        rows=[
            [
                row.get("keyword") or row.get("词语") or row.get("word"),
                row.get("frequency") or row.get("出现频次"),
                row.get("percentage") or row.get("百分比"),
            ]
            for row in rows
        ],
    )


def _asin_worksheet(value: Any) -> SellerSpriteWorksheet:
    """构造拓展流量词的原始 ASIN 辅助工作表。"""
    return SellerSpriteWorksheet(
        name="Asin",
        columns=["ASIN"],
        number_formats=[None],
        rows=[[asin] for asin in split_association_traffic_asins(value)],
    )


def _main_sheet_title(*, scenario: str, site: str, period: str, params: dict[str, Any], rows: list[dict[str, Any]]) -> str:
    if scenario == "keyword-conversion-rate":
        keyword = _keyword_conversion_rate_first_keyword(params)
        title = f"{site.upper()}-{keyword}({len(rows)})"
    elif scenario == "real-time-bidding":
        asin = str(params.get("asin") or "ASIN").strip().upper()
        query_time = _real_time_bidding_query_time(
            (rows[0].get("queryTime") if rows else "")
            or params.get("queryTime")
            or ""
        )
        timestamp = re.sub(r"\D", "", query_time)[:14] or "latest"
        title = f"{site.upper()}-{asin}-{timestamp}"
    elif scenario == "keyword-miner":
        keyword = params.get("keyword") or params.get("q") or "keyword"
        title = f"{site.upper()}-{keyword}({len(rows)})_"
    elif scenario == "keyword-reverse":
        asin = params.get("asin") or params.get("q") or "ASIN"
        title = f"{site.upper()}-{asin}-Keywords({len(rows)})_"
    elif scenario == "keyword-research":
        title = f"Keywords({len(rows)})"
    elif scenario == "aba-research":
        title = f"ABAKeyword({len(rows)})"
    elif scenario == "association-traffic":
        asins = split_association_traffic_asins(params.get("asins") or params.get("asin"))
        first_asin = asins[0] if asins else "ASIN"
        # 官网批量导出的主表名以 ``(31`` 收尾，本地保持相同可见命名。
        title = f"Related-{first_asin}-batch({len(asins)})(31"
    elif scenario == "traffic-extend":
        asins = split_association_traffic_asins(params.get("asins") or params.get("asin"))
        first_asin = asins[0] if asins else "ASIN"
        title = f"{site.upper()}-{first_asin}({len(asins)})__"
    elif scenario == "product-research":
        title = f"Product-{site.upper()}-{_period_label(period)}"
    elif scenario == "competitor-lookup":
        title = f"Competitor-{site.upper()}-{_period_label(period)}"
    elif scenario == "market-research":
        title = f"Market-research-{site.upper()}-{_period_label(period)}"
    else:
        title = scenario
    return _safe_sheet_title(title)


# 列宽来自官方 KeywordResearch-US-202606-667951.xlsx，用于保证本地导出视觉一致。
KEYWORD_RESEARCH_COLUMN_WIDTHS = [
    28,
    13,
    14,
    13,
    11,
    14,
    12,
    13,
    13,
    13,
    11,
    13,
    13,
    11,
    13,
    11,
    13,
    13,
    13,
    13,
    12.6637168141593,
    13.0530973451327,
    18,
    13,
    13,
    13,
    74,
    120,
]


# 列宽来自官方 ABAKeywordTrend-US-2026第29周-690875.xlsx。
ABA_RESEARCH_COLUMN_WIDTHS = [
    18.3362831858407,
    17.3362831858407,
    12,
    9.15929203539823,
    17.9115044247788,
    15,
    13,
    15,
    15,
    15,
    13,
    15,
    15,
    15,
    13,
    36,
    25.6637168141593,
    46,
    120,
]


# 列宽来自官方 RelatedProducts-US-B098T9ZFB5-batch(5)-260723.xlsx。
ASSOCIATION_TRAFFIC_COLUMN_WIDTHS = [
    14, 15, 15, 15, 15, 13, 35, 14, 13, 13, 19, 15, 10, 20, 13, 19, 10, 13, 13,
    15, 11, 13.353982300885, 11, 10, 13, 11, 13, 13, 13, 15, 13, 11, 11.0088495575221,
    9, 11, 13, 18, 13, 38, 12, 15, 13, 13, 10, 13, 13, 13.5132743362832, 13, 13,
    13, 13, 13, 13, 13, 13.5132743362832, 12.7610619469027,
]


# 列宽来自官方 ExpandKeywords-US-B089K9L3VY-batch(5)-202607-922297.xlsx。
TRAFFIC_EXTEND_COLUMN_WIDTHS = [
    26.5044247787611, 16.6725663716814, 13, 16.6725663716814, 16.6725663716814,
    13, 13, 13, 13, 16.1681415929204, 15.6725663716814, 13, 13, 13, 13, 13,
    14.6725663716814, 13.8495575221239, 13, 13, 18.6725663716814,
    12.353982300885, 16.6725663716814, 13, 13, 13, 13, 13, 13, 13, 13, 13,
    118.353982300885,
]

# 列宽按官方 KeywordConversionRate-US-wireless charger stand(123)-2026 Week 29.xlsx 对齐。
KEYWORD_CONVERSION_RATE_COLUMN_WIDTHS = [
    26.5044247787611, 16.6725663716814, 13, 16.6725663716814, 16.6725663716814,
    13, 13, 13, 16.1681415929204, 15.6725663716814, 13, 13, 13, 13, 13, 13,
    14.6725663716814, 13.8495575221239, 13, 13, 18.6725663716814,
    12.353982300885, 16.6725663716814, 13, 13, 13, 13, 13, 13, 13, 13, 13,
    118.353982300885,
]

# 官网实时查竞价主表宽度较统一，竞价列保留可完整显示长表头的宽度。
REAL_TIME_BIDDING_COLUMN_WIDTHS = [
    26, 16, 21, 13, 13, 13, 13,
    *([24] * 36),
    15, 15, 38,
]


def _keyword_conversion_rate_first_keyword(params: dict[str, Any]) -> str:
    """提取官网工作表名使用的首个关键词词组。"""
    value = (
        params.get("keywords")
        or params.get("keywordList")
        or params.get("keyword")
        or params.get("q")
        or "keyword"
    )
    if isinstance(value, (list, tuple, set)):
        value = next(iter(value), "keyword")
    first = re.split(r"[,，\r\n\t]+", str(value), maxsplit=1)[0]
    return re.sub(r"\s+", " ", first).strip() or "keyword"


def _safe_sheet_title(value: str) -> str:
    title = "".join(char for char in value if char not in r"[]:*?/\\")
    return (title or "seller-sprite")[:31]


def _period_label(period: str) -> str:
    text = str(period or "")
    if text in {"30d", "nearly", "latest30", "last30", ""}:
        return "Last-30-days"
    return text.replace("-", "")


def _column_width(title: str) -> int:
    if any(key in title for key in ["标题", "详细参数", "卖家信息"]):
        return 48
    if any(key in title for key in ["链接", "主图", "前十ASIN"]):
        return 38
    if any(key in title for key in ["类目路径", "尺寸"]):
        return 32
    if any(key in title for key in ["ASIN", "SKU", "品牌"]):
        return 18
    return max(12, min(22, len(str(title)) * 2 + 4))


def _json_object_lines(value: Any) -> Any:
    if _is_blank(value):
        return ""
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except json.JSONDecodeError:
            return value
    if not isinstance(value, dict):
        return value
    return "\n".join(f"{key}:{item}" for key, item in value.items())


def _amazon_product_url(value: Any, row: dict[str, Any]) -> str:
    if str(value or "").startswith("http"):
        return str(value)
    asin = value or row.get("asin")
    return f"https://{_amazon_domain(row)}/dp/{asin}" if asin else ""


def _amazon_seller_url(value: Any, row: dict[str, Any]) -> str:
    if not value:
        return ""
    return f"https://{_amazon_domain(row)}/gp/help/seller/at-a-glance.html?seller={value}"


def _amazon_bestseller_url(row: dict[str, Any]) -> str:
    """生成官方关联流量导出中的大类 Best Sellers 链接。"""
    # 官网美国站使用产品组短路径；未命中时用 BSR 节点 ID，仍可落到对应榜单。
    us_category_paths = {
        "Beauty & Personal Care": "beauty",
        "Clothing, Shoes & Jewelry": "fashion",
        "Health & Household": "hpc",
        "Home & Kitchen": "home-garden",
        "Patio, Lawn & Garden": "lawn-garden",
    }
    domain = _amazon_domain(row)
    label = str(row.get("bsrLabel") or "")
    category = us_category_paths.get(label) if domain == "www.amazon.com" else None
    category = category or row.get("bsrId")
    return f"https://{domain}/gp/bestsellers/{category}" if category else ""


def _amazon_domain(row: dict[str, Any]) -> str:
    station = str(row.get("station") or "").upper()
    market_id = row.get("marketId")
    if station == "JAPAN" or market_id == 6:
        return "www.amazon.co.jp"
    if station == "GERMANY" or market_id == 4:
        return "www.amazon.de"
    if station == "UNITED_KINGDOM" or market_id == 3:
        return "www.amazon.co.uk"
    if station == "CANADA" or market_id == 7:
        return "www.amazon.ca"
    if station == "FRANCE" or market_id == 5:
        return "www.amazon.fr"
    if station == "ITALY" or market_id in {8, 35691}:
        return "www.amazon.it"
    if station == "SPAIN" or market_id in {9, 44551}:
        return "www.amazon.es"
    if station == "INDIA" or market_id in {10, 44571}:
        return "www.amazon.in"
    if station == "MEXICO" or market_id in {11, 771770}:
        return "www.amazon.com.mx"
    return "www.amazon.com"


def _departments_join(value: Any) -> str:
    if not isinstance(value, list):
        return ""
    return "&".join(str(item.get("label")) for item in value if isinstance(item, dict) and item.get("label"))


def _percentage(value: Any) -> str:
    if _is_blank(value):
        return ""
    return f"{float(value) * 100:.2f}%"


def _date_millis(value: Any) -> str:
    if _is_blank(value):
        return ""
    return datetime.fromtimestamp(float(value) / 1000, tz=timezone.utc).strftime("%Y-%m-%d")


def _real_time_bidding_query_time(value: Any) -> str:
    """将详情任务时间统一为官方工作簿的秒级显示格式。"""
    if _is_blank(value):
        return ""
    text = str(value).strip()
    if re.fullmatch(r"\d{12,13}", text):
        return datetime.fromtimestamp(
            float(text) / 1000,
            tz=_timezone_for_site("CN"),
        ).strftime("%Y/%m/%d %H:%M:%S")
    normalized = text.replace("T", " ").replace("Z", "")
    match = re.match(
        r"(\d{4})[-/](\d{2})[-/](\d{2})[ ]+(\d{2}):(\d{2}):(\d{2})",
        normalized,
    )
    if not match:
        return text
    return (
        f"{match.group(1)}/{match.group(2)}/{match.group(3)} "
        f"{match.group(4)}:{match.group(5)}:{match.group(6)}"
    )


def _keyword_reverse_updated_time(value: Any, *, site: str) -> str:
    if _is_blank(value):
        return ""
    timestamp = float(value) / 1000
    china_time = datetime.fromtimestamp(timestamp, tz=_timezone_for_site("CN"))
    site_time = datetime.fromtimestamp(timestamp, tz=_timezone_for_site(site))
    site_label = SITE_TIME_LABELS.get(str(site).upper(), str(site).upper())
    return f"中{china_time.strftime('%m.%d %H:%M')}\n{site_label}{site_time.strftime('%m.%d %H:%M')}"


def _rank_page(value: Any) -> str:
    if not isinstance(value, dict):
        return ""
    page = value.get("page")
    index = value.get("index")
    page_size = value.get("pageSize")
    if _is_blank(page):
        return ""
    if _is_blank(index) or _is_blank(page_size):
        return f"第{page}页"
    return f"第{page}页,{index}/{page_size}"


def _rank_position(value: Any) -> Any:
    if not isinstance(value, dict):
        return "前3页无排名" if _is_blank(value) else value
    position = value.get("position")
    return "前3页无排名" if _is_blank(position) else position


def _seller_nation(value: Any) -> str:
    if _is_blank(value):
        return ""
    text = str(value)
    return "CN(HK)" if text == "HK" else text


def _seller_address(value: Any) -> str:
    """将接口中的 HTML 换行地址归一为官方导出的单行文本。"""
    if _is_blank(value):
        return ""
    text = re.sub(r"<br\s*/?>", " ", str(value), flags=re.IGNORECASE)
    return " ".join(text.split())


def _bid_range(row: dict[str, Any], *, site: str) -> str:
    bid_min = row.get("bidMin")
    bid_max = row.get("bidMax")
    if _is_blank(bid_min) or _is_blank(bid_max):
        return "-"
    currency = currency_label(site)
    return f"{currency}{float(bid_min):.2f}-{currency}{float(bid_max):.2f}"


def _aba_bid_range(row: dict[str, Any], *, site: str) -> str:
    """按 ABA 官方模板拼接建议竞价范围。"""
    bid_min = row.get("bidMin")
    bid_max = row.get("bidMax")
    if _is_blank(bid_min) or _is_blank(bid_max):
        return "-"
    currency = currency_label(site)
    return f"{currency}{float(bid_min):.2f} - {currency}{float(bid_max):.2f}"


def _aba_period_lines(row: dict[str, Any], suffix: str, *, value_kind: str) -> str:
    """把 ABA 的上周、4 周前和 12 周前指标拼为官方多行文本。"""
    periods = (("上周", "w1"), ("4周前", "w4"), ("12周前", "w12"))
    lines = []
    for label, prefix in periods:
        value = row.get(f"{prefix}{suffix}")
        lines.append(f"{label}: {_aba_metric_text(value, value_kind=value_kind)}")
    return "\n".join(lines)


def _aba_metric_text(value: Any, *, value_kind: str) -> str:
    """格式化 ABA 排名、变化值和百分比。"""
    if _is_blank(value):
        return "-"
    number = float(value)
    if value_kind == "percentage":
        sign = "+" if number > 0 else ""
        return f"{sign}{number:.2f}%"
    integer = int(number)
    sign = "+" if value_kind == "signed" and integer > 0 else ""
    return f"{sign}{integer:,}"


def _aba_share_lines(value: Any, *, field: str) -> str:
    """拼接前三 ASIN 占比及合计，接口百分比值保持原始口径。"""
    items = value if isinstance(value, list) else []
    rates = []
    for index in range(3):
        item = items[index] if index < len(items) and isinstance(items[index], dict) else {}
        raw_rate = item.get(field)
        # 历史响应中的转化字段可能使用 conversionShareRate，兼容后仍按页面数值输出。
        if field == "conversionRate" and _is_blank(raw_rate):
            raw_rate = item.get("conversionShareRate")
        try:
            rates.append(float(raw_rate or 0))
        except (TypeError, ValueError):
            rates.append(0.0)
    lines = [f"TOP{index}: {rate:.2f}%" for index, rate in enumerate(rates, start=1)]
    lines.append(f"合计: {sum(rates):.2f}%")
    return "\n".join(lines)


def _aba_top_asins(value: Any) -> str:
    """按官方分隔符拼接点击前三 ASIN。"""
    if not isinstance(value, list):
        return ""
    return "、".join(
        str(item.get("asin"))
        for item in value[:3]
        if isinstance(item, dict) and item.get("asin")
    )


def _aba_list_text(value: Any, *, separator: str) -> str:
    """兼容字符串和对象列表并保持接口顺序。"""
    if _is_blank(value):
        return ""
    if not isinstance(value, list):
        return str(value)
    parts = []
    for item in value:
        if isinstance(item, dict):
            text = item.get("label") or item.get("name") or item.get("brand") or item.get("value")
        else:
            text = item
        if not _is_blank(text):
            parts.append(str(text))
    return separator.join(parts)


SITE_TIME_LABELS = {
    "US": "美",
    "JP": "日",
    "DE": "德",
    "UK": "英",
    "FR": "法",
    "IT": "意",
    "ES": "西",
    "CA": "加",
    "IN": "印",
    "MX": "墨",
}

SITE_TIMEZONES = {
    "CN": ("Asia/Shanghai", timezone(timedelta(hours=8))),
    "US": ("America/Los_Angeles", timezone(timedelta(hours=-7))),
    "JP": ("Asia/Tokyo", timezone(timedelta(hours=9))),
    "DE": ("Europe/Berlin", timezone(timedelta(hours=1))),
    "UK": ("Europe/London", timezone.utc),
    "FR": ("Europe/Paris", timezone(timedelta(hours=1))),
    "IT": ("Europe/Rome", timezone(timedelta(hours=1))),
    "ES": ("Europe/Madrid", timezone(timedelta(hours=1))),
    "CA": ("America/Los_Angeles", timezone(timedelta(hours=-7))),
    "IN": ("Asia/Kolkata", timezone(timedelta(hours=5, minutes=30))),
    "MX": ("America/Mexico_City", timezone(timedelta(hours=-6))),
}


def _timezone_for_site(site: str):
    name, fallback = SITE_TIMEZONES.get(str(site).upper(), SITE_TIMEZONES["CN"])
    try:
        return ZoneInfo(name)
    except ZoneInfoNotFoundError:
        return fallback


def _asin_list(value: Any) -> str:
    if not isinstance(value, list):
        return ""
    return ",".join(str(item.get("asin")) for item in value if isinstance(item, dict) and item.get("asin"))


def _list_join(value: Any) -> str:
    if not isinstance(value, list):
        return "" if _is_blank(value) else str(value)
    parts = []
    for item in value:
        if isinstance(item, str):
            parts.append(item)
        elif isinstance(item, dict):
            parts.append(str(item.get("label") or item.get("name") or item.get("code") or item.get("value") or json.dumps(item, ensure_ascii=False)))
        elif item is not None:
            parts.append(str(item))
    return "/".join(part for part in parts if part)


BADGE_LABELS = {
    "NATURAL_SEARCHING": "自然搜索词",
    "naturalSearching": "自然搜索词",
    "AMAZON_CHOICE": "AC推荐词",
    "AMAZON_CHOICH": "AC推荐词",
    "amazonChoice": "AC推荐词",
    "EDITORIAL_RECOMMENDATIONS": "ER推荐词",
    "editorialRecommendations": "ER推荐词",
    "FOUR_STAR": "4星推荐词",
    "fourStar": "4星推荐词",
    "HIGHLY_RATED": "HR推荐词",
    "highlyRated": "HR推荐词",
    "SPONSOR_BRAND": "品牌广告词",
    "sponsorBrand": "品牌广告词",
    "SPONSOR_VIDEO": "视频广告词",
    "sponsorVideo": "视频广告词",
    "ADS": "SP广告词",
    "ads": "SP广告词",
}

TRAFFIC_SOURCE_LABELS = {
    "SEARCH": "自然搜索",
    "search": "自然搜索",
    "OFFICIAL": "亚马逊推荐",
    "official": "亚马逊推荐",
    "AD": "PPC广告",
    "ad": "PPC广告",
}

TRAFFIC_KEYWORD_TYPE_LABELS = {
    "PRIMARY": "主要流量词",
    "primary": "主要流量词",
    "PRECISE": "精准流量词",
    "precise": "精准流量词",
    "PRECISE_LONG_TAIL": "转化流失词",
    "preciseLongTail": "转化流失词",
}

CONVERSION_KEYWORD_TYPE_LABELS = {
    "EXCELLENT": "转化优质词",
    "excellent": "转化优质词",
    "STABLE": "转化平稳词",
    "stable": "转化平稳词",
    "LOST": "转化流失词",
    "lost": "转化流失词",
    "INVALID": "无效曝光词",
    "invalid": "无效曝光词",
}

# 关联类型 code 与官网筛选及官方导出中文值保持一致。
ASSOCIATION_RELATION_LABELS = {
    "VAV": "看了又看",
    "CSI": "相似产品",
    "AVP": "看了还看",
    "BAV": "看了却买",
    "MIB": "捆绑销售",
    "FBT": "组合购买",
    "MIE": "更多相关",
    "BAB": "买了又买",
    "COB": "品牌推荐",
    "SP": "商品广告",
    "FSA": "四星产品",
    "BCA": "品牌广告",
}


def _enum_list_join(value: Any, labels: dict[str, str]) -> str:
    if _is_blank(value):
        return ""
    values = value if isinstance(value, list) else [value]
    parts = []
    for item in values:
        key = str(item)
        parts.append(labels.get(key, key))
    return "/".join(part for part in parts if part)


def _is_blank(value: Any) -> bool:
    return value is None or value == ""


def _is_number(value: Any) -> bool:
    try:
        float(value)
    except (TypeError, ValueError):
        return False
    return True
