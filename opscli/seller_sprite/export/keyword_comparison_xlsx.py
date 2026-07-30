"""卖家精灵流量词对比动态 XLSX 导出。"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from opscli.seller_sprite.domain.exceptions import SellerSpriteConfigError
from opscli.seller_sprite.domain.models import SellerSpriteExportResult


# 官网流量词类型枚举与中文工作簿标签的固定映射。
TRAFFIC_KEYWORD_TYPE_LABELS = {
    "PRIMARY": "主要流量词",
    "PRECISE": "精准流量词",
    "PRECISE_LONG_TAIL": "精准长尾词",
}
# 所有 ASIN 动态对比列之后追加的公共业务指标及接口字段。
COMMON_COLUMNS = [
    ("有效竞品数", "competitors"),
    ("ABA排名(周)", "searchesRank"),
    ("月搜索量", "searches"),
    ("月购买量", "purchases"),
    ("购买率", "purchaseRate"),
    ("展示量", "impressions"),
    ("点击量", "clicks"),
    ("商品数", "products"),
    ("需供比", "supplyDemandRatio"),
]


def export_keyword_comparison_to_xlsx(
    *,
    rows: list[dict[str, Any]],
    output_path: Path,
    site: str,
    own_asin: str,
    asin_list: list[str],
) -> SellerSpriteExportResult:
    """按最终畅销变体顺序生成流量词对比业务工作簿。

    参数：
        rows: 第一页流量词业务行。
        output_path: 本地 XLSX 输出路径。
        site: 查询站点。
        own_asin: 用户自己的 ASIN。
        asin_list: 页面最终提交的畅销变体 ASIN 顺序。

    返回：
        已生成工作簿的文件元数据。

    异常：
        SellerSpriteConfigError: 依赖缺失或最终 ASIN 列表不合法时抛出。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise SellerSpriteConfigError("缺少 openpyxl 依赖，无法导出 XLSX") from exc

    normalized_own_asin = str(own_asin or "").strip().upper()
    normalized_asins = _normalize_asin_list(asin_list)
    if normalized_own_asin not in normalized_asins:
        raise SellerSpriteConfigError("流量词对比导出缺少自己的 ASIN")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _main_sheet_title(site, normalized_asins)
    sheet.sheet_view.showGridLines = False

    headers = ["关键词", "关键词翻译"]
    for asin in normalized_asins:
        headers.extend(
            [
                f"{asin}(我的)" if asin == normalized_own_asin else asin,
                f"{asin}流量词类型",
            ]
        )
    headers.extend(title for title, _ in COMMON_COLUMNS)

    thin_side = Side(style="thin", color="FFD9D9D9")
    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )
    header_fill = PatternFill("solid", fgColor="FFE98A00")
    header_font = Font(name="等线", size=10, color="FFFFFFFF")
    data_font = Font(name="等线", size=10)

    for column_index, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(rows, start=2):
        competitors = _competitor_map(row.get("competitorList"))
        values: list[Any] = [row.get("keyword"), row.get("keywordCn")]
        for asin in normalized_asins:
            competitor = competitors.get(asin, {})
            values.extend(
                [
                    _traffic_percentage(competitor.get("trafficPercentage")),
                    _traffic_keyword_types(competitor.get("trafficKeywordTypes")),
                ]
            )
        values.extend(row.get(field) for _, field in COMMON_COLUMNS)
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = data_font
            cell.border = border
            if column_index <= 2:
                horizontal = "left"
            elif column_index <= 2 + len(normalized_asins) * 2:
                horizontal = "center" if column_index % 2 == 0 else "right"
            else:
                horizontal = "right"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center")
        sheet.row_dimensions[row_index].height = 20
    sheet.row_dimensions[1].height = 20

    widths = (
        [25, 22.6719]
        + [15.671875] * (len(normalized_asins) * 2)
        + [12] * 7
        + [9, 13]
    )
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    asin_sheet = workbook.create_sheet("ASIN")
    asin_sheet["A1"] = "asin"
    asin_sheet["A2"] = f"{normalized_own_asin}(我的)"
    for row_index, asin in enumerate(normalized_asins, start=3):
        asin_sheet.cell(row=row_index, column=1, value=asin)

    workbook.save(output_path)
    resolved = output_path.resolve()
    return SellerSpriteExportResult(
        path=str(resolved),
        filename=resolved.name,
        url=resolved.as_uri(),
    )


def _normalize_asin_list(values: list[str]) -> list[str]:
    """校验并稳定去重最终参与对比的 ASIN 列表。"""
    asins: list[str] = []
    for value in values:
        asin = str(value or "").strip().upper()
        if not re.fullmatch(r"[A-Z0-9]{10}", asin):
            raise SellerSpriteConfigError(f"流量词对比导出 ASIN 格式无效：{asin}")
        if asin not in asins:
            asins.append(asin)
    if not asins:
        raise SellerSpriteConfigError("流量词对比导出缺少 ASIN 列表")
    return asins


def _competitor_map(value: Any) -> dict[str, dict[str, Any]]:
    """将行内竞品对比数据按 ASIN 建立索引。"""
    if not isinstance(value, list):
        return {}
    return {
        str(item.get("asin") or "").strip().upper(): item
        for item in value
        if isinstance(item, dict) and item.get("asin")
    }


def _traffic_percentage(value: Any) -> str | None:
    """将接口比例转换为官网工作簿使用的百分比文本。"""
    if value is None or value == "":
        return None
    try:
        percentage = float(value) * 100
    except (TypeError, ValueError):
        return str(value)
    if 0 < percentage < 0.01:
        return "<0.01%"
    return f"{percentage:.2f}%"


def _traffic_keyword_types(value: Any) -> str | None:
    """将流量词类型枚举转换为中文标签。"""
    if not value:
        return None
    values = value if isinstance(value, list) else [value]
    labels = [TRAFFIC_KEYWORD_TYPE_LABELS.get(str(item), str(item)) for item in values]
    return "、".join(labels)


def _main_sheet_title(site: str, asin_list: list[str]) -> str:
    """生成符合 Excel 31 字符限制的官方语义主表名称。"""
    title = f"{str(site).upper()}-流量占比对比-{''.join(asin_list)}"
    return title[:31]
