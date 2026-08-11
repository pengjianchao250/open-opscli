"""Canopy 场景执行、脱敏、上传和凭据状态回写测试。"""

import asyncio
import json
from pathlib import Path

import httpx
import pytest
from openpyxl import load_workbook

from opscli.beta.canopy.config import CanopySettings
from opscli.beta.canopy.domain.models import CanopyScenarioRequest
from opscli.beta.canopy.services import api_manager as api_manager_module
from opscli.beta.canopy.services.api_manager import CanopyApiManager


def _run(coro):
    return asyncio.run(coro)


class DisabledUploadClient:
    """测试用上传客户端，默认关闭真实文件上传。"""

    enabled = False

    def __init__(self, *args, **kwargs):
        pass


class FakeCredentialPool:
    def __init__(self):
        self.successes = []
        self.failures = []

    def report_success(self, lease, *, runtime=None):
        self.successes.append((lease, runtime))

    def report_failure(self, lease, **kwargs):
        self.failures.append((lease, kwargs))


def test_canopy_http_error_redacts_api_key_from_payload() -> None:
    """Canopy 错误响应即使回显 Key，也不能进入异常或运行状态。"""
    response = httpx.Response(
        401,
        json={"error": "invalid secret-canopy-key", "api_key": "secret-canopy-key"},
    )

    error = api_manager_module._api_error_from_response(
        response,
        api_key="secret-canopy-key",
    )

    assert "secret-canopy-key" not in str(error.response_payload)
    assert "secret-canopy-key" not in str(error.response_excerpt)
    assert "<REDACTED>" in str(error.response_payload)


def test_manager_writes_params_raw_result_and_xlsx_without_api_key(monkeypatch, tmp_path: Path):
    captured = {}

    async def fake_request_canopy_api(*, path, params, api_key, timeout_seconds):
        captured.update({"path": path, "params": params, "api_key": api_key, "timeout_seconds": timeout_seconds})
        return {
            "success": True,
            "debug": "secret-canopy-key should not leak",
            "data": {
                "reviews": [
                    {
                        "asin": "B0B3JBVDYP",
                        "reviewTitle": "Great",
                        "reviewText": "Works well",
                        "verifiedPurchase": True,
                    }
                ]
            },
        }

    monkeypatch.setattr(api_manager_module, "request_canopy_api", fake_request_canopy_api)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DisabledUploadClient, raising=False)
    manager = CanopyApiManager(settings=CanopySettings(output_dir=tmp_path))

    result = _run(
        manager.run(
            CanopyScenarioRequest(
                scenario="product-reviews",
                domain="US",
                params={"asin": "B0B3JBVDYP", "rating": "FIVE_STAR"},
                path="/api/amazon/product/reviews",
                method="GET",
                title="Amazon 商品评论",
                api_key="secret-canopy-key",
                timeout_seconds=3,
                job_id="canopy-review-regression",
            )
        )
    )

    root_dir = tmp_path / "canopy-review-regression"
    assert captured["path"] == "/api/amazon/product/reviews"
    assert captured["params"]["domain"] == "US"
    assert captured["api_key"] == "secret-canopy-key"
    assert result.row_count == 1
    assert (root_dir / "params.json").exists()
    assert (root_dir / "raw.json").exists()
    assert (root_dir / "result.json").exists()
    assert result.export is not None
    assert result.export.path.endswith("canopy-review-regression.xlsx")
    assert result.export.format == "xlsx"

    combined = "\n".join(
        [
            (root_dir / "params.json").read_text(encoding="utf-8"),
            (root_dir / "raw.json").read_text(encoding="utf-8"),
            (root_dir / "result.json").read_text(encoding="utf-8"),
        ]
    )
    assert "secret-canopy-key" not in combined

    params_payload = json.loads((root_dir / "params.json").read_text(encoding="utf-8"))
    raw_payload = json.loads((root_dir / "raw.json").read_text(encoding="utf-8"))
    assert params_payload["normalized_params"]["asin"] == "B0B3JBVDYP"
    assert raw_payload["request_params"]["domain"] == "US"

    workbook = load_workbook(result.export.path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    assert "ASIN" in headers
    assert "评论标题" in headers
    assert sheet.cell(row=2, column=headers.index("ASIN") + 1).value == "B0B3JBVDYP"


def test_manager_reports_pool_account_success(monkeypatch, tmp_path: Path):
    async def fake_request_canopy_api(**kwargs):
        return {"success": True, "data": {"product": {"asin": "B0B3JBVDYP"}}}

    monkeypatch.setattr(api_manager_module, "request_canopy_api", fake_request_canopy_api)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DisabledUploadClient)
    pool = FakeCredentialPool()
    manager = CanopyApiManager(
        settings=CanopySettings(output_dir=tmp_path),
        credential_pool=pool,
    )

    _run(
        manager.run(
            CanopyScenarioRequest(
                scenario="product",
                domain="US",
                params={"asin": "B0B3JBVDYP"},
                path="/api/amazon/product",
                api_key="secret-canopy-key",
                credential_account_id=15,
                credential_account_name="primary",
                credential_secret_version=3,
                job_id="canopy-pool-success",
            )
        )
    )

    assert len(pool.successes) == 1
    assert pool.successes[0][0].account_id == 15
    assert pool.successes[0][0].secret_version == 3
    assert pool.failures == []


class DummyUploadClient:
    """测试用上传客户端，记录 beta 导出文件上传参数。"""

    instances = []

    def __init__(self, *args, **kwargs):
        self.uploads = []
        self.enabled = True
        self.__class__.instances.append(self)

    def upload(self, path, *, purpose, folder=None, public=None, metadata=None):
        self.uploads.append(
            {
                "path": str(path),
                "purpose": purpose,
                "folder": folder,
                "public": public,
                "metadata": metadata,
            }
        )

        class Upload:
            url = "https://ops.example.com/download/canopy-review.xlsx"

        return Upload()


def test_manager_uploads_export_and_returns_download_url(monkeypatch, tmp_path: Path):
    """beta 导出应复用公共上传能力，成功后返回远端下载链接。"""
    DummyUploadClient.instances = []

    async def fake_request_canopy_api(**kwargs):
        return {
            "success": True,
            "data": {
                "reviews": [
                    {"asin": "B0CDQFTWQ2", "reviewTitle": "Good", "reviewText": "Works"}
                ]
            },
        }

    monkeypatch.setattr(api_manager_module, "request_canopy_api", fake_request_canopy_api)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DummyUploadClient, raising=False)
    manager = CanopyApiManager(settings=CanopySettings(output_dir=tmp_path))

    result = _run(
        manager.run(
            CanopyScenarioRequest(
                scenario="product-reviews",
                domain="US",
                params={"asin": "B0CDQFTWQ2", "page": 1},
                path="/api/amazon/product/reviews",
                method="GET",
                title="Amazon 商品评论",
                api_key="secret-canopy-key",
                timeout_seconds=60,
                job_id="canopy-review-upload",
            )
        )
    )

    assert result.export is not None
    assert result.export.url == "https://ops.example.com/download/canopy-review.xlsx"
    assert len(DummyUploadClient.instances) == 1
    upload = DummyUploadClient.instances[0].uploads[0]
    assert upload["purpose"] == "beta_canopy_export"
    assert upload["folder"] == "beta/canopy/export"
    assert upload["public"] == "1"
    assert upload["metadata"]["job_id"] == "canopy-review-upload"
    assert upload["metadata"]["scenario"] == "product-reviews"
    saved = json.loads((tmp_path / "canopy-review-upload" / "result.json").read_text(encoding="utf-8"))
    assert saved["export"]["url"] == "https://ops.example.com/download/canopy-review.xlsx"


def test_manager_rejects_non_xls_export_before_http(monkeypatch, tmp_path: Path):
    called = False

    async def fake_request_canopy_api(**kwargs):
        nonlocal called
        called = True
        return {"success": True}

    monkeypatch.setattr(api_manager_module, "request_canopy_api", fake_request_canopy_api)
    manager = CanopyApiManager(settings=CanopySettings(output_dir=tmp_path))

    with pytest.raises(Exception) as exc_info:
        _run(
            manager.run(
                CanopyScenarioRequest(
                    scenario="product",
                    domain="US",
                    params={"asin": "B0B3JBVDYP"},
                    path="/api/amazon/product",
                    api_key="secret-canopy-key",
                    job_id="canopy-invalid-export",
                    export_format="xlsx",
                )
            )
        )

    assert "不支持的导出格式" in str(exc_info.value)
    assert called is False
    assert not (tmp_path / "canopy-invalid-export").exists()
