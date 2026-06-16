from pathlib import Path

from openpyxl import load_workbook

from opscli.beta.canopy.export.xlsx import export_rows_to_xlsx, response_to_export_rows


def test_response_to_export_rows_extracts_nested_reviews():
    rows = response_to_export_rows(
        {
            "success": True,
            "data": {
                "reviews": [
                    {"asin": "B0B3JBVDYP", "reviewTitle": "Good", "reviewText": "Works"},
                    {"asin": "B0B3JBVDYP", "reviewTitle": "Bad", "reviewText": "Broken"},
                ],
                "totalResults": 2,
            },
        }
    )

    assert len(rows) == 2
    assert rows[0]["success"] is True
    assert rows[0]["rowSource"] == "reviews"
    assert rows[0]["asin"] == "B0B3JBVDYP"
    assert rows[1]["reviewTitle"] == "Bad"


def test_response_to_export_rows_extracts_canopy_reviews_paginated_shape():
    rows = response_to_export_rows(
        {
            "success": True,
            "data": {
                "amazonProduct": {
                    "asin": "B0B3JBVDYP",
                    "title": "Rich's Backyard Blackened Seasoning",
                    "brand": "RICH'S BACKYARD",
                    "rating": 4.8,
                    "ratingsTotal": 34,
                    "reviewsPaginated": {
                        "reviews": [
                            {
                                "id": "R1",
                                "title": "Delicious.",
                                "body": "Great flavor.",
                                "rating": 5,
                                "helpfulVotes": 3,
                                "isVerifiedPurchase": True,
                                "profileName": "Alice",
                                "profileId": "AGUEST1",
                                "profileUrl": "https://www.amazon.com/gp/profile/AGUEST1",
                                "date": "Reviewed in the United States on June 1, 2026",
                                "imageUrls": ["https://example.test/review-1.jpg"],
                            },
                            {
                                "id": "R2",
                                "title": "Best blackening seasoning yet",
                                "body": "So easy to use.",
                                "rating": 5,
                                "helpfulVotes": 1,
                                "verifiedPurchase": False,
                                "reviewerName": "Bob",
                            },
                        ],
                        "pageInfo": {"currentPage": 1, "totalPages": 3, "totalResults": 26},
                    },
                    "topReviews": [{"id": "TOP", "title": "Top review"}],
                }
            },
        }
    )

    assert len(rows) == 2
    assert rows[0]["success"] is True
    assert rows[0]["rowSource"] == "reviewsPaginated.reviews"
    assert rows[0]["asin"] == "B0B3JBVDYP"
    assert rows[0]["productTitle"] == "Rich's Backyard Blackened Seasoning"
    assert rows[0]["brand"] == "RICH'S BACKYARD"
    assert rows[0]["productRating"] == 4.8
    assert rows[0]["ratingsTotal"] == 34
    assert rows[0]["reviewId"] == "R1"
    assert rows[0]["reviewTitle"] == "Delicious."
    assert rows[0]["reviewText"] == "Great flavor."
    assert rows[0]["verifiedPurchase"] is True
    assert rows[0]["reviewerName"] == "Alice"
    assert rows[0]["reviewerId"] == "AGUEST1"
    assert rows[0]["currentPage"] == 1
    assert rows[0]["totalPages"] == 3
    assert rows[0]["totalResults"] == 26
    assert rows[1]["reviewTitle"] == "Best blackening seasoning yet"
    assert rows[1]["verifiedPurchase"] is False


def test_response_to_export_rows_falls_back_to_canopy_top_reviews():
    rows = response_to_export_rows(
        {
            "success": True,
            "data": {
                "amazonProduct": {
                    "asin": "B0B3JBVDYP",
                    "title": "Test Product",
                    "topReviews": [{"id": "TOP", "title": "Only top review", "body": "Fallback"}],
                }
            },
        }
    )

    assert len(rows) == 1
    assert rows[0]["rowSource"] == "topReviews"
    assert rows[0]["reviewId"] == "TOP"
    assert rows[0]["reviewTitle"] == "Only top review"
    assert rows[0]["reviewText"] == "Fallback"


def test_response_to_export_rows_does_not_fallback_to_top_reviews_when_paginated_reviews_empty():
    rows = response_to_export_rows(
        {
            "success": True,
            "data": {
                "amazonProduct": {
                    "asin": "B0B3JBVDYP",
                    "reviewsPaginated": {
                        "reviews": [],
                        "pageInfo": {"currentPage": 1, "totalPages": 0, "totalResults": 0},
                    },
                    "topReviews": [{"id": "TOP", "title": "Unfiltered top review"}],
                }
            },
        }
    )

    assert rows == []


def test_response_to_export_rows_exports_single_product_dict():
    rows = response_to_export_rows(
        {
            "success": True,
            "data": {
                "product": {
                    "asin": "B0B3JBVDYP",
                    "title": "Test Product",
                    "price": {"value": 29.99, "currency": "USD"},
                }
            },
        }
    )

    assert len(rows) == 1
    assert rows[0]["success"] is True
    assert rows[0]["rowSource"] == "product"
    assert rows[0]["title"] == "Test Product"


def test_response_to_export_rows_exports_single_amazon_product_dict():
    rows = response_to_export_rows(
        {
            "success": True,
            "data": {
                "amazonProduct": {
                    "asin": "B0B3JBVDYP",
                    "title": "Test Product",
                    "brand": "RICH'S BACKYARD",
                }
            },
        }
    )

    assert len(rows) == 1
    assert rows[0]["success"] is True
    assert rows[0]["rowSource"] == "amazonProduct"
    assert rows[0]["asin"] == "B0B3JBVDYP"
    assert rows[0]["title"] == "Test Product"
    assert "amazonProduct" not in rows[0]


def test_export_rows_to_xlsx_writes_headers_and_serializes_nested_values(tmp_path: Path):
    export = export_rows_to_xlsx(
        rows=[
            {
                "asin": "B0B3JBVDYP",
                "title": "Test Product",
                "price": {"value": 29.99, "currency": "USD"},
            }
        ],
        output_path=tmp_path / "canopy.xlsx",
        scenario="product",
        domain="US",
        params={"asin": "B0B3JBVDYP"},
    )

    workbook = load_workbook(export.path)
    sheet = workbook.active

    assert export.filename == "canopy.xlsx"
    assert sheet.cell(row=1, column=1).value == "ASIN"
    assert sheet.cell(row=1, column=2).value == "标题"
    assert sheet.cell(row=1, column=3).value == "价格"
    assert sheet.cell(row=2, column=1).value == "B0B3JBVDYP"
    assert sheet.cell(row=2, column=3).value == '{"value": 29.99, "currency": "USD"}'
