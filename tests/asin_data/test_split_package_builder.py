from pathlib import Path

from openpyxl import load_workbook

from opscli.asin_data.services.split_package_builder import (
    FILE_BI,
    SHEET_BASIC,
    SHEET_BULLETS,
    SHEET_CRAWLER,
    SHEET_IMAGES,
    SHEET_LISTING,
    SHEET_PRODUCT,
    SHEET_QA,
    SHEET_REVIEWS,
    build_split_package,
    write_basic_workbook,
)


def test_basic_workbook_uses_listing_basic_when_crawler_source_missing(tmp_path):
    output_path = tmp_path / "basic.xlsx"
    asin_result = {
        "asin": "B0TEST1234",
        "site": "US",
        "bi_report_data": {
            "sources": {
                "listing_basic": {
                    "rows": [
                        {
                            "\u5546\u54c1\u6807\u9898": "Listing Endpoint Title",
                            "\u54c1\u724c": "ListingBrand",
                            "\u7c7b\u76ee": "Home & Kitchen,Furniture,Bed Frames",
                            "\u4e3b\u56fe\u94fe\u63a5": "https://example.test/main.jpg",
                            "\u5176\u4ed6\u9644\u56fe\u94fe\u63a5": [
                                "https://example.test/side-1.jpg",
                                "https://example.test/side-2.jpg",
                            ],
                            "\u4e94\u70b9\u63cf\u8ff0": ["Listing bullet 1", "Listing bullet 2"],
                            "\u5173\u952e\u8bcd\u641c\u7d22": "storage bed frame",
                            "\u5382\u5546\u5efa\u8bae\u96f6\u552e\u4ef7": 299.99,
                            "\u5382\u5546\u5efa\u8bae\u96f6\u552e\u4ef7\u5e01\u79cd": "USD",
                        }
                    ]
                }
            }
        },
    }

    write_basic_workbook(output_path, asin_result)

    wb = load_workbook(output_path)
    assert SHEET_BASIC not in wb.sheetnames

    listing_rows = list(wb[SHEET_LISTING].iter_rows(min_row=2, values_only=True))
    listing_values = {field: value for _row_type, field, value in listing_rows}
    assert listing_values["\u5546\u54c1\u6807\u9898"] == "Listing Endpoint Title"
    assert listing_values["\u54c1\u724c"] == "ListingBrand"
    assert listing_values["\u7c7b\u76ee"] == "Home & Kitchen,Furniture,Bed Frames"
    assert listing_values["\u4e3b\u56fe\u94fe\u63a5"] == "https://example.test/main.jpg"
    assert "\u5173\u952e\u8bcd\u641c\u7d22" not in listing_values
    assert listing_values["\u4e94\u70b9\u63cf\u8ff0"] == "Listing bullet 1\nListing bullet 2"
    assert listing_values["\u5382\u5546\u5efa\u8bae\u96f6\u552e\u4ef7"] == 299.99
    assert listing_values["\u5382\u5546\u5efa\u8bae\u96f6\u552e\u4ef7\u5e01\u79cd"] == "USD"
    assert {SHEET_PRODUCT, SHEET_BULLETS, SHEET_IMAGES, SHEET_QA, SHEET_REVIEWS}.isdisjoint(wb.sheetnames)


def test_split_package_can_build_bi_only_without_zip(tmp_path):
    package = build_split_package(
        output_root=tmp_path,
        asin_results=[
            {
                "asin": "B0TEST1234",
                "bi_report_data": {
                    "sources": {
                        "sales_traffic": {
                            "label": "sales",
                            "rows": [{"ASIN": "B0TEST1234", "orderQty": 4}],
                        }
                    }
                },
            }
        ],
        summary={"summary": {"asin_count": 1}},
        file_keys=("bi",),
        include_zip=False,
    )

    item = package["items"][0]
    assert item["files"].keys() == {"bi"}
    assert Path(item["files"]["bi"]).name == FILE_BI
    assert Path(item["files"]["bi"]).exists()
    assert package["zip_path"] is None
    assert not (tmp_path / "B0TEST1234-asin-data-package.zip").exists()
    assert not (Path(item["dir"]) / "01-基础数据.xlsx").exists()


def test_listing_sheet_omits_input_metadata_from_frontend_basic_data(tmp_path):
    output_path = tmp_path / "basic.xlsx"
    asin_result = {
        "asin": "B0TEST1234",
        "site": "US",
        "frontend_data": {
            "\u57fa\u7840\u6570\u636e": {
                "\u641c\u7d22\u5173\u952e\u8bcd": "twin bed frame platform",
                "ASIN": "B0TEST1234",
                "\u7ad9\u70b9": "US",
                "\u8f93\u5165\u5173\u952e\u8bcd": "",
                "\u8f93\u5165\u5173\u952e\u8bcd\u5217\u8868": "",
                "\u5173\u952e\u8bcd\u6570\u91cf": 0,
                "\u5173\u952e\u8bcd\u6765\u6e90": "\u672a\u63d0\u4f9b",
                "\u8f93\u5165\u884c\u53f7": 1,
                "\u6765\u6e90\u6587\u4ef6": "--asin",
                "\u4fdd\u7559\u5b57\u6bb5": "keep me",
            }
        },
        "bi_report_data": {
            "sources": {
                "listing_basic": {
                    "rows": [
                        {
                            "ASIN": "B0TEST1234",
                            "\u7ad9\u70b9": "US",
                            "\u5173\u952e\u8bcd\u641c\u7d22": "listing endpoint terms",
                            "\u4ea7\u54c1\u6807\u9898": "Listing Title",
                        }
                    ]
                }
            }
        },
    }

    write_basic_workbook(output_path, asin_result)

    wb = load_workbook(output_path)
    listing_rows = list(wb[SHEET_LISTING].iter_rows(min_row=2, values_only=True))
    listing_values = {field: value for _row_type, field, value in listing_rows}
    omitted_fields = {
        "\u5173\u952e\u8bcd\u641c\u7d22",
        "generic_keyword.value",
        "asin",
        "\u8f93\u5165\u5173\u952e\u8bcd",
        "\u8f93\u5165\u5173\u952e\u8bcd\u5217\u8868",
        "\u5173\u952e\u8bcd\u6570\u91cf",
        "\u5173\u952e\u8bcd\u6765\u6e90",
        "\u8f93\u5165\u884c\u53f7",
        "\u6765\u6e90\u6587\u4ef6",
    }

    assert omitted_fields.isdisjoint(listing_values)
    assert listing_values["\u641c\u7d22\u5173\u952e\u8bcd"] == "twin bed frame platform"
    assert listing_values["ASIN"] == "B0TEST1234"
    assert listing_values["\u7ad9\u70b9"] == "US"
    assert listing_values["\u4ea7\u54c1\u6807\u9898"] == "Listing Title"
    assert listing_values["\u4fdd\u7559\u5b57\u6bb5"] == "keep me"


def test_listing_sheet_omits_redundant_normalized_fields_when_template_fields_exist(tmp_path):
    output_path = tmp_path / "basic.xlsx"
    asin_result = {
        "asin": "B0TEST1234",
        "site": "US",
        "bi_report_data": {
            "sources": {
                "listing_basic": {
                    "rows": [
                        {
                            "\u5546\u54c1\u6807\u9898": "Normalized Title",
                            "\u4ea7\u54c1\u6807\u9898": "Template Title",
                            "\u54c1\u724c": "NormalizedBrand",
                            "\u54c1\u724c\u540d": "TemplateBrand",
                            "\u7c7b\u76ee": "Home & Kitchen,Furniture,Bed Frames",
                            "\u4e3b\u56fe\u94fe\u63a5": "https://example.test/main-normalized.jpg",
                            "\u4e3b\u56fe": "https://example.test/main-template.jpg",
                            "\u5176\u4ed6\u9644\u56fe\u94fe\u63a5": [
                                "https://example.test/side-normalized-1.jpg",
                                "https://example.test/side-normalized-2.jpg",
                            ],
                            "\u526f\u56fe1": "https://example.test/side-template-1.jpg",
                            "\u4e94\u70b9\u63cf\u8ff0": ["Listing bullet 1", "Listing bullet 2"],
                            "\u5e97\u94fa/\u90e8\u95e8": "Internal Dept",
                            "\u8d1f\u8d23\u4eba": "Owner",
                            "listid": 123,
                        }
                    ]
                }
            }
        },
    }

    write_basic_workbook(output_path, asin_result)

    wb = load_workbook(output_path)
    listing_rows = list(wb[SHEET_LISTING].iter_rows(min_row=2, values_only=True))
    listing_values = {field: value for _row_type, field, value in listing_rows}
    omitted_fields = {
        "\u5546\u54c1\u6807\u9898",
        "\u54c1\u724c",
        "\u4e3b\u56fe\u94fe\u63a5",
        "\u5176\u4ed6\u9644\u56fe\u94fe\u63a5",
        "\u5e97\u94fa/\u90e8\u95e8",
        "\u8d1f\u8d23\u4eba",
        "listid",
    }

    assert omitted_fields.isdisjoint(listing_values)
    assert listing_values["\u4ea7\u54c1\u6807\u9898"] == "Template Title"
    assert listing_values["\u54c1\u724c\u540d"] == "TemplateBrand"
    assert listing_values["\u7c7b\u76ee"] == "Home & Kitchen,Furniture,Bed Frames"
    assert listing_values["\u4e3b\u56fe"] == "https://example.test/main-template.jpg"
    assert listing_values["\u526f\u56fe1"] == "https://example.test/side-template-1.jpg"
    assert listing_values["\u4e94\u70b9\u63cf\u8ff0"] == "Listing bullet 1\nListing bullet 2"
    assert {SHEET_PRODUCT, SHEET_BULLETS, SHEET_IMAGES, SHEET_QA, SHEET_REVIEWS}.isdisjoint(wb.sheetnames)


def test_basic_workbook_prefers_listing_content_and_hides_conflicting_crawler_fields(tmp_path):
    output_path = tmp_path / "basic.xlsx"
    asin_result = {
        "asin": "B0TEST1234",
        "site": "US",
        "bi_report_data": {
            "sources": {
                "listing_basic": {
                    "rows": [
                        {
                            "\u5546\u54c1\u6807\u9898": "Listing Endpoint Title",
                            "\u4e3b\u56fe\u94fe\u63a5": "https://example.test/listing-main.jpg",
                            "\u5176\u4ed6\u9644\u56fe\u94fe\u63a5": [
                                "https://example.test/listing-side-1.jpg",
                                "https://example.test/listing-side-2.jpg",
                            ],
                            "\u4e94\u70b9\u63cf\u8ff0": ["Listing bullet 1", "Listing bullet 2"],
                        }
                    ]
                },
                "crawler_details": {
                    "rows": [
                        {
                            "ASIN": "B0TEST1234",
                            "\u5546\u54c1\u6807\u9898": "Crawler Title",
                            "listing": "Crawler Listing Title",
                            "f_listing": "Crawler f_listing Title",
                            "title": "Crawler Raw Title",
                            "\u4e94\u70b9\u63cf\u8ff0": ["Crawler bullet"],
                            "five_point_description": ["Crawler bullet raw"],
                            "f_five_point_description": ["Crawler bullet f"],
                            "\u5546\u54c1\u63cf\u8ff0": "Crawler Description",
                            "description": "Crawler raw description",
                            "f_description": "Crawler f description",
                            "a_description": "Crawler A+ description",
                            "\u4e3b\u56fe": "https://example.test/crawler-main-cn.jpg",
                            "image": "https://example.test/crawler-main.jpg",
                            "f_image": "https://example.test/crawler-main-f.jpg",
                            "\u526f\u56fe": ["https://example.test/crawler-side-cn.jpg"],
                            "subplot": ["https://example.test/crawler-side.jpg"],
                            "f_subplot": ["https://example.test/crawler-side-f.jpg"],
                            "A+\u56fe\u7247": ["https://example.test/crawler-aplus-cn.jpg"],
                            "a_image": ["https://example.test/crawler-aplus.jpg"],
                            "f_a_image": ["https://example.test/crawler-aplus-f.jpg"],
                            "product_details": {
                                "Brand": "CrawlerBrand",
                                "Description": "Crawler product detail description",
                                "Images": ["https://example.test/product-detail-image.jpg"],
                            },
                            "qa": [{"question": "Q", "answer": "A"}],
                            "review_list": [{"rating": 5, "content": "Good"}],
                            "keep_crawler_metric": 7,
                        }
                    ]
                },
            }
        },
    }

    write_basic_workbook(output_path, asin_result)

    wb = load_workbook(output_path)
    crawler_headers = [cell.value for cell in next(wb[SHEET_CRAWLER].iter_rows(max_row=1))]
    conflicting_fields = {
        "\u5546\u54c1\u6807\u9898",
        "listing",
        "f_listing",
        "title",
        "\u4e94\u70b9\u63cf\u8ff0",
        "five_point_description",
        "f_five_point_description",
        "\u5546\u54c1\u63cf\u8ff0",
        "description",
        "f_description",
        "\u4e3b\u56fe",
        "image",
        "f_image",
        "\u526f\u56fe",
        "subplot",
        "f_subplot",
        "product_details",
        "f_product_details",
    }
    assert conflicting_fields.isdisjoint(crawler_headers)
    assert "a_description" in crawler_headers
    assert "A+\u56fe\u7247" in crawler_headers
    assert "a_image" in crawler_headers
    assert "f_a_image" in crawler_headers
    assert "qa" in crawler_headers
    assert "review_list" in crawler_headers
    assert "keep_crawler_metric" in crawler_headers

    listing_rows = list(wb[SHEET_LISTING].iter_rows(min_row=2, values_only=True))
    listing_values = {field: value for _row_type, field, value in listing_rows}
    assert listing_values["\u4e94\u70b9\u63cf\u8ff0"] == "Listing bullet 1\nListing bullet 2"
    assert {SHEET_PRODUCT, SHEET_BULLETS, SHEET_IMAGES, SHEET_QA, SHEET_REVIEWS}.isdisjoint(wb.sheetnames)


def test_basic_workbook_keeps_listing_bullets_and_crawler_aplus_without_detail_sheets(tmp_path):
    output_path = tmp_path / "basic.xlsx"
    asin_result = {
        "asin": "B0TEST1234",
        "site": "US",
        "bi_report_data": {
            "sources": {
                "listing_basic": {
                    "rows": [
                        {
                            "ASIN": "B0TEST1234",
                            "\u4e94\u70b9\u63cf\u8ff0": ["Listing bullet 1", "Listing bullet 2"],
                            "title_differentiation": "Listing title differentiation",
                        }
                    ]
                },
                "crawler_details": {
                    "rows": [
                        {
                            "ASIN": "B0TEST1234",
                            "a_description": "Crawler A+ description",
                            "a_image": ["https://example.test/crawler-aplus.jpg"],
                            "qa": [{"question": "Q", "answer": "A"}],
                            "review_list": [{"rating": 5, "content": "Good"}],
                            "product_details": {"Brand": "CrawlerBrand"},
                            "image": "https://example.test/crawler-main.jpg",
                            "five_point_description": ["Crawler bullet"],
                        }
                    ]
                },
            }
        },
    }

    write_basic_workbook(output_path, asin_result)

    wb = load_workbook(output_path)
    removed_sheets = {SHEET_PRODUCT, SHEET_BULLETS, SHEET_IMAGES, SHEET_QA, SHEET_REVIEWS}
    assert removed_sheets.isdisjoint(wb.sheetnames)
    assert wb.sheetnames == [SHEET_LISTING, SHEET_CRAWLER]

    listing_rows = list(wb[SHEET_LISTING].iter_rows(min_row=2, values_only=True))
    listing_values = {field: value for _row_type, field, value in listing_rows}
    assert listing_values["\u4e94\u70b9\u63cf\u8ff0"] == "Listing bullet 1\nListing bullet 2"
    assert listing_values["title_differentiation"] == "Listing title differentiation"

    crawler_headers = [cell.value for cell in next(wb[SHEET_CRAWLER].iter_rows(max_row=1))]
    assert "a_description" in crawler_headers
    assert "a_image" in crawler_headers
    assert "qa" in crawler_headers
    assert "review_list" in crawler_headers
    assert "product_details" not in crawler_headers
    assert "image" not in crawler_headers
    assert "five_point_description" not in crawler_headers
