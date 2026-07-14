import json
from pathlib import Path
from types import SimpleNamespace

import httpx
from openpyxl import load_workbook

from opscli.asin_data.services import category_top as category_top_module
from opscli.asin_data.services.category_top import (
    AsinCategoryTopClient,
    AsinCategoryTopService,
)


class DummyAuthClient:
    def build_request_auth(self, alias):
        return {"Authorization": f"Bearer {alias}-jwt"}, {"sid": "demo"}


def test_category_top_client_fetches_remote_rows():
    calls = {}

    def fake_get(url, **kwargs):
        calls["url"] = url
        calls["kwargs"] = kwargs
        return httpx.Response(
            200,
            json={
                "code": 200,
                "data": [
                    {
                        "排名": 1,
                        "ASIN": "B0TEST1234",
                        "平台类目": "Bed Frames",
                    }
                ],
            },
        )

    client = AsinCategoryTopClient(
        auth_client=DummyAuthClient(),
        http_get=fake_get,
        ops_url="https://ops.example.com/",
    )

    result = client.fetch(
        category="Bed Frames",
        date_from="2026-07-01",
        date_to="2026-07-13",
        limit=10,
    )

    assert calls["url"] == "https://ops.example.com/dataMetrics/v1/asin-report-files/internal-category-top10"
    assert calls["kwargs"]["params"] == {
        "category": "Bed Frames",
        "date_from": "2026-07-01",
        "date_to": "2026-07-13",
        "limit": 10,
    }
    assert calls["kwargs"]["headers"]["Authorization"] == "Bearer ops-jwt"
    assert result["row_count"] == 1
    assert result["rows"][0]["ASIN"] == "B0TEST1234"


def test_category_top_service_writes_and_uploads_xlsx_artifact(tmp_path):
    calls = {"source_keys": [], "uploads": [], "listing_account_type_by_source": {}}

    class DummyTopClient:
        def fetch(self, **kwargs):
            calls["top_kwargs"] = kwargs
            return {
                "status": "success",
                "endpoint": "/top",
                "params": kwargs,
                "row_count": 2,
                "rows": [
                    {
                        "排名": 1,
                        "ASIN": "B0TEST1234",
                        "销量": 20,
                        "渠道": "AochuangVC-US",
                    },
                    {
                        "排名": 2,
                        "ASIN": "B0TEST5678",
                        "销量": 10,
                        "渠道": "Amazon-CA",
                    },
                ],
            }

    class DummyBiClient:
        def fetch(self, **kwargs):
            source_key = kwargs["source_keys"][0]
            calls["source_keys"].append(source_key)
            calls["listing_account_type_by_source"][source_key] = kwargs.get("listing_account_type_by_asin")
            rows = [
                {"ASIN": "B0TEST1234", "source": source_key, "value": "A"},
                {"ASIN": "B0TEST5678", "source": source_key, "value": "B"},
            ]
            return {
                "status": "success",
                "asins": kwargs["asins"],
                "count": len(kwargs["asins"]),
                "sources": {
                    source_key: {
                        "key": source_key,
                        "label": source_key,
                        "endpoint": f"/{source_key}",
                        "status": "success",
                        "row_count": len(rows),
                        "rows": rows,
                    }
                },
            }

    class DummyUploadClient:
        def upload(self, path, **kwargs):
            calls["uploads"].append({"path": Path(path), **kwargs})
            return SimpleNamespace(
                url="https://example.oss/asin-data/internal-category-top-asin-data.xlsx",
                raw={"code": 200},
            )

    service = AsinCategoryTopService(
        top_client=DummyTopClient(),
        bi_report_data_client_factory=DummyBiClient,
        file_upload_client_factory=DummyUploadClient,
    )

    result = service.run(
        category="Bed Frames",
        date_from="2026-07-01",
        date_to="2026-07-13",
        limit=2,
        output_dir=str(tmp_path),
        run_id="run-1",
        upload=True,
        return_content=False,
    )

    assert result["success"] is True
    assert result["metadata"]["protocol"] == "asin_data_ai_response"
    assert result["metadata"]["tool"] == "asin_data_category_top"
    assert result["metadata"]["data_scope"] == "internal_category_top"
    assert result["run"]["run_id"] == "run-1"
    assert result["summary"]["asin_count"] == 2
    assert result["summary"]["file_url"] == "https://example.oss/asin-data/internal-category-top-asin-data.xlsx"
    artifact = result["items"][0]["artifacts"][0]
    assert artifact["file_key"] == "category_top_xlsx"
    assert artifact["type"] == "xlsx"
    assert artifact["report_filename"] == "internal-category-top-asin-data.xlsx"
    assert artifact["uri"] == "https://example.oss/asin-data/internal-category-top-asin-data.xlsx"
    assert result["items"][0]["datasets"][0]["source_key"] == "category_top"
    assert result["items"][0]["datasets"][1]["source_key"] == "listing_basic"
    assert result["items"][0]["datasets"][2]["source_key"] == "crawler_details"
    assert "rows" not in result["items"][0]["datasets"][0]
    assert sorted(calls["source_keys"]) == ["crawler_details", "listing_basic"]
    assert calls["listing_account_type_by_source"]["listing_basic"] == {"B0TEST1234": 2}
    assert calls["listing_account_type_by_source"]["crawler_details"] is None
    assert calls["uploads"][0]["purpose"] == "asin_data_category_top_xlsx"
    assert calls["uploads"][0]["path"].suffix == ".xlsx"
    assert calls["uploads"][0]["folder"] == "asin-data"
    assert calls["uploads"][0]["metadata"]["category"] == "Bed Frames"

    workbook_path = Path(artifact["local_path"])
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["类目top10", "刊登数据", "爬虫数据"]
    finally:
        workbook.close()

    document_path = workbook_path.with_name("internal-category-top-asin-data.json")
    document = json.loads(document_path.read_text(encoding="utf-8"))
    assert document["metadata"]["protocol"] == "asin_data_ai_response"
    assert document["summary"]["status"] == "success"
    assert document["items"][0]["asin"] == "B0TEST1234"
    assert document["items"][0]["site"] == "US"
    assert document["items"][1]["site"] == "CA"
    assert document["items"][0]["artifacts"][0]["uri"] == artifact["uri"]
    datasets = {dataset["source_key"]: dataset for dataset in document["items"][0]["datasets"]}
    assert datasets["category_top"]["rows"][0]["ASIN"] == "B0TEST1234"
    assert datasets["listing_basic"]["rows"][0]["source"] == "listing_basic"
    assert datasets["crawler_details"]["rows"][0]["source"] == "crawler_details"


def test_category_top_service_no_upload_still_writes_local_xlsx(tmp_path):
    class DummyTopClient:
        def fetch(self, **kwargs):
            return {
                "status": "success",
                "rows": [{"排名": 1, "ASIN": "B0TEST1234", "渠道": "AochuangVC-US"}],
            }

    service = AsinCategoryTopService(top_client=DummyTopClient())

    result = service.run(
        category="Bed Frames",
        limit=1,
        output_dir=str(tmp_path),
        run_id="no-upload",
        upload=False,
        enrich=False,
        return_content=False,
    )

    artifact = result["items"][0]["artifacts"][0]
    workbook_path = Path(artifact["local_path"])
    assert result["summary"]["file_url"] is None
    assert artifact["type"] == "xlsx"
    assert artifact["uri"] is None
    assert workbook_path.name == "internal-category-top-asin-data.xlsx"
    assert workbook_path.exists()

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["类目top10", "刊登数据", "爬虫数据"]
    finally:
        workbook.close()


def test_category_top_maps_chinese_channel_country_to_listing_site_code():
    rows = [
        {
            "ASIN": "B086M58PQ3",
            "channel": "傲创VC-美国",
        }
    ]

    assert category_top_module._site_by_asin(rows, default_site="US") == {"B086M58PQ3": "US"}
