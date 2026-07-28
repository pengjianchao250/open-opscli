from openpyxl import load_workbook

from opscli.asin_data.services.category_top_workbook import (
    SHEET_CATEGORY_TOP,
    SHEET_CRAWLER,
    SHEET_LISTING,
    write_category_top_workbook,
)


def _dataset(source_key, rows):
    return {"source_key": source_key, "rows": rows}


def test_category_top_workbook_writes_fixed_sheets_and_field_ownership(tmp_path):
    output_path = tmp_path / "category-top.xlsx"
    document = {
        "items": [
            {
                "asin": "B0TEST1234",
                "datasets": [
                    _dataset(
                        "category_top",
                        [{"排名": 1, "ASIN": "B0TEST1234", "渠道": "AochuangVC-US"}],
                    ),
                    _dataset(
                        "listing_basic",
                        [
                            {
                                "ASIN": "B0TEST1234",
                                "商品标题": "Listing title 1",
                                "商品亮点": "Quiet metal platform",
                            }
                        ],
                    ),
                    _dataset(
                        "crawler_details",
                        [
                            {
                                "ASIN": "B0TEST1234",
                                "listing": "Crawler title",
                                "five_point_description": ["Crawler bullet"],
                                "description": "Crawler description",
                                "image": "https://example.test/main.jpg",
                                "subplot": ["https://example.test/side.jpg"],
                                "product_details": {"Brand": "CrawlerBrand"},
                                "a_image": ["https://example.test/aplus.jpg"],
                                "a_description": "A+ description",
                                "qa": [{"question": "Q", "answer": "A"}],
                                "review_list": [{"rating": 5, "content": "Good"}],
                                "rating": 4.8,
                            }
                        ],
                    ),
                ],
            },
            {
                "asin": "B0TEST5678",
                "datasets": [
                    _dataset(
                        "category_top",
                        [{"排名": 2, "ASIN": "B0TEST5678", "渠道": "AochuangVC-US"}],
                    ),
                    _dataset(
                        "listing_basic",
                        [{"ASIN": "B0TEST5678", "商品标题": "Listing title 2"}],
                    ),
                    _dataset(
                        "crawler_details",
                        [
                            {
                                "ASIN": "B0TEST5678",
                                "a_image": [],
                                "a_description": "",
                                "qa": [],
                                "review_list": [],
                            }
                        ],
                    ),
                ],
            },
        ]
    }

    write_category_top_workbook(output_path, document)

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == [SHEET_CATEGORY_TOP, SHEET_LISTING, SHEET_CRAWLER]
        assert workbook.sheetnames == ["类目top10", "刊登数据", "爬虫数据"]
        assert all(workbook[name].max_row == 3 for name in workbook.sheetnames)

        top_rows = list(workbook[SHEET_CATEGORY_TOP].iter_rows(values_only=True))
        top_headers = list(top_rows[0])
        top_asin_index = top_headers.index("ASIN")
        assert [row[top_asin_index] for row in top_rows[1:]] == ["B0TEST1234", "B0TEST5678"]

        listing_rows = list(workbook[SHEET_LISTING].iter_rows(values_only=True))
        listing_headers = list(listing_rows[0])
        highlight_index = listing_headers.index("商品亮点")
        assert listing_rows[1][highlight_index] == "Quiet metal platform"
        assert listing_rows[2][highlight_index] is None

        crawler_headers = set(next(workbook[SHEET_CRAWLER].iter_rows(values_only=True)))
        conflicts = {
            "listing",
            "five_point_description",
            "description",
            "image",
            "subplot",
            "product_details",
        }
        assert conflicts.isdisjoint(crawler_headers)
        assert {"a_image", "a_description", "qa", "review_list", "rating"}.issubset(
            crawler_headers
        )
    finally:
        workbook.close()


def test_category_top_workbook_keeps_one_row_when_enrichment_is_missing(tmp_path):
    output_path = tmp_path / "category-top-missing.xlsx"
    document = {
        "items": [
            {
                "asin": "B0TEST1234",
                "datasets": [
                    _dataset("category_top", [{"排名": 1, "ASIN": "B0TEST1234"}]),
                    _dataset("listing_basic", []),
                    _dataset("crawler_details", []),
                ],
            }
        ]
    }

    write_category_top_workbook(output_path, document)

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        assert workbook[SHEET_LISTING].max_row == 2
        assert workbook[SHEET_CRAWLER].max_row == 2
        listing_rows = list(workbook[SHEET_LISTING].iter_rows(values_only=True))
        listing_headers = list(listing_rows[0])
        assert listing_rows[1][listing_headers.index("ASIN")] == "B0TEST1234"
        assert "商品亮点" in listing_headers
    finally:
        workbook.close()
