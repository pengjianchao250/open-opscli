import json
from zipfile import ZipFile
from types import SimpleNamespace
from pathlib import Path

import pytest
from openpyxl import load_workbook

from opscli.asin_data.services.collector import AsinDataCollector, DirectOpsRunner, load_legacy_collector
from opscli.asin_data.services.merged_report_renderer import render_merged_report_text
from opscli.asin_data.services.report_file_submitter import AsinReportFileSubmitter
from opscli.asin_data.services.report_files import AsinReportFileNotFoundError


class DummyUploadClient:
    def __init__(self):
        self.called_with = None
        self.calls = []

    def upload(self, path, **kwargs):
        upload_path = Path(path)
        self.called_with = {"path": upload_path, **kwargs}
        self.calls.append(self.called_with)
        url = f"https://p-amazon-task-test.oss-cn-hongkong.aliyuncs.com/asin-data/2026/06/1781083839-{upload_path.name}"
        return SimpleNamespace(
            url=url,
            raw={"code": 200, "data": {"url": url}},
        )


class DummyReportFileClient:
    def __init__(self, url: str = "https://example.oss.aliyuncs.com/asin-data/report-file.txt"):
        self.url = url
        self.called_with = []

    def fetch(self, *, asin: str, site: str):
        self.called_with.append({"asin": asin, "site": site})
        return SimpleNamespace(url=self.url, record={"file_url": self.url, "asin": asin, "site": site})


class MissingReportFileClient(DummyReportFileClient):
    def fetch(self, *, asin: str, site: str):
        self.called_with.append({"asin": asin, "site": site})
        return SimpleNamespace(url=None, record={"asin": asin, "site": site})


class DummyBiReportDataClient:
    def __init__(self):
        self.called_with = None
        self.calls = []
        self.date_calls = []
        self.source_key_calls = []
        self.site_calls = []

    def fetch(
        self,
        *,
        asins,
        start_date=None,
        end_date=None,
        source_keys=None,
        site_by_asin=None,
        default_site="US",
    ):
        normalized_asins = [str(asin).upper() for asin in asins]
        normalized_source_keys = list(source_keys) if source_keys is not None else None
        self.called_with = normalized_asins
        self.calls.append(normalized_asins)
        self.date_calls.append({"start_date": start_date, "end_date": end_date})
        self.source_key_calls.append(normalized_source_keys)
        self.site_calls.append({"site_by_asin": dict(site_by_asin or {}), "default_site": default_site})
        sources = {
            "status": "success",
            "asins": normalized_asins,
            "count": len(normalized_asins),
            "sources": {
                "sales_traffic": {
                    "key": "sales_traffic",
                    "label": "销售/库存/广告/流量数据",
                    "endpoint": "/dataMetrics/v1/asin-report-files/sales-traffic-data",
                    "status": "success",
                    "row_count": len(normalized_asins),
                    "rows": [
                        {
                            "asin": asin,
                            "productName": "Test Product",
                            "orderQty": 4,
                            "salesAmount": 99.5,
                        }
                        for asin in normalized_asins
                    ],
                    "raw": {"code": 0},
                },
                "listing_basic": {
                    "key": "listing_basic",
                    "label": "刊登基础数据",
                    "endpoint": "https://bi.api.xenkee.com/listing/amazonlisdet",
                    "status": "success",
                    "row_count": len(normalized_asins),
                    "rows": [
                        {
                            "asin": asin,
                            "ASIN": asin,
                            "平台SKU": "TEST-SKU",
                            "商品标题": "Listing Endpoint Title",
                            "品牌": "ListingBrand",
                            "关键词搜索": "storage bed frame search terms",
                            "generic_keyword.value": "storage bed frame search terms",
                            "listid": 3526099,
                        }
                        for asin in normalized_asins
                    ],
                    "raw": {"code": 200},
                },
                "sp_search_term": {
                    "key": "sp_search_term",
                    "label": "SP广告搜索词数据",
                    "endpoint": "/api/v1/sp-search-term/query",
                    "status": "success",
                    "row_count": len(normalized_asins),
                    "rows": [{"asin": asin, "searchTerm": "bed frame", "clicks": 9} for asin in normalized_asins],
                    "raw": {"code": 0},
                },
                "deals": {
                    "key": "deals",
                    "label": "活动数据",
                    "endpoint": "/dataMetrics/v1/asin-report-files/deals-data",
                    "status": "success",
                    "row_count": len(normalized_asins),
                    "rows": [{"asin": asin, "dealName": "Prime Day"} for asin in normalized_asins],
                    "raw": {"code": 0},
                },
                "turnover_inventory": {
                    "key": "turnover_inventory",
                    "label": "物控版库存数据",
                    "endpoint": "/dataMetrics/v1/asin-report-files/turnover-inventory-data",
                    "status": "success",
                    "row_count": len(normalized_asins),
                    "rows": [{"asin": asin, "availableInventory": 12} for asin in normalized_asins],
                    "raw": {"code": 0},
                },
                "crawler_details": {
                    "key": "crawler_details",
                    "label": "爬虫ASIN详情数据",
                    "endpoint": "/dataMetrics/v1/asin-report-files/crawler-details",
                    "status": "success",
                    "row_count": len(normalized_asins),
                    "rows": [
                        {
                            "asin": asin,
                            "商品标题": "Crawler Endpoint Title",
                            "品牌": "EndpointBrand",
                            "五点描述": ["Endpoint bullet"],
                        }
                        for asin in normalized_asins
                    ],
                    "raw": {"code": 0},
                },
            },
        }
        if normalized_source_keys is not None:
            sources["sources"] = {
                key: source for key, source in sources["sources"].items() if key in normalized_source_keys
            }
        return sources


def test_collect_single_asin_without_input_file(tmp_path: Path):
    result = AsinDataCollector().collect(
        asin="B0TEST1234",
        site="US",
        keywords=["bed frame", "storage bed"],
        output_dir=str(tmp_path),
        run_id="single-asin",
        skip_query=True,
        skip_seller_sprite=True,
        skip_amazon=True,
        skip_rufus=True,
    )

    output_dir = Path(result["output_dir"])
    manifest = json.loads((output_dir / "manifest.json").read_text(encoding="utf-8"))
    frontend = json.loads((output_dir / "frontend-data.json").read_text(encoding="utf-8"))

    assert result["summary"]["asin_count"] == 1
    assert manifest["summary"]["input_count"] == 1
    assert manifest["options"]["skip_query"] is True
    record = frontend[next(key for key, value in frontend.items() if isinstance(value, list))][0]
    base_section = next(value for value in record.values() if isinstance(value, dict) and value.get("ASIN"))
    assert base_section["ASIN"] == "B0TEST1234"


def test_collect_uploads_split_package_zip_and_returns_aliyun_url(tmp_path: Path):
    upload_client = DummyUploadClient()

    result = AsinDataCollector(file_upload_client=upload_client).collect(
        asin="B0TEST1234",
        site="US",
        keywords=["bed frame"],
        output_dir=str(tmp_path),
        run_id="single-asin-upload",
        skip_query=True,
        skip_seller_sprite=True,
        skip_amazon=True,
        skip_rufus=True,
        skip_bi_report_data=True,
        upload=True,
    )

    output_dir = Path(result["output_dir"])
    manifest = json.loads((output_dir / "manifest.json").read_text(encoding="utf-8"))

    report_path = output_dir / "B0TEST1234-asin-data-report.txt"
    package_path = output_dir / "B0TEST1234-asin-data-package.zip"
    package_url = "https://p-amazon-task-test.oss-cn-hongkong.aliyuncs.com/asin-data/2026/06/1781083839-B0TEST1234-asin-data-package.zip"

    assert result["aliyun_url"] == package_url
    assert result["upload"]["url"] == package_url
    assert result["upload"]["path"] == package_path.as_posix()
    assert result["upload"]["upload_path"] == package_path.as_posix()
    assert manifest["files"]["frontend_data"] == (output_dir / "frontend-data.json").as_posix()
    assert manifest["files"]["frontend_html"] == (output_dir / "frontend-data.html").as_posix()
    assert manifest["files"]["asin_report_txt"] == report_path.as_posix()
    assert manifest["files"]["asin_data_package_zip"] == package_path.as_posix()
    assert manifest["files"]["asin_data_package_upload_url"] == package_url
    assert manifest["files"]["asin_report_upload_url"] == package_url
    assert "frontend_data_url" not in manifest["files"]
    assert "frontend_json_url" not in manifest["files"]
    assert "frontend_html_url" not in manifest["files"]
    package_upload = next(call for call in upload_client.calls if call["path"] == package_path)
    with ZipFile(package_path) as archive:
        names = archive.namelist()
    assert "B0TEST1234/01-基础数据.xlsx" in names
    assert "B0TEST1234/02-BI数据.xlsx" in names
    assert "B0TEST1234/06-Rufus数据分析.md" in names
    assert not any(name.lower().endswith(".json") for name in names)
    frontend_html = (output_dir / "frontend-data.html").read_text(encoding="utf-8-sig")
    assert "<!doctype html>" in frontend_html
    assert "B0TEST1234" in frontend_html
    report_bytes = report_path.read_bytes()
    assert report_bytes.startswith(b"\xef\xbb\xbf")
    assert not report_bytes.startswith(b"\xef\xbb\xbf\xef\xbb\xbf")
    report_text = report_path.read_text(encoding="utf-8-sig")
    assert report_text.startswith("# ASIN 取数汇总报告 - B0TEST1234")
    assert "## 运行信息" not in report_text
    for heading in (
        "## BI 数据",
        "## 爬虫基础数据",
        "### 五点描述",
        "## 卖家精灵关键词挖掘",
        "## 卖家精灵 AI 全景分析",
        "## Rufus 数据",
        "## 数据链接",
    ):
        assert heading in report_text
    assert package_upload["purpose"] == "asin_data_split_package_zip"
    assert package_upload["folder"] == "asin-data"
    assert package_upload["public"] == "1"


def test_collect_fetches_bi_report_data_and_merges_single_asin_report(tmp_path: Path):
    bi_client = DummyBiReportDataClient()

    result = AsinDataCollector(bi_report_data_client=bi_client).collect(
        asin="B0TEST1234",
        site="US",
        keywords=["bed frame"],
        output_dir=str(tmp_path),
        run_id="single-asin-bi-report-data",
        skip_query=False,
        skip_sales_query=True,
        skip_crawler_query=True,
        skip_seller_sprite=True,
        skip_amazon=True,
        skip_rufus=True,
        upload=False,
        fetch_report_files=False,
    )

    output_dir = Path(result["output_dir"])
    manifest = json.loads((output_dir / "manifest.json").read_text(encoding="utf-8"))
    frontend = json.loads((output_dir / "frontend-data.json").read_text(encoding="utf-8"))
    jsonl_record = json.loads((output_dir / "asin-data.jsonl").read_text(encoding="utf-8").splitlines()[0])
    report_text = (output_dir / "B0TEST1234-asin-data-report.txt").read_text(encoding="utf-8-sig")

    assert bi_client.called_with == ["B0TEST1234"]
    assert manifest["options"]["skip_bi_report_data"] is False
    assert manifest["bi_report_data"]["status"] == "success"
    assert manifest["bi_report_data"]["sources"]["listing_basic"]["row_count"] == 1
    assert manifest["bi_report_data"]["sources"]["sales_traffic"]["row_count"] == 1
    assert jsonl_record["bi_report_data"]["sources"]["sales_traffic"]["rows"][0]["salesAmount"] == 99.5
    assert jsonl_record["bi_report_data"]["sources"]["listing_basic"]["rows"][0]["generic_keyword.value"] == "storage bed frame search terms"
    basic_workbook = output_dir / "asin-data-packages" / "B0TEST1234" / "01-基础数据.xlsx"
    wb = load_workbook(basic_workbook, read_only=True, data_only=True)
    try:
        assert "基础汇总" not in wb.sheetnames
        listing_values = {
            row[1]: row[2]
            for row in wb["刊登数据"].iter_rows(min_row=2, values_only=True)
            if row and len(row) >= 3 and row[1]
        }
    finally:
        wb.close()
    assert listing_values["商品标题"] == "Listing Endpoint Title"
    assert listing_values["品牌"] == "ListingBrand"
    assert "### 销售/库存/广告/流量数据" in report_text
    assert "## 刊登基础数据" in report_text
    assert "- 关键词搜索: storage bed frame search terms" in report_text
    assert "### SP广告搜索词数据" in report_text
    assert "### 活动数据" in report_text
    assert "### 物控版库存数据" in report_text
    assert report_text.count("#### 数据表") >= 4
    assert "| asin | productName | orderQty | salesAmount |" in report_text
    assert "| B0TEST1234 | Test Product | 4 | 99.5 |" in report_text
    assert "| asin | searchTerm | clicks |" in report_text
    assert "| B0TEST1234 | bed frame | 9 |" in report_text
    assert "#### 完整 JSON（压缩格式）" in report_text
    assert "- 销售额: 99.5" in report_text
    assert "- 标题: Crawler Endpoint Title" in report_text
    assert "- 品牌: EndpointBrand" in report_text
    assert "/dataMetrics/v1/asin-report-files" not in report_text
    assert "- 接口:" not in report_text

    record = frontend[next(key for key, value in frontend.items() if isinstance(value, list))][0]
    base_section = next(value for value in record.values() if isinstance(value, dict) and value.get("ASIN"))
    assert base_section["BI接口数据"]["状态"] == "success"
    assert base_section["BI接口数据"]["数据源"]["销售/库存/广告/流量数据"]["行数"] == 1


def test_collect_passes_sales_date_range_to_bi_report_data_client(tmp_path: Path):
    bi_client = DummyBiReportDataClient()

    AsinDataCollector(bi_report_data_client=bi_client).collect(
        asin="B0TEST1234",
        site="US",
        output_dir=str(tmp_path),
        run_id="single-asin-bi-report-date-range",
        sales_start="2026-07-01",
        sales_end="2026-07-08",
        skip_query=True,
        skip_seller_sprite=True,
        skip_amazon=True,
        skip_rufus=True,
        upload=False,
        fetch_report_files=False,
    )

    assert bi_client.date_calls == [{"start_date": "2026-07-01", "end_date": "2026-07-08"}]
    assert bi_client.site_calls == [{"site_by_asin": {"B0TEST1234": "US"}, "default_site": "US"}]


def test_collect_fetches_bi_report_data_per_asin_for_input_batch(tmp_path: Path):
    input_path = tmp_path / "asins.csv"
    input_path.write_text("asin,site,keyword\nB0TEST1234,US,bed frame\nB0TEST5678,US,desk\n", encoding="utf-8")
    bi_client = DummyBiReportDataClient()

    result = AsinDataCollector(bi_report_data_client=bi_client).collect(
        input=str(input_path),
        output_dir=str(tmp_path),
        run_id="batch-asin-bi-report-data",
        skip_query=False,
        skip_sales_query=True,
        skip_crawler_query=True,
        skip_seller_sprite=True,
        skip_amazon=True,
        skip_rufus=True,
        upload=False,
        fetch_report_files=False,
    )

    output_dir = Path(result["output_dir"])
    manifest = json.loads((output_dir / "manifest.json").read_text(encoding="utf-8"))
    jsonl_records = [
        json.loads(line)
        for line in (output_dir / "asin-data.jsonl").read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]

    assert bi_client.calls == [["B0TEST1234"], ["B0TEST5678"]]
    assert bi_client.site_calls == [
        {"site_by_asin": {"B0TEST1234": "US", "B0TEST5678": "US"}, "default_site": "US"},
        {"site_by_asin": {"B0TEST1234": "US", "B0TEST5678": "US"}, "default_site": "US"},
    ]
    assert manifest["bi_report_data"]["request_mode"] == "per_asin"
    assert manifest["bi_report_data"]["per_asin"]["B0TEST1234"]["sources"]["sales_traffic"]["row_count"] == 1
    assert manifest["bi_report_data"]["per_asin"]["B0TEST5678"]["sources"]["sales_traffic"]["row_count"] == 1
    assert [record["bi_report_data"]["sources"]["sales_traffic"]["rows"][0]["asin"] for record in jsonl_records] == [
        "B0TEST1234",
        "B0TEST5678",
    ]


def test_collect_batches_bi_only_report_data_for_input_batch(tmp_path: Path):
    input_path = tmp_path / "asins.csv"
    input_path.write_text("asin,site,keyword\nB0TEST1234,US,bed frame\nB0TEST5678,US,desk\n", encoding="utf-8")
    bi_client = DummyBiReportDataClient()

    result = AsinDataCollector(bi_report_data_client=bi_client).collect(
        input=str(input_path),
        output_dir=str(tmp_path),
        run_id="batch-bi-only-report-data",
        bi_report_source_keys=["sales_traffic", "deals"],
        skip_query=False,
        skip_sales_query=True,
        skip_crawler_query=True,
        skip_seller_sprite=True,
        skip_amazon=True,
        skip_rufus=True,
        upload=False,
        fetch_report_files=False,
    )

    output_dir = Path(result["output_dir"])
    manifest = json.loads((output_dir / "manifest.json").read_text(encoding="utf-8"))
    jsonl_records = [
        json.loads(line)
        for line in (output_dir / "asin-data.jsonl").read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]

    assert bi_client.calls == [["B0TEST1234", "B0TEST5678"]]
    assert bi_client.source_key_calls == [["sales_traffic", "deals"]]
    assert bi_client.site_calls == [{"site_by_asin": {"B0TEST1234": "US", "B0TEST5678": "US"}, "default_site": "US"}]
    assert manifest["bi_report_data"]["request_mode"] == "batch"
    assert set(manifest["bi_report_data"]["sources"]) == {"sales_traffic", "deals"}
    assert manifest["bi_report_data"]["sources"]["sales_traffic"]["row_count"] == 2
    assert [record["bi_report_data"]["sources"]["sales_traffic"]["rows"][0]["asin"] for record in jsonl_records] == [
        "B0TEST1234",
        "B0TEST5678",
    ]


def test_report_file_submitter_builds_items_from_collect_result(tmp_path: Path):
    upload_client = DummyUploadClient()

    result = AsinDataCollector(file_upload_client=upload_client).collect(
        asin="B0TEST1234",
        site="US",
        keywords=["bed frame"],
        output_dir=str(tmp_path),
        run_id="single-asin-submit",
        skip_query=True,
        skip_seller_sprite=True,
        skip_amazon=True,
        skip_rufus=True,
        skip_bi_report_data=True,
        upload=True,
    )

    items = AsinReportFileSubmitter().build_items(result, report_date="2026-06-10")

    assert len(items) == 1
    item = items[0]
    assert item["asin"] == "B0TEST1234"
    assert item["site"] == "US"
    assert item["report_type"] == "asin_data_split_package_zip"
    assert item["report_date"] == "2026-06-10"
    assert item["status"] == "success"
    assert item["file_url"] == "https://p-amazon-task-test.oss-cn-hongkong.aliyuncs.com/asin-data/2026/06/1781083839-B0TEST1234-asin-data-package.zip"
    assert item["file_name"] == "B0TEST1234-asin-data-package.zip"
    assert item["file_ext"] == "zip"
    assert item["mime_type"] == "application/zip"
    assert len(item["content_hash"]) == 64
    assert item["content_text"] is None
    assert item["frontend_json"] is None
    assert item["raw_record_json"] is None


def test_collect_fetches_report_file_url_and_attaches_frontend_data(tmp_path: Path):
    report_file_client = DummyReportFileClient()

    result = AsinDataCollector(report_file_client=report_file_client).collect(
        asin="B0TEST1234",
        site="US",
        keywords=["bed frame"],
        output_dir=str(tmp_path),
        run_id="single-asin-report-file",
        skip_query=True,
        skip_seller_sprite=True,
        skip_amazon=True,
        skip_rufus=True,
        upload=False,
        fetch_report_files=True,
    )

    output_dir = Path(result["output_dir"])
    manifest = json.loads((output_dir / "manifest.json").read_text(encoding="utf-8"))
    frontend = json.loads((output_dir / "frontend-data.json").read_text(encoding="utf-8"))
    jsonl_record = json.loads((output_dir / "asin-data.jsonl").read_text(encoding="utf-8").splitlines()[0])

    assert report_file_client.called_with == [{"asin": "B0TEST1234", "site": "US"}]
    assert result["report_file_url"] == report_file_client.url
    assert result["aliyun_url"] == report_file_client.url
    assert result["report_files"]["items"][0]["url"] == report_file_client.url
    assert manifest["options"]["fetch_report_files"] is True
    assert manifest["report_files"]["items"][0]["url"] == report_file_client.url
    assert jsonl_record["asin_report_file"]["url"] == report_file_client.url

    record = frontend[next(key for key, value in frontend.items() if isinstance(value, list))][0]
    base_section = next(value for value in record.values() if isinstance(value, dict) and value.get("ASIN"))
    assert base_section["取数报告地址"] == report_file_client.url
    assert base_section["取数报告状态"] == "success"


def test_collect_raises_when_required_report_file_url_missing(tmp_path: Path):
    report_file_client = MissingReportFileClient()

    with pytest.raises(AsinReportFileNotFoundError, match="取数服务异常"):
        AsinDataCollector(report_file_client=report_file_client).collect(
            asin="B0TEST1234",
            site="US",
            keywords=["bed frame"],
            output_dir=str(tmp_path),
            run_id="single-asin-report-file-missing",
            skip_query=True,
            skip_seller_sprite=True,
            skip_amazon=True,
            skip_rufus=True,
            upload=False,
            fetch_report_files=True,
        )

    output_dir = tmp_path / "single-asin-report-file-missing"
    assert report_file_client.called_with == [{"asin": "B0TEST1234", "site": "US"}]
    assert not (output_dir / "frontend-data.json").exists()


def test_rufus_frontend_text_removes_newlines():
    legacy = load_legacy_collector()
    assert hasattr(legacy, "clean_rufus_text"), legacy.__file__
    assert legacy.clean_rufus_text("a\nb\\nc") == "a b c"

    data = legacy.localize_rufus_data(
        {
            "status": "success",
            "country": "US",
            "questions": ["问题一\n问题二", "问题三\\n问题四"],
            "answers": [
                {
                    "index": 1,
                    "question": "问题\n标题",
                    "related_products": ["产品A\n产品B"],
                    "answer": "第一行\n第二行\\n第三行",
                    "recommended_asins": ["B0TEST\n1234"],
                    "summary": "总结\r\n内容",
                }
            ],
            "report_path": "output\\amazon-rufus\\report.md",
        }
    )

    assert data["问题列表"] == ["问题一 问题二", "问题三 问题四"]
    assert data["报告路径"] == "output/amazon-rufus/report.md"
    answer = data["数据"][0]
    assert answer["问题"] == "问题 标题"
    assert answer["相关产品"] == ["产品A 产品B"]
    assert answer["答案"] == "第一行 第二行 第三行"
    assert answer["推荐ASIN"] == ["B0TEST 1234"]
    assert answer["总结"] == "总结 内容"


def test_merged_report_restores_rufus_answer_tables():
    report_text = render_merged_report_text(
        {
            "asin": "B0TEST1234",
            "site": "US",
            "rufus": {
                "status": "success",
                "questions": ["分析标题"],
                "answers": [
                    {
                        "index": 1,
                        "question": "分析标题",
                        "text": "答案正文",
                        "summaryText": "答案总结",
                        "blocks": [
                            {
                                "type": "container",
                                "children": [
                                    {
                                        "type": "text",
                                        "children": [
                                            {"type": "text", "children": "标题层"},
                                            {"type": "link", "children": "产品链接"},
                                        ],
                                    },
                                    {
                                        "type": "table",
                                        "children": [
                                            {
                                                "type": "tableRow",
                                                "children": [
                                                    {"type": "text", "children": "维度"},
                                                    {"type": "text", "children": "问题"},
                                                    {"type": "text", "children": "建议改为"},
                                                ],
                                            },
                                            {
                                                "type": "tableRow",
                                                "children": [
                                                    {"type": "text", "children": "标题"},
                                                    {"type": "text", "children": "缺少关键词"},
                                                    {"type": "text", "children": "补充 core | keyword"},
                                                ],
                                            },
                                        ],
                                    },
                                ],
                            }
                        ],
                    }
                ],
            },
        },
        summary={"run_id": "rufus-display-table", "files": {}},
    )

    assert "#### Rufus 展示内容" in report_text
    assert "标题层 产品链接" in report_text
    assert "| 维度 | 问题 | 建议改为 |" in report_text
    assert "| --- | --- | --- |" in report_text
    assert "| 标题 | 缺少关键词 | 补充 core \\| keyword |" in report_text
    assert "层级汇总表" not in report_text
    assert "| 层级 | 节点类型 | 内容摘要 |" not in report_text


def test_collect_requires_input_or_asin(tmp_path: Path):
    collector = AsinDataCollector()

    with pytest.raises(ValueError, match="--input or --asin"):
        collector.collect(output_dir=str(tmp_path), run_id="missing-source")


def test_collect_rejects_input_and_asin_together(tmp_path: Path):
    input_path = tmp_path / "asins.csv"
    input_path.write_text("asin,site\nB0TEST1234,US\n", encoding="utf-8")
    collector = AsinDataCollector()

    with pytest.raises(ValueError, match="--input or --asin"):
        collector.collect(input=str(input_path), asin="B0TEST1234", output_dir=str(tmp_path), run_id="conflict")


def test_direct_runner_passes_rufus_parallel_options_to_manager():
    legacy = load_legacy_collector()
    captured = {}

    class FakeRufusManager:
        def get_backend(self, **kwargs):
            captured.update(kwargs)
            return {"asin": kwargs["asin"], "answers": []}

    runner = DirectOpsRunner(legacy, rufus_manager=FakeRufusManager())

    result = runner.run_or_plan(
        source="rufus.get_backend",
        command=[
            "__direct__",
            "amazon-rufus",
            "get-backend",
            "B0TEST1234",
            "US",
            "--skills-dir",
            ".agents/skills",
            "--timeout",
            "180",
            "--parallel",
            "--concurrency",
            "2",
            "--retry",
            "1",
            "--strict-answer",
            "--no-upload-payload",
        ],
        dry_run=False,
        raw_output_path=None,
        command_log=None,
        error_log=None,
    )

    assert result["stderr"] == ""
    assert result["status"] == "success"
    assert captured["parallel"] is True
    assert captured["concurrency"] == 2
    assert captured["retry"] == 1
    assert captured["strict_answer"] is True
