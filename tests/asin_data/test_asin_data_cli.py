import json
import threading
import time
from pathlib import Path
from types import SimpleNamespace

from typer.testing import CliRunner
from openpyxl import Workbook

from opscli.asin_data import cli as asin_cli
from opscli.asin_data.services.live_data import upload_live_split_files
from opscli.asin_data.services.report_files import AsinReportFileNotFoundError


runner = CliRunner()


def write_test_workbook(path: Path, sheet_name: str, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


class DummyReportFileClient:
    def __init__(self, url: str | None = "https://example.oss.aliyuncs.com/asin-data/report.txt"):
        self.url = url
        self.called_with = None

    def fetch(self, *, asin: str, site: str):
        self.called_with = {"asin": asin, "site": site}
        return SimpleNamespace(
            asin=asin,
            site=site,
            url=self.url,
            record={"file_url": self.url, "asin": asin, "site": site},
            raw={"code": 0, "data": {"file_url": self.url}},
        )


class DummyFileUploadClient:
    def __init__(self):
        self.calls = []

    def upload(self, path, **kwargs):
        upload_path = Path(path)
        upload_name = kwargs.get("filename") or upload_path.name
        self.calls.append({"path": upload_path, **kwargs})
        return SimpleNamespace(
            url=f"https://example.oss.aliyuncs.com/asin-data/{upload_name}",
            raw={"code": 200, "data": {"url": f"https://example.oss.aliyuncs.com/asin-data/{upload_name}"}},
        )


def test_report_url_outputs_json(monkeypatch):
    client = DummyReportFileClient()
    monkeypatch.setattr(asin_cli, "AsinReportFileClient", lambda: client)

    result = runner.invoke(
        asin_cli.app,
        ["report-url", "--asin", " b0test1234 ", "--site", "us"],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["success"] is True
    assert payload["command"] == "asin-data report-url"
    assert payload["data"]["asin"] == "B0TEST1234"
    assert payload["data"]["site"] == "US"
    assert payload["data"]["report_file_url"] == "https://example.oss.aliyuncs.com/asin-data/report.txt"
    assert payload["data"]["record"]["file_url"] == "https://example.oss.aliyuncs.com/asin-data/report.txt"
    assert client.called_with == {"asin": "B0TEST1234", "site": "US"}


def test_report_url_url_only_prints_url(monkeypatch):
    client = DummyReportFileClient()
    monkeypatch.setattr(asin_cli, "AsinReportFileClient", lambda: client)

    result = runner.invoke(
        asin_cli.app,
        ["report-url", "--asin", "B0TEST1234", "--site", "US", "--url-only"],
    )

    assert result.exit_code == 0
    assert result.stdout.strip() == "https://example.oss.aliyuncs.com/asin-data/report.txt"


def test_report_url_outputs_not_found_error(monkeypatch):
    client = DummyReportFileClient(url=None)
    monkeypatch.setattr(asin_cli, "AsinReportFileClient", lambda: client)

    result = runner.invoke(
        asin_cli.app,
        ["report-url", "--asin", "B0TEST1234", "--site", "US"],
    )

    assert result.exit_code == 1
    payload = json.loads(result.stdout)
    assert payload["success"] is False
    assert payload["command"] == "asin-data report-url"
    assert payload["error"]["code"] == "ASIN_REPORT_FILE_NOT_FOUND"
    assert payload["error"]["asin"] == "B0TEST1234"
    assert payload["error"]["site"] == "US"


def test_collect_outputs_doc_aligned_json_and_passes_options(monkeypatch):
    class DummyCollector:
        def __init__(self):
            self.called_with = None

        def collect(self, **kwargs):
            self.called_with = kwargs
            return {
                "success": True,
                "output_dir": "output/asin-data/run-1",
                "summary": {"asin_count": 1},
                "manifest": {"run_id": "run-1"},
            }

    collector = DummyCollector()
    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: collector)

    result = runner.invoke(
        asin_cli.app,
        [
            "collect",
            "--input",
            "asins.csv",
            "--run-id",
            "run-1",
            "--skip-amazon",
            "--skip-query",
            "--skip-bi-report-data",
            "--rufus-question",
            "Q1 {{asin}}",
            "--rufus-question",
            "Q2 {{asin}}",
            "--rufus-parallel",
            "--rufus-concurrency",
            "2",
            "--rufus-retry",
            "1",
            "--rufus-strict-answer",
            "--sales-field-mode",
            "compatible",
            "--crawler-field-mode",
            "compatible",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["success"] is True
    assert payload["command"] == "asin-data collect"
    assert payload["data"]["output_dir"] == "output/asin-data/run-1"
    assert collector.called_with["input"] == "asins.csv"
    assert collector.called_with["run_id"] == "run-1"
    assert collector.called_with["skip_amazon"] is True
    assert collector.called_with["skip_query"] is True
    assert collector.called_with["skip_bi_report_data"] is True
    assert collector.called_with["fetch_report_files"] is True
    assert collector.called_with["upload"] is True
    assert collector.called_with["rufus_questions"] == ["Q1 {{asin}}", "Q2 {{asin}}"]
    assert collector.called_with["rufus_parallel"] is True
    assert collector.called_with["rufus_concurrency"] == 2
    assert collector.called_with["rufus_retry"] == 1
    assert collector.called_with["rufus_strict_answer"] is True
    assert collector.called_with["sales_field_mode"] == "compatible"
    assert collector.called_with["crawler_field_mode"] == "compatible"


def test_collect_accepts_single_asin_and_keywords(monkeypatch):
    class DummyCollector:
        def __init__(self):
            self.called_with = None

        def collect(self, **kwargs):
            self.called_with = kwargs
            return {
                "success": True,
                "output_dir": "output/asin-data/run-single",
                "summary": {"asin_count": 1},
                "manifest": {"run_id": "run-single"},
            }

    collector = DummyCollector()
    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: collector)

    result = runner.invoke(
        asin_cli.app,
        [
            "collect",
            "--asin",
            "B0TEST1234",
            "--site",
            "US",
            "--keyword",
            "bed frame",
            "--keyword",
            "storage bed",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["success"] is True
    assert collector.called_with["input"] is None
    assert collector.called_with["asin"] == "B0TEST1234"
    assert collector.called_with["site"] == "US"
    assert collector.called_with["keywords"] == ["bed frame", "storage bed"]
    assert collector.called_with["skip_bi_report_data"] is False
    assert collector.called_with["skip_crawler_query"] is True
    assert collector.called_with["fetch_report_files"] is True


def test_collect_can_enable_legacy_crawler_query(monkeypatch):
    class DummyCollector:
        def __init__(self):
            self.called_with = None

        def collect(self, **kwargs):
            self.called_with = kwargs
            return {
                "success": True,
                "output_dir": "output/asin-data/run-single",
                "summary": {"asin_count": 1},
                "manifest": {"run_id": "run-single"},
            }

    collector = DummyCollector()
    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: collector)

    result = runner.invoke(
        asin_cli.app,
        [
            "collect",
            "--asin",
            "B0TEST1234",
            "--site",
            "US",
            "--legacy-crawler-query",
        ],
    )

    assert result.exit_code == 0
    assert collector.called_with["skip_crawler_query"] is False


def test_collect_can_disable_report_file_lookup(monkeypatch):
    class DummyCollector:
        def __init__(self):
            self.called_with = None

        def collect(self, **kwargs):
            self.called_with = kwargs
            return {
                "success": True,
                "output_dir": "output/asin-data/run-single",
                "summary": {"asin_count": 1},
                "manifest": {"run_id": "run-single"},
            }

    collector = DummyCollector()
    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: collector)

    result = runner.invoke(
        asin_cli.app,
        [
            "collect",
            "--asin",
            "B0TEST1234",
            "--site",
            "US",
            "--no-fetch-report-files",
        ],
    )

    assert result.exit_code == 0
    assert collector.called_with["fetch_report_files"] is False


def test_live_data_returns_inline_frontend_data_and_uses_live_defaults(monkeypatch, tmp_path: Path):
    class DummyCollector:
        def __init__(self):
            self.called_with = None

        def collect(self, **kwargs):
            self.called_with = kwargs
            output_dir = tmp_path / "live-run"
            output_dir.mkdir()
            basic_path = output_dir / "asin-data-packages" / "B0TEST1234" / "01-基础数据.xlsx"
            bi_path = output_dir / "asin-data-packages" / "B0TEST1234" / "02-BI数据.xlsx"
            write_test_workbook(basic_path, "基础汇总", [["数据项", "值"], ["ASIN", "B0TEST1234"]])
            write_test_workbook(bi_path, "销售数据", [["asin", "orderQty"], ["B0TEST1234", 4]])
            frontend_data = {
                "运行信息": {"运行ID": "live-run"},
                "数据": [{"基础数据": {"ASIN": "B0TEST1234"}}],
            }
            (output_dir / "frontend-data.json").write_text(
                json.dumps(frontend_data, ensure_ascii=False),
                encoding="utf-8",
            )
            return {
                "success": True,
                "output_dir": output_dir.as_posix(),
                "summary": {"asin_count": 1},
                "manifest": {
                    "run_id": "live-run",
                    "asin_data_package": {
                        "items": [
                            {
                                "asin": "B0TEST1234",
                                "files": {
                                    "basic": basic_path.as_posix(),
                                    "bi": bi_path.as_posix(),
                                },
                            }
                        ]
                    },
                },
            }

    collector = DummyCollector()
    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: collector)

    result = runner.invoke(
        asin_cli.app,
        [
            "live-data",
            "--asin",
            "B0TEST1234",
            "--site",
            "US",
            "--sales-start",
            "2026-07-01",
            "--sales-end",
            "2026-07-08",
            "--pretty",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["success"] is True
    assert payload["command"] == "asin-data live-data"
    assert payload["data"]["frontend_data"]["数据"][0]["基础数据"]["ASIN"] == "B0TEST1234"
    assert payload["data"]["split_files"]["B0TEST1234"]["basic"]["content"]["基础汇总"] == [
        ["数据项", "值"],
        ["ASIN", "B0TEST1234"],
    ]
    assert payload["data"]["split_files"]["B0TEST1234"]["bi"]["content"]["销售数据"] == [
        ["asin", "orderQty"],
        ["B0TEST1234", 4],
    ]
    assert collector.called_with["asin"] == "B0TEST1234"
    assert collector.called_with["sales_start"] == "2026-07-01"
    assert collector.called_with["sales_end"] == "2026-07-08"
    assert collector.called_with["skip_query"] is True
    assert collector.called_with["skip_bi_report_data"] is False
    assert collector.called_with["skip_seller_sprite"] is True
    assert collector.called_with["skip_keyword_miner"] is True
    assert collector.called_with["skip_listing_analysis"] is True
    assert collector.called_with["skip_amazon"] is True
    assert collector.called_with["skip_rufus"] is True
    assert collector.called_with["fetch_report_files"] is False
    assert collector.called_with["upload"] is False
    assert collector.called_with["bi_report_source_keys"] is None


def test_live_data_listing_scope_only_returns_basic_file(monkeypatch, tmp_path: Path):
    class DummyCollector:
        def __init__(self):
            self.called_with = None

        def collect(self, **kwargs):
            self.called_with = kwargs
            output_dir = tmp_path / "live-run-listing"
            output_dir.mkdir()
            basic_path = output_dir / "asin-data-packages" / "B0TEST1234" / "01-基础数据.xlsx"
            bi_path = output_dir / "asin-data-packages" / "B0TEST1234" / "02-BI数据.xlsx"
            write_test_workbook(basic_path, "基础汇总", [["数据项", "值"], ["ASIN", "B0TEST1234"]])
            write_test_workbook(bi_path, "销售数据", [["asin", "orderQty"], ["B0TEST1234", 4]])
            (output_dir / "frontend-data.json").write_text(
                json.dumps({"数据": [{"基础数据": {"ASIN": "B0TEST1234"}}]}, ensure_ascii=False),
                encoding="utf-8",
            )
            return {
                "success": True,
                "output_dir": output_dir.as_posix(),
                "summary": {"asin_count": 1},
                "manifest": {
                    "run_id": "live-run-listing",
                    "asin_data_package": {
                        "items": [
                            {
                                "asin": "B0TEST1234",
                                "files": {
                                    "basic": basic_path.as_posix(),
                                    "bi": bi_path.as_posix(),
                                },
                            }
                        ]
                    },
                },
            }

    collector = DummyCollector()
    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: collector)

    result = runner.invoke(
        asin_cli.app,
        [
            "live-data",
            "--asin",
            "B0TEST1234",
            "--data-scope",
            "listing",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["data"]["data_scope"] == "listing"
    assert payload["data"]["split_files"]["B0TEST1234"].keys() == {"basic"}
    assert collector.called_with["bi_report_source_keys"] == ("listing_basic",)


def test_live_data_listing_basic_scope_alias_returns_basic_file(monkeypatch, tmp_path: Path):
    class DummyCollector:
        def __init__(self):
            self.called_with = None

        def collect(self, **kwargs):
            self.called_with = kwargs
            output_dir = tmp_path / "live-run-listing-basic"
            output_dir.mkdir()
            basic_path = output_dir / "asin-data-packages" / "B0TEST1234" / "01-基础数据.xlsx"
            bi_path = output_dir / "asin-data-packages" / "B0TEST1234" / "02-BI数据.xlsx"
            write_test_workbook(basic_path, "基础汇总", [["数据项", "值"], ["ASIN", "B0TEST1234"]])
            write_test_workbook(bi_path, "销售数据", [["asin", "orderQty"], ["B0TEST1234", 4]])
            (output_dir / "frontend-data.json").write_text(
                json.dumps({"数据": [{"基础数据": {"ASIN": "B0TEST1234"}}]}, ensure_ascii=False),
                encoding="utf-8",
            )
            return {
                "success": True,
                "output_dir": output_dir.as_posix(),
                "summary": {"asin_count": 1},
                "manifest": {
                    "run_id": "live-run-listing-basic",
                    "asin_data_package": {
                        "items": [
                            {
                                "asin": "B0TEST1234",
                                "files": {
                                    "basic": basic_path.as_posix(),
                                    "bi": bi_path.as_posix(),
                                },
                            }
                        ]
                    },
                },
            }

    collector = DummyCollector()
    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: collector)

    result = runner.invoke(
        asin_cli.app,
        [
            "live-data",
            "--asin",
            "B0TEST1234",
            "--data-scope",
            "listing_basic",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["data"]["data_scope"] == "listing_basic"
    assert payload["data"]["split_files"]["B0TEST1234"].keys() == {"basic"}
    assert collector.called_with["bi_report_source_keys"] == ("listing_basic",)


def test_live_data_basic_scope_returns_complete_basic_file(monkeypatch, tmp_path: Path):
    class DummyCollector:
        def __init__(self):
            self.called_with = None

        def collect(self, **kwargs):
            self.called_with = kwargs
            output_dir = tmp_path / "live-run-basic"
            output_dir.mkdir()
            basic_path = output_dir / "asin-data-packages" / "B0TEST1234" / "01-基础数据.xlsx"
            bi_path = output_dir / "asin-data-packages" / "B0TEST1234" / "02-BI数据.xlsx"
            write_test_workbook(basic_path, "爬虫数据", [["asin", "title"], ["B0TEST1234", "Crawler title"]])
            write_test_workbook(bi_path, "销售数据", [["asin", "orderQty"], ["B0TEST1234", 4]])
            (output_dir / "frontend-data.json").write_text(
                json.dumps({"数据": [{"基础数据": {"ASIN": "B0TEST1234"}}]}, ensure_ascii=False),
                encoding="utf-8",
            )
            return {
                "success": True,
                "output_dir": output_dir.as_posix(),
                "summary": {"asin_count": 1},
                "manifest": {
                    "run_id": "live-run-basic",
                    "asin_data_package": {
                        "items": [
                            {
                                "asin": "B0TEST1234",
                                "files": {
                                    "basic": basic_path.as_posix(),
                                    "bi": bi_path.as_posix(),
                                },
                            }
                        ]
                    },
                },
            }

    collector = DummyCollector()
    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: collector)

    result = runner.invoke(
        asin_cli.app,
        [
            "live-data",
            "--asin",
            "B0TEST1234",
            "--data-scope",
            "basic",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["data"]["data_scope"] == "basic"
    assert payload["data"]["split_files"]["B0TEST1234"].keys() == {"basic"}
    assert payload["data"]["split_files"]["B0TEST1234"]["basic"]["content"]["爬虫数据"] == [
        ["asin", "title"],
        ["B0TEST1234", "Crawler title"],
    ]
    assert collector.called_with["bi_report_source_keys"] == ("listing_basic", "crawler_details")


def test_live_data_bi_scope_only_returns_bi_file(monkeypatch, tmp_path: Path):
    class DummyCollector:
        def __init__(self):
            self.called_with = None

        def collect(self, **kwargs):
            self.called_with = kwargs
            output_dir = tmp_path / "live-run-bi"
            output_dir.mkdir()
            basic_path = output_dir / "asin-data-packages" / "B0TEST1234" / "01-基础数据.xlsx"
            bi_path = output_dir / "asin-data-packages" / "B0TEST1234" / "02-BI数据.xlsx"
            write_test_workbook(basic_path, "基础汇总", [["数据项", "值"], ["ASIN", "B0TEST1234"]])
            write_test_workbook(bi_path, "销售数据", [["asin", "orderQty"], ["B0TEST1234", 4]])
            (output_dir / "frontend-data.json").write_text(
                json.dumps({"数据": [{"基础数据": {"ASIN": "B0TEST1234"}}]}, ensure_ascii=False),
                encoding="utf-8",
            )
            return {
                "success": True,
                "output_dir": output_dir.as_posix(),
                "summary": {"asin_count": 1},
                "manifest": {
                    "run_id": "live-run-bi",
                    "asin_data_package": {
                        "items": [
                            {
                                "asin": "B0TEST1234",
                                "files": {
                                    "basic": basic_path.as_posix(),
                                    "bi": bi_path.as_posix(),
                                },
                            }
                        ]
                    },
                },
            }

    collector = DummyCollector()
    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: collector)

    result = runner.invoke(
        asin_cli.app,
        [
            "live-data",
            "--asin",
            "B0TEST1234",
            "--data-scope",
            "bi",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["data"]["data_scope"] == "bi"
    assert payload["data"]["split_files"]["B0TEST1234"].keys() == {"bi"}
    assert collector.called_with["bi_report_source_keys"] == (
        "sales_traffic",
        "sp_search_term",
        "deals",
        "turnover_inventory",
    )


def test_live_data_can_upload_basic_and_bi_xlsx(monkeypatch, tmp_path: Path):
    class DummyCollector:
        def collect(self, **kwargs):
            output_dir = tmp_path / "live-run-upload"
            output_dir.mkdir()
            basic_path = output_dir / "asin-data-packages" / "B0TEST1234" / "01-基础数据.xlsx"
            bi_path = output_dir / "asin-data-packages" / "B0TEST1234" / "02-BI数据.xlsx"
            write_test_workbook(basic_path, "基础汇总", [["数据项", "值"], ["ASIN", "B0TEST1234"]])
            write_test_workbook(bi_path, "销售数据", [["asin", "orderQty"], ["B0TEST1234", 4]])
            (output_dir / "frontend-data.json").write_text(
                json.dumps({"数据": [{"基础数据": {"ASIN": "B0TEST1234"}}]}, ensure_ascii=False),
                encoding="utf-8",
            )
            return {
                "success": True,
                "output_dir": output_dir.as_posix(),
                "summary": {"asin_count": 1},
                "manifest": {
                    "run_id": "live-run-upload",
                    "asin_data_package": {
                        "items": [
                            {
                                "asin": "B0TEST1234",
                                "files": {
                                    "basic": basic_path.as_posix(),
                                    "bi": bi_path.as_posix(),
                                },
                            }
                        ]
                    },
                },
            }

    upload_client = DummyFileUploadClient()
    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: DummyCollector())
    monkeypatch.setattr(asin_cli, "FileUploadClient", lambda: upload_client)

    result = runner.invoke(
        asin_cli.app,
        [
            "live-data",
            "--asin",
            "B0TEST1234",
            "--site",
            "US",
            "--upload-xlsx",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    files = payload["data"]["split_files"]["B0TEST1234"]
    assert files["basic"]["file_url"] == "https://example.oss.aliyuncs.com/asin-data/B0TEST1234-basic-live-data.xlsx"
    assert files["bi"]["file_url"] == "https://example.oss.aliyuncs.com/asin-data/B0TEST1234-bi-live-data.xlsx"
    assert payload["data"]["split_file_uploads"]["files_uploaded"] == 2
    assert sorted(call["metadata"]["file_key"] for call in upload_client.calls) == ["basic", "bi"]
    assert sorted(call["metadata"]["source_filename"] for call in upload_client.calls) == [
        "01-基础数据.xlsx",
        "02-BI数据.xlsx",
    ]
    assert {call["purpose"] for call in upload_client.calls} == {"asin_data_live_xlsx"}


def test_live_data_url_only_uploads_without_parsing_inline_content(monkeypatch, tmp_path: Path):
    class DummyCollector:
        def __init__(self):
            self.called_with = None

        def collect(self, **kwargs):
            self.called_with = kwargs
            output_dir = tmp_path / "live-run-url-only"
            output_dir.mkdir()
            basic_path = output_dir / "asin-data-packages" / "B0TEST1234" / "01-基础数据.xlsx"
            bi_path = output_dir / "asin-data-packages" / "B0TEST1234" / "02-BI数据.xlsx"
            basic_path.parent.mkdir(parents=True, exist_ok=True)
            basic_path.write_text("not a real xlsx", encoding="utf-8")
            bi_path.write_text("not a real xlsx", encoding="utf-8")
            return {
                "success": True,
                "output_dir": output_dir.as_posix(),
                "summary": {"asin_count": 1},
                "manifest": {
                    "run_id": "live-run-url-only",
                    "asin_data_package": {
                        "items": [
                            {
                                "asin": "B0TEST1234",
                                "files": {
                                    "basic": basic_path.as_posix(),
                                    "bi": bi_path.as_posix(),
                                },
                            }
                        ]
                    },
                },
            }

    upload_client = DummyFileUploadClient()
    collector = DummyCollector()
    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: collector)
    monkeypatch.setattr(asin_cli, "FileUploadClient", lambda: upload_client)

    result = runner.invoke(
        asin_cli.app,
        [
            "live-data",
            "--asin",
            "B0TEST1234",
            "--site",
            "US",
            "--upload-xlsx",
            "--return-mode",
            "url_only",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert collector.called_with is not None
    assert collector.called_with["split_file_keys"] == ("basic", "bi")
    assert collector.called_with["build_split_package_zip"] is False
    assert collector.called_with["write_frontend_outputs"] is False
    assert payload["data"]["return_mode"] == "url_only"
    assert "frontend_data" not in payload["data"]
    assert "split_files" not in payload["data"]
    assert payload["data"]["split_file_urls"] == {
        "B0TEST1234": {
            "basic": "https://example.oss.aliyuncs.com/asin-data/B0TEST1234-basic-live-data.xlsx",
            "bi": "https://example.oss.aliyuncs.com/asin-data/B0TEST1234-bi-live-data.xlsx",
        }
    }
    assert payload["data"]["split_file_uploads"]["files_uploaded"] == 2


def test_basic_command_returns_json_and_records_usage(monkeypatch):
    calls = {}
    usage_events = []

    class DummyQueryService:
        def fetch_basic(self, **kwargs):
            calls["kwargs"] = kwargs
            return {
                "asins": ["B086M58PQ3"],
                "site": "US",
                "sources": {"listing_basic": {"rows": [{"ASIN": "B086M58PQ3"}]}},
            }

    monkeypatch.setattr(asin_cli, "AsinDataQueryService", DummyQueryService)
    monkeypatch.setattr(asin_cli, "append_usage_event", lambda **kwargs: usage_events.append(kwargs))

    result = runner.invoke(
        asin_cli.app,
        ["basic", "--asin", "B086M58PQ3", "--site", "US", "--source", "listing"],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["success"] is True
    assert payload["command"] == "asin-data basic"
    assert payload["data"]["sources"]["listing_basic"]["rows"][0]["ASIN"] == "B086M58PQ3"
    assert calls["kwargs"] == {
        "asins": ["B086M58PQ3"],
        "site": "US",
        "sources": ["listing"],
    }
    assert usage_events[0]["command"] == "basic"
    assert usage_events[0]["status"] == "success"
    assert usage_events[0]["params"]["asins"] == ["B086M58PQ3"]


def test_bi_command_accepts_json_asin_list_and_domains(monkeypatch):
    calls = {}

    class DummyQueryService:
        def fetch_bi(self, **kwargs):
            calls["kwargs"] = kwargs
            return {
                "asins": kwargs["asins"],
                "site": kwargs["site"],
                "date_from": kwargs["date_from"],
                "date_to": kwargs["date_to"],
                "domains": kwargs["domains"],
                "sources": {},
            }

    monkeypatch.setattr(asin_cli, "AsinDataQueryService", DummyQueryService)
    monkeypatch.setattr(asin_cli, "append_usage_event", lambda **kwargs: None)

    result = runner.invoke(
        asin_cli.app,
        [
            "bi",
            "--asins",
            '["B086M58PQ3","B0TEST1234"]',
            "--site",
            "DE",
            "--date-from",
            "2026-07-01",
            "--date-to",
            "2026-07-15",
            "--domain",
            "sales_traffic",
            "--domain",
            "deals",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["command"] == "asin-data bi"
    assert calls["kwargs"] == {
        "asins": ["B086M58PQ3", "B0TEST1234"],
        "site": "DE",
        "date_from": "2026-07-01",
        "date_to": "2026-07-15",
        "domains": ["sales_traffic", "deals"],
    }


def test_category_top_command_returns_only_category_top_json(monkeypatch):
    calls = {}

    class DummyQueryService:
        def fetch_category_top(self, **kwargs):
            calls["kwargs"] = kwargs
            return {
                "category": kwargs["category"],
                "site": kwargs["site"],
                "date_from": kwargs["date_from"],
                "date_to": kwargs["date_to"],
                "limit": kwargs["limit"],
                "row_count": 1,
                "category_top": [{"ASIN": "B0TEST1234", "排名": 1}],
            }

    monkeypatch.setattr(asin_cli, "AsinDataQueryService", DummyQueryService)
    monkeypatch.setattr(asin_cli, "append_usage_event", lambda **kwargs: None)

    result = runner.invoke(
        asin_cli.app,
        [
            "category-top",
            "--category",
            "Bed Frames",
            "--date-from",
            "2026-07-01",
            "--date-to",
            "2026-07-13",
            "--limit",
            "5",
            "--site",
            "DE",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["success"] is True
    assert payload["command"] == "asin-data category-top"
    assert payload["data"]["category_top"] == [{"ASIN": "B0TEST1234", "排名": 1}]
    assert set(payload["data"]) == {
        "category",
        "site",
        "date_from",
        "date_to",
        "limit",
        "row_count",
        "category_top",
    }
    assert calls["kwargs"] == {
        "category": "Bed Frames",
        "data_type": "asin",
        "date_from": "2026-07-01",
        "date_to": "2026-07-13",
        "limit": 5,
        "site": "DE",
    }


def test_category_top_traffic_command_allows_all_categories(monkeypatch):
    calls = {}

    class DummyQueryService:
        def fetch_category_top(self, **kwargs):
            calls["kwargs"] = kwargs
            return {
                "data_type": "traffic",
                "category": None,
                "date_from": kwargs["date_from"],
                "date_to": kwargs["date_to"],
                "row_count": 1,
                "category_total": 3289,
                "category_names": ["3D Wall Panels"],
                "ranking_metric": "page_views",
                "top_n": 10,
                "category_traffic": [{"category": "3D Wall Panels"}],
            }

    monkeypatch.setattr(asin_cli, "AsinDataQueryService", DummyQueryService)
    monkeypatch.setattr(asin_cli, "append_usage_event", lambda **kwargs: None)

    result = runner.invoke(
        asin_cli.app,
        [
            "category-top",
            "--data-type",
            "traffic",
            "--date-from",
            "2026-07-01",
            "--date-to",
            "2026-07-27",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["data"]["category_traffic"] == [{"category": "3D Wall Panels"}]
    assert calls["kwargs"] == {
        "category": None,
        "data_type": "traffic",
        "date_from": "2026-07-01",
        "date_to": "2026-07-27",
        "limit": 10,
        "site": "US",
    }


def test_basic_command_records_error_usage(monkeypatch):
    usage_events = []

    class DummyQueryService:
        def fetch_basic(self, **kwargs):
            raise ValueError("ASIN 格式无效")

    monkeypatch.setattr(asin_cli, "AsinDataQueryService", DummyQueryService)
    monkeypatch.setattr(asin_cli, "append_usage_event", lambda **kwargs: usage_events.append(kwargs))

    result = runner.invoke(asin_cli.app, ["basic", "--asin", "bad"])

    assert result.exit_code == 1
    payload = json.loads(result.stdout)
    assert payload["success"] is False
    assert usage_events[0]["status"] == "error"
    assert usage_events[0]["error"]["message"] == "ASIN 格式无效"


def test_live_data_ai_ready_returns_dataset_manifest_without_inline_content(monkeypatch, tmp_path: Path):
    class DummyCollector:
        def __init__(self):
            self.called_with = None

        def collect(self, **kwargs):
            self.called_with = kwargs
            output_dir = tmp_path / "live-run-ai-ready"
            output_dir.mkdir()
            bi_path = output_dir / "asin-data-packages" / "B0TEST1234" / "02-BI数据.xlsx"
            wb = Workbook()
            wb.remove(wb.active)
            for sheet_name, rows in [
                ("sales", [["ASIN", "orderQty"], ["B0TEST1234", 4]]),
                ("sp", [["channel", "asin_group", "search_term"], ["US", "B0TEST1234", "bed"]]),
                ("deals", [["说明"], ["无数据"]]),
                ("inventory", [["ASIN", "stock"], ["B0TEST1234", 8]]),
            ]:
                ws = wb.create_sheet(sheet_name)
                for row in rows:
                    ws.append(row)
            bi_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(bi_path)
            return {
                "success": True,
                "output_dir": output_dir.as_posix(),
                "summary": {"asin_count": 1},
                "manifest": {
                    "run_id": "live-run-ai-ready",
                    "asin_data_package": {
                        "items": [
                            {
                                "asin": "B0TEST1234",
                                "site": "US",
                                "files": {
                                    "bi": bi_path.as_posix(),
                                },
                            }
                        ]
                    },
                },
            }

    collector = DummyCollector()
    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: collector)

    result = runner.invoke(
        asin_cli.app,
        [
            "live-data",
            "--asin",
            "B0TEST1234",
            "--site",
            "US",
            "--data-scope",
            "bi",
            "--sales-start",
            "2026-07-02",
            "--sales-end",
            "2026-07-08",
            "--return-mode",
            "ai_ready",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    data = payload["data"]
    assert collector.called_with is not None
    assert collector.called_with["split_file_keys"] == ("bi",)
    assert collector.called_with["build_split_package_zip"] is False
    assert collector.called_with["write_frontend_outputs"] is False
    assert data["metadata"]["protocol"] == "asin_data_ai_response"
    assert data["metadata"]["request"]["return_mode"] == "ai_ready"
    assert "frontend_data" not in data
    assert "split_files" not in data
    datasets = {dataset["source_key"]: dataset for dataset in data["items"][0]["datasets"]}
    assert set(datasets) == {"sales_traffic", "sp_search_term", "deals", "turnover_inventory"}
    assert len(datasets["sp_search_term"]["preview_rows"]) == 1
    assert data["split_file_paths"]["B0TEST1234"]["bi"]["file_path"].endswith("02-BI数据.xlsx")


def test_upload_live_split_files_uploads_xlsx_in_parallel(tmp_path: Path):
    paths: dict[tuple[str, str], Path] = {}
    split_files: dict[str, dict[str, dict[str, str]]] = {}
    for asin in ("B0TEST1234", "B0TEST5678"):
        asin_files: dict[str, dict[str, str]] = {}
        for file_key in ("basic", "bi"):
            path = tmp_path / asin / f"{file_key}.xlsx"
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(file_key, encoding="utf-8")
            paths[(asin, file_key)] = path
            asin_files[file_key] = {
                "asin": asin,
                "file_key": file_key,
                "file_path": path.as_posix(),
            }
        split_files[asin] = asin_files

    class SlowUploadClient:
        def __init__(self):
            self.active = 0
            self.max_active = 0
            self.lock = threading.Lock()
            self.calls = []

        def upload(self, path, **kwargs):
            with self.lock:
                self.active += 1
                self.max_active = max(self.max_active, self.active)
            try:
                time.sleep(0.05)
            finally:
                with self.lock:
                    self.active -= 1
            upload_path = Path(path)
            upload_name = kwargs.get("filename") or upload_path.name
            self.calls.append({"path": upload_path, **kwargs})
            return SimpleNamespace(
                url=f"https://example.oss.aliyuncs.com/asin-data/{upload_name}",
                raw={"code": 200},
            )

    upload_client = SlowUploadClient()

    result = upload_live_split_files(
        split_files,
        run_id="parallel-upload",
        file_upload_client=upload_client,
    )

    assert result["files_uploaded"] == 4
    assert upload_client.max_active >= 2
    assert split_files["B0TEST1234"]["basic"]["file_url"].endswith("B0TEST1234-basic-live-data.xlsx")
    assert split_files["B0TEST5678"]["bi"]["file_url"].endswith("B0TEST5678-bi-live-data.xlsx")
    assert [item["asin"] for item in result["items"]] == ["B0TEST1234", "B0TEST5678"]
    assert sorted(call["filename"] for call in upload_client.calls) == [
        "B0TEST1234-basic-live-data.xlsx",
        "B0TEST1234-bi-live-data.xlsx",
        "B0TEST5678-basic-live-data.xlsx",
        "B0TEST5678-bi-live-data.xlsx",
    ]
    assert not (tmp_path / "B0TEST1234" / "B0TEST1234-basic-live-data.xlsx").exists()


def test_collect_can_submit_report_file_records(monkeypatch):
    class DummyCollector:
        def __init__(self):
            self.called_with = None

        def collect(self, **kwargs):
            self.called_with = kwargs
            return {
                "success": True,
                "output_dir": "output/asin-data/run-single",
                "summary": {"asin_count": 1},
                "manifest": {"run_id": "run-single"},
                "upload": {"url": "https://example.oss.aliyuncs.com/report.txt"},
                "aliyun_url": "https://example.oss.aliyuncs.com/report.txt",
            }

    class DummySubmitter:
        def __init__(self):
            self.called_with = None

        def submit(self, collect_result, **kwargs):
            self.called_with = {"collect_result": collect_result, **kwargs}
            return {"submitted": True, "count": 1, "inserted": 1}

    collector = DummyCollector()
    submitter = DummySubmitter()
    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: collector)
    monkeypatch.setattr(asin_cli, "AsinReportFileClient", lambda endpoint=None: SimpleNamespace(endpoint=endpoint))
    monkeypatch.setattr(asin_cli, "AsinReportFileSubmitter", lambda client: submitter)

    result = runner.invoke(
        asin_cli.app,
        [
            "collect",
            "--asin",
            "B0TEST1234",
            "--site",
            "US",
            "--submit-report-files",
            "--report-date",
            "2026-06-10",
            "--report-type",
            "asin_data_report_txt",
            "--report-source",
            "asin_data_daily_report",
            "--register-endpoint",
            "https://ops.example.com/dataMetrics/v1/asin-report-files",
            "--no-include-report-content",
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["success"] is True
    assert payload["data"]["report_file_submit"]["submitted"] is True
    assert collector.called_with["fetch_report_files"] is False
    assert submitter.called_with["report_date"] == "2026-06-10"
    assert submitter.called_with["report_type"] == "asin_data_report_txt"
    assert submitter.called_with["source"] == "asin_data_daily_report"
    assert submitter.called_with["include_content"] is False


def test_collect_url_only_prints_aliyun_url(monkeypatch):
    class DummyCollector:
        def collect(self, **kwargs):
            return {
                "success": True,
                "output_dir": "output/asin-data/run-single",
                "summary": {"asin_count": 1},
                "manifest": {"run_id": "run-single"},
                "upload": {"url": "https://example.oss.aliyuncs.com/asin-data/frontend-data.json"},
                "aliyun_url": "https://example.oss.aliyuncs.com/asin-data/frontend-data.json",
            }

    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: DummyCollector())

    result = runner.invoke(
        asin_cli.app,
        [
            "collect",
            "--asin",
            "B0TEST1234",
            "--site",
            "US",
            "--keyword",
            "bed frame",
            "--url-only",
        ],
    )

    assert result.exit_code == 0
    assert result.stdout.strip() == "https://example.oss.aliyuncs.com/asin-data/frontend-data.json"


def test_collect_outputs_error_payload(monkeypatch):
    class DummyCollector:
        def collect(self, **kwargs):
            raise ValueError("bad input")

    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: DummyCollector())

    result = runner.invoke(asin_cli.app, ["collect", "--input", "missing.csv"])

    assert result.exit_code == 1
    payload = json.loads(result.stdout)
    assert payload["success"] is False
    assert payload["command"] == "asin-data collect"
    assert payload["error"]["code"] == "ASIN_DATA_ERROR"
    assert payload["error"]["message"] == "bad input"


def test_collect_outputs_report_file_not_found_error_payload(monkeypatch):
    class DummyCollector:
        def collect(self, **kwargs):
            raise AsinReportFileNotFoundError(asin="B0TEST1234", site="US")

    monkeypatch.setattr(asin_cli, "AsinDataCollector", lambda: DummyCollector())

    result = runner.invoke(asin_cli.app, ["collect", "--asin", "B0TEST1234", "--site", "US"])

    assert result.exit_code == 1
    payload = json.loads(result.stdout)
    assert payload["success"] is False
    assert payload["command"] == "asin-data collect"
    assert payload["error"]["code"] == "ASIN_REPORT_FILE_NOT_FOUND"
    assert "取数服务异常" in payload["error"]["message"]
    assert payload["error"]["asin"] == "B0TEST1234"
    assert payload["error"]["site"] == "US"


def test_yicopy_keyword_engine_cli_writes_keyword_reverse_file(monkeypatch, tmp_path: Path):
    """asin-data yicopy 命令应支持 -a/-o 并写出销词数组。"""

    calls = {}

    class DummyEngine:
        async def run(self, sources, options):
            calls["sources"] = sources
            calls["options"] = options
            return {
                "status": "succeeded",
                "keywordRows": [
                    {
                        "keyword": "wireless mouse",
                        "titleFrequency": 0,
                        "bulletsFrequency": 0,
                        "totalFrequency": 0,
                    }
                ],
                "summary": {"asinCount": 1, "keywordReverseCount": 1},
            }

    monkeypatch.setattr(asin_cli, "YicopyKeywordEngine", lambda: DummyEngine(), raising=False)
    output_file = tmp_path / "yicopy.json"

    result = runner.invoke(
        asin_cli.app,
        [
            "yicopy-keyword-engine",
            "-a",
            "B0TEST1234",
            "-o",
            str(output_file),
            "--max-prefixes-per-asin",
            "1",
            "--pretty",
        ],
    )

    assert result.exit_code == 0, result.stdout
    assert json.loads(output_file.read_text(encoding="utf-8")) == [
        {
            "keyword": "wireless mouse",
            "titleFrequency": 0,
            "bulletsFrequency": 0,
            "totalFrequency": 0,
        }
    ]
    payload = json.loads(result.stdout)
    assert payload["success"] is True
    assert payload["command"] == "asin-data yicopy-keyword-engine"
    data = payload["data"]
    assert data["status"] == "succeeded"
    assert data["output_file"] == str(output_file)
    assert data["row_count"] == 1
    assert data["metadata"]["protocol"] == "asin_data_ai_response"
    assert data["metadata"]["tool"] == "asin-data yicopy-keyword-engine"
    assert data["metadata"]["data_scope"] == "yicopy_keyword_reverse"
    assert data["metadata"]["request"]["asin"] == ["B0TEST1234"]
    assert data["run"]["output_dir"] == output_file.parent.as_posix()
    assert data["summary"]["keywordReverseCount"] == 1
    item = data["items"][0]
    assert item["asin"] == "B0TEST1234"
    assert item["artifacts"][0]["file_key"] == "yicopy_keyword_reverse"
    assert item["artifacts"][0]["local_path"] == output_file.as_posix()
    assert item["datasets"][0]["source_key"] == "yicopy_keyword_reverse"
    assert item["datasets"][0]["preview_rows"][0]["keyword"] == "wireless mouse"
    assert data["diagnostics"] == []
    assert calls["sources"] == ["B0TEST1234"]
    assert calls["options"].max_prefixes_per_asin == 1
