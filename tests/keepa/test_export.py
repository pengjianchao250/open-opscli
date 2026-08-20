import json
from pathlib import Path
from collections import defaultdict
from types import SimpleNamespace

from openpyxl import load_workbook

from opscli.keepa.api.scenarios import SCENARIOS
from opscli.keepa.export.xlsx import export_rows_to_xlsx
from opscli.keepa.export.json import export_rows_to_json


SCENARIO_ROWS = {
    "product": [{"asin": "B0088PUEPK", "title": "Test Product", "unknownField": "raw"}],
    "product-search": ["B0088PUEPK"],
    "product-finder": [{"asin": "B0088PUEPK", "brand": "Brand", "current_SALES": 123}],
    "category-search": [{"catId": 172282, "name": "Electronics", "unknownCategoryField": "raw"}],
    "category-lookup": [{"categoryId": 172282, "name": "Electronics", "children": [1, 2]}],
    "seller": [{"sellerId": "A2L77EE7U53NWQ", "sellerName": "Test Seller", "ratingCount": 5}],
    "seller-finder": ["A2L77EE7U53NWQ"],
    "top-seller": ["A2L77EE7U53NWQ"],
    "bestsellers": ["B0088PUEPK"],
    "deals": [{"asin": "B0088PUEPK", "dealId": "deal-1", "rawDealField": {"price": 1299}}],
    "lightning-deals": [{"asin": "B0088PUEPK", "lightningStart": 7588958}],
}


def test_top_seller_string_rows_export_as_seller_ids(tmp_path: Path):
    export = export_rows_to_xlsx(
        rows=["A2L77EE7U53NWQ"],
        output_path=tmp_path / "top-sellers.xlsx",
        scenario="top-seller",
        site="US",
    )

    workbook = load_workbook(export.path)
    sheet = workbook.active

    assert sheet.cell(row=1, column=1).value == "Seller ID"
    assert sheet.cell(row=2, column=1).value == "A2L77EE7U53NWQ"


def test_product_search_string_rows_still_export_as_asins(tmp_path: Path):
    export = export_rows_to_xlsx(
        rows=["B0088PUEPK"],
        output_path=tmp_path / "asins.xlsx",
        scenario="product-search",
        site="US",
    )

    workbook = load_workbook(export.path)
    sheet = workbook.active

    assert sheet.cell(row=1, column=1).value == "ASIN"
    assert sheet.cell(row=2, column=1).value == "B0088PUEPK"


def test_known_headers_are_translated_and_unknown_headers_remain_raw(tmp_path: Path):
    export = export_rows_to_xlsx(
        rows=[{"asin": "B0088PUEPK", "title": "Test Product", "unknownField": "raw"}],
        output_path=tmp_path / "dynamic.xlsx",
        scenario="product",
        site="US",
    )

    workbook = load_workbook(export.path)
    sheet = workbook.active

    assert [cell.value for cell in sheet[1]][:3] == ["ASIN", "标题", "unknownField"]
    assert [sheet.cell(row=2, column=index).value for index in range(1, 4)] == ["B0088PUEPK", "Test Product", "raw"]


def test_nested_values_are_serialized_without_domain_formatting(tmp_path: Path):
    export = export_rows_to_xlsx(
        rows=[{"asin": "B0088PUEPK", "stats": {"current": [1299]}, "isAdultProduct": False}],
        output_path=tmp_path / "nested.xlsx",
        scenario="product",
        site="US",
    )

    workbook = load_workbook(export.path)
    sheet = workbook.active

    assert [cell.value for cell in sheet[1]][:3] == ["ASIN", "统计数据", "成人产品"]
    assert sheet.cell(row=2, column=2).value == '{"current": [1299]}'
    assert sheet.cell(row=2, column=3).value is False


def test_xlsx_export_writes_extra_sheets(tmp_path: Path):
    export = export_rows_to_xlsx(
        rows=[{"asin": "B0088PUEPK", "title": "Test Product"}],
        output_path=tmp_path / "product.xlsx",
        scenario="product",
        site="US",
        extra_sheets={
            "csv_history": [
                {
                    "asin": "B0088PUEPK",
                    "csvName": "AMAZON",
                    "priceAmount": 12.99,
                }
            ]
        },
    )

    workbook = load_workbook(export.path)

    assert "csv_history" in workbook.sheetnames
    sheet = workbook["csv_history"]
    assert [cell.value for cell in sheet[1]][:3] == ["ASIN", "csvName", "priceAmount"]
    assert sheet.cell(row=2, column=3).value == 12.99


def test_json_export_matches_formatted_xlsx_sheets(tmp_path: Path):
    rows = [{"asin": "B0088PUEPK", "title": "Test Product", "stats": {"current": [1299]}}]
    extra_sheets = {
        "csv_history": [{"asin": "B0088PUEPK", "csvName": "AMAZON", "priceAmount": 12.99}]
    }

    xlsx_export = export_rows_to_xlsx(
        rows=rows,
        output_path=tmp_path / "product.xlsx",
        scenario="product",
        site="US",
        extra_sheets=extra_sheets,
    )
    json_export = export_rows_to_json(
        rows=rows,
        output_path=tmp_path / "product.json",
        scenario="product",
        site="US",
        extra_sheets=extra_sheets,
    )

    workbook = load_workbook(xlsx_export.path, read_only=True, data_only=True)
    payload = json.loads(Path(json_export.path).read_text(encoding="utf-8"))

    assert payload["schema_version"] == "1.0"
    assert list(payload["sheets"]) == ["Sheet1", "Sheet2"]
    assert payload["sheets"]["Sheet1"]["name"] == workbook.worksheets[0].title
    assert payload["sheets"]["Sheet1"]["columns"] == list(next(workbook.worksheets[0].values))
    assert payload["sheets"]["Sheet1"]["rows"] == [list(row) for row in workbook.worksheets[0].values][1:]
    assert payload["sheets"]["Sheet2"]["name"] == "csv_history"
    assert payload["sheets"]["Sheet2"]["rows"] == [["B0088PUEPK", "AMAZON", 12.99]]
    assert json_export.format == "json"
    assert json_export.mime_type == "application/json"


def test_json_and_xlsx_use_the_same_unique_names_after_sheet_title_sanitizing(tmp_path: Path):
    extra_sheets = {
        "a/b": [{"asin": "B0088PUEPK"}],
        "ab": [{"asin": "B0088PUEPK"}],
    }
    xlsx_export = export_rows_to_xlsx(
        rows=[{"asin": "B0088PUEPK"}],
        output_path=tmp_path / "collision.xlsx",
        scenario="product",
        extra_sheets=extra_sheets,
    )
    json_export = export_rows_to_json(
        rows=[{"asin": "B0088PUEPK"}],
        output_path=tmp_path / "collision.json",
        scenario="product",
        extra_sheets=extra_sheets,
    )

    workbook = load_workbook(xlsx_export.path, read_only=True)
    payload = json.loads(Path(json_export.path).read_text(encoding="utf-8"))

    assert workbook.sheetnames == [sheet["name"] for sheet in payload["sheets"].values()]
    assert workbook.sheetnames[-2:] == ["ab", "ab1"]


def test_xlsx_export_uses_streaming_workbook(monkeypatch, tmp_path: Path):
    captured: dict[str, object] = {}

    class FakeSheet:
        def __init__(self, title: str):
            self.title = title
            self.rows = []
            self.freeze_panes = None
            self.column_dimensions = defaultdict(SimpleNamespace)

        def append(self, row):
            self.rows.append(row)

        def cell(self, *args, **kwargs):
            raise AssertionError("streaming export should append rows instead of writing cells one by one")

    class FakeWorkbook:
        def __init__(self, **kwargs):
            assert kwargs == {"write_only": True}
            captured["workbook_kwargs"] = kwargs
            self.sheets = []

        def create_sheet(self, title: str):
            sheet = FakeSheet(title)
            self.sheets.append(sheet)
            return sheet

        def save(self, path):
            captured["sheets"] = self.sheets
            Path(path).write_bytes(b"fake-xlsx")

    monkeypatch.setattr("openpyxl.Workbook", FakeWorkbook)
    monkeypatch.setattr(
        "openpyxl.cell.WriteOnlyCell",
        lambda sheet, value: SimpleNamespace(value=value, font=None, fill=None),
    )
    monkeypatch.setattr("openpyxl.utils.get_column_letter", lambda index: f"C{index}")

    export = export_rows_to_xlsx(
        rows=[{"asin": "B0088PUEPK", "title": "Test Product"}],
        output_path=tmp_path / "streaming.xlsx",
        scenario="product",
        site="US",
        extra_sheets={"details": [{"asin": "B0088PUEPK", "value": {"nested": True}}]},
    )

    sheets = captured["sheets"]
    assert export.filename == "streaming.xlsx"
    assert len(sheets) == 2
    assert sheets[0].rows[0][0].value == "ASIN"
    assert sheets[0].rows[1] == ["B0088PUEPK", "Test Product"]
    assert sheets[1].rows[1][1] == '{"nested": true}'


def test_all_keepa_scenarios_can_export_xlsx(tmp_path: Path):
    missing_samples = set(SCENARIOS) - set(SCENARIO_ROWS)
    assert missing_samples == set()

    for scenario in SCENARIOS:
        export = export_rows_to_xlsx(
            rows=SCENARIO_ROWS[scenario],
            output_path=tmp_path / f"{scenario}.xlsx",
            scenario=scenario,
            site="US",
        )
        workbook = load_workbook(export.path)
        sheet = workbook.active

        assert sheet.max_row == 2
        assert sheet.max_column >= 1
        assert sheet.cell(row=1, column=1).value
        assert sheet.cell(row=2, column=1).value is not None
