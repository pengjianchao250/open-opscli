from __future__ import annotations

from openpyxl import load_workbook

from opscli.keepa.export.xlsx import export_rows_to_xlsx
from opscli.keepa.tracking_formatter import (
    format_notification_export,
    format_tracking_export,
)


def test_tracking_formatter_formats_rules_notifications_and_unknown_nested_values():
    formatted = format_tracking_export(
        [
            {
                "asin": "B003IEUAZK",
                "createDate": 7588958,
                "mainDomainId": 1,
                "thresholdValues": [
                    {
                        "thresholdValueCSV": [7588958, 1999],
                        "domain": 1,
                        "csvType": 0,
                        "isDrop": True,
                    }
                ],
                "notifyIf": [{"domain": 1, "csvType": 0, "notifyIfType": 1}],
                "notificationType": [False, False, False, False, False, True, False],
                "notificationCSV": [1, 0, 5, 1, 7588960],
                "newNested": {"items": [{"code": "new"}]},
            }
        ],
        site="US",
    )

    row = formatted.trackings[0]
    assert row["createDateUtc"] == "2025-06-06T02:38:00Z"
    assert row["thresholdCount"] == 1
    assert row["notifyIfCount"] == 1
    assert row["notificationTypeJoined"] == "api"
    assert formatted.thresholds[0]["thresholdValue"] == 1999
    assert formatted.thresholds[0]["thresholdKeepaTimeUtc"] == "2025-06-06T02:38:00Z"
    assert formatted.notify_if[0]["notifyIfType"] == 1
    assert formatted.notification_csv[0]["notificationKeepaTimeUtc"] == "2025-06-06T02:40:00Z"
    assert formatted.notification_csv[0]["notificationCauseLabel"] == "DESIRED_PRICE"
    assert formatted.nested_values == [
        {"id": "B003IEUAZK", "field": "newNested", "path": "newNested.items[0].code", "value": "new"}
    ]
    assert all(
        not isinstance(item, (dict, list))
        for output in (formatted.trackings, formatted.thresholds, formatted.notify_if, formatted.notification_csv)
        for row in output
        for item in row.values()
    )


def test_notification_formatter_formats_prices_channels_images_and_time():
    formatted = format_notification_export(
        [
            {
                "asin": "B003IEUAZK",
                "title": "Example",
                "image": "61vXG3m1KWL.jpg",
                "createDate": 7588958,
                "domainId": 1,
                "notificationDomainId": 1,
                "currentPrices": [4899, 4750, -1, 15230, 6999, -1] + [None] * 13 + [125],
                "sentNotificationVia": [False, False, False, False, False, True, False],
                "newNested": {"source": "keepa"},
            }
        ],
        site="US",
    )

    row = formatted.notifications[0]
    assert row["imageUrl"].endswith("/61vXG3m1KWL.jpg")
    assert row["createDateUtc"] == "2025-06-06T02:38:00Z"
    assert row["sentNotificationViaJoined"] == "api"
    assert formatted.current_prices[0]["amount"] == 48.99
    assert formatted.current_prices[0]["currencyCode"] == "USD"
    shipping_row = next(item for item in formatted.current_prices if item["priceTypeIndex"] == 19)
    assert shipping_row["priceType"] == "USED_NEW_SHIPPING"
    assert shipping_row["amount"] == 1.25
    assert formatted.current_prices[2]["amount"] is None
    assert formatted.sent_via[5]["channel"] == "api"
    assert formatted.sent_via[5]["enabled"] is True
    assert formatted.nested_values == [
        {"id": "B003IEUAZK", "field": "newNested", "path": "newNested.source", "value": "keepa"}
    ]
    assert all(
        not isinstance(item, (dict, list))
        for output in (formatted.notifications, formatted.current_prices, formatted.sent_via)
        for row in output
        for item in row.values()
    )


def test_tracking_formatter_exports_all_detail_sheets(tmp_path):
    formatted = format_tracking_export(
        [
            {
                "asin": "B003IEUAZK",
                "thresholdValues": [{"thresholdValueCSV": [7588958, 1999]}],
                "notifyIf": [{"domain": 1, "csvType": 0, "notifyIfType": 1}],
                "notificationCSV": [1, 0, 5, 1, 7588960],
            }
        ]
    )
    export = export_rows_to_xlsx(
        rows=formatted.trackings,
        output_path=tmp_path / "tracking.xlsx",
        scenario="tracking",
        extra_sheets=formatted.extra_sheets(),
    )

    workbook = load_workbook(export.path)
    assert {
        "tracking_thresholds",
        "tracking_notify_if",
        "tracking_notification_csv",
    }.issubset(workbook.sheetnames)
