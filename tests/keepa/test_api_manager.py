import asyncio
import json
from pathlib import Path

from openpyxl import load_workbook

from opscli.keepa.accounts import KeepaApiKey
from opscli.keepa.config import KeepaSettings
from opscli.keepa.domain.models import KeepaScenarioRequest
from opscli.keepa.services import api_manager as api_manager_module
from opscli.keepa.services.api_manager import KeepaApiManager


def _run(coro):
    return asyncio.run(coro)


class DummyApiKeyProvider:
    def get_default(self, *, refresh=False):
        return KeepaApiKey(name="default", api_key="keepa-test-key", source="test")


class DummyKeepaClient:
    requests = []

    def __init__(self, *, api_key):
        self.api_key = api_key

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        return None

    async def token_status(self):
        return {
            "timestamp": 1000,
            "tokensLeft": 50,
            "refillIn": 300000,
            "refillRate": 5,
        }

    async def get_json(self, endpoint, params):
        self.__class__.requests.append({"endpoint": endpoint, "params": params})
        return {
            "timestamp": 2000,
            "tokensLeft": 49,
            "tokensConsumed": 1,
            "products": [{"asin": "B0088PUEPK", "title": "Test Product", "lastUpdate": 7588958}],
        }


class DisabledUploadClient:
    def __init__(self, *args, **kwargs):
        self.enabled = False


class DummyUploadClient:
    instances = []

    def __init__(self, *args, **kwargs):
        self.enabled = True
        self.uploads = []
        self.__class__.instances.append(self)

    def upload(self, path, *, purpose, folder=None, public=None, metadata=None):
        self.uploads.append(
            {
                "path": path,
                "purpose": purpose,
                "folder": folder,
                "public": public,
                "metadata": metadata,
            }
        )

        class Result:
            url = "https://ops.example.com/uploads/keepa.xlsx"

        return Result()


def test_manager_writes_params_raw_result_and_xlsx_export(monkeypatch, tmp_path: Path):
    DummyKeepaClient.requests = []
    monkeypatch.setattr(api_manager_module, "KeepaApiClient", DummyKeepaClient)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DisabledUploadClient)
    settings = KeepaSettings(output_dir=tmp_path, api_key=None, reserve_tokens=10)
    manager = KeepaApiManager(settings=settings, api_key_provider=DummyApiKeyProvider())

    result = _run(
        manager.run(
            KeepaScenarioRequest(
                scenario="product",
                site="US",
                params={"asin": "B0088PUEPK", "stats": 30, "history": False},
                job_id="keepa-offline-regression",
            )
        )
    )

    root_dir = tmp_path / "keepa-offline-regression"
    assert result.row_count == 1
    assert (root_dir / "params.json").exists()
    assert (root_dir / "raw.json").exists()
    assert (root_dir / "result.json").exists()
    assert result.export is not None
    assert result.export.path.endswith("keepa-offline-regression.xlsx")
    assert result.export.format == "xlsx"

    params_payload = json.loads((root_dir / "params.json").read_text(encoding="utf-8"))
    raw_payload = json.loads((root_dir / "raw.json").read_text(encoding="utf-8"))

    assert params_payload["normalized_params"]["asin"] == "B0088PUEPK"
    assert raw_payload["request_params"]["history"] is False
    assert result.data[0]["lastUpdateUtc"] == "2025-06-06T02:38:00Z"
    assert DummyKeepaClient.requests[0]["endpoint"] == "product"

    workbook = load_workbook(result.export.path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    assert headers[:3] == ["ASIN", "标题", "最近更新(Keepa分钟)"]
    assert sheet.cell(row=2, column=1).value == "B0088PUEPK"
    assert sheet.cell(row=2, column=2).value == "Test Product"
    assert sheet.cell(row=2, column=3).value == 7588958


def test_manager_submits_complete_success_result_to_collection_storage(
    monkeypatch, tmp_path: Path
):
    DummyKeepaClient.requests = []
    submissions = []
    monkeypatch.setattr(api_manager_module, "KeepaApiClient", DummyKeepaClient)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DisabledUploadClient)
    settings = KeepaSettings(output_dir=tmp_path, api_key=None, reserve_tokens=10)

    def submitter(*, request, result):
        assert Path(result.result_path).is_file()
        assert result.export is not None
        assert Path(result.export.path).is_file()
        submissions.append((request, result))
        return True

    manager = KeepaApiManager(
        settings=settings,
        api_key_provider=DummyApiKeyProvider(),
        collection_submitter=submitter,
    )
    request = KeepaScenarioRequest(
        scenario="product",
        site="US",
        params={"asin": "B0088PUEPK", "history": False},
        job_id="keepa-collection-submit",
    )

    result = _run(manager.run(request))

    assert submissions == [(request, result)]


def test_manager_uploads_export_to_keepa_export_folder(monkeypatch, tmp_path: Path):
    DummyKeepaClient.requests = []
    DummyUploadClient.instances = []
    monkeypatch.setattr(api_manager_module, "KeepaApiClient", DummyKeepaClient)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DummyUploadClient)
    settings = KeepaSettings(output_dir=tmp_path, api_key=None, reserve_tokens=10)
    manager = KeepaApiManager(settings=settings, api_key_provider=DummyApiKeyProvider())

    result = _run(
        manager.run(
            KeepaScenarioRequest(
                scenario="product",
                site="US",
                params={"asin": "B0088PUEPK", "stats": 30, "history": False},
                job_id="keepa-upload-regression",
            )
        )
    )

    upload = DummyUploadClient.instances[0].uploads[0]
    assert upload["purpose"] == "keepa_export"
    assert upload["folder"] == "keepa/export"
    assert upload["public"] == "1"
    assert upload["metadata"]["job_id"] == "keepa-upload-regression"
    assert upload["metadata"]["filename"] == "keepa-upload-regression.xlsx"
    assert result.export.url == "https://ops.example.com/uploads/keepa.xlsx"
    assert result.warnings == []


def test_manager_writes_formatted_json_export_when_requested(monkeypatch, tmp_path: Path):
    DummyKeepaClient.requests = []
    monkeypatch.setattr(api_manager_module, "KeepaApiClient", DummyKeepaClient)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DisabledUploadClient)
    settings = KeepaSettings(output_dir=tmp_path, api_key=None, reserve_tokens=10)
    manager = KeepaApiManager(settings=settings, api_key_provider=DummyApiKeyProvider())

    result = _run(
        manager.run(
            KeepaScenarioRequest(
                scenario="product",
                site="US",
                params={"asin": "B0088PUEPK", "stats": 30, "history": False},
                job_id="keepa-json-regression",
                export_format="json",
            )
        )
    )

    export_path = tmp_path / "keepa-json-regression" / "keepa-json-regression.json"
    payload = json.loads(export_path.read_text(encoding="utf-8"))

    assert DummyKeepaClient.requests
    assert result.export is not None
    assert result.export.path == str(export_path.resolve())
    assert result.export.format == "json"
    assert payload["sheets"]["Sheet1"]["columns"][:3] == [
        "ASIN",
        "标题",
        "最近更新(Keepa分钟)",
    ]
    assert payload["sheets"]["Sheet1"]["rows"][0][:3] == [
        "B0088PUEPK",
        "Test Product",
        7588958,
    ]


def test_product_finder_formats_search_insights_sheets(monkeypatch, tmp_path: Path):
    class ProductFinderClient(DummyKeepaClient):
        async def get_json(self, endpoint, params):
            self.__class__.requests.append({"endpoint": endpoint, "params": params})
            return {
                "timestamp": 2000,
                "tokensLeft": 49,
                "tokensConsumed": 1,
                "asinList": ["B0088PUEPK"],
                "searchInsights": {
                    "avgBuyBox": 1299,
                    "avgRating": 45,
                    "isFBAPercent": 78.3,
                    "relatedCategories": [172282],
                    "topBrandsWithCounts": {"Brand A": 10},
                    "topSellersWithCounts": {"A2L77EE7U53NWQ": 7},
                },
            }

    ProductFinderClient.requests = []
    monkeypatch.setattr(api_manager_module, "KeepaApiClient", ProductFinderClient)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DisabledUploadClient)
    settings = KeepaSettings(output_dir=tmp_path, api_key=None, reserve_tokens=10)
    manager = KeepaApiManager(settings=settings, api_key_provider=DummyApiKeyProvider())

    result = _run(
        manager.run(
            KeepaScenarioRequest(
                scenario="product-finder",
                site="US",
                params={
                    "stats": 1,
                    "queryName": "portable charger",
                    "current_SALES_gte": 1,
                    },
                    job_id="keepa-search-insights-regression",
                    force=True,
                )
        )
    )

    request_params = ProductFinderClient.requests[0]["params"]
    selection = json.loads(request_params["selection"])

    assert request_params["stats"] == 1
    assert "stats" not in selection
    assert selection["current_SALES_gte"] == 1

    workbook = load_workbook(result.export.path)
    assert "search_insights" in workbook.sheetnames
    assert "search_insight_brands" in workbook.sheetnames
    assert "search_insight_sellers" in workbook.sheetnames
    assert "search_insight_categories" in workbook.sheetnames
    main_headers = [cell.value for cell in workbook.active[1]]
    assert "searchInsights" not in main_headers

    sheet = workbook["search_insights"]
    headers = [cell.value for cell in sheet[1]]
    assert "avgBuyBoxAmount" in headers
    amount_column = headers.index("avgBuyBoxAmount") + 1
    rating_column = headers.index("avgRatingStars") + 1
    assert sheet.cell(row=2, column=amount_column).value == 12.99
    assert sheet.cell(row=2, column=rating_column).value == 4.5


def test_product_stats_formatting_sheets_are_exported(monkeypatch, tmp_path: Path):
    class ProductStatsClient(DummyKeepaClient):
        async def get_json(self, endpoint, params):
            self.__class__.requests.append({"endpoint": endpoint, "params": params})
            current = [-1] * 36
            current[0] = 1299
            current[16] = 45
            current[17] = 456
            return {
                "timestamp": 2000,
                "tokensLeft": 49,
                "tokensConsumed": 1,
                "products": [
                    {
                        "asin": "B0088PUEPK",
                        "title": "Stats Product",
                        "stats": {
                            "current": current,
                            "min": [[7588958, 999]],
                            "buyBoxStats": {
                                "A2L77EE7U53NWQ": {
                                    "avgPrice": 1299,
                                    "lastSeen": 7588958,
                                    "percentageWon": 80,
                                }
                            },
                            "retrievedOfferCount": 10,
                            "sellerIdsLowestFBA": ["A2L77EE7U53NWQ"],
                        },
                    }
                ],
            }

    ProductStatsClient.requests = []
    monkeypatch.setattr(api_manager_module, "KeepaApiClient", ProductStatsClient)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DisabledUploadClient)
    settings = KeepaSettings(output_dir=tmp_path, api_key=None, reserve_tokens=10)
    manager = KeepaApiManager(settings=settings, api_key_provider=DummyApiKeyProvider())

    result = _run(
        manager.run(
            KeepaScenarioRequest(
                scenario="product",
                site="US",
                params={"asin": "B0088PUEPK", "stats": 30},
                job_id="keepa-product-stats-regression",
            )
        )
    )

    workbook = load_workbook(result.export.path)
    assert "stats_price_types" in workbook.sheetnames
    assert "stats_extremes" in workbook.sheetnames
    assert "stats_buy_box_sellers" in workbook.sheetnames
    assert "stats_offer_snapshot" in workbook.sheetnames

    headers = [cell.value for cell in workbook.active[1]]
    price_column = headers.index("statsCurrentAmazonPrice") + 1
    rating_column = headers.index("statsCurrentRating") + 1
    assert workbook.active.cell(row=2, column=price_column).value == 12.99
    assert workbook.active.cell(row=2, column=rating_column).value == 4.5


def test_bestsellers_formats_ranked_asin_rows(monkeypatch, tmp_path: Path):
    class BestSellersClient(DummyKeepaClient):
        async def get_json(self, endpoint, params):
            self.__class__.requests.append({"endpoint": endpoint, "params": params})
            return {
                "timestamp": 2000,
                "tokensLeft": 49,
                "tokensConsumed": 1,
                "bestSellersList": {
                    "domainId": 1,
                    "categoryId": 172282,
                    "lastUpdate": 7588958,
                    "asinList": ["B000000001", "B000000002"],
                },
            }

    BestSellersClient.requests = []
    monkeypatch.setattr(api_manager_module, "KeepaApiClient", BestSellersClient)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DisabledUploadClient)
    settings = KeepaSettings(output_dir=tmp_path, api_key=None, reserve_tokens=10)
    manager = KeepaApiManager(settings=settings, api_key_provider=DummyApiKeyProvider())

    result = _run(
        manager.run(
            KeepaScenarioRequest(
                scenario="bestsellers",
                site="US",
                params={"category": "172282"},
                job_id="keepa-bestsellers-regression",
                force=True,
            )
        )
    )

    assert result.row_count == 2
    assert result.data[0]["bestSellerRank"] == 1
    assert result.data[1]["bestSellerRank"] == 2

    workbook = load_workbook(result.export.path)
    assert "best_sellers_list" in workbook.sheetnames
    headers = [cell.value for cell in workbook.active[1]]
    assert "bestSellerRank" in headers
    assert workbook.active.cell(row=2, column=headers.index("ASIN") + 1).value == "B000000001"
    assert workbook.active.cell(row=2, column=headers.index("bestSellerRank") + 1).value == 1


def test_deals_formats_metric_sheet(monkeypatch, tmp_path: Path):
    class DealsClient(DummyKeepaClient):
        async def get_json(self, endpoint, params):
            self.__class__.requests.append({"endpoint": endpoint, "params": params})
            return {
                "timestamp": 2000,
                "tokensLeft": 49,
                "tokensConsumed": 1,
                "deals": {
                    "dr": [
                        {
                            "asin": "B0088PUEPK",
                            "title": "<b>Deal Product</b>",
                            "image": [97, 98, 99, 46, 106, 112, 103],
                            "lastUpdate": 7588958,
                            "warehouseCondition": 3,
                            "current": [1299, 1399, -1, 12345] + [-1] * 12 + [45, 456, 1499],
                            "currentSince": [7588958, 7588959],
                            "deltaPercent": [[-8, -10, -1, 5]],
                        }
                    ]
                },
            }

    DealsClient.requests = []
    monkeypatch.setattr(api_manager_module, "KeepaApiClient", DealsClient)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DisabledUploadClient)
    settings = KeepaSettings(output_dir=tmp_path, api_key=None, reserve_tokens=10)
    manager = KeepaApiManager(settings=settings, api_key_provider=DummyApiKeyProvider())

    result = _run(
        manager.run(
            KeepaScenarioRequest(
                scenario="deals",
                site="US",
                params={"selection": {"page": 0}},
                job_id="keepa-deals-regression",
            )
        )
    )

    assert result.row_count == 1
    assert result.data[0]["titleText"] == "Deal Product"
    assert result.data[0]["currentAmazonPrice"] == 12.99
    assert result.data[0]["currentRating"] == 4.5

    workbook = load_workbook(result.export.path)
    assert "deal_metrics" in workbook.sheetnames
    headers = [cell.value for cell in workbook.active[1]]
    assert "imageUrl" in headers
    assert "currentAmazonPrice" in headers
    assert workbook.active.cell(row=2, column=headers.index("currentAmazonPrice") + 1).value == 12.99


def test_product_search_reuses_product_object_formatter(monkeypatch, tmp_path: Path):
    class ProductSearchClient(DummyKeepaClient):
        async def get_json(self, endpoint, params):
            return {
                "products": [
                    {
                        "asin": "B000000001",
                        "images": [{"variant": "MAIN", "l": "main.jpg"}],
                        "categoryTree": [{"catId": 1, "name": "Root"}],
                    }
                ]
            }

    monkeypatch.setattr(api_manager_module, "KeepaApiClient", ProductSearchClient)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DisabledUploadClient)
    manager = KeepaApiManager(
        settings=KeepaSettings(output_dir=tmp_path, api_key=None, reserve_tokens=10),
        api_key_provider=DummyApiKeyProvider(),
    )

    result = _run(
        manager.run(
            KeepaScenarioRequest(
                scenario="product-search",
                site="US",
                params={"term": "camera"},
                job_id="keepa-product-search-formatting",
                force=True,
            )
        )
    )

    assert result.data[0]["categoryPathName"] == "Root"
    assert "images" not in result.data[0]
    workbook = load_workbook(result.export.path)
    assert "images" in workbook.sheetnames
    assert "category_tree" in workbook.sheetnames


def test_product_search_asins_only_keeps_asin_rows(monkeypatch, tmp_path: Path):
    class AsinsOnlyClient(DummyKeepaClient):
        async def get_json(self, endpoint, params):
            return {"asinList": ["B000000001", "B000000002"]}

    monkeypatch.setattr(api_manager_module, "KeepaApiClient", AsinsOnlyClient)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DisabledUploadClient)
    manager = KeepaApiManager(
        settings=KeepaSettings(output_dir=tmp_path, api_key=None, reserve_tokens=10),
        api_key_provider=DummyApiKeyProvider(),
    )

    result = _run(
        manager.run(
            KeepaScenarioRequest(
                scenario="product-search",
                site="US",
                params={"term": "camera", "asins_only": True},
                job_id="keepa-product-search-asins",
                force=True,
            )
        )
    )

    assert result.data == ["B000000001", "B000000002"]
    workbook = load_workbook(result.export.path)
    assert workbook.active.cell(row=1, column=1).value == "ASIN"
    assert workbook.active.cell(row=2, column=1).value == "B000000001"


def test_category_seller_and_lightning_formatters_are_integrated(monkeypatch, tmp_path: Path):
    responses = {
        "category": {
            "categories": {
                "281052": {
                    "domainId": 1,
                    "catId": 281052,
                    "name": "Digital Cameras",
                    "children": [3017941],
                    "relatedCategories": [502394],
                    "topBrands": ["Sony"],
                }
            },
            "categoryParents": {
                "502394": {
                    "domainId": 1,
                    "catId": 502394,
                    "name": "Camera & Photo",
                    "children": [281052],
                }
            },
        },
        "seller": {
            "sellers": {
                "SELLER1": {
                    "sellerId": "SELLER1",
                    "ratingCount": [1, 2, 3, 4],
                    "positiveRating": [90, 91, 92, 93],
                    "asinList": ["B000000001"],
                    "asinListLastSeen": [7588958],
                }
            }
        },
        "lightningdeal": {
            "lightningDeals": [
                {
                    "asin": "B000000001",
                    "dealId": "deal-1",
                    "dealPrice": 1299,
                    "variation": [{"dimension": "Color", "value": "Black"}],
                }
            ]
        },
    }

    class ResponseObjectClient(DummyKeepaClient):
        async def get_json(self, endpoint, params):
            return responses[endpoint]

    monkeypatch.setattr(api_manager_module, "KeepaApiClient", ResponseObjectClient)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DisabledUploadClient)
    manager = KeepaApiManager(
        settings=KeepaSettings(output_dir=tmp_path, api_key=None, reserve_tokens=10),
        api_key_provider=DummyApiKeyProvider(),
    )
    requests = (
        ("category-lookup", {"category": "281052", "parents": True}, {"category_children", "category_related", "category_brands", "category_parents", "category_parent_children"}),
        ("seller", {"seller": "SELLER1", "storefront": True}, {"seller_ratings", "seller_storefront"}),
        ("lightning-deals", {"asin": "B000000001"}, {"lightning_variations"}),
    )

    for index, (scenario, params, expected_sheets) in enumerate(requests):
        result = _run(
            manager.run(
                KeepaScenarioRequest(
                    scenario=scenario,
                    site="US",
                    params=params,
                    job_id=f"keepa-response-object-{index}",
                    force=True,
                )
            )
        )
        workbook = load_workbook(result.export.path)
        assert expected_sheets <= set(workbook.sheetnames)


def test_manager_blocks_low_quota_without_force(monkeypatch, tmp_path: Path):
    class LowQuotaClient(DummyKeepaClient):
        async def token_status(self):
            return {"timestamp": 1000, "tokensLeft": 1, "refillIn": 300000, "refillRate": 5}

    monkeypatch.setattr(api_manager_module, "KeepaApiClient", LowQuotaClient)
    monkeypatch.setattr(api_manager_module, "FileUploadClient", DisabledUploadClient)
    settings = KeepaSettings(output_dir=tmp_path, api_key=None, reserve_tokens=10)
    manager = KeepaApiManager(settings=settings, api_key_provider=DummyApiKeyProvider())

    try:
        _run(
            manager.run(
                KeepaScenarioRequest(
                    scenario="product",
                    site="US",
                    params={"asin": "B0088PUEPK"},
                    job_id="low-quota",
                )
            )
        )
    except Exception as exc:
        assert "可用额度不足" in str(exc)
    else:
        raise AssertionError("expected quota precheck failure")

    assert (tmp_path / "low-quota" / "params.json").exists()
