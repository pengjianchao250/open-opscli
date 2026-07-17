from pathlib import Path

from openpyxl import load_workbook

from opscli.scrape_do.export.xlsx import extract_rows, export_rows_to_xlsx


def test_extract_rows_normalizes_pdp_payload():
    rows = extract_rows(
        "amazon-pdp",
        {
            "asin": "B0C7BKZ883",
            "brand": "Gogoonike",
            "name": "Laptop Stand",
            "price": 14.99,
            "list_price": 39.99,
            "currency": "USD",
            "rating": 4.6,
            "total_ratings": 2712,
            "is_prime": True,
            "best_seller_rankings": [{"category": "Laptop Stands", "rank": 2}],
            "technical_details": {"Material": "Aluminum"},
        },
    )

    assert rows == [
        {
            "asin": "B0C7BKZ883",
            "brand": "Gogoonike",
            "title": "Laptop Stand",
            "price": 14.99,
            "list_price": 39.99,
            "currency": "USD",
            "rating": 4.6,
            "total_ratings": 2712,
            "is_prime": True,
            "best_seller_rankings": '[{"category":"Laptop Stands","rank":2}]',
            "technical_details": '{"Material":"Aluminum"}',
        }
    ]


def test_extract_rows_normalizes_offer_listing_payload():
    rows = extract_rows(
        "amazon-offer-listing",
        {
            "asin": "B0DGJ7HYG1",
            "offers": [
                {
                    "sellerId": "SELLER1",
                    "merchantName": "6ave",
                    "condition": "New",
                    "listingPrice": {"currencyCode": "USD", "amount": 196.98},
                    "shipping": {"currencyCode": "USD", "amount": 0},
                    "isBuyBoxWinner": False,
                    "isFulfilledByAmazon": False,
                    "primeInformation": {"isPrime": False},
                    "quantity": 30,
                }
            ],
        },
    )

    assert rows[0]["asin"] == "B0DGJ7HYG1"
    assert rows[0]["seller_id"] == "SELLER1"
    assert rows[0]["listing_price"] == 196.98
    assert rows[0]["shipping_price"] == 0
    assert rows[0]["total_price"] == 196.98


def test_extract_rows_normalizes_search_payload():
    rows = extract_rows(
        "amazon-search",
        {
            "keyword": "laptop stands",
            "page": 2,
            "products": [
                {
                    "asin": "B0CBL1TQMP",
                    "title": "Portable Stand",
                    "price": {"currencyCode": "USD", "amount": 18.99},
                    "rating": {"value": 4.4, "count": 2340},
                    "isSponsored": True,
                    "isPrime": True,
                    "position": 2,
                    "badge": None,
                }
            ],
        },
    )

    assert rows[0]["keyword"] == "laptop stands"
    assert rows[0]["page"] == 2
    assert rows[0]["asin"] == "B0CBL1TQMP"
    assert rows[0]["price"] == 18.99
    assert rows[0]["rating"] == 4.4
    assert rows[0]["rating_count"] == 2340
    assert rows[0]["is_sponsored"] is True


def test_export_rows_to_xlsx_writes_file(tmp_path: Path):
    export = export_rows_to_xlsx(
        rows=[{"asin": "B0C7BKZ883", "title": "Laptop Stand"}],
        output_path=tmp_path / "job.xlsx",
        scenario="amazon-pdp",
        site="US",
        params={"asin": "B0C7BKZ883"},
    )

    assert Path(export.path).exists()
    assert export.filename == "job.xlsx"
    assert export.url.startswith("file:")


def test_export_rows_to_xlsx_writes_raw_fields_and_array_sheets(tmp_path: Path):
    export = export_rows_to_xlsx(
        rows=[{"asin": "B0C7BKZ883", "title": "Laptop Stand"}],
        output_path=tmp_path / "job.xlsx",
        scenario="amazon-pdp",
        site="US",
        params={"asin": "B0C7BKZ883"},
        raw_payload={
            "asin": "B0C7BKZ883",
            "name": "Laptop Stand",
            "rating": 4.6,
            "reviews": [
                {
                    "review_id": "R1",
                    "author": "Alice",
                    "date": "Reviewed in the United States on April 17, 2026",
                    "verified_purchase": True,
                }
            ],
            "images": ["https://example.com/1.jpg"],
            "technical_details": {"Material": "Aluminum"},
        },
    )

    workbook = load_workbook(export.path, read_only=True)

    assert "amazon-pdp-US" in workbook.sheetnames
    assert "Raw Fields" in workbook.sheetnames
    assert "Raw Reviews" in workbook.sheetnames
    assert "Raw Images" in workbook.sheetnames

    raw_fields = workbook["Raw Fields"]
    raw_field_rows = list(raw_fields.iter_rows(values_only=True))
    assert raw_field_rows[0] == ("字段", "类型", "值")
    assert ("asin", "str", "B0C7BKZ883") in raw_field_rows
    assert any(row[0] == "technical_details" and row[1] == "dict" and '"Material":"Aluminum"' in row[2] for row in raw_field_rows)

    raw_reviews = workbook["Raw Reviews"]
    review_rows = list(raw_reviews.iter_rows(values_only=True))
    assert review_rows[0] == ("asin", "review_id", "author", "date", "verified_purchase")
    assert review_rows[1] == ("B0C7BKZ883", "R1", "Alice", "Reviewed in the United States on April 17, 2026", True)

    raw_images = workbook["Raw Images"]
    image_rows = list(raw_images.iter_rows(values_only=True))
    assert image_rows[0] == ("asin", "value")
    assert image_rows[1] == ("B0C7BKZ883", "https://example.com/1.jpg")
