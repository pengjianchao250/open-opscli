"""共享采集结果 Parser 工具测试。"""

import pytest
from openpyxl import Workbook

from opscli.shared.collection_storage.parser_utils import (
    CollectionParseError,
    xlsx_datasets,
)


def test_xlsx_records_pad_trailing_empty_cells_omitted_by_openpyxl(tmp_path):
    export_path = tmp_path / "keepa.xlsx"
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("csv_history")
    worksheet.append(["ASIN", "rank", "count"])
    worksheet.append(["B0B56CHMSC", None, None])
    workbook.save(export_path)

    datasets = xlsx_datasets(
        export_path,
        source_name="Keepa",
        business_key_fields=("asin",),
    )

    records = tuple(datasets[0].records)

    assert records[0].payload == {
        "ASIN": "B0B56CHMSC",
        "rank": None,
        "count": None,
    }
    assert records[0].business_key == "B0B56CHMSC"


def test_xlsx_records_still_reject_values_beyond_declared_columns(tmp_path):
    export_path = tmp_path / "invalid.xlsx"
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("invalid")
    worksheet.append(["ASIN", "rank"])
    worksheet.append(["B0B56CHMSC", 1, "unexpected"])
    workbook.save(export_path)

    datasets = xlsx_datasets(export_path, source_name="Keepa")

    with pytest.raises(CollectionParseError, match="行列数量不一致"):
        tuple(datasets[0].records)
