"""Google Trends XLSX 导出测试。"""

import json
from pathlib import Path

from openpyxl import load_workbook

from opscli.google_trends.export.xlsx import export_rows_to_xlsx


def test_xlsx_keeps_serpapi_nested_fields(tmp_path: Path):
    """XLSX 应保留 Trending Now 的分类和趋势拆解字段。"""
    output_path = tmp_path / "trending.xlsx"

    export_rows_to_xlsx(
        rows=[
            {
                "rank": 1,
                "search_term": "game 7 nba finals",
                "search_volume": 2_000_000,
                "categories": [{"id": 17, "name": "Sports"}],
                "trend_breakdown": ["nba finals game 7", "game 7"],
            }
        ],
        output_path=output_path,
        scenario="trending-now",
        geo="US",
        params={},
    )

    sheet = load_workbook(output_path).active
    headers = [cell.value for cell in sheet[1]]
    values = dict(zip(headers, [cell.value for cell in sheet[2]]))

    assert json.loads(values["categories"])[0]["name"] == "Sports"
    assert json.loads(values["trend_breakdown"])[0] == "nba finals game 7"


def test_xlsx_uses_q_in_trends_sheet_title(tmp_path: Path):
    """Trends 工作表标题应包含新的 q 查询参数。"""
    output_path = tmp_path / "trends.xlsx"

    export_rows_to_xlsx(
        rows=[{"date": "2026-01-01", "flashlight": 42}],
        output_path=output_path,
        scenario="trends",
        geo="US",
        params={"q": "flashlight"},
    )

    assert "flashlight" in load_workbook(output_path).active.title
