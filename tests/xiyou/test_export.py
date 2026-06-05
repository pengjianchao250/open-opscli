from pathlib import Path

from openpyxl import load_workbook

from opscli.xiyou.export.xlsx import export_rows_to_xlsx


def test_xlsx_export_writes_known_and_extra_columns(tmp_path: Path):
    output = tmp_path / "xiyou.xlsx"

    result = export_rows_to_xlsx(
        rows=[
            {
                "flowRank": 1,
                "product": {
                    "asin": "B00TEST123",
                    "title": "Demo Product",
                    "price": 9.99,
                },
                "flow": {"score": 12345},
                "customField": "kept",
            }
        ],
        output_path=output,
        target="asin",
        site="US",
        period="week",
    )

    assert result.path == str(output.resolve())
    assert result.url == output.resolve().as_uri()
    assert output.exists()

    workbook = load_workbook(output)
    sheet = workbook.active
    assert sheet.title == "Xiyou-US-asin-week"
    assert sheet.cell(row=1, column=1).value == "排名"
    assert sheet.cell(row=1, column=2).value == "ASIN"
    assert sheet.cell(row=1, column=21).value == "customField"
    assert sheet.cell(row=2, column=2).value == "B00TEST123"
    assert sheet.cell(row=2, column=3).value == "Demo Product"
    assert sheet.cell(row=2, column=8).value == 9.99
    assert sheet.cell(row=2, column=12).value == 12345
    assert sheet.cell(row=2, column=21).value == "kept"


def test_xlsx_export_writes_keyword_headers_for_empty_rows(tmp_path: Path):
    output = tmp_path / "keyword.xlsx"

    export_rows_to_xlsx(
        rows=[],
        output_path=output,
        target="keyword",
        site="US",
        period="month",
    )

    workbook = load_workbook(output)
    sheet = workbook.active
    assert sheet.cell(row=1, column=1).value == "排名"
    assert sheet.cell(row=1, column=2).value == "关键词"
