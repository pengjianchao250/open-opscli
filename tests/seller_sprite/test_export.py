import json
from pathlib import Path

from openpyxl import load_workbook

from opscli.seller_sprite.export.keyword_comparison_xlsx import (
    build_keyword_comparison_worksheets,
    export_keyword_comparison_to_xlsx,
)
from opscli.seller_sprite.export.xlsx import build_export_worksheets, export_rows_to_xlsx


def test_real_time_bidding_export_merges_sp_sb_into_official_columns(
    tmp_path: Path,
):
    output = tmp_path / "real-time-bidding.xlsx"
    row = {
        "keyword": "phone stand",
        "keywordCn": "手机支架",
        "queryTime": "2025-06-30 00:00:00",
        "weekSearchNum": 56539,
        "rank": 2471,
        "displayNum": 1483771,
        "clickNum": 17832,
        "autoSponsor": {
            "EXACT": {"value": 0.53, "min": 0.42, "max": 0.68},
            "BROAD": {"value": 0.49, "min": 0.37, "max": 0.62},
            "PHRASE": {"value": 0.50, "min": 0.39, "max": 0.65},
        },
        "manualSponsor": {
            "EXACT": {"value": 0.47, "min": 0.35, "max": 0.60},
            "BROAD": {"value": 0.45, "min": 0.34, "max": 0.57},
            "PHRASE": {"value": 0.46, "min": 0.34, "max": 0.59},
        },
        "sponsorBrand": {
            "EXACT": {"value": 0.71, "min": 0.56, "max": 0.91},
            "BROAD": {"value": 0.67, "min": 0.51, "max": 0.84},
            "PHRASE": {"value": 0.69, "min": 0.54, "max": 0.88},
        },
        "sponsorBrandVideo": {
            "EXACT": {"value": 0.83, "min": 0.64, "max": 1.05},
            "BROAD": {"value": 0.78, "min": 0.60, "max": 0.99},
            "PHRASE": {"value": 0.80, "min": 0.62, "max": 1.02},
        },
        "abaConcentrationDegree": 0.2784,
        "cvsShareRate": 0.2127,
        "topClickAsin": [
            {"asin": "B07F8S18D5"},
            {"asin": "B092J6LZPF"},
            {"asin": "B0CMM89Y6Z"},
        ],
    }

    export_rows_to_xlsx(
        rows=[row, *({"keyword": f"keyword {index}"} for index in range(100))],
        output_path=output,
        scenario="real-time-bidding",
        site="US",
        params={"asin": "B07Z82895W"},
    )

    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == ["US-B07Z82895W-20250630000000"]
    sheet = workbook.active
    assert sheet.max_row == 101
    assert sheet.max_column == 46
    headers_path = (
        Path(__file__).resolve().parents[2]
        / "opscli/seller_sprite/reference/scenarios/real-time-bidding/official-headers.json"
    )
    assert [cell.value for cell in sheet[1]] == json.loads(
        headers_path.read_text(encoding="utf-8")
    )
    assert sheet["H2"].value == "$0.53"
    assert sheet["Z2"].value == "$0.71"
    assert sheet["AI2"].value == "$0.83"
    assert sheet["AR2"].value == 0.2784
    assert sheet["AS2"].value == 0.2127
    assert sheet["AT2"].value == "B07F8S18D5,B092J6LZPF,B0CMM89Y6Z"
    assert sheet["A1"].fill.fgColor.rgb == "FFE98A00"
    assert sheet["A1"].font.name == "宋体"
    assert sheet["A1"].font.color.rgb == "FFFFFFFF"
    assert "Notes" not in workbook.sheetnames


def test_keyword_conversion_rate_export_writes_first_page_business_sheet(
    tmp_path: Path,
):
    output = tmp_path / "keyword-conversion-rate.xlsx"
    rows = [
        {
            "keyword": "wireless charger stand",
            "keywordCn": "无线充电器支架",
            "weekIndex": 80,
            "searches": 3449,
            "clicks": 1172,
            "purchases": 180,
            "searchConvRate": 0.0522,
            "clickConvRate": 0.1536,
            "exactPpc": {"min": 1.04, "value": 1.33, "max": 1.86},
            "exactCpa": {"min": 6.77, "value": 8.66, "max": 12.11},
            "avgProductPrice": {"min": 8.54, "value": 24.99, "max": 159.99},
            "exactAcos": {"min": 1.0139, "value": 0.3465, "max": 0.0541},
            "exactBudget": {"value": 13.30},
            "clickingRate": 0.3069,
            "conversionRate": 0.2487,
            "top3Asins": [
                {
                    "asin": "B000000001",
                    "clickRate": 0.1284,
                    "conversionRate": 0.1295,
                },
                {
                    "asin": "B000000002",
                    "clickRate": 0.0929,
                    "conversionRate": 0.0674,
                },
                {
                    "asin": "B000000003",
                    "clickRate": 0.0856,
                    "conversionRate": 0.0518,
                },
            ],
            "gkDatas": [
                {"asin": f"B0000000{index:02d}"}
                for index in range(10)
            ],
        }
    ] + [
        {"keyword": f"keyword {index}"}
        for index in range(100)
    ]

    export_rows_to_xlsx(
        rows=rows,
        output_path=output,
        scenario="keyword-conversion-rate",
        site="US",
        period="W",
        params={"keywords": ["wireless charger stand"]},
    )

    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == ["US-wireless charger stand(100)"]
    sheet = workbook.active
    assert sheet.max_row == 101
    assert [cell.value for cell in sheet[1]] == [
        "关键词",
        "关键词翻译",
        "时间节点",
        "周搜索量",
        "周点击量",
        "周购买量",
        "搜索转化率",
        "点击转化率",
        "PPC竞价-最低($)",
        "PPC竞价-推荐($)",
        "PPC竞价-最高($)",
        "CPA-最低($)",
        "CPA-推荐($)",
        "CPA-最高($)",
        "产品均价-最低($)",
        "产品均价-平均($)",
        "产品均价-最高($)",
        "ACOS-最低",
        "ACOS-推荐",
        "ACOS-最高",
        "广告预算($)",
        "点击总占比",
        "转化总占比",
        "#1 前三ASIN",
        "#1 点击共享",
        "#1 转化共享",
        "#2 前三ASIN",
        "#2 点击共享",
        "#2 转化共享",
        "#3 前三ASIN",
        "#3 点击共享",
        "#3 转化共享",
        "搜索结果前10ASIN",
    ]
    assert sheet["A2"].value == "wireless charger stand"
    assert sheet["C2"].value == 80
    assert sheet["I2"].value == 1.04
    assert sheet["R2"].value == 1.0139
    assert sheet["X2"].value == "B000000001"
    assert sheet["Y2"].value == 0.1284
    assert sheet["AG2"].value == ",".join(
        f"B0000000{index:02d}" for index in range(10)
    )
    assert sheet["G2"].number_format == "0.00%"
    assert sheet["R2"].number_format == "0.00%"
    assert sheet["I2"].number_format == "#,##0.00_ "
    assert sheet.freeze_panes == "A2"
    assert "Notes" not in workbook.sheetnames


def test_traffic_extend_export_matches_three_sheet_business_contract(tmp_path: Path):
    output = tmp_path / "traffic-extend.xlsx"
    rows = [
        {
            "keywords": "phone stand",
            "keywordCn": "手机支架",
            "ac": 79,
            "trafficPercentage": 0.2737,
            "badges": ["NATURAL_SEARCHING", "ADS"],
            "calculatedWeeklySearches": 341698,
            "relationVariationsItems": [
                {"asin": "B089K9L3VY"},
                {"asin": "B07F8S18D5"},
            ],
            "searchesRank": 2488,
            "searches": 201870,
            "purchases": 8680,
            "purchaseRate": 0.043,
            "impressions": 4726178,
            "clicks": 77948,
            "cprExact": 197,
            "titleDensityExact": 28,
            "products": 93795,
            "supplyDemandRatio": 2.15,
            "latest7daysAds": 429,
            "top3ClickingRate": 0.2107,
            "top3ConversionRate": 0.1712,
            "bid": 1.12,
            "bidMin": 0.88,
            "bidMax": 1.31,
            "clickTop3s": [
                {"asin": "B000000001", "clickRate": 0.1, "conversionRate": 0.08},
                {"asin": "B000000002", "clickRate": 0.07, "conversionRate": 0.05},
                {"asin": "B000000003", "clickRate": 0.04, "conversionShareRate": 0.03},
            ],
            "gkDatas": [{"asin": "B000000001"}, {"asin": "B000000004"}],
        }
    ] + [{"keywords": f"phone accessory {index}"} for index in range(100)]

    export_rows_to_xlsx(
        rows=rows,
        output_path=output,
        scenario="traffic-extend",
        site="US",
        params={"asins": ["B089K9L3VY", "B07F8S18D5"]},
    )
    worksheets = build_export_worksheets(
        rows=rows,
        scenario="traffic-extend",
        site="US",
        params={"asins": ["B089K9L3VY", "B07F8S18D5"]},
    )

    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == [
        "US-B089K9L3VY(2)__",
        "Unique Words",
        "Asin",
    ]
    sheet = workbook["US-B089K9L3VY(2)__"]
    assert sheet.max_row == 101
    headers_path = (
        Path(__file__).resolve().parents[2]
        / "opscli/seller_sprite/reference/scenarios/traffic-extend/official-headers.json"
    )
    assert [cell.value for cell in sheet[1]] == json.loads(
        headers_path.read_text(encoding="utf-8")
    )
    assert sheet["A2"].value == "phone stand"
    assert sheet["G2"].value == 2
    assert sheet["H2"].value == "B089K9L3VY,B07F8S18D5"
    assert sheet["V2"].value == "$1.12"
    assert sheet["W2"].value == "$0.88-$1.31"
    assert sheet["D2"].number_format == "0.00%"
    assert sheet["L2"].number_format == "0.00%"
    assert sheet["A1"].font.color.rgb == "FFFFFFFF"
    assert sheet.column_dimensions["A"].width == 26.5044247787611
    assert sheet.column_dimensions["AG"].width == 118.353982300885
    assert sheet["X2"].value == "B000000001"
    assert sheet["AF2"].value == 0.03
    assert sheet["AG2"].value == "B000000001,B000000004"

    unique_words = workbook["Unique Words"]
    assert [cell.value for cell in unique_words[1]][:3] == ["词语", "出现频次", "百分比"]
    assert unique_words["A2"].value == "phone"
    assert unique_words["B2"].value == 100
    assert unique_words["C2"].number_format == "0.00%"
    assert workbook["Asin"]["A1"].value == "ASIN"
    assert workbook["Asin"]["A2"].value == "B089K9L3VY"
    assert workbook["Asin"]["A3"].value == "B07F8S18D5"
    assert "Notes" not in workbook.sheetnames
    assert [worksheet.name for worksheet in worksheets] == workbook.sheetnames
    assert worksheets[0].columns == [cell.value for cell in sheet[1]]
    assert worksheets[0].rows[0] == [cell.value for cell in sheet[2]]
    assert worksheets[1].rows[0][:2] == [
        unique_words["A2"].value,
        unique_words["B2"].value,
    ]
    assert round(worksheets[1].rows[0][2], 15) == round(unique_words["C2"].value, 15)


def test_xlsx_export_writes_template_headers_without_notes(tmp_path: Path):
    output = tmp_path / "seller-sprite.xlsx"

    result = export_rows_to_xlsx(
        rows=[
            {
                "asin": "B00TEST123",
                "title": "Demo Product",
                "amazonChoice": True,
                "coupon": "5%",
            }
        ],
        output_path=output,
        scenario="competitor-lookup",
        site="US",
    )

    assert result.path == str(output)
    assert result.url == output.resolve().as_uri()
    assert output.exists()

    workbook = load_workbook(output)
    sheet = workbook.active
    assert sheet.cell(row=1, column=1).value == "ASIN"
    assert sheet.cell(row=1, column=6).value == "商品标题"
    assert sheet.cell(row=1, column=25).value == "Coupon"
    assert sheet.cell(row=1, column=54).value == "AC关键词"
    assert sheet.cell(row=2, column=1).value == "B00TEST123"
    assert sheet.cell(row=2, column=25).value == "5%"
    assert "Notes" not in workbook.sheetnames


def test_keyword_export_writes_unique_words_sheet(tmp_path: Path):
    output = tmp_path / "keyword.xlsx"

    export_rows_to_xlsx(
        rows=[{"keyword": "flashlight", "keywordCn": "手电筒", "amazonChoice": True}],
        output_path=output,
        scenario="keyword-miner",
        site="JP",
        params={"keyword": "flashlight"},
        high_frequency_rows=[{"keyword": "flashlight", "frequency": 10, "percentage": 0.5}],
    )

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["JP-flashlight(1)_", "Unique Words"]
    unique_words = workbook["Unique Words"]
    assert unique_words.cell(row=1, column=1).value == "词语"
    assert unique_words.cell(row=2, column=1).value == "flashlight"
    assert unique_words.cell(row=2, column=2).value == 10


def test_keyword_comparison_export_builds_dynamic_business_sheets(tmp_path: Path):
    output = tmp_path / "CompareKeywords-US-B0949DWJCV.xlsx"
    rows = [
        {
                "keyword": "phone stand",
                "keywordCn": "手机支架",
                "competitors": 3,
                "searchesRank": 2488,
                "searches": 201870,
                "purchases": 8680,
                "purchaseRate": 0.043,
                "impressions": 99123,
                "clicks": 4321,
                "products": 93795,
                "supplyDemandRatio": 2.15,
                "competitorList": [
                    {
                        "asin": "B0949DWJCV",
                        "trafficPercentage": 0.0686,
                        "trafficKeywordTypes": ["PRIMARY"],
                    },
                    {
                        "asin": "B0744DM3Y3",
                        "trafficPercentage": 0.00001,
                        "trafficKeywordTypes": ["PRECISE", "PRECISE_LONG_TAIL"],
                    },
                ],
        }
    ]

    export_keyword_comparison_to_xlsx(
        rows=rows,
        output_path=output,
        site="US",
        own_asin="B0949DWJCV",
        asin_list=["B0949DWJCV", "B0744DM3Y3"],
    )

    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "US-流量占比对比-B0949DWJCVB0744DM3Y3",
        "ASIN",
    ]
    assert "Notes" not in workbook.sheetnames
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == [
        "关键词",
        "关键词翻译",
        "B0949DWJCV(我的)",
        "B0949DWJCV流量词类型",
        "B0744DM3Y3",
        "B0744DM3Y3流量词类型",
        "有效竞品数",
        "ABA排名(周)",
        "月搜索量",
        "月购买量",
        "购买率",
        "展示量",
        "点击量",
        "商品数",
        "需供比",
    ]
    assert sheet["C2"].value == "6.86%"
    assert sheet["D2"].value == "主要流量词"
    assert sheet["E2"].value == "<0.01%"
    assert sheet["F2"].value == "精准流量词、精准长尾词"
    assert sheet["K2"].value == 0.043
    assert sheet.freeze_panes is None
    assert sheet.auto_filter.ref is None
    assert sheet.sheet_view.showGridLines is False
    assert sheet.row_dimensions[1].height == 20
    assert sheet.row_dimensions[2].height == 20
    assert sheet.column_dimensions["A"].width == 25
    assert sheet["A1"].fill.fgColor.rgb == "FFE98A00"
    assert sheet["A1"].font.name == "等线"
    assert sheet["A1"].font.sz == 10
    assert sheet.merged_cells.ranges == set()
    asin_sheet = workbook["ASIN"]
    assert [asin_sheet.cell(row=index, column=1).value for index in range(1, 5)] == [
        "asin",
        "B0949DWJCV(我的)",
        "B0949DWJCV",
        "B0744DM3Y3",
    ]
    worksheets = build_keyword_comparison_worksheets(
        rows=rows,
        site="US",
        own_asin="B0949DWJCV",
        asin_list=["B0949DWJCV", "B0744DM3Y3"],
    )
    assert [worksheet.name for worksheet in worksheets] == workbook.sheetnames
    assert worksheets[0].columns == [cell.value for cell in sheet[1]]
    assert worksheets[0].rows[0] == [cell.value for cell in sheet[2]]
    assert worksheets[1].rows == [
        ["B0949DWJCV(我的)"],
        ["B0949DWJCV"],
        ["B0744DM3Y3"],
    ]


def test_keyword_research_export_matches_required_workbook_contract(tmp_path: Path):
    output = tmp_path / "keyword-research.xlsx"
    row = {
        "keyword": "fathers day gifts",
        "keywordCn": "父亲节礼物",
        "searchRank": 1,
        "searches": 10124360,
        "searchesCr": 4.9992,
        "purchases": 443447,
        "purchaseRate": 0.0438,
        "impressions": 267193010,
        "clicks": 4135940,
        "products": 1566,
        "supplyDemandRatio": 6465.11,
        "spr": 10052,
        "titleDensity": 0,
        "monopolyClickRate": 0.2878,
        "cvsShareRate": 0.0129,
        "goodsValue": 0.171,
        "avgPrice": "$19.99",
        "avgReviews": 20341,
        "avgRating": 4.5,
        "bidMin": 1.03,
        "bid": 1.32,
        "bidMax": 1.6,
        "searchMonthCv": 6347161,
        "searchMonthCr": 1.6804,
        "searchNearlyCv": 119308,
        "searchNearlyCr": 2.5434,
        "departments": "Clothing, Shoes & Jewelry; Home & Kitchen; Garden & Outdoor",
        "gkDatas": [{"asin": "B0H6WT6Q8C"}, {"asin": "B0H4G1XJQD"}],
    }

    export_rows_to_xlsx(
        rows=[row],
        output_path=output,
        scenario="keyword-research",
        site="US",
        period="2026-06",
    )

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Keywords(1)"]
    sheet = workbook["Keywords(1)"]
    assert [cell.value for cell in sheet[1]] == [
        "关键词", "关键词翻译", "ABA排名", "月搜索量", "搜索增长率", "月购买量", "购买率",
        "展示量", "点击量", "商品数", "需供比", "SPR", "标题密度", "点击总占比", "转化总占比",
        "货流值", "均价", "评分数", "评分值", "PPC竞价-最低($)", "PPC竞价-推荐($)",
        "PPC竞价-最高($)", "同比增长值", "同比增长率", "近3个月增长值", "近3个月增长率",
        "所属类目", "前10ASIN",
    ]
    assert sheet.freeze_panes == "A2"
    assert sheet["A1"].fill.fgColor.rgb == "FFE98A00"
    assert sheet["A1"].font.name == "Calibri"
    assert sheet["A1"].font.sz == 10
    assert sheet["A1"].font.bold is False
    assert sheet.column_dimensions["A"].width == 28
    assert sheet.column_dimensions["AB"].width == 120
    assert sheet["G2"].value == 0.0438
    assert sheet["Q2"].value == "$19.99"
    assert sheet["AB2"].value == "B0H6WT6Q8C,B0H4G1XJQD"


def test_aba_research_export_matches_official_main_sheet_without_notes(tmp_path: Path):
    output = tmp_path / "aba-research.xlsx"
    row = {
        "keyword": "obsession",
        "keywordCn": "痴迷",
        "searches": 34822,
        "searchRank": 1,
        "w1SearchRank": 1,
        "w4SearchRank": 506,
        "w12SearchRank": 74457,
        "w1RankGrowthValue": 0,
        "w4RankGrowthValue": 505,
        "w12RankGrowthValue": 74456,
        "w1RankGrowthRate": 0,
        "w4RankGrowthRate": 99.8,
        "w12RankGrowthRate": 100,
        "bid": 1.23,
        "bidMin": 0.84,
        "bidMax": 1.62,
        "impressions": 150000,
        "clicks": 32000,
        "cprExact": 12,
        "titleDensityExact": 4,
        "top3AsinDtoList": [
            {"asin": "B000000001", "clickRate": 93.81, "conversionRate": 92.5},
            {"asin": "B000000002", "clickRate": 1.78, "conversionShareRate": 2.25},
            {"asin": "B000000003", "clickRate": 1.64, "conversionRate": 1.5},
        ],
        "top3Brands": [{"brand": "Brand A"}, {"brand": "Brand B"}],
        "departments": [{"label": "Home & Kitchen"}, {"label": "Office Products"}],
        "gkDatas": [{"asin": "B000000001"}, {"asin": "B000000004"}],
    }

    export_rows_to_xlsx(
        rows=[row],
        output_path=output,
        scenario="aba-research",
        site="US",
        period="2026第29周(07/12~07/18)",
    )

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["ABAKeyword(1)"]
    sheet = workbook["ABAKeyword(1)"]
    assert [cell.value for cell in sheet[1]] == [
        "关键词", "关键词翻译", "周搜索量", "现排名", "历史排名", "周变化量", "周变化率",
        "PPC价格", "建议竞价范围", "展示量", "点击量", "SPR", "标题密度", "点击占比",
        "转化占比", "点击前三ASIN", "点击前三品牌", "所属类目", "前10ASIN",
    ]
    assert sheet.max_column == 19
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is None
    assert sheet["A1"].fill.fgColor.rgb == "FFE98A00"
    assert sheet["A1"].font.color.rgb == "FFFFFFFF"
    assert sheet.row_dimensions[1].height == 20
    assert sheet.row_dimensions[2].height == 20
    assert sheet.column_dimensions["A"].width == 18.3362831858407
    assert sheet.column_dimensions["S"].width == 120
    assert sheet["C2"].value == 34822
    assert sheet["C2"].number_format == "#,##0_ "
    assert sheet["E2"].value == "上周: 1\n4周前: 506\n12周前: 74,457"
    assert sheet["F2"].value == "上周: 0\n4周前: +505\n12周前: +74,456"
    assert sheet["G2"].value == "上周: 0.00%\n4周前: +99.80%\n12周前: +100.00%"
    assert sheet["H2"].value == "$1.23"
    assert sheet["I2"].value == "$0.84 - $1.62"
    assert sheet["N2"].value == "TOP1: 93.81%\nTOP2: 1.78%\nTOP3: 1.64%\n合计: 97.23%"
    assert sheet["O2"].value == "TOP1: 92.50%\nTOP2: 2.25%\nTOP3: 1.50%\n合计: 96.25%"
    assert sheet["P2"].value == "B000000001、B000000002、B000000003"
    assert sheet["Q2"].value == "Brand A、Brand B"
    assert sheet["R2"].value == "Home & Kitchen; Office Products"
    assert sheet["S2"].value == "B000000001,B000000004"


def test_association_traffic_export_matches_official_main_sheet_without_notes(tmp_path: Path):
    output = tmp_path / "association-traffic.xlsx"
    asins = ["B098T9ZFB5", "B09JW5FNVX", "B0B71DH45N", "B07MHHM31K", "B08RYQR1CJ"]

    export_rows_to_xlsx(
        rows=[
            {
                "asin": "B0D9XRB6YF",
                "count": 1,
                "relationAsinDtoList": [{"asin": "B0DKH8RXV2"}],
                "relationList": ["VAV"],
                "sku": "Color: Green",
                "brand": "Narwey",
                "title": "Demo Product",
                "bigImageUrl": "https://example.com/demo.jpg",
                "parent": "B0FFH4FL98",
                "nodeLabelPath": "Beauty & Personal Care:Toiletry Bags",
                "bsrLabel": "Beauty & Personal Care",
                "bsrId": "3760911",
                "marketId": 1,
                "bsrRank": 2305,
                "bsrRankCv": 803,
                "bsrRankCr": 25.84,
                "subcategories": [{"label": "Toiletry Bags", "rank": 9}],
                "totalUnits": 7151,
                "totalUnitsGrowth": 1.35,
                "totalAmount": 121495.49,
                "amzUnit": 1000,
                "subTotalAmount": 19570,
                "price": 16.99,
                "questions": 34,
                "profit": 52.51,
                "fba": 5.52,
                "reviews": 4049,
                "reviewsRate": 1.43,
                "rating": 4.7,
                "reviewsIncreasement": 102,
                "availableDate": 1724706720000,
                "sellerType": "FBA",
                "deliveryPrice": -1,
                "lqs": 100,
                "variations": 40,
                "sellers": 1,
                "sellerName": "Narwey®",
                "sellerNation": "US",
                "sellerDto": {"businessAddress": "UPPER MARLBORO<br/>MD<br/>20774<br/>US"},
                "ebc": "Y",
                "video": "Y",
                "weight": "0.79 pounds",
                "weightTag": "358.34 g",
                "dimensions": "10.8 x 5.5 x 8.1 inches",
                "dimensionsTag": "27.43 x 13.97 x 20.57 cm",
                "pkgWeight": "0.93 pounds",
                "pkgWeightTag": "421.84 g",
                "pkgDimensions": "11.6 x 8.7 x 1.9 inches",
                "pkgDimensionsTag": "29.46 x 22.10 x 4.83 cm",
                "pkgDimensionType": "大号标准尺寸",
                "createdTime": 1784766462000,
            }
        ],
        output_path=output,
        scenario="association-traffic",
        site="US",
        params={"asins": asins},
    )

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Related-B098T9ZFB5-batch(5)(31"]
    sheet = workbook.active
    assert sheet.max_column == 56
    headers_path = (
        Path(__file__).resolve().parents[2]
        / "opscli/seller_sprite/reference/scenarios/association-traffic/official-headers.json"
    )
    official_headers = json.loads(headers_path.read_text(encoding="utf-8"))
    assert [cell.value for cell in sheet[1]] == official_headers
    assert [sheet.cell(row=1, column=index).value for index in (20, 22, 23, 26, 33)] == [
        "月销售额($)",
        "子体销售额($)",
        "价格($)",
        "FBA运费($)",
        "买家运费($)",
    ]
    assert sheet.freeze_panes == "A2"
    assert sheet["A1"].fill.fgColor.rgb == "FFE98A00"
    assert sheet.column_dimensions["A"].width == 14
    assert sheet.column_dimensions["G"].width == 35
    assert sheet.column_dimensions["BD"].width == 12.7610619469027
    assert sheet["A2"].hyperlink.target == "https://www.amazon.com/dp/B0D9XRB6YF"
    assert sheet["J2"].hyperlink.target == "https://www.amazon.com/dp/B0FFH4FL98"
    assert sheet["L2"].hyperlink.target == "https://www.amazon.com/gp/bestsellers/beauty"
    assert sheet["C2"].value == "B0DKH8RXV2"
    assert sheet["D2"].value == "看了又看"
    assert sheet["O2"].value == 0.2584
    assert sheet["S2"].value == 0.0135
    assert sheet["Y2"].value == 0.5251
    assert sheet["AB2"].value == 0.0143
    assert sheet["AH2"].value == "10.0"
    assert sheet["AM2"].value == "UPPER MARLBORO MD 20774 US"
    assert sheet["BD2"].value == "2026-07-23"


def test_association_traffic_export_keeps_official_batch_sheet_title(tmp_path: Path):
    output = tmp_path / "association-traffic-375.xlsx"

    export_rows_to_xlsx(
        rows=[{"asin": "B0D9XRB6YF"} for _ in range(375)],
        output_path=output,
        scenario="association-traffic",
        site="US",
        params={
            "asins": [
                "B098T9ZFB5",
                "B09JW5FNVX",
                "B0B71DH45N",
                "B07MHHM31K",
                "B08RYQR1CJ",
            ]
        },
    )

    workbook = load_workbook(output, read_only=True)
    assert workbook.sheetnames == ["Related-B098T9ZFB5-batch(5)(31"]
