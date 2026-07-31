"""SellerSprite 官网 XLSX golden fixtures 完整性测试。"""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


FIXTURE_ROOT = (
    Path(__file__).resolve().parents[1]
    / "fixtures"
    / "seller_sprite"
    / "official_exports"
)


def test_official_export_fixture_index_matches_immutable_files():
    """清单必须覆盖全部官方原件，并固定大小、哈希和 XLSX 基础结构。"""
    index = json.loads((FIXTURE_ROOT / "index.json").read_text(encoding="utf-8"))
    entries = index["fixtures"]
    indexed_paths = {entry["path"] for entry in entries}
    actual_paths = {
        path.relative_to(FIXTURE_ROOT).as_posix()
        for path in FIXTURE_ROOT.rglob("*.xlsx")
    }

    assert index["immutable"] is True
    assert indexed_paths == actual_paths
    assert len(entries) == 5

    for entry in entries:
        fixture_path = FIXTURE_ROOT / entry["path"]
        content = fixture_path.read_bytes()

        assert len(content) == entry["size"]
        assert hashlib.sha256(content).hexdigest() == entry["sha256"]
        with ZipFile(fixture_path) as archive:
            assert archive.testzip() is None
            assert "[Content_Types].xml" in archive.namelist()
            assert "xl/workbook.xml" in archive.namelist()


def test_real_time_bidding_official_fixture_defines_export_contract():
    """官方原件必须直接约束业务表名称、46 列表头和代表性值类型。"""
    fixture_path = (
        FIXTURE_ROOT
        / "real-time-bidding"
        / "cpcSuggestBid-US-B07Z82895W(1)-20260731162033.xlsx"
    )
    headers_path = (
        Path(__file__).resolve().parents[2]
        / "opscli"
        / "seller_sprite"
        / "reference"
        / "scenarios"
        / "real-time-bidding"
        / "official-headers.json"
    )
    workbook = load_workbook(fixture_path, data_only=True, read_only=True)

    assert workbook.sheetnames == [
        "US-B07Z82895W-20250630000000",
        "Notes",
    ]
    sheet = workbook[workbook.sheetnames[0]]
    assert sheet.max_row == 198
    assert sheet.max_column == 46
    assert [cell.value for cell in sheet[1]] == json.loads(
        headers_path.read_text(encoding="utf-8")
    )
    assert isinstance(sheet["A2"].value, str)
    assert isinstance(sheet["C2"].value, str)
    assert isinstance(sheet["D2"].value, float)
    assert isinstance(sheet["H2"].value, str)
    assert isinstance(sheet["AR2"].value, float)
    assert isinstance(sheet["AT2"].value, str)
