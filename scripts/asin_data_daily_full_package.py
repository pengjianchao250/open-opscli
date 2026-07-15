from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import json
import re
import shutil
import subprocess
import sys
import time
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Iterable
from uuid import NAMESPACE_URL, uuid5
from zoneinfo import ZoneInfo

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover - import error is surfaced when needed.
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = Path("output") / "asin-data"
DEFAULT_SELLER_SKILL_SCRIPT = (
    Path.home()
    / ".codex"
    / "skills"
    / "ops-seller-sprite-asin-keyword-pack"
    / "scripts"
    / "seller_sprite_asin_keyword_pack.py"
)
DEFAULT_RUFUS_SKILL_ROOT = Path.home() / ".codex" / "skills" / "ops-rufus-listing-score-reask"
DEFAULT_RUFUS_BATCH_SCRIPT = DEFAULT_RUFUS_SKILL_ROOT / "scripts" / "run_rufus_batch.py"
DEFAULT_RUFUS_SCORE_SCRIPT = DEFAULT_RUFUS_SKILL_ROOT / "scripts" / "score_rufus_report.py"
DEFAULT_SENSITIVE_SCRIPT = DEFAULT_RUFUS_SKILL_ROOT / "scripts" / "sensitive_terms.py"

PACKAGE_ROOT_NAME = "asin-data-packages"
FILE_BASIC = "01-\u57fa\u7840\u6570\u636e.xlsx"
FILE_BI = "02-BI\u6570\u636e.xlsx"
FILE_SELLER_REVERSE = "03-\u5356\u5bb6\u7cbe\u7075\u5173\u952e\u8bcd\u53cd\u67e5\u6570\u636e.xlsx"
FILE_SELLER_MINER_01 = "04-\u5356\u5bb6\u7cbe\u7075\u5173\u952e\u8bcd\u6316\u639801.xlsx"
FILE_SELLER_MINER_02 = "04-\u5356\u5bb6\u7cbe\u7075\u5173\u952e\u8bcd\u6316\u639802.xlsx"
FILE_SELLER_COMPETITOR_01 = "05-\u5356\u5bb6\u7cbe\u7075\u7ade\u54c101.xlsx"
FILE_SELLER_COMPETITOR_02 = "05-\u5356\u5bb6\u7cbe\u7075\u7ade\u54c102.xlsx"
FILE_RUFUS = "06-Rufus\u6570\u636e\u5206\u6790.md"

SELLER_FILES = [
    FILE_SELLER_REVERSE,
    FILE_SELLER_MINER_01,
    FILE_SELLER_MINER_02,
    FILE_SELLER_COMPETITOR_01,
    FILE_SELLER_COMPETITOR_02,
]

OUTPUT_FILES = [
    FILE_BASIC,
    FILE_BI,
    *SELLER_FILES,
    FILE_RUFUS,
]


@dataclass(frozen=True)
class CacheHit:
    asin: str
    source: str
    path: Path
    file_paths: dict[str, Path]
    newest_mtime: float
    oldest_mtime: float


def normalize_asin(value: str) -> str:
    match = re.search(r"\b[A-Z0-9]{10}\b", value.strip().upper())
    if not match:
        raise ValueError(f"Invalid ASIN: {value}")
    return match.group(0)


def now_run_id(prefix: str = "asin-data-daily-full") -> str:
    return f"{prefix}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"


def default_report_date() -> str:
    return datetime.now(ZoneInfo("Asia/Shanghai")).date().isoformat()


def load_asins(input_path: Path, asin_column: str = "asin") -> list[str]:
    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        return load_asins_csv(input_path, asin_column)
    if suffix == ".json":
        data = json.loads(input_path.read_text(encoding="utf-8-sig"))
        if isinstance(data, list):
            rows = data
        elif isinstance(data, dict):
            rows = data.get("rows") or data.get("data") or []
        else:
            rows = []
        return unique_asins(row.get(asin_column) or row.get("ASIN") for row in rows if isinstance(row, dict))
    if suffix == ".jsonl":
        rows = []
        for line in input_path.read_text(encoding="utf-8-sig").splitlines():
            if line.strip():
                rows.append(json.loads(line))
        return unique_asins(row.get(asin_column) or row.get("ASIN") for row in rows if isinstance(row, dict))
    if suffix in {".xlsx", ".xlsm"}:
        return load_asins_xlsx(input_path, asin_column)
    raise ValueError(f"Unsupported input file type: {input_path}")


def load_asins_csv(input_path: Path, asin_column: str) -> list[str]:
    with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            return []
        column = find_column(reader.fieldnames, asin_column)
        return unique_asins(row.get(column, "") for row in reader)


def load_asins_xlsx(input_path: Path, asin_column: str) -> list[str]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read XLSX input files")
    wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows, [])]
        column = find_column(headers, asin_column)
        index = headers.index(column)
        return unique_asins(row[index] if index < len(row) else "" for row in rows)
    finally:
        wb.close()


def find_column(headers: Iterable[str], wanted: str) -> str:
    normalized = {str(header).strip().lower(): str(header).strip() for header in headers}
    for candidate in (wanted, "asin", "ASIN"):
        value = normalized.get(candidate.lower())
        if value:
            return value
    raise ValueError(f"ASIN column not found. Expected {wanted!r}; got {list(headers)!r}")


def unique_asins(values: Iterable[Any]) -> list[str]:
    seen: set[str] = set()
    asins: list[str] = []
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        asin = normalize_asin(text)
        if asin not in seen:
            seen.add(asin)
            asins.append(asin)
    return asins


def cutoff_time(cache_days: int) -> float:
    return (datetime.now() - timedelta(days=cache_days)).timestamp()


def is_recent(path: Path, cutoff: float) -> bool:
    try:
        return path.stat().st_mtime >= cutoff
    except FileNotFoundError:
        return False


def find_recent_file_set(
    *,
    roots: list[Path],
    asin: str,
    filenames: list[str],
    cache_days: int,
    source: str,
) -> CacheHit | None:
    cutoff = cutoff_time(cache_days)
    candidates: list[CacheHit] = []
    for root in roots:
        if not root.exists():
            continue
        for candidate in root.rglob(asin):
            if not candidate.is_dir() or candidate.name.upper() != asin:
                continue
            file_paths = {name: candidate / name for name in filenames}
            if not all(path.exists() and is_recent(path, cutoff) for path in file_paths.values()):
                continue
            mtimes = [path.stat().st_mtime for path in file_paths.values()]
            candidates.append(
                CacheHit(
                    asin=asin,
                    source=source,
                    path=candidate,
                    file_paths=file_paths,
                    newest_mtime=max(mtimes),
                    oldest_mtime=min(mtimes),
                )
            )
    if not candidates:
        return None
    return max(candidates, key=lambda item: (item.oldest_mtime, item.newest_mtime))


def passing_score(score_path: Path, threshold: int) -> bool:
    try:
        data = json.loads(score_path.read_text(encoding="utf-8-sig"))
    except Exception:
        return False
    title = data.get("title") if isinstance(data.get("title"), dict) else {}
    bullets = data.get("bullet_points") if isinstance(data.get("bullet_points"), dict) else {}
    title_score = int(title.get("score") or 0)
    bullet_score = int(bullets.get("score") or 0)
    return bool(data.get("passed")) and title_score >= threshold and bullet_score >= threshold


def sensitive_hits_for_score(
    score_path: Path,
    *,
    sensitive_excel: Path | None = None,
    sensitive_script: Path = DEFAULT_SENSITIVE_SCRIPT,
    country: str = "US",
) -> list[dict[str, Any]]:
    if sensitive_excel is None:
        return []
    if not sensitive_excel.exists():
        raise RuntimeError(f"Sensitive terms Excel not found: {sensitive_excel}")
    if not sensitive_script.exists():
        raise RuntimeError(f"Sensitive terms script not found: {sensitive_script}")
    module = load_sensitive_module(sensitive_script)
    terms = module.load_terms(sensitive_excel, country)
    score = load_json(score_path)
    title = score.get("title") if isinstance(score.get("title"), dict) else {}
    bullets = score.get("bullet_points") if isinstance(score.get("bullet_points"), dict) else {}
    fields = {
        "title": str(title.get("content") or ""),
        "bullet_points": "\n".join(str(item) for item in (bullets.get("content") or [])),
    }
    hits: list[dict[str, Any]] = []
    for field, text in fields.items():
        for hit in module.scan_text(text, terms):
            item = dict(hit)
            item["field"] = field
            hits.append(item)
    return hits


def load_sensitive_module(path: Path) -> Any:
    spec = importlib.util.spec_from_file_location("daily_sensitive_terms", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load sensitive terms module: {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def apply_sensitive_hits_to_score(score_data: dict[str, Any], hits: list[dict[str, Any]]) -> None:
    if not hits:
        return
    grouped: dict[str, list[str]] = {}
    for hit in hits:
        field = str(hit.get("field") or "")
        term = str(hit.get("term") or hit.get("word") or hit.get("keyword") or hit)
        grouped.setdefault(field, []).append(term)
    for field, terms in grouped.items():
        key = "title" if field == "title" else "bullet_points"
        target = score_data.get(key)
        if not isinstance(target, dict):
            continue
        target["pass"] = False
        deductions = target.setdefault("deductions", [])
        if isinstance(deductions, list):
            deductions.append("Sensitive terms hit: " + ", ".join(sorted(set(terms))))


def quality_gate_passes(
    score_path: Path,
    *,
    threshold: int,
    sensitive_excel: Path | None = None,
    sensitive_script: Path = DEFAULT_SENSITIVE_SCRIPT,
    country: str = "US",
) -> bool:
    if not passing_score(score_path, threshold):
        return False
    return not sensitive_hits_for_score(
        score_path,
        sensitive_excel=sensitive_excel,
        sensitive_script=sensitive_script,
        country=country,
    )


def find_recent_rufus_cache(
    *,
    roots: list[Path],
    asin: str,
    cache_days: int,
    threshold: int,
    sensitive_excel: Path | None = None,
    sensitive_script: Path = DEFAULT_SENSITIVE_SCRIPT,
    country: str = "US",
) -> CacheHit | None:
    cutoff = cutoff_time(cache_days)
    candidates: list[CacheHit] = []
    for root in roots:
        if not root.exists():
            continue
        for report in root.rglob("final-round.md"):
            if report.parent.name.upper() != asin:
                continue
            score = report.parent / "final-round-quality-score.json"
            if not score.exists():
                continue
            if not is_recent(report, cutoff) or not is_recent(score, cutoff):
                continue
            if not quality_gate_passes(
                score,
                threshold=threshold,
                sensitive_excel=sensitive_excel,
                sensitive_script=sensitive_script,
                country=country,
            ):
                continue
            mtimes = [report.stat().st_mtime, score.stat().st_mtime]
            candidates.append(
                CacheHit(
                    asin=asin,
                    source="rufus",
                    path=report.parent,
                    file_paths={FILE_RUFUS: report, "score": score},
                    newest_mtime=max(mtimes),
                    oldest_mtime=min(mtimes),
                )
            )
    if not candidates:
        return None
    return max(candidates, key=lambda item: (item.oldest_mtime, item.newest_mtime))


def run_command(
    command: list[str],
    *,
    cwd: Path,
    timeout: int,
    dry_run: bool,
) -> dict[str, Any]:
    started_at = datetime.now().isoformat(timespec="seconds")
    if dry_run:
        return {
            "command": command,
            "cwd": str(cwd),
            "started_at": started_at,
            "finished_at": started_at,
            "returncode": 0,
            "status": "planned",
            "stdout": "",
            "stderr": "",
        }
    try:
        proc = subprocess.run(
            command,
            cwd=str(cwd),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
        )
    except subprocess.TimeoutExpired as exc:
        stdout = exc.stdout if isinstance(exc.stdout, str) else (exc.stdout or b"").decode("utf-8", errors="replace")
        stderr = exc.stderr if isinstance(exc.stderr, str) else (exc.stderr or b"").decode("utf-8", errors="replace")
        return {
            "command": command,
            "cwd": str(cwd),
            "started_at": started_at,
            "finished_at": datetime.now().isoformat(timespec="seconds"),
            "returncode": -1,
            "status": "timeout",
            "stdout_tail": stdout[-4000:],
            "stderr_tail": stderr[-4000:],
            "error": f"TIMEOUT after {timeout}s",
        }
    return {
        "command": command,
        "cwd": str(cwd),
        "started_at": started_at,
        "finished_at": datetime.now().isoformat(timespec="seconds"),
        "returncode": proc.returncode,
        "status": "success" if proc.returncode == 0 else "failed",
        "stdout_tail": proc.stdout[-4000:],
        "stderr_tail": proc.stderr[-4000:],
    }


def collect_base_package_for_asin(args: argparse.Namespace, asin: str, run_root: Path) -> dict[str, Any]:
    collect_output_dir = run_root / "base-collect"
    collect_run_id = f"{args.run_id}-{asin}"
    command = [
        sys.executable,
        "-m",
        "opscli.cli",
        "asin-data",
        "collect",
        "--asin",
        asin,
        "--site",
        args.site,
        "--run-id",
        collect_run_id,
        "--output-dir",
        str(collect_output_dir),
        "--no-fetch-report-files",
    ]
    if not args.base_collect_full:
        command.extend(
            [
                "--skip-seller-sprite",
                "--skip-amazon",
                "--skip-rufus",
                "--skip-sales-query",
                "--skip-crawler-query",
            ]
        )
    if args.no_upload:
        command.append("--no-upload")
    if not args.no_submit_report_files and not args.no_upload:
        command.extend(["--submit-report-files", "--report-date", args.report_date])
    command.append("--pretty")
    result = run_command(command, cwd=REPO_ROOT, timeout=args.collect_timeout, dry_run=args.dry_run)
    write_json(run_root / "base-collect" / f"{asin}-command.json", result)
    if result["returncode"] != 0:
        raise RuntimeError(f"Base asin-data collect failed for {asin}. See {run_root / 'base-collect' / f'{asin}-command.json'}")
    result["asin"] = asin
    result["collect_run_id"] = collect_run_id
    result["base_package_dir"] = str(collect_output_dir / collect_run_id / PACKAGE_ROOT_NAME)
    return result


def collect_base_packages(args: argparse.Namespace, asins: list[str], run_root: Path) -> dict[str, Any]:
    results: list[dict[str, Any]] = []
    base_package_dirs: dict[str, str] = {}
    for asin in asins:
        result = collect_base_package_for_asin(args, asin, run_root)
        results.append(result)
        base_package_dirs[asin] = str(Path(result["base_package_dir"]))
    summary = {
        "mode": "single_asin",
        "count": len(results),
        "results": results,
        "base_package_dirs": base_package_dirs,
    }
    write_json(run_root / "base-collect-command.json", summary)
    return summary


def refresh_seller_sprite(args: argparse.Namespace, asins: list[str], run_root: Path) -> dict[str, Any] | None:
    if not asins:
        return None
    command = [
        sys.executable,
        str(args.seller_sprite_script),
        "--asins",
        *asins,
        "--output-dir",
        str(run_root / "seller-sprite-refresh"),
        "--site",
        args.site,
        "--period",
        args.seller_sprite_period,
        "--page-size",
        str(args.seller_sprite_page_size),
        "--continue-on-error",
    ]
    result = run_command(command, cwd=REPO_ROOT, timeout=args.seller_sprite_timeout, dry_run=args.dry_run)
    write_json(run_root / "seller-sprite-refresh-command.json", result)
    if result["returncode"] != 0:
        raise RuntimeError(f"SellerSprite refresh failed. See {run_root / 'seller-sprite-refresh-command.json'}")
    return result


def score_rufus_report(args: argparse.Namespace, report: Path, asin: str) -> dict[str, Any]:
    command = [
        sys.executable,
        str(args.rufus_score_script),
        "--report",
        str(report),
        "--asin",
        asin,
        "--threshold",
        str(args.rufus_threshold),
    ]
    result = run_command(command, cwd=REPO_ROOT, timeout=args.rufus_score_timeout, dry_run=args.dry_run)
    if result["returncode"] != 0:
        raise RuntimeError(f"Rufus score failed for {asin}. Report: {report}")
    return result


def score_path_for_report(report: Path) -> Path:
    return report.with_name(f"{report.stem}-quality-score.json")


def load_json(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    return data if isinstance(data, dict) else {}


def compact_prompt_text(value: Any, limit: int = 900) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def build_bullet_group_prompt(asin: str, score_data: dict[str, Any]) -> str:
    bullet_result = score_data.get("bullet_points") if isinstance(score_data.get("bullet_points"), dict) else {}
    deductions = "; ".join(str(item) for item in (bullet_result.get("deductions") or [])) or "clarity and compliance need improvement"
    current = bullet_result.get("content") if isinstance(bullet_result.get("content"), list) else []
    current_text = "\n".join(f"{index}. {item}" for index, item in enumerate(current[:5], start=1))
    if not current_text:
        current_text = "The previous answer did not return a complete optimized five-bullet set."
    return (
        f"Analyze ASIN {asin} bullet points again and provide a cleaner optimized five-bullet set.\n\n"
        f"Previous optimized bullets:\n{compact_prompt_text(current_text)}\n\n"
        f"Problems to fix: {compact_prompt_text(deductions, 650)}\n\n"
        "Requirements:\n"
        "1. Output exactly 5 bullets.\n"
        "2. Rank by buyer benefit, one core selling point per bullet.\n"
        "3. Start with the buyer benefit, then support it with visible product facts.\n"
        "4. Keep each optimized bullet under 180 English characters.\n"
        "5. Avoid repetition, keyword stuffing, exaggerated claims, review claims, refunds, promotions, unsupported quantities, and unverified features.\n"
        "6. Output only this Markdown table:\n\n"
        "| 五点序号 | 当前五点 | 存在的问题 | 优化建议 | 优化后五点内容 |"
    )


def report_path_from_stdout(stdout: str, cwd: Path) -> Path | None:
    patterns = [
        r'"report_path"\s*:\s*"([^"]+)"',
        r"Rufus answer report saved:\s*(.+?\.md)",
        r"Rufus .{0,80}[:\uff1a]\s*(.+?\.md)",
    ]
    for pattern in patterns:
        match = re.search(pattern, stdout, flags=re.IGNORECASE | re.DOTALL)
        if not match:
            continue
        raw = match.group(1).strip().strip('"')
        path = Path(raw)
        if not path.is_absolute():
            path = cwd / path
        return path
    return None


def normalize_question_section(text: str, qid: int) -> str:
    text = text.strip()
    if re.search(r"(?m)^##\s+", text):
        return re.sub(r"(?m)^##\s+.*$", f"## \u7b2c {qid} \u9898", text, count=1)
    return f"## \u7b2c {qid} \u9898\n\n{text}"


def run_rufus_followup_prompt(args: argparse.Namespace, asin: str, prompt: str, dst: Path, qid: int) -> dict[str, Any]:
    command = [
        args.opscli_bin,
        "amazon-rufus",
        "get-backend",
        asin,
        args.site,
        "--pretty",
        "--timeout",
        str(args.rufus_timeout),
        "--no-upload-payload",
        "-q",
        prompt,
    ]
    result = run_command(command, cwd=REPO_ROOT, timeout=args.rufus_subprocess_timeout, dry_run=args.dry_run)
    write_json(dst.with_suffix(".command.json"), result)
    if result["returncode"] != 0:
        raise RuntimeError(f"Rufus follow-up failed for {asin} Q{qid}. See {dst.with_suffix('.command.json')}")
    if args.dry_run:
        return result
    src = report_path_from_stdout(str(result.get("stdout_tail") or ""), REPO_ROOT)
    if src is None or not src.exists():
        raise RuntimeError(f"Rufus follow-up did not return a readable report path for {asin} Q{qid}.")
    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_text(normalize_question_section(src.read_text(encoding="utf-8", errors="replace"), qid), encoding="utf-8")
    return result


def first_reask_prompt(score_data: dict[str, Any], content_type: str) -> str:
    for item in score_data.get("reask_plan") or []:
        if isinstance(item, dict) and item.get("content_type") == content_type and item.get("prompt"):
            return str(item["prompt"])
    return ""


def run_rufus_followups_for_score(args: argparse.Namespace, asin_dir: Path, asin: str, score_data: dict[str, Any], round_no: int) -> None:
    title = score_data.get("title") if isinstance(score_data.get("title"), dict) else {}
    bullets = score_data.get("bullet_points") if isinstance(score_data.get("bullet_points"), dict) else {}
    if not title.get("pass"):
        prompt = first_reask_prompt(score_data, "title")
        if not prompt:
            prompt = (
                f"Re-optimize the title for ASIN {asin}. Keep it natural, buyer-search oriented, compliant, and based only on visible product facts. "
                "Output only this table: | 当前标题内容 | 存在的问题 | 优化建议 | 优化后标题内容 |"
            )
        run_rufus_followup_prompt(args, asin, prompt, asin_dir / f"followup-r{round_no}-title.md", 1)
    if not bullets.get("pass"):
        run_rufus_followup_prompt(args, asin, build_bullet_group_prompt(asin, score_data), asin_dir / f"followup-r{round_no}-bullets.md", 2)


def newest_existing(paths: Iterable[Path]) -> Path | None:
    existing = [path for path in paths if path.exists()]
    if not existing:
        return None
    return max(existing, key=lambda path: path.stat().st_mtime)


def assemble_rufus_round(asin_dir: Path, round_no: int) -> Path:
    parts: list[str] = []
    title = newest_existing(asin_dir / f"followup-r{idx}-title.md" for idx in range(1, round_no + 1)) or asin_dir / "initial-q01.md"
    bullets = newest_existing(asin_dir / f"followup-r{idx}-bullets.md" for idx in range(1, round_no + 1)) or asin_dir / "initial-q02.md"
    sources = [title, bullets, *(asin_dir / f"initial-q{qid:02d}.md" for qid in range(3, 8))]
    for qid, source in enumerate(sources, start=1):
        if not source.exists():
            raise RuntimeError(f"Missing Rufus source section: {source}")
        parts.append(normalize_question_section(source.read_text(encoding="utf-8", errors="replace"), qid))
    report = asin_dir / f"round-{round_no}.md"
    report.write_text("\n\n".join(parts).rstrip() + "\n", encoding="utf-8")
    return report


def finalize_rufus_refresh(args: argparse.Namespace, refresh_dir: Path, asin: str) -> None:
    asin_dir = refresh_dir / asin
    current_report = asin_dir / "initial-round.md"
    if not current_report.exists():
        raise RuntimeError(f"Rufus initial report missing for {asin}: {current_report}")
    for round_no in range(0, args.rufus_max_followups + 1):
        score_rufus_report(args, current_report, asin)
        score_path = score_path_for_report(current_report)
        sensitive_hits = sensitive_hits_for_score(
            score_path,
            sensitive_excel=args.sensitive_excel,
            sensitive_script=args.sensitive_script,
            country=args.site,
        )
        if passing_score(score_path, args.rufus_threshold) and not sensitive_hits:
            shutil.copy2(current_report, asin_dir / "final-round.md")
            shutil.copy2(score_path, asin_dir / "final-round-quality-score.json")
            return
        if round_no >= args.rufus_max_followups:
            raise RuntimeError(f"Rufus score did not pass after {args.rufus_max_followups} follow-up rounds for {asin}: {score_path}")
        score_data = load_json(score_path)
        apply_sensitive_hits_to_score(score_data, sensitive_hits)
        run_rufus_followups_for_score(args, asin_dir, asin, score_data, round_no + 1)
        current_report = assemble_rufus_round(asin_dir, round_no + 1)


def refresh_rufus(args: argparse.Namespace, asins: list[str], run_root: Path) -> dict[str, Any] | None:
    if not asins:
        return None
    if args.rufus_refresh_command:
        command = render_refresh_command(args.rufus_refresh_command, asins=asins, output_dir=run_root / "rufus-refresh", site=args.site)
    else:
        command = [
            sys.executable,
            str(args.rufus_batch_script),
            "--asins",
            ",".join(asins),
            "--country",
            args.site,
            "--asin-workers",
            str(args.rufus_asin_workers),
            "--question-workers",
            str(args.rufus_question_workers),
            "--retry-question-workers",
            str(args.rufus_retry_question_workers),
            "--timeout",
            str(args.rufus_timeout),
            "--retry-timeout",
            str(args.rufus_retry_timeout),
            "--subprocess-timeout",
            str(args.rufus_subprocess_timeout),
            "--retry-subprocess-timeout",
            str(args.rufus_retry_subprocess_timeout),
            "--retries",
            str(args.rufus_retries),
            "--output-dir",
            str(run_root / "rufus-refresh"),
        ]
    result = run_command(command, cwd=REPO_ROOT, timeout=args.rufus_batch_timeout, dry_run=args.dry_run)
    write_json(run_root / "rufus-refresh-command.json", result)
    if result["returncode"] != 0:
        raise RuntimeError(f"Rufus refresh failed. See {run_root / 'rufus-refresh-command.json'}")

    if not args.dry_run and not args.rufus_refresh_command:
        refresh_dir = run_root / "rufus-refresh"
        for asin in asins:
            finalize_rufus_refresh(args, refresh_dir, asin)
    return result


def render_refresh_command(template: str, *, asins: list[str], output_dir: Path, site: str) -> list[str]:
    rendered = (
        template.replace("{asins_csv}", ",".join(asins))
        .replace("{asins_space}", " ".join(asins))
        .replace("{output_dir}", str(output_dir))
        .replace("{site}", site)
    )
    return split_command_line(rendered)


def split_command_line(command: str) -> list[str]:
    import shlex

    if sys.platform.startswith("win"):
        return shlex.split(command, posix=False)
    return shlex.split(command)


def find_base_file(base_package_dir: Path, asin: str, target_name: str, prefix: str) -> Path | None:
    asin_dir = base_package_dir / asin
    candidate = asin_dir / target_name
    if candidate.exists():
        return candidate
    if asin_dir.exists():
        matches = sorted(path for path in asin_dir.glob(f"{prefix}*") if path.is_file())
        if matches:
            return matches[0]
    return None


def ensure_placeholder_xlsx(path: Path, asin: str, label: str) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required to create placeholder XLSX files")
    wb = Workbook()
    ws = wb.active
    ws.title = "No data"
    ws.append(["ASIN", "File", "Status"])
    ws.append([asin, label, "No source file found during daily package assembly"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def copy_or_placeholder(
    source: Path | None,
    target: Path,
    asin: str,
    label: str,
    *,
    allow_placeholder: bool,
) -> dict[str, Any]:
    target.parent.mkdir(parents=True, exist_ok=True)
    if source and source.exists():
        shutil.copy2(source, target)
        return {"status": "copied", "source": str(source), "target": str(target)}
    if not allow_placeholder:
        raise FileNotFoundError(
            f"Required base package file is missing for {asin}: {label}. "
            "Run asin-data collect without --dry-run, or pass --allow-missing-base-placeholders only for local layout tests."
        )
    ensure_placeholder_xlsx(target, asin, label)
    return {"status": "placeholder", "source": None, "target": str(target)}


def build_readme(run_id: str, asins: list[str], generated_at: str) -> str:
    lines = [
        "# ASIN Data Packages",
        "",
        f"- Run ID: {run_id}",
        f"- Generated At: {generated_at}",
        "",
        "Each ASIN folder contains:",
        "",
    ]
    lines.extend(f"- `{name}`" for name in OUTPUT_FILES)
    lines.extend(["", "## ASINs", ""])
    lines.extend(f"- {asin}" for asin in asins)
    lines.append("")
    return "\n".join(lines)


def write_package_zip(zip_path: Path, *, staging_root: Path, package_dir: Path, asin: str | None = None) -> None:
    if zip_path.exists():
        zip_path.unlink()
    if asin:
        roots = [package_dir / asin]
    else:
        roots = [package_dir]
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for root in roots:
            if root.is_file():
                archive.write(root, root.relative_to(staging_root).as_posix())
                continue
            for path in sorted(root.rglob("*")):
                if path.is_file():
                    archive.write(path, path.relative_to(staging_root).as_posix())


def build_final_package(
    *,
    run_root: Path,
    run_id: str,
    asins: list[str],
    base_package_dir: Path | None = None,
    base_package_dirs: dict[str, Path] | None = None,
    seller_hits: dict[str, CacheHit],
    rufus_hits: dict[str, CacheHit],
    allow_base_placeholders: bool = False,
) -> dict[str, Any]:
    staging_root = run_root / "_daily-full-package"
    package_dir = staging_root / PACKAGE_ROOT_NAME
    if staging_root.exists():
        resolved = staging_root.resolve()
        if run_root.resolve() not in resolved.parents:
            raise RuntimeError(f"Refusing to remove package staging dir outside run root: {staging_root}")
        shutil.rmtree(staging_root)
    package_dir.mkdir(parents=True, exist_ok=True)

    items: list[dict[str, Any]] = []
    for asin in asins:
        asin_dir = package_dir / asin
        asin_dir.mkdir(parents=True, exist_ok=True)
        file_records: dict[str, Any] = {}
        source_base_package_dir = (base_package_dirs or {}).get(asin) or base_package_dir
        if source_base_package_dir is None:
            raise RuntimeError(f"Base package directory is missing for {asin}")
        file_records[FILE_BASIC] = copy_or_placeholder(
            find_base_file(source_base_package_dir, asin, FILE_BASIC, "01-"),
            asin_dir / FILE_BASIC,
            asin,
            FILE_BASIC,
            allow_placeholder=allow_base_placeholders,
        )
        file_records[FILE_BI] = copy_or_placeholder(
            find_base_file(source_base_package_dir, asin, FILE_BI, "02-"),
            asin_dir / FILE_BI,
            asin,
            FILE_BI,
            allow_placeholder=allow_base_placeholders,
        )
        seller_hit = seller_hits[asin]
        for name in SELLER_FILES:
            target = asin_dir / name
            shutil.copy2(seller_hit.file_paths[name], target)
            file_records[name] = {"status": "copied", "source": str(seller_hit.file_paths[name]), "target": str(target)}
        rufus_hit = rufus_hits[asin]
        target_rufus = asin_dir / FILE_RUFUS
        shutil.copy2(rufus_hit.file_paths[FILE_RUFUS], target_rufus)
        file_records[FILE_RUFUS] = {"status": "copied", "source": str(rufus_hit.file_paths[FILE_RUFUS]), "target": str(target_rufus)}
        items.append({"asin": asin, "dir": str(asin_dir), "files": file_records})

    generated_at = datetime.now().isoformat(timespec="seconds")
    (package_dir / "README.md").write_text(build_readme(run_id, asins, generated_at), encoding="utf-8")
    zip_path = run_root / "asin-data-packages.zip"
    write_package_zip(zip_path, staging_root=staging_root, package_dir=package_dir)
    asin_zips: dict[str, dict[str, Any]] = {}
    for asin in asins:
        asin_zip_path = run_root / f"{asin}-asin-data-package.zip"
        write_package_zip(asin_zip_path, staging_root=staging_root, package_dir=package_dir, asin=asin)
        asin_zips[asin] = {
            "asin": asin,
            "zip_path": str(asin_zip_path),
            "size_bytes": asin_zip_path.stat().st_size,
        }
    return {
        "package_dir": str(package_dir),
        "staging_root": str(staging_root),
        "zip_path": str(zip_path),
        "asin_zips": asin_zips,
        "items": items,
        "generated_at": generated_at,
    }


def upload_zip(zip_path: Path, *, run_id: str, asins: list[str], summary: dict[str, Any], dry_run: bool) -> dict[str, Any] | None:
    if dry_run:
        return {"status": "planned", "path": str(zip_path)}
    from opscli.shared.file_uploads import FileUploadClient

    upload = FileUploadClient().upload(
        zip_path,
        purpose="asin_data_split_package_zip",
        folder="asin-data",
        public="1",
        metadata={
            "run_id": run_id,
            "asin_count": len(asins),
            "asins": asins,
            "report_type": "asin_data_split_package_zip",
            "report_filename": zip_path.name,
            "package_filename": zip_path.name,
            "source_filename": zip_path.name,
            "upload_filename": zip_path.name,
            "summary": summary.get("summary"),
        },
    )
    path = zip_path.as_posix()
    return {
        "url": upload.url,
        "path": path,
        "upload_path": path,
        "purpose": "asin_data_split_package_zip",
        "folder": "asin-data",
        "raw": upload.raw,
    }


def upload_package_zips(args: argparse.Namespace, package: dict[str, Any], asins: list[str], manifest: dict[str, Any]) -> dict[str, Any]:
    uploads: dict[str, Any] = {"by_asin": {}, "combined": None}
    for asin in asins:
        asin_zip = package.get("asin_zips", {}).get(asin)
        if not isinstance(asin_zip, dict):
            raise RuntimeError(f"Missing ASIN zip for upload: {asin}")
        uploads["by_asin"][asin] = upload_zip(
            Path(str(asin_zip["zip_path"])),
            run_id=args.run_id,
            asins=[asin],
            summary=manifest,
            dry_run=args.dry_run,
        )
    if args.upload_combined_zip:
        uploads["combined"] = upload_zip(
            Path(str(package["zip_path"])),
            run_id=args.run_id,
            asins=asins,
            summary=manifest,
            dry_run=args.dry_run,
        )
    return uploads


def submit_final_report_files(
    args: argparse.Namespace,
    *,
    package: dict[str, Any],
    uploads: dict[str, Any],
    asins: list[str],
) -> dict[str, Any] | None:
    if args.no_submit_report_files or args.no_upload:
        return None
    items = build_final_report_file_items(args, package=package, uploads=uploads, asins=asins)
    if args.dry_run:
        return {
            "submitted": False,
            "reason": "dry_run",
            "endpoint": args.register_endpoint,
            "items": items,
        }
    from opscli.asin_data.services.report_files import AsinReportFileClient

    client = AsinReportFileClient(endpoint=args.register_endpoint) if args.register_endpoint else AsinReportFileClient()
    response = client.upsert(
        items=items,
        request_id=args.run_id,
        source=args.report_source,
        idempotency_key=args.run_id,
    )
    return {
        "submitted": True,
        "endpoint": client.endpoint,
        "item_count": len(items),
        "response": response,
    }


def build_final_report_file_items(
    args: argparse.Namespace,
    *,
    package: dict[str, Any],
    uploads: dict[str, Any],
    asins: list[str],
) -> list[dict[str, Any]]:
    by_asin = uploads.get("by_asin") if isinstance(uploads.get("by_asin"), dict) else {}
    items: list[dict[str, Any]] = []
    for asin in asins:
        zip_info = package.get("asin_zips", {}).get(asin)
        if not isinstance(zip_info, dict):
            raise RuntimeError(f"Missing ASIN zip for report submit: {asin}")
        zip_path = Path(str(zip_info["zip_path"]))
        upload = by_asin.get(asin)
        file_url = upload.get("url") if isinstance(upload, dict) else None
        file_bytes = zip_path.read_bytes() if zip_path.exists() and not args.dry_run else b""
        content_hash = hashlib.sha256(file_bytes).hexdigest() if file_bytes else None
        items.append(
            {
                "report_uuid": str(uuid5(NAMESPACE_URL, f"{asin}:{args.site}:{args.report_type}:{args.report_date}")),
                "run_id": args.run_id,
                "asin": asin,
                "site": args.site,
                "report_type": args.report_type,
                "source": args.report_source,
                "status": "success",
                "report_date": args.report_date,
                "file_name": zip_path.name,
                "file_ext": "zip",
                "mime_type": "application/zip",
                "file_path": zip_path.as_posix(),
                "file_url": file_url,
                "content_hash": content_hash,
                "file_size_bytes": zip_path.stat().st_size if zip_path.exists() else None,
                "meta_json": {
                    "run_id": args.run_id,
                    "source": "asin_data_daily_full_package",
                    "package_root": PACKAGE_ROOT_NAME,
                    "package_files": OUTPUT_FILES,
                    "combined_zip_path": package.get("zip_path"),
                },
            }
        )
    return items


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    tmp.replace(path)


def parse_path_list(values: list[str]) -> list[Path]:
    roots: list[Path] = []
    for value in values:
        for item in re.split(r"[;,]", value):
            item = item.strip().strip('"')
            if item:
                roots.append(Path(item).expanduser())
    return roots


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Daily ASIN package collector with SellerSprite/Rufus 7-day cache reuse.")
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--site", default="US")
    parser.add_argument("--run-id", default="")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--asin-column", default="asin")
    parser.add_argument("--report-date", default="", help="Report date used by collect --submit-report-files and final package submit. Defaults to Asia/Shanghai today.")
    parser.add_argument("--cache-days", type=int, default=7)
    parser.add_argument("--seller-cache-root", action="append", default=[])
    parser.add_argument("--rufus-cache-root", action="append", default=[])
    parser.add_argument("--seller-sprite-script", type=Path, default=DEFAULT_SELLER_SKILL_SCRIPT)
    parser.add_argument("--seller-sprite-period", default="30d")
    parser.add_argument("--seller-sprite-page-size", type=int, default=100)
    parser.add_argument("--opscli-bin", default="opscli")
    parser.add_argument("--rufus-batch-script", type=Path, default=DEFAULT_RUFUS_BATCH_SCRIPT)
    parser.add_argument("--rufus-score-script", type=Path, default=DEFAULT_RUFUS_SCORE_SCRIPT)
    parser.add_argument("--rufus-refresh-command", default="", help="Optional full score/re-ask command. Supports {asins_csv}, {asins_space}, {output_dir}, {site}.")
    parser.add_argument("--rufus-threshold", type=int, default=80)
    parser.add_argument("--rufus-max-followups", type=int, default=3)
    parser.add_argument("--sensitive-excel", type=Path, default=None, help="Optional sensitive terms Excel. When set, Rufus final title/bullets must have zero hits.")
    parser.add_argument("--sensitive-script", type=Path, default=DEFAULT_SENSITIVE_SCRIPT)
    parser.add_argument("--rufus-asin-workers", type=int, default=2)
    parser.add_argument("--rufus-question-workers", type=int, default=2)
    parser.add_argument("--rufus-retry-question-workers", type=int, default=1)
    parser.add_argument("--rufus-timeout", type=int, default=600)
    parser.add_argument("--rufus-retry-timeout", type=int, default=720)
    parser.add_argument("--rufus-subprocess-timeout", type=int, default=720)
    parser.add_argument("--rufus-retry-subprocess-timeout", type=int, default=840)
    parser.add_argument("--rufus-retries", type=int, default=2)
    parser.add_argument("--collect-timeout", type=int, default=3600)
    parser.add_argument("--seller-sprite-timeout", type=int, default=21600)
    parser.add_argument("--rufus-batch-timeout", type=int, default=21600)
    parser.add_argument("--rufus-score-timeout", type=int, default=600)
    parser.add_argument(
        "--base-collect-full",
        action="store_true",
        help="Run per-ASIN base collect without skip flags. Default skips SellerSprite/Amazon/Rufus because this wrapper refreshes them separately.",
    )
    parser.add_argument(
        "--allow-missing-base-placeholders",
        action="store_true",
        help="Only for local layout tests. By default, missing 01/02 base files fail the run.",
    )
    parser.add_argument("--report-type", default="asin_data_split_package_zip")
    parser.add_argument("--report-source", default="asin_data_daily_full_package")
    parser.add_argument("--register-endpoint", default=None)
    parser.add_argument("--no-submit-report-files", action="store_true")
    parser.add_argument("--upload-combined-zip", action="store_true")
    parser.add_argument("--no-upload", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    args.input = args.input.expanduser().resolve()
    args.output_dir = args.output_dir.expanduser()
    if args.sensitive_excel is not None:
        args.sensitive_excel = args.sensitive_excel.expanduser().resolve()
    args.sensitive_script = args.sensitive_script.expanduser().resolve()
    args.run_id = args.run_id or now_run_id()
    args.report_date = args.report_date or default_report_date()
    run_root = args.output_dir / args.run_id
    run_root.mkdir(parents=True, exist_ok=True)

    asins = load_asins(args.input, args.asin_column)
    if not asins:
        raise SystemExit("No ASINs found in input.")

    seller_roots = parse_path_list(args.seller_cache_root) or [args.output_dir, run_root / "seller-sprite-refresh"]
    rufus_roots = parse_path_list(args.rufus_cache_root) or [args.output_dir, Path.cwd() / "outputs", run_root / "rufus-refresh"]
    manifest: dict[str, Any] = {
        "run_id": args.run_id,
        "site": args.site,
        "input": str(args.input),
        "output_dir": str(args.output_dir),
        "run_root": str(run_root),
        "cache_days": args.cache_days,
        "asins": asins,
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "seller_cache_roots": [str(path) for path in seller_roots],
        "rufus_cache_roots": [str(path) for path in rufus_roots],
        "sensitive_excel": str(args.sensitive_excel) if args.sensitive_excel else None,
        "report_date": args.report_date,
        "steps": [],
    }
    write_json(run_root / "daily-full-package-manifest.json", manifest)

    collect_result = collect_base_packages(args, asins, run_root)
    manifest["steps"].append({"name": "base_collect", "result": collect_result})

    seller_hits = {
        asin: hit
        for asin in asins
        if (hit := find_recent_file_set(roots=seller_roots, asin=asin, filenames=SELLER_FILES, cache_days=args.cache_days, source="seller_sprite"))
    }
    seller_missing = [asin for asin in asins if asin not in seller_hits]
    if seller_missing:
        refresh_seller_sprite(args, seller_missing, run_root)
        refresh_roots = [run_root / "seller-sprite-refresh", *seller_roots]
        for asin in seller_missing:
            hit = find_recent_file_set(roots=refresh_roots, asin=asin, filenames=SELLER_FILES, cache_days=args.cache_days, source="seller_sprite")
            if hit:
                seller_hits[asin] = hit
    still_missing_seller = [asin for asin in asins if asin not in seller_hits]
    if still_missing_seller:
        raise RuntimeError(f"SellerSprite files still missing for ASINs: {', '.join(still_missing_seller)}")

    rufus_hits = {
        asin: hit
        for asin in asins
        if (
            hit := find_recent_rufus_cache(
                roots=rufus_roots,
                asin=asin,
                cache_days=args.cache_days,
                threshold=args.rufus_threshold,
                sensitive_excel=args.sensitive_excel,
                sensitive_script=args.sensitive_script,
                country=args.site,
            )
        )
    }
    rufus_missing = [asin for asin in asins if asin not in rufus_hits]
    if rufus_missing:
        refresh_rufus(args, rufus_missing, run_root)
        refresh_roots = [run_root / "rufus-refresh", *rufus_roots]
        for asin in rufus_missing:
            hit = find_recent_rufus_cache(
                roots=refresh_roots,
                asin=asin,
                cache_days=args.cache_days,
                threshold=args.rufus_threshold,
                sensitive_excel=args.sensitive_excel,
                sensitive_script=args.sensitive_script,
                country=args.site,
            )
            if hit:
                rufus_hits[asin] = hit
    still_missing_rufus = [asin for asin in asins if asin not in rufus_hits]
    if still_missing_rufus:
        raise RuntimeError(f"Rufus scored final files still missing for ASINs: {', '.join(still_missing_rufus)}")

    base_package_dirs = {asin: Path(path) for asin, path in collect_result["base_package_dirs"].items()}
    package = build_final_package(
        run_root=run_root,
        run_id=args.run_id,
        asins=asins,
        base_package_dirs=base_package_dirs,
        seller_hits=seller_hits,
        rufus_hits=rufus_hits,
        allow_base_placeholders=args.allow_missing_base_placeholders,
    )
    manifest["package"] = package
    manifest["cache"] = {
        "seller_sprite": {asin: {"path": str(hit.path), "oldest_mtime": hit.oldest_mtime} for asin, hit in seller_hits.items()},
        "rufus": {asin: {"path": str(hit.path), "oldest_mtime": hit.oldest_mtime} for asin, hit in rufus_hits.items()},
        "seller_sprite_refreshed_asins": seller_missing,
        "rufus_refreshed_asins": rufus_missing,
    }

    upload = None
    report_file_submit = None
    if not args.no_upload:
        upload = upload_package_zips(args, package, asins, manifest)
        manifest["upload"] = upload
        manifest["aliyun_urls"] = {
            asin: item.get("url")
            for asin, item in (upload.get("by_asin") or {}).items()
            if isinstance(item, dict) and item.get("url")
        }
        report_file_submit = submit_final_report_files(args, package=package, uploads=upload, asins=asins)
        if report_file_submit is not None:
            manifest["report_file_submit"] = report_file_submit
    manifest["finished_at"] = datetime.now().isoformat(timespec="seconds")
    manifest["status"] = "success"
    write_json(run_root / "daily-full-package-manifest.json", manifest)
    print(
        json.dumps(
            {
                "success": True,
                "manifest": str(run_root / "daily-full-package-manifest.json"),
                "zip_path": package["zip_path"],
                "asin_zips": package["asin_zips"],
                "upload": upload,
                "report_file_submit": report_file_submit,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
